#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
局域网文件共享服务器。
支持在局域网内上传、下载、预览和协作编辑文件。
"""

from flask import Flask, request, render_template_string, send_file, redirect, url_for, flash, session, jsonify, after_this_request, stream_with_context
from flask.wrappers import Request as FlaskRequest
from werkzeug.exceptions import RequestEntityTooLarge
import os
import logging
from werkzeug.utils import secure_filename as werkzeug_secure_filename
import socket
from datetime import datetime
import zipfile
import io
import shutil
import re
import tempfile
import json
import hashlib
import hmac
import uuid
import mimetypes
import difflib
import queue
from werkzeug.serving import make_server
from werkzeug.wrappers import Response
import configparser
import sys
from docx import Document as WordDocument
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table as DocxTable
from docx.text.paragraph import Paragraph as DocxParagraph
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import ExcelWriter

# 自定义Request类，绕过大小限制
class UnrestrictedRequest(FlaskRequest):
    """自定义Request类，完全绕过文件大小限制"""
    @property
    def max_content_length(self):
        """动态返回 None，完全不限制。"""
        return None
    
    @property  
    def max_form_memory_size(self):
        """动态返回 None，完全不限制。"""
        return None

def secure_filename(filename):
    """
    安全化文件名，支持中文等Unicode字符
    - 移除路径分隔符和危险字符
    - 保留中文、英文、数字、下划线、连字符和点号
    """
    filename = filename.replace('\\', '').replace('/', '')
    filename = re.sub(r'[<>:"|?*\x00-\x1f]', '', filename)
    # 去除首尾空格
    filename = filename.strip()
    # 如果文件名为空或只有扩展名，使用默认名称
    if not filename or filename.startswith('.'):
        filename = 'unnamed_file' + filename
    return filename

# ==================== 读取配置文件 ====================
def load_config():
    """加载配置文件"""
    config = configparser.ConfigParser()
    
    # 获取程序所在目录
    if getattr(sys, 'frozen', False):
        # 如果是打包后的exe
        base_path = os.path.dirname(sys.executable)
    else:
        # 如果是Python脚本
        base_path = os.path.dirname(os.path.abspath(__file__))
    
    config_file = os.path.join(base_path, 'config.ini')
    
    # 如果配置文件不存在，创建默认配置
    if not os.path.exists(config_file):
        print(f"⚠️  配置文件不存在，创建默认配置: {config_file}")
        create_default_config(config_file)
    
    config.read(config_file, encoding='utf-8')
    return config, base_path

def create_default_config(config_file):
    """创建默认配置文件"""
    config = configparser.ConfigParser()
    
    config['SERVER'] = {
        'PORT': '5000',
        'ALLOW_LAN': 'True'
    }
    
    config['PATHS'] = {
        'UPLOAD_FOLDER': 'shared_files',
        'TEMP_FOLDER': 'temp_uploads',
        'STATIC_FOLDER': 'static'
    }
    
    config['FILES'] = {
        'MAX_FILE_SIZE_MB': '100',
        'MAX_FOLDER_FILES': '1000',
        'AUTO_EXTRACT_ZIP': 'False',
        'ALLOWED_EXTENSIONS': 'txt,pdf,png,jpg,jpeg,gif,bmp,webp,svg,ico,doc,docx,xls,xlsx,ppt,pptx,zip,rar,7z,tar,gz,py,java,cpp,c,h,js,ts,html,css,json,jsonl,xml,yaml,yml,md,csv,log,sql,sh,bat,mp4,avi,mkv,mov,wmv,flv,webm,mp3,wav,flac,aac,ogg,m4a,exe,apk,iso,dmg,deb,rpm,msi,pkg'
    }
    
    config['ADVANCED'] = {
        'CHUNK_SIZE_MB': '10',
        'TASKS_FILE': 'tasks.json',
        'REGISTRATIONS_FILE': 'user_registrations.json'
    }
    
    config['DISPLAY'] = {
        'HIDE_SYSTEM_FOLDERS': 'True',
        'HIDDEN_ITEMS': '$RECYCLE.BIN,$Recycle.Bin,System Volume Information,RECYCLER,Config.Msi,Recovery,ProgramData,hiberfil.sys,pagefile.sys,swapfile.sys,DumpStack.log.tmp,System32,Windows,Program Files,Program Files (x86),PerfLogs',
        'AUTO_REFRESH': 'True',
        'REFRESH_INTERVAL': '5'
    }
    
    with open(config_file, 'w', encoding='utf-8') as f:
        config.write(f)
    
    print(f"已创建默认配置文件: {config_file}")

# 加载配置
config, BASE_PATH = load_config()

# 读取配置项
SERVER_PORT = config.getint('SERVER', 'PORT', fallback=5000)
ALLOW_LAN = config.getboolean('SERVER', 'ALLOW_LAN', fallback=True)

UPLOAD_FOLDER_NAME = config.get('PATHS', 'UPLOAD_FOLDER', fallback='shared_files')
TEMP_FOLDER_NAME = config.get('PATHS', 'TEMP_FOLDER', fallback='temp_uploads')
STATIC_FOLDER_NAME = config.get('PATHS', 'STATIC_FOLDER', fallback='static')

# 缓存绝对路径，避免重复计算
UPLOAD_FOLDER_ABS = None  # 将在文件夹创建后初始化
# 处理相对路径
if not os.path.isabs(UPLOAD_FOLDER_NAME):
    UPLOAD_FOLDER = os.path.join(BASE_PATH, UPLOAD_FOLDER_NAME)
else:
    UPLOAD_FOLDER = UPLOAD_FOLDER_NAME

if not os.path.isabs(TEMP_FOLDER_NAME):
    TEMP_FOLDER = os.path.join(BASE_PATH, TEMP_FOLDER_NAME)
else:
    TEMP_FOLDER = TEMP_FOLDER_NAME

if not os.path.isabs(STATIC_FOLDER_NAME):
    STATIC_FOLDER = os.path.join(BASE_PATH, STATIC_FOLDER_NAME)
else:
    STATIC_FOLDER = STATIC_FOLDER_NAME

MAX_FILE_SIZE = config.getint('FILES', 'MAX_FILE_SIZE_MB', fallback=100) * 1024 * 1024
MAX_FOLDER_FILES = config.getint('FILES', 'MAX_FOLDER_FILES', fallback=1000)
AUTO_EXTRACT_ZIP = config.getboolean('FILES', 'AUTO_EXTRACT_ZIP', fallback=False)

# 解析允许的文件扩展名
extensions_str = config.get('FILES', 'ALLOWED_EXTENSIONS', fallback='')
if extensions_str.strip():
    ALLOWED_EXTENSIONS = set(ext.strip() for ext in extensions_str.split(','))
else:
    ALLOWED_EXTENSIONS = set()  # 空集合表示允许所有文件
CHUNK_SIZE_MB = config.getint('ADVANCED', 'CHUNK_SIZE_MB', fallback=10)
TASKS_FILE = config.get('ADVANCED', 'TASKS_FILE', fallback='tasks.json')
REGISTRATIONS_FILE = config.get('ADVANCED', 'REGISTRATIONS_FILE', fallback='user_registrations.json')

# 显示配置
HIDE_SYSTEM_FOLDERS = config.getboolean('DISPLAY', 'HIDE_SYSTEM_FOLDERS', fallback=True)
HIDDEN_ITEMS_STR = config.get('DISPLAY', 'HIDDEN_ITEMS', 
                              fallback='$RECYCLE.BIN,$Recycle.Bin,System Volume Information,RECYCLER,Config.Msi,Recovery,ProgramData,hiberfil.sys,pagefile.sys,swapfile.sys,DumpStack.log.tmp,System32,Windows,Program Files,Program Files (x86),PerfLogs')
HIDDEN_ITEMS = set(item.strip().lower() for item in HIDDEN_ITEMS_STR.split(',') if item.strip())
HIDDEN_ITEMS.update({'__macosx', '.ds_store'})
AUTO_REFRESH = config.getboolean('DISPLAY', 'AUTO_REFRESH', fallback=True)
REFRESH_INTERVAL = config.getint('DISPLAY', 'REFRESH_INTERVAL', fallback=5)

if not os.path.isabs(TASKS_FILE):
    TASKS_FILE = os.path.join(BASE_PATH, TASKS_FILE)
if not os.path.isabs(REGISTRATIONS_FILE):
    REGISTRATIONS_FILE = os.path.join(BASE_PATH, REGISTRATIONS_FILE)

# 打印配置信息
print("=" * 50)
print("📋 配置信息:")
print(f"   端口: {SERVER_PORT}")
print(f"   允许局域网访问: {ALLOW_LAN}")
print(f"   共享目录: {UPLOAD_FOLDER}")
print(f"   临时文件目录: {TEMP_FOLDER}")
print(f"   静态资源目录: {STATIC_FOLDER}")
print(f"   最大文件大小: {MAX_FILE_SIZE // (1024*1024)}MB")
print(f"   分块大小: {CHUNK_SIZE_MB}MB")
print(f"   隐藏系统文件: {'是' if HIDE_SYSTEM_FOLDERS else '否'}")
if HIDE_SYSTEM_FOLDERS:
    print(f"   hidden_items: {len(HIDDEN_ITEMS)}")
print("=" * 50)

ONLINE_SESSION_COOKIE_NAME = 'lanfs_client_id'
FLASK_SESSION_COOKIE_NAME = 'lanfs_session'
SECRET_KEY_FILE = os.path.join(BASE_PATH, '.flask_secret_key')
ONLINE_PRESENCE_TTL_SECONDS = 45
REALTIME_STREAM_INTERVAL_SECONDS = 2.0
PASSWORD_MIN_LENGTH = 6
PASSWORD_HASH_ITERATIONS = 200000
ACCOUNT_KEY_PREFIX = 'user:'
NICKNAME_MAX_LENGTH = 4

def load_or_create_secret_key():
    """Docstring."""
    try:
        if os.path.exists(SECRET_KEY_FILE):
            with open(SECRET_KEY_FILE, 'r', encoding='utf-8') as f:
                secret = f.read().strip()
                if secret:
                    return secret
    except Exception as e:
        print(f"⚠️  读取会话密钥失败，将重新生成: {e}")

    secret = os.urandom(32).hex()
    try:
        with open(SECRET_KEY_FILE, 'w', encoding='utf-8') as f:
            f.write(secret)
    except Exception as e:
        print(f"⚠️  保存会话密钥失败，本次使用临时密钥: {e}")
    return secret

app = Flask(__name__, static_folder=STATIC_FOLDER, static_url_path='/static')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['SESSION_COOKIE_NAME'] = FLASK_SESSION_COOKIE_NAME
# 不设置 MAX_CONTENT_LENGTH，避免大文件上传被 Flask 默认限制
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0
app.secret_key = load_or_create_secret_key()
app.request_class = UnrestrictedRequest  # 使用自定义 Request 类完全绕过大小限制
# 在线用户管理
import threading
import time
from datetime import datetime

online_users = {}  # {session_id: {'username': str, 'ip': str}}
user_lock = threading.Lock()
current_activities = []  # [{'username': str, 'action': str, 'filename': str, 'timestamp': float}]
activity_lock = threading.Lock()

# 注册信息存储（使用配置文件中的路径）
registration_lock = threading.Lock()

# 管理员系统（存储在 user_registrations.json 中）
admin_lock = threading.Lock()
admin_requests_lock = threading.Lock()

# 管理员申请持久化文件
ADMIN_REQUESTS_FILE = os.path.join(BASE_PATH, 'admin_requests.json')

def load_admin_requests():
    """Docstring."""
    if os.path.exists(ADMIN_REQUESTS_FILE):
        try:
            with open(ADMIN_REQUESTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️  加载管理员申请记录失败: {e}")
            return {}
    return {}

def save_admin_requests():
    """Docstring."""
    try:
        with open(ADMIN_REQUESTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(admin_requests, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  保存管理员申请记录失败: {e}")

admin_requests = load_admin_requests()  # {request_id: {'username': str, 'ip': str, 'timestamp': str}}
print(f"👑 已加载 {len(admin_requests)} 个管理员申请记录")

def get_admin_users():
    """从注册信息中获取管理员用户名集合"""
    admins = set()
    with registration_lock:
        for client_id, user_info in registered_users.items():
            if user_info.get('is_admin', False):
                admins.add(user_info['username'])
    return admins

def set_user_admin(username, is_admin=True):
    """Docstring."""
    with registration_lock:
        for client_id, user_info in registered_users.items():
            if user_info['username'] == username:
                user_info['is_admin'] = is_admin
                if is_admin:
                    user_info['admin_granted_at'] = datetime.now().isoformat()
                else:
                    user_info.pop('admin_granted_at', None)
                save_registrations(registered_users)
                print(f"已更新用户管理员状态: {username} -> {is_admin}")
                return True
    return False

def is_user_admin(username):
    """Docstring."""
    with registration_lock:
        for client_id, user_info in registered_users.items():
            if user_info['username'] == username:
                return user_info.get('is_admin', False)
    return False

INVALID_USERNAME_PATTERN = re.compile(r'^[?？�]+$')
STRICT_NICKNAME_PATTERN = re.compile(r'^[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]{1,4}$')

def normalize_username(value, allow_legacy=False):
    """Docstring."""
    username = str(value or '').strip()
    if not username:
        return ''
    if INVALID_USERNAME_PATTERN.fullmatch(username):
        return ''
    if STRICT_NICKNAME_PATTERN.fullmatch(username):
        return username
    if allow_legacy and 1 <= len(username) <= 20:
        return username
    return ''

def get_account_storage_key(username, allow_legacy=False):
    """Docstring."""
    normalized = normalize_username(username, allow_legacy=allow_legacy)
    if not normalized:
        return ''
    return f"{ACCOUNT_KEY_PREFIX}{normalized.casefold()}"

def hash_password(password, salt=None):
    """Docstring."""
    password_text = str(password or '')
    if salt is None:
        salt = os.urandom(16).hex()
    password_hash = hashlib.pbkdf2_hmac(
        'sha256',
        password_text.encode('utf-8'),
        bytes.fromhex(salt),
        PASSWORD_HASH_ITERATIONS
    ).hex()
    return salt, password_hash

def verify_password(password, user_info):
    """Docstring."""
    if not isinstance(user_info, dict):
        return False
    password_salt = str(user_info.get('password_salt') or '')
    password_hash = str(user_info.get('password_hash') or '')
    if not password_salt or not password_hash:
        return False
    _, candidate_hash = hash_password(password, password_salt)
    return hmac.compare_digest(candidate_hash, password_hash)

def merge_registration_info(existing_info, incoming_info):
    """Docstring."""
    merged = dict(existing_info or {})
    incoming = dict(incoming_info or {})

    for field in ('username', 'ip'):
        if incoming.get(field):
            merged[field] = incoming[field]

    for field in ('registered_at', 'admin_granted_at'):
        current_value = str(merged.get(field) or '')
        incoming_value = str(incoming.get(field) or '')
        if incoming_value and (not current_value or incoming_value < current_value):
            merged[field] = incoming_value

    for field in ('last_login_at', 'password_updated_at'):
        current_value = str(merged.get(field) or '')
        incoming_value = str(incoming.get(field) or '')
        if incoming_value and (not current_value or incoming_value > current_value):
            merged[field] = incoming_value

    if incoming.get('is_admin'):
        merged['is_admin'] = True

    if incoming.get('password_hash') and incoming.get('password_salt'):
        merged['password_hash'] = incoming['password_hash']
        merged['password_salt'] = incoming['password_salt']

    return merged

def get_user_record_by_username(username):
    """Docstring."""
    account_key = get_account_storage_key(username)
    if not account_key:
        return '', None

    user_info = registered_users.get(account_key)
    if user_info:
        return account_key, user_info

    normalized = normalize_username(username)
    for client_id, user_info in registered_users.items():
        if normalize_username(user_info.get('username')) == normalized:
            return client_id, user_info
    return account_key, None

def load_registrations():
    """Load persisted account registrations."""
    if os.path.exists(REGISTRATIONS_FILE):
        try:
            with open(REGISTRATIONS_FILE, 'r', encoding='utf-8') as f:
                registrations = json.load(f)
            if not isinstance(registrations, dict):
                return {}

            cleaned_registrations = {}
            removed_clients = []
            migrated_clients = []
            for client_id, user_info in registrations.items():
                if not isinstance(user_info, dict):
                    removed_clients.append(client_id)
                    continue

                username = normalize_username(user_info.get('username'), allow_legacy=True)
                if not username:
                    removed_clients.append(client_id)
                    continue

                account_key = get_account_storage_key(username, allow_legacy=True)
                if not account_key:
                    removed_clients.append(client_id)
                    continue

                cleaned_info = dict(user_info)
                cleaned_info['username'] = username
                cleaned_info['ip'] = cleaned_info.get('ip') or (
                    client_id if not str(client_id).startswith(ACCOUNT_KEY_PREFIX) else ''
                )
                cleaned_info['password_hash'] = str(cleaned_info.get('password_hash') or '')
                cleaned_info['password_salt'] = str(cleaned_info.get('password_salt') or '')

                if account_key in cleaned_registrations:
                    cleaned_registrations[account_key] = merge_registration_info(
                        cleaned_registrations[account_key],
                        cleaned_info
                    )
                else:
                    cleaned_registrations[account_key] = cleaned_info

                if client_id != account_key:
                    migrated_clients.append(client_id)

            if removed_clients or migrated_clients or cleaned_registrations != registrations:
                save_registrations(cleaned_registrations)
                if removed_clients:
                    print(
                        f"[account-migration] removed {len(removed_clients)} invalid registrations: "
                        f"{', '.join(map(str, removed_clients))}"
                    )

            if migrated_clients:
                print(
                    f"[account-migration] migrated {len(migrated_clients)} legacy registrations "
                    "to username-based accounts"
                )

            return cleaned_registrations
        except Exception as e:
            print(f"[account-storage] failed to load registrations: {e}")
            return {}
    return {}

def save_registrations(registrations):
    """Persist account registrations."""
    try:
        with open(REGISTRATIONS_FILE, 'w', encoding='utf-8') as f:
            json.dump(registrations, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"[account-storage] failed to save registrations: {e}")

def get_client_id():
    """Return a stable identifier for the current logged-in account when available."""
    return get_account_storage_key(normalize_username(session.get('username'))) or (request.remote_addr or '')

# 加载已注册的用户信息
registered_users = load_registrations()
print(f"[startup] loaded {len(registered_users)} registered accounts")

# 加载管理员列表并打印
admin_count = len(get_admin_users())
print(f"[startup] loaded {admin_count} admins")

# 任务管理系统（使用配置文件中的路径）
tasks = {}  # {task_id: {'type': 'upload'|'download', 'status': 'running'|'paused'|'completed'|'error', ...}}
tasks_lock = threading.Lock()

def load_tasks():
    """加载任务列表"""
    if os.path.exists(TASKS_FILE):
        try:
            with open(TASKS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️  加载任务列表失败: {e}")
            return {}
    return {}

def save_tasks():
    """保存任务列表"""
    try:
        with open(TASKS_FILE, 'w', encoding='utf-8') as f:
            json.dump(tasks, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  保存任务列表失败: {e}")

# 加载任务列表
tasks = load_tasks()
print(f"[startup] loaded {len(tasks)} tasks")

# 分享链接管理系统
SHARE_LINKS_FILE = os.path.join(BASE_PATH, 'share_links.json')
share_links = {}  # {link_id: {'file_path': str, 'password': str|None, 'expires': timestamp, 'downloads': int, 'max_downloads': int|None}}
share_links_lock = threading.Lock()

def load_share_links():
    """加载分享链接"""
    if os.path.exists(SHARE_LINKS_FILE):
        try:
            with open(SHARE_LINKS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"⚠️  加载分享链接失败: {e}")
            return {}
    return {}

def save_share_links():
    """保存分享链接"""
    try:
        with open(SHARE_LINKS_FILE, 'w', encoding='utf-8') as f:
            json.dump(share_links, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"⚠️  保存分享链接失败: {e}")

share_links = load_share_links()
print(f"[startup] loaded {len(share_links)} share links")

def cleanup_old_tasks():
    """Remove completed or failed tasks older than 24 hours."""
    with tasks_lock:
        current_time = time.time()
        tasks_to_delete = []
        
        for task_id, task in tasks.items():
            if task['status'] in ['completed', 'error']:
                # 获取任务更新时间
                task_time = task.get('updated_at', task.get('created_at', 0))
                if isinstance(task_time, str):
                    try:
                        task_time = datetime.fromisoformat(task_time).timestamp()
                    except:
                        task_time = 0
                
                # 如果超过24小时，标记为删除
                if current_time - task_time > 86400:  # 24小时
                    tasks_to_delete.append(task_id)
        
        # 删除任务
        deleted_count = 0
        for task_id in tasks_to_delete:
            try:
                # 清理临时文件
                for filename in os.listdir(TEMP_FOLDER):
                    if filename.startswith(task_id + '_'):
                        filepath = os.path.join(TEMP_FOLDER, filename)
                        try:
                            os.remove(filepath)
                        except:
                            pass
                
                # 删除任务
                del tasks[task_id]
                deleted_count += 1
            except:
                pass
        
        if deleted_count > 0:
            print(f"[cleanup] removed {deleted_count} expired tasks")
            save_tasks()

def periodic_cleanup():
    """Periodically clean stale background state."""
    while True:
        time.sleep(3600)
        try:
            cleanup_old_tasks()
            cleanup_expired_excel_collaboration()
            cleanup_expired_document_collaboration()
            cleanup_expired_text_realtime()
        except Exception as e:
            print(f"[cleanup] failed: {e}")

cleanup_thread = threading.Thread(target=periodic_cleanup, daemon=True)
cleanup_thread.start()
print("[startup] cleanup thread started")

# 不再使用自动清理离线用户的功能，改为手动通知离线

# 添加CORS支持（允许测试工具访问）
@app.after_request
def add_cors_headers(response):
    """Docstring."""
    response.headers['Access-Control-Allow-Origin'] = '*'
    response.headers['Access-Control-Allow-Methods'] = 'GET, POST, OPTIONS, DELETE'
    response.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization'
    response.headers['Access-Control-Max-Age'] = '3600'
    if not request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    if request.cookies.get('session'):
        response.delete_cookie('session', path='/')
    return response

@app.route('/', methods=['OPTIONS'])
def handle_options():
    """Handle CORS preflight requests."""
    response = app.make_default_options_response()
    return response

# 用户管理相关路由
# 移除了自动更新活跃状态的逻辑

def get_page_session_id():
    """Docstring."""
    raw_value = (
        request.headers.get('X-Page-Session-Id')
        or request.values.get('page_session_id')
        or request.args.get('page_session_id')
        or ''
    )
    page_session_id = re.sub(r'[^A-Za-z0-9._:-]', '', str(raw_value).strip())[:128]
    return page_session_id or (request.remote_addr or '')

def _prune_expired_online_users_locked(now=None):
    """Docstring."""
    current_time = now if now is not None else time.time()
    expired_keys = [
        presence_id
        for presence_id, info in online_users.items()
        if current_time - float(info.get('last_seen', current_time)) > ONLINE_PRESENCE_TTL_SECONDS
    ]
    for presence_id in expired_keys:
        online_users.pop(presence_id, None)

def sync_current_online_presence():
    """Docstring."""
    username = normalize_username(session.get('username'))
    client_ip = request.remote_addr
    presence_id = get_page_session_id()
    if not username or not client_ip or not presence_id:
        return None

    now = time.time()
    with user_lock:
        _prune_expired_online_users_locked(now)
        online_users[presence_id] = {
            'username': username,
            'ip': client_ip,
            'last_seen': now
        }
    return presence_id

@app.route('/login', methods=['POST'])
def login():
    """Handle account login and registration."""
    redirect_mode = request.args.get('redirect') == '1'

    def build_login_error(message, status_code=400):
        if redirect_mode:
            flash(message, 'danger')
            return redirect(url_for('index'))
        return jsonify({'success': False, 'message': message}), status_code

    username = normalize_username(request.form.get('username'))
    if not username:
        return build_login_error(f'昵称必须为 1 到 {NICKNAME_MAX_LENGTH} 个中文字符。')

    password = str(request.form.get('password') or '')
    mode = (request.form.get('mode') or 'login').strip().lower()
    if mode not in {'login', 'register'}:
        mode = 'login'
    if len(password) < PASSWORD_MIN_LENGTH:
        return build_login_error(f'密码长度不能少于 {PASSWORD_MIN_LENGTH} 个字符。')

    now = datetime.now().isoformat()
    client_ip = request.remote_addr or ''

    with registration_lock:
        account_key, user_info = get_user_record_by_username(username)
        if mode == 'register':
            if user_info and user_info.get('password_hash'):
                return build_login_error('该昵称已存在，请直接登录。', 409)

            password_salt, password_hash = hash_password(password)
            account_info = dict(user_info or {})
            account_info['username'] = username
            account_info['ip'] = client_ip
            account_info['registered_at'] = account_info.get('registered_at') or now
            account_info['last_login_at'] = now
            account_info['password_salt'] = password_salt
            account_info['password_hash'] = password_hash
            account_info['password_updated_at'] = now
            registered_users[account_key] = account_info
            save_registrations(registered_users)
            user_info = account_info
        else:
            if not user_info or not user_info.get('password_hash'):
                return build_login_error('该昵称不存在，请先注册。', 404)
            if not verify_password(password, user_info):
                return build_login_error('昵称或密码错误。', 401)

            account_info = dict(user_info)
            account_info['username'] = username
            account_info['ip'] = client_ip
            account_info['last_login_at'] = now
            registered_users[account_key] = account_info
            save_registrations(registered_users)
            user_info = account_info

    session['username'] = username
    sync_current_online_presence()
    add_activity(username, 'online', None)
    
    if redirect_mode:
        flash('登录成功。', 'success')
        return redirect(url_for('index'))
    return jsonify({
        'success': True,
        'username': username,
        'message': '注册成功并已登录。' if mode == 'register' else '登录成功。'
    })

@app.route('/check_registration', methods=['GET'])
def check_registration():
    """Return the current login state for the active session."""
    username = normalize_username(session.get('username'))
    if not username:
        return jsonify({'success': True, 'registered': False, 'logged_in': False})

    with registration_lock:
        account_key, user_info = get_user_record_by_username(username)
        if not user_info or not user_info.get('password_hash'):
            session.pop('username', None)
            return jsonify({'success': True, 'registered': False, 'logged_in': False})

    sync_current_online_presence()
    return jsonify({
        'success': True,
        'registered': True,
        'logged_in': True,
        'username': username,
        'message': f'欢迎回来，{username}。'
    })

@app.route('/logout', methods=['POST'])
def logout():
    """Log out the current session."""
    username = normalize_username(session.get('username'))
    if username:
        add_activity(username, 'offline', None)
    session.pop('username', None)
    presence_id = get_page_session_id()
    with user_lock:
        online_users.pop(presence_id, None)
    return jsonify({'success': True, 'message': '已退出登录。'})

@app.route('/get_online_users', methods=['GET'])
def get_online_users():
    """获取在线用户列表（去重，每个用户只显示一次）"""
    sync_current_online_presence()
    with user_lock:
        _prune_expired_online_users_locked()
        unique_users = {}
        for info in online_users.values():
            username = info['username']
            if username not in unique_users:
                unique_users[username] = {
                    'username': username,
                    'ip': info['ip']
                }
        users = list(unique_users.values())
    return jsonify({'count': len(users), 'users': users})

@app.route('/offline', methods=['POST'])
def user_offline():
    """Mark the current page session as offline."""
    presence_id = get_page_session_id()
    with user_lock:
        if presence_id in online_users:
            username = online_users[presence_id]['username']
            del online_users[presence_id]
            add_activity(username, 'offline', None)
            return jsonify({'success': True, 'message': '已下线。'})
    return jsonify({'success': False, 'message': '未找到在线会话。'}), 404

@app.route('/api/check_admin', methods=['GET'])
def check_admin():
    """Docstring."""
    admin_state = get_admin_state_snapshot()
    return jsonify({
        'is_admin': admin_state['is_admin'],
        'is_host': admin_state['is_host'],
        'username': admin_state['username'],
        'client_ip': admin_state['client_ip'],
        'local_ip': admin_state['local_ip'],
        'requests': admin_state['requests']
    })

@app.route('/api/request_admin', methods=['POST'])
def request_admin():
    """Submit an admin approval request for the current user."""
    username = get_current_username()
    if username == '匿名用户':
        return jsonify({'success': False, 'message': '请先登录。'}), 401
    
    if is_user_admin(username):
        return jsonify({'success': False, 'message': '您已经是管理员了'}), 400
    
    # 创建申请
    request_id = str(uuid.uuid4())
    with admin_requests_lock:
        admin_requests[request_id] = {
            'username': username,
            'ip': request.remote_addr,
            'timestamp': datetime.now().isoformat(),
            'status': 'pending'
        }
        save_admin_requests()
    print(f"[admin-request] {username} requested admin access ({request_id})")
    
    return jsonify({
        'success': True,
        'message': '管理员申请已提交，请等待主机审批',
        'request_id': request_id
    })

@app.route('/api/admin_requests', methods=['GET'])
def get_admin_requests():
    """获取管理员申请列表（仅主机可见）"""
    # 检查是否来自主机
    local_ip = get_local_ip()
    client_ip = request.remote_addr
    if client_ip not in ['127.0.0.1', 'localhost', '::1', local_ip]:
        return jsonify({'success': False, 'message': '仅主机可访问'}), 403
    
    with admin_requests_lock:
        pending_requests = {
            rid: req for rid, req in admin_requests.items()
            if req['status'] == 'pending'
        }
    
    return jsonify({
        'success': True,
        'requests': pending_requests
    })

@app.route('/api/admin_approve/<request_id>', methods=['POST'])
def approve_admin(request_id):
    """Approve a pending admin request."""
    # 检查是否来自主机
    local_ip = get_local_ip()
    client_ip = request.remote_addr
    if client_ip not in ['127.0.0.1', 'localhost', '::1', local_ip]:
        return jsonify({'success': False, 'message': '仅主机可操作'}), 403
    
    with admin_requests_lock:
        if request_id not in admin_requests:
            return jsonify({'success': False, 'message': '申请不存在。'}), 404
        
        req = admin_requests[request_id]
        username = req['username']
        
        set_user_admin(username, True)
        
        req['status'] = 'approved'
        req['approved_at'] = datetime.now().isoformat()
        
        del admin_requests[request_id]
        save_admin_requests()
    print(f"[admin-request] approved {username}")
    
    return jsonify({
        'success': True,
        'message': f'已批准 {username} 成为管理员。'
    })

@app.route('/api/admin_reject/<request_id>', methods=['POST'])
def reject_admin(request_id):
    """Reject a pending admin request."""
    # 检查是否来自主机
    local_ip = get_local_ip()
    client_ip = request.remote_addr
    if client_ip not in ['127.0.0.1', 'localhost', '::1', local_ip]:
        return jsonify({'success': False, 'message': '仅主机可操作'}), 403
    
    with admin_requests_lock:
        if request_id not in admin_requests:
            return jsonify({'success': False, 'message': '申请不存在。'}), 404
        
        req = admin_requests[request_id]
        username = req['username']
        
        req['status'] = 'rejected'
        req['rejected_at'] = datetime.now().isoformat()
        
        del admin_requests[request_id]
        save_admin_requests()
    print(f"[admin-request] rejected {username}")
    
    return jsonify({
        'success': True,
        'message': f'已拒绝 {username} 的申请。'
    })

@app.route('/api/set_admin/<username>', methods=['POST'])
def manual_set_admin(username):
    """Grant admin permission manually from the host machine."""
    # 检查是否来自主机
    local_ip = get_local_ip()
    client_ip = request.remote_addr
    if client_ip not in ['127.0.0.1', 'localhost', '::1', local_ip]:
        return jsonify({'success': False, 'message': '仅主机可操作'}), 403
    
    # 设置为管理员
    success = set_user_admin(username, True)
    
    if success:
        print(f"[admin-request] manually granted admin to {username}")
        return jsonify({
            'success': True,
            'message': f'已授予 {username} 管理员权限。'
        })
    else:
        return jsonify({
            'success': False,
            'message': f'用户 {username} 不存在或尚未注册。'
        }), 404

@app.route('/get_activities', methods=['GET'])
def get_activities():
    """Docstring."""
    sync_current_online_presence()
    return jsonify({
        'activities': get_recent_activities_snapshot(20),
        'active_tasks': get_active_tasks_snapshot()
    })

@app.route('/api/realtime_stream', methods=['GET'])
def realtime_stream():
    """Docstring."""
    def generate_events():
        last_payload = None
        while True:
            try:
                sync_current_online_presence()
                payload = {
                    'current_username': get_current_username(),
                    'online_users': get_online_users_snapshot(),
                    'activities': get_recent_activities_snapshot(10),
                    'active_tasks': get_active_tasks_snapshot(),
                    'admin': get_admin_state_snapshot()
                }
                serialized = json.dumps(payload, ensure_ascii=False, separators=(',', ':'))
                if serialized != last_payload:
                    yield f"event: snapshot\ndata: {serialized}\n\n"
                    last_payload = serialized
                else:
                    yield ": heartbeat\n\n"
                time.sleep(REALTIME_STREAM_INTERVAL_SECONDS)
            except GeneratorExit:
                break
            except Exception as e:
                error_payload = json.dumps({'message': str(e)}, ensure_ascii=False)
                yield f"event: error\ndata: {error_payload}\n\n"
                time.sleep(REALTIME_STREAM_INTERVAL_SECONDS)

    response = Response(stream_with_context(generate_events()), mimetype='text/event-stream')
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    return response

def add_activity(username, action, filename):
    """添加活动记录"""
    with activity_lock:
        current_activities.append({
            'username': username,
            'action': action,
            'filename': filename,
            'timestamp': time.time()
        })
        if len(current_activities) > 100:
            current_activities.pop(0)

def get_online_users_snapshot():
    """Docstring."""
    with user_lock:
        _prune_expired_online_users_locked()
        unique_users = {}
        for info in online_users.values():
            username = normalize_username(info.get('username'))
            if not username or username in unique_users:
                continue
            unique_users[username] = {
                'username': username,
                'ip': info.get('ip', '')
            }
        return list(unique_users.values())

def get_recent_activities_snapshot(limit=10):
    """Docstring."""
    with activity_lock:
        recent = current_activities[-limit:] if len(current_activities) > limit else current_activities[:]
    return list(reversed(recent))

def get_active_tasks_snapshot():
    """Docstring."""
    active_tasks = []
    with tasks_lock:
        for task_id, task in tasks.items():
            if task.get('status') not in ['running', 'paused']:
                continue
            if task.get('type') != 'upload':
                continue

            progress = 0
            if task.get('total_chunks', 0) > 0:
                progress = round((task.get('uploaded_chunks', 0) / task['total_chunks']) * 100, 1)
                if progress > 99 and progress < 100:
                    progress = 99.9

            active_tasks.append({
                'task_id': task_id,
                'username': task.get('username', '未知用户'),
                'type': 'upload',
                'filename': task.get('filename', ''),
                'upload_path': task.get('upload_path', ''),
                'progress': progress,
                'status': task.get('status', 'running'),
                'uploaded_chunks': task.get('uploaded_chunks', 0),
                'total_chunks': task.get('total_chunks', 0)
            })
    return active_tasks

def get_admin_state_snapshot():
    """Docstring."""
    username = normalize_username(session.get('username'))
    client_ip = request.remote_addr
    local_ip = get_local_ip()
    is_host = client_ip in ['127.0.0.1', 'localhost', '::1', local_ip]

    is_admin = False
    if username:
        if is_host:
            if not is_user_admin(username):
                set_user_admin(username, True)
                print(f"自动授予本机用户管理员权限: {username} (IP: {client_ip})")
            is_admin = True
        else:
            is_admin = is_user_admin(username)

    pending_requests = {}
    if is_host:
        with admin_requests_lock:
            pending_requests = {
                rid: req for rid, req in admin_requests.items()
                if req.get('status') == 'pending'
            }

    return {
        'is_admin': is_admin,
        'is_host': is_host,
        'username': username,
        'client_ip': client_ip,
        'local_ip': local_ip,
        'requests': pending_requests
    }

class NoLimitMiddleware:
    def __init__(self, app):
        self.app = app

    def __call__(self, environ, start_response):
        if environ.get('REQUEST_METHOD') == 'POST':
            # 移除或修改可能触发大小检查的环境变量
            environ['werkzeug.request.max_content_length'] = None
            # 确保没有限制
            if 'HTTP_CONTENT_LENGTH' in environ:
                pass
        return self.app(environ, start_response)

# 应用中间件
app.wsgi_app = NoLimitMiddleware(app.wsgi_app)

# 确保上传文件夹存在并设置权限
def ensure_upload_folder():
    """Docstring."""
    if not os.path.exists(UPLOAD_FOLDER):
        try:
            os.makedirs(UPLOAD_FOLDER, exist_ok=True)
            print(f"✅ 创建上传文件夹: {UPLOAD_FOLDER}")
        except Exception as e:
            print(f"❌ 创建文件夹失败: {e}")
            raise
    
    if not os.path.exists(TEMP_FOLDER):
        try:
            os.makedirs(TEMP_FOLDER, exist_ok=True)
            print(f"✅ 创建临时文件夹: {TEMP_FOLDER}")
        except Exception as e:
            print(f"⚠️  创建临时文件夹失败: {e}")
    
    test_file = os.path.join(UPLOAD_FOLDER, '.write_test')
    try:
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        print(f"✅ 文件夹写入权限正常: {UPLOAD_FOLDER}")
    except PermissionError:
        print(f"⚠️  警告: 文件夹可能没有写入权限: {UPLOAD_FOLDER}")
        print(f"   请以管理员身份运行程序，或检查文件夹权限")
    except Exception as e:
        print(f"⚠️  警告: 无法写入文件夹: {e}")

# 初始化上传文件夹
ensure_upload_folder()

UPLOAD_FOLDER_ABS = os.path.abspath(UPLOAD_FOLDER)

def allowed_file(filename):
    """检查文件扩展名是否允许。"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def save_file_safely(file, filepath):
    """Save an uploaded file safely."""
    try:
        directory = os.path.dirname(filepath)
        if directory and not os.path.exists(directory):
            os.makedirs(directory, exist_ok=True)
        
        if not os.access(directory if directory else UPLOAD_FOLDER, os.W_OK):
            return False, '目标目录没有写入权限。'
        
        file.save(filepath)
        
        print(f"[upload] saved {os.path.basename(filepath)}")
        return True, None
        
    except PermissionError as e:
        return False, f'权限不足: {str(e)}'
    except OSError as e:
        if "No space" in str(e) or "磁盘空间不足" in str(e):
            return False, '磁盘空间不足。'
        return False, f'文件系统错误: {str(e)}'
    except Exception as e:
        return False, f'保存文件失败: {str(e)}'

def extract_zip_safely(zip_path, extract_to):
    """Extract a zip file safely."""
    try:
        if not zipfile.is_zipfile(zip_path):
            return False, 0, '不是有效的 ZIP 文件。'
        
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            file_list = zip_ref.namelist()
            file_count = len(file_list)

            zip_ref.extractall(extract_to)

        try:
            os.remove(zip_path)
        except:
            pass
        return True, file_count, None
        
    except zipfile.BadZipFile:
        return False, 0, 'ZIP 文件损坏或格式不正确。'
    except PermissionError as e:
        return False, 0, f'权限不足: {str(e)}'
    except Exception as e:
        return False, 0, f'解压失败: {str(e)}'

def get_file_icon(filename):
    """Return a display icon for the given filename."""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    
    if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg', 'ico']:
        return '🖼️'
    elif ext in ['mp4', 'avi', 'mkv', 'mov', 'wmv', 'flv', 'webm']:
        return '🎬'
    elif ext in ['mp3', 'wav', 'flac', 'aac', 'ogg', 'm4a']:
        return '🎵'
    elif ext in ['zip', 'rar', '7z', 'tar', 'gz']:
        return '📦'
    elif ext in ['doc', 'docx']:
        return '📄'
    elif ext in ['xls', 'xlsx']:
        return '📊'
    elif ext in ['ppt', 'pptx']:
        return '📽️'
    elif ext == 'pdf':
        return '📕'
    elif ext in ['py', 'java', 'cpp', 'c', 'h', 'js', 'ts', 'html', 'css']:
        return '💻'
    elif ext in ['txt', 'md', 'json', 'xml', 'yaml', 'yml', 'log', 'csv']:
        return '📝'
    
    elif ext in ['exe', 'msi', 'apk', 'dmg', 'deb', 'rpm', 'pkg']:
        return '⚙️'
    elif ext == 'iso':
        return '💿'
    elif ext in ['sh', 'bat', 'cmd', 'ps1']:
        return '📜'
    
    elif ext in ['sql', 'db', 'sqlite', 'mdb']:
        return '🗄️'
    else:
        return '📁'

def get_file_size(size):
    """Format a file size in a human-readable form."""
    for unit in ['B', 'KB', 'MB', 'GB']:
        if size < 1024.0:
            return f"{size:.1f} {unit}"
        size /= 1024.0
    return f"{size:.1f} TB"

def validate_path(path, base_folder=None):
    """Validate that a path stays inside the shared root."""
    try:
        abs_path = os.path.abspath(path)
        
        if base_folder is None:
            abs_base = UPLOAD_FOLDER_ABS
        else:
            abs_base = os.path.abspath(base_folder)
        
        if not abs_path.startswith(abs_base):
            return False, None, '非法路径：超出允许范围。'
        
        return True, abs_path, None
    except Exception as e:
        return False, None, f'路径校验失败: {str(e)}'

def get_current_username():
    """Return the current logged-in username or a guest label."""
    return normalize_username(session.get('username')) or '匿名用户'

def is_macos_metadata_file(filename):
    """判断文件是否为 macOS 生成的元数据副本，例如 .DS_Store 或 ._xxx。"""
    normalized = str(filename or '').replace('\\', '/')
    basename = os.path.basename(normalized)
    if not basename:
        return False
    lowered = basename.lower()
    return lowered == '.ds_store' or basename.startswith('._')

def should_hide_shared_item(item_name):
    """Docstring."""
    if is_macos_metadata_file(item_name):
        return True
    return HIDE_SYSTEM_FOLDERS and str(item_name or '').lower() in HIDDEN_ITEMS

def resolve_macos_metadata_target(relative_path):
    """Docstring."""
    normalized = str(relative_path or '').replace('\\', '/').strip('/')
    basename = os.path.basename(normalized)
    if not basename.startswith('._') or len(basename) <= 2:
        return None

    parent = os.path.dirname(normalized).replace('\\', '/')
    actual_name = basename[2:]
    actual_relative = f'{parent}/{actual_name}' if parent else actual_name
    success, actual_path, _ = safe_join_path(app.config['UPLOAD_FOLDER'], actual_relative)
    if success and os.path.exists(actual_path) and os.path.isfile(actual_path):
        return actual_relative.replace('\\', '/')
    return None

TEXT_PREVIEW_EXTENSIONS = {
    'txt', 'py', 'js', 'html', 'css', 'json', 'xml', 'md', 'csv', 'log',
    'sql', 'sh', 'bat', 'java', 'cpp', 'c', 'h', 'ts', 'yaml', 'yml',
    'ini', 'conf', 'ps1'
}
EXCEL_EDITABLE_EXTENSIONS = {'xlsx', 'xlsm', 'xls'}
WORD_EDITABLE_EXTENSIONS = {'docx'}
DIRECT_PREVIEW_EXTENSIONS = {
    'jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg', 'ico',
    'pdf',
    'mp4', 'webm', 'ogg', 'mp3', 'wav', 'flac', 'aac', 'm4a'
}
INLINE_STREAM_EXTENSIONS = {'mp4', 'webm', 'ogg', 'mp3', 'wav', 'flac', 'aac', 'm4a'}
TEXT_EDIT_ENCODINGS = ('utf-8', 'utf-8-sig', 'gb18030', 'gbk', 'big5')
MAX_EDITABLE_TEXT_FILE_SIZE = 2 * 1024 * 1024
TEXT_PREVIEW_FALLBACK_BYTES = 100 * 1024
MAX_EDITABLE_EXCEL_FILE_SIZE = 5 * 1024 * 1024
MAX_EDITABLE_DOCX_FILE_SIZE = 10 * 1024 * 1024
DEFAULT_EXCEL_PREVIEW_ROWS = 200
DEFAULT_EXCEL_PREVIEW_COLS = 50
HEAVY_EXCEL_PREVIEW_ROWS = 120
HEAVY_EXCEL_PREVIEW_COLS = 25
MAX_EXCEL_PREVIEW_ROWS = DEFAULT_EXCEL_PREVIEW_ROWS
MAX_EXCEL_PREVIEW_COLS = DEFAULT_EXCEL_PREVIEW_COLS
MAX_EXCEL_PREVIEW_SHEETS = 10
MAX_EXCEL_SEARCH_RESULTS = 30
EXCEL_SEARCH_BLOCK_ROWS = 200
MAX_EXCEL_STRUCTURE_AMOUNT = 200
EXCEL_MAX_ROWS = 1048576
EXCEL_MAX_COLS = 16384
INVALID_EXCEL_SHEET_NAME_RE = re.compile(r'[:\\/?*\[\]]')
EXCEL_COLLAB_HEARTBEAT_TTL = 18
EXCEL_COLLAB_LOCK_TTL = 18
EXCEL_COLLAB_POLL_INTERVAL_MS = 3500
DOCUMENT_COLLAB_HEARTBEAT_TTL = 18
DOCUMENT_COLLAB_LOCK_TTL = 18
DOCUMENT_COLLAB_POLL_INTERVAL_MS = 3200
TEXT_REALTIME_CLIENT_TTL = 25
TEXT_REALTIME_HISTORY_LIMIT = 80
TEXT_REALTIME_STREAM_HEARTBEAT = 15
TEXT_REALTIME_STREAM_QUEUE_SIZE = 32
TEXT_REALTIME_EDIT_DEBOUNCE_MS = 260

excel_collaboration_lock = threading.Lock()
excel_collaboration_state = {}
document_collaboration_lock = threading.Lock()
document_collaboration_state = {}
text_realtime_lock = threading.Lock()
text_realtime_state = {}
document_realtime_stream_lock = threading.Lock()
document_realtime_streams = {}
excel_com_available_cache = None

def safe_join_path(base, *paths):
    """
    安全地连接路径并验证
    
    Args:
        base: 基础路径
        *paths: 要连接的路径片段
    
    Returns:
        (success, absolute_path, error_message)
    """
    try:
        if not paths or all(not p for p in paths):
            joined_path = base
        else:
            joined_path = os.path.join(base, *[p for p in paths if p])
        
        return validate_path(joined_path, base)
    except Exception as e:
        return False, None, f'路径连接失败: {str(e)}'

def get_local_ip():
    """获取本机局域网 IP 地址。"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except:
        return '127.0.0.1'

def is_request_from_host():
    """Docstring."""
    client_ip = request.remote_addr
    if not client_ip:
        return False
    return client_ip in {'127.0.0.1', 'localhost', '::1', get_local_ip()}

def user_can_edit_files():
    """Docstring."""
    username = normalize_username(session.get('username'))
    if not username:
        return False
    if is_request_from_host():
        if not is_user_admin(username):
            set_user_admin(username, True)
        return True
    return is_user_admin(username)

def get_excel_editor_client_id():
    """为当前浏览器会话返回稳定的 Excel 协作客户端 ID。"""
    client_id = session.get('excel_editor_client_id')
    if not client_id:
        client_id = uuid.uuid4().hex
        session['excel_editor_client_id'] = client_id
    return client_id

def get_document_editor_client_id():
    """为文本与 DOCX 编辑页返回稳定的协作客户端 ID。"""
    client_id = session.get('document_editor_client_id')
    if not client_id:
        client_id = uuid.uuid4().hex
        session['document_editor_client_id'] = client_id
    return client_id

def get_document_collaboration_file_key(filepath):
    """将文本或 DOCX 文件路径标准化为协作状态键。"""
    return os.path.normcase(os.path.abspath(filepath))

def build_text_document_target():
    """Docstring."""
    return {
        'kind': 'text_document',
        'key': 'text_document',
        'label': '全文'
    }

def build_docx_paragraph_target_key(block_id):
    """为 DOCX 段落生成稳定的目标键。"""
    return f'paragraph::{str(block_id or "").strip()}'

def build_docx_table_cell_target_key(block_id, row, col):
    """Docstring."""
    return f'table_cell::{str(block_id or "").strip()}::{int(row)}::{int(col)}'

def normalize_document_collaboration_target(target_data, editor_type=''):
    """Docstring."""
    editor_type = str(editor_type or '').strip().lower()
    if editor_type == 'text':
        target = build_text_document_target()
        if isinstance(target_data, dict):
            label = str(target_data.get('label') or '').strip()
            if label:
                target['label'] = label[:120]
        return target

    if editor_type != 'docx' or not isinstance(target_data, dict):
        return None

    kind = str(target_data.get('kind') or '').strip().lower()
    block_id = str(target_data.get('block_id') or '').strip()
    label = str(target_data.get('label') or '').strip()[:160]
    if kind == 'docx_paragraph':
        if not block_id:
            return None
        return {
            'kind': kind,
            'block_id': block_id,
            'key': build_docx_paragraph_target_key(block_id),
            'label': label or block_id
        }

    if kind == 'docx_image_meta':
        field = str(target_data.get('field') or '').strip().lower()
        if not block_id or field not in {'title', 'description'}:
            return None
        field_label = '标题' if field == 'title' else '说明'
        return {
            'kind': kind,
            'block_id': block_id,
            'field': field,
            'key': build_docx_image_meta_target_key(block_id, field),
            'label': label or f'{block_id} / {field_label}'
        }

    if kind != 'docx_table_cell':
        return None

    try:
        row = int(target_data.get('row'))
        col = int(target_data.get('col'))
    except (TypeError, ValueError):
        return None

    if not block_id or row < 0 or col < 0:
        return None

    return {
        'kind': kind,
        'block_id': block_id,
        'row': row,
        'col': col,
        'key': build_docx_table_cell_target_key(block_id, row, col),
        'label': label or f'{block_id} / {get_column_letter(col + 1)}{row + 1}'
    }

def cleanup_expired_document_collaboration(now=None):
    """清理过期的文本与 DOCX 协同会话"""
    with document_collaboration_lock:
        _cleanup_expired_document_collaboration_locked(now)

def _cleanup_expired_document_collaboration_locked(now=None):
    """Docstring."""
    current_time = time.time() if now is None else float(now)
    file_keys_to_delete = []

    for file_key, file_state in document_collaboration_state.items():
        sessions = file_state.get('sessions', {})
        locks = file_state.get('locks', {})

        expired_sessions = [
            client_id for client_id, session_data in sessions.items()
            if current_time - float(session_data.get('updated_at', 0)) > DOCUMENT_COLLAB_HEARTBEAT_TTL
        ]
        for client_id in expired_sessions:
            sessions.pop(client_id, None)

        expired_locks = [
            lock_key for lock_key, lock_data in locks.items()
            if current_time - float(lock_data.get('updated_at', 0)) > DOCUMENT_COLLAB_LOCK_TTL
        ]
        for lock_key in expired_locks:
            locks.pop(lock_key, None)

        active_client_ids = set(sessions.keys())
        stale_locks = [
            lock_key for lock_key, lock_data in locks.items()
            if lock_data.get('client_id') not in active_client_ids
        ]
        for lock_key in stale_locks:
            locks.pop(lock_key, None)

        if not sessions and not locks:
            file_keys_to_delete.append(file_key)

    for file_key in file_keys_to_delete:
        document_collaboration_state.pop(file_key, None)

def _release_document_collaboration_locks_locked(file_state, client_id, except_lock_key=None):
    """Docstring."""
    locks = file_state.get('locks', {})
    owned_keys = [
        lock_key for lock_key, lock_data in locks.items()
        if lock_data.get('client_id') == client_id and lock_key != except_lock_key
    ]
    for lock_key in owned_keys:
        locks.pop(lock_key, None)

def build_document_collaboration_snapshot(filepath, client_id):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    now = time.time()

    with document_collaboration_lock:
        _cleanup_expired_document_collaboration_locked(now)
        file_state = document_collaboration_state.setdefault(file_key, {'sessions': {}, 'locks': {}})

        editors = []
        for session_data in file_state.get('sessions', {}).values():
            active_target = session_data.get('active_target')
            editors.append({
                'client_id': session_data.get('client_id'),
                'username': session_data.get('username') or '匿名用户',
                'is_self': session_data.get('client_id') == client_id,
                'editor_type': session_data.get('editor_type') or '',
                'active_target': active_target,
                'active_target_label': active_target.get('label') if isinstance(active_target, dict) else '',
                'updated_at': session_data.get('updated_at', now)
            })

        editors.sort(key=lambda item: (not item['is_self'], item['username'], item['client_id'] or ''))

        locks = []
        own_lock = None
        for lock_data in file_state.get('locks', {}).values():
            payload = {
                'client_id': lock_data.get('client_id'),
                'username': lock_data.get('username') or '匿名用户',
                'editor_type': lock_data.get('editor_type') or '',
                'target': lock_data.get('target'),
                'key': lock_data.get('key'),
                'label': (lock_data.get('target') or {}).get('label', ''),
                'updated_at': lock_data.get('updated_at', now),
                'is_self': lock_data.get('client_id') == client_id
            }
            locks.append(payload)
            if payload['is_self']:
                own_lock = payload

    return {
        'client_id': client_id,
        'mtime_ns': get_file_mtime_token(filepath),
        'editors': editors,
        'locks': locks,
        'own_lock': own_lock
    }

def sync_document_collaboration_presence(filepath, username, client_id, ip_address, editor_type, active_target=None, lock_target=None, release_lock=False):
    """Docstring."""
    editor_type = str(editor_type or '').strip().lower()
    file_key = get_document_collaboration_file_key(filepath)
    now = time.time()
    active_target = normalize_document_collaboration_target(active_target, editor_type)
    lock_target = normalize_document_collaboration_target(lock_target, editor_type)
    lock_denied = None

    with document_collaboration_lock:
        _cleanup_expired_document_collaboration_locked(now)
        file_state = document_collaboration_state.setdefault(file_key, {'sessions': {}, 'locks': {}})
        sessions = file_state['sessions']

        sessions[client_id] = {
            'client_id': client_id,
            'username': username or '匿名用户',
            'ip': ip_address or '',
            'editor_type': editor_type,
            'active_target': active_target,
            'updated_at': now
        }

        if release_lock or not lock_target:
            _release_document_collaboration_locks_locked(file_state, client_id)
        else:
            desired_lock_key = lock_target['key']
            _release_document_collaboration_locks_locked(file_state, client_id, desired_lock_key)
            current_lock = file_state['locks'].get(desired_lock_key)

            if current_lock and current_lock.get('client_id') != client_id:
                lock_denied = {
                    'client_id': current_lock.get('client_id'),
                    'username': current_lock.get('username') or '匿名用户',
                    'key': current_lock.get('key'),
                    'label': (current_lock.get('target') or {}).get('label', '')
                }
            else:
                file_state['locks'][desired_lock_key] = {
                    'client_id': client_id,
                    'username': username or '匿名用户',
                    'ip': ip_address or '',
                    'editor_type': editor_type,
                    'key': desired_lock_key,
                    'target': lock_target,
                    'updated_at': now
                }

    snapshot = build_document_collaboration_snapshot(filepath, client_id)
    snapshot['lock_denied'] = lock_denied
    snapshot['lock_acquired'] = lock_denied is None if lock_target and not release_lock else True
    return snapshot

def release_document_collaboration_presence(filepath, client_id, remove_session=False):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    now = time.time()

    with document_collaboration_lock:
        _cleanup_expired_document_collaboration_locked(now)
        file_state = document_collaboration_state.get(file_key)
        if not file_state:
            return

        _release_document_collaboration_locks_locked(file_state, client_id)
        if remove_session:
            file_state.get('sessions', {}).pop(client_id, None)

        if not file_state.get('sessions') and not file_state.get('locks'):
            document_collaboration_state.pop(file_key, None)

def find_document_locked_targets_for_changes(filepath, target_keys, client_id):
    """Docstring."""
    target_keys = {str(item) for item in (target_keys or set()) if str(item)}
    if not target_keys:
        return []

    file_key = get_document_collaboration_file_key(filepath)
    now = time.time()

    with document_collaboration_lock:
        _cleanup_expired_document_collaboration_locked(now)
        file_state = document_collaboration_state.get(file_key)
        if not file_state:
            return []

        blockers = []
        for lock_data in file_state.get('locks', {}).values():
            if lock_data.get('client_id') == client_id:
                continue
            if lock_data.get('key') in target_keys:
                blockers.append({
                    'username': lock_data.get('username') or '匿名用户',
                    'key': lock_data.get('key'),
                    'label': (lock_data.get('target') or {}).get('label', '')
                })
        return blockers

def register_document_realtime_stream(filepath, client_id, editor_type):
    """为指定文件注册实时事件流订阅"""
    file_key = get_document_collaboration_file_key(filepath)
    stream_id = uuid.uuid4().hex
    stream_queue = queue.Queue(maxsize=TEXT_REALTIME_STREAM_QUEUE_SIZE)

    with document_realtime_stream_lock:
        streams = document_realtime_streams.setdefault(file_key, {})
        streams[stream_id] = {
            'client_id': client_id,
            'editor_type': editor_type,
            'queue': stream_queue,
            'updated_at': time.time()
        }

    return stream_id, stream_queue

def unregister_document_realtime_stream(filepath, stream_id):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    with document_realtime_stream_lock:
        streams = document_realtime_streams.get(file_key)
        if not streams:
            return
        streams.pop(stream_id, None)
        if not streams:
            document_realtime_streams.pop(file_key, None)

def publish_document_realtime_event(filepath, event_type, payload, exclude_client_id=None):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    event_data = {
        'type': event_type,
        **(payload or {})
    }

    with document_realtime_stream_lock:
        streams = list((document_realtime_streams.get(file_key) or {}).values())

    for stream in streams:
        if exclude_client_id and stream.get('client_id') == exclude_client_id:
            continue
        stream_queue = stream.get('queue')
        if not stream_queue:
            continue
        try:
            stream_queue.put_nowait(event_data)
        except queue.Full:
            try:
                stream_queue.get_nowait()
            except queue.Empty:
                pass
            try:
                stream_queue.put_nowait(event_data)
            except queue.Full:
                pass

def cleanup_expired_text_realtime(now=None):
    """Docstring."""
    current_time = time.time() if now is None else float(now)
    file_keys_to_delete = []

    with text_realtime_lock:
        for file_key, state in text_realtime_state.items():
            clients = state.get('clients', {})
            expired_clients = [
                client_id for client_id, client_state in clients.items()
                if current_time - float(client_state.get('updated_at', 0)) > TEXT_REALTIME_CLIENT_TTL
            ]
            for client_id in expired_clients:
                clients.pop(client_id, None)

            if not clients and current_time - float(state.get('last_activity', 0)) > TEXT_REALTIME_CLIENT_TTL:
                file_keys_to_delete.append(file_key)

        for file_key in file_keys_to_delete:
            text_realtime_state.pop(file_key, None)

def build_text_change_spans(base_text, modified_text):
    """将文本差异归一化为连续替换区间。"""
    base_text = '' if base_text is None else str(base_text)
    modified_text = '' if modified_text is None else str(modified_text)
    matcher = difflib.SequenceMatcher(None, base_text, modified_text)
    changes = []
    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == 'equal':
            continue
        changes.append({
            'start': i1,
            'end': i2,
            'replacement': modified_text[j1:j2]
        })
    return changes

def apply_text_change_spans(base_text, changes):
    """Docstring."""
    result = '' if base_text is None else str(base_text)
    offset = 0
    for change in changes or []:
        start = int(change.get('start', 0))
        end = int(change.get('end', start))
        replacement = str(change.get('replacement', ''))
        actual_start = max(0, start + offset)
        actual_end = max(actual_start, end + offset)
        result = f"{result[:actual_start]}{replacement}{result[actual_end:]}"
        offset += len(replacement) - (end - start)
    return result

def _is_text_change_insert(change):
    """判断文本变更是否为纯插入"""
    return int(change.get('start', 0)) == int(change.get('end', 0))

def _text_change_touches_cluster(change, cluster_start, cluster_end):
    """判断区间是否与当前冲突簇相交"""
    change_start = int(change.get('start', 0))
    change_end = int(change.get('end', change_start))
    if cluster_start == cluster_end and change_start == change_end == cluster_start:
        return True
    return change_start < cluster_end and change_end > cluster_start

def merge_text_versions(base_text, ours_text, theirs_text):
    """基于共同基线合并两份文本，冲突区优先保留 ours"""
    base_text = '' if base_text is None else str(base_text)
    ours_text = '' if ours_text is None else str(ours_text)
    theirs_text = '' if theirs_text is None else str(theirs_text)

    ours_changes = build_text_change_spans(base_text, ours_text)
    theirs_changes = build_text_change_spans(base_text, theirs_text)

    if not ours_changes:
        return theirs_text, False
    if not theirs_changes:
        return ours_text, False

    merged_changes = []
    i = 0
    j = 0
    merged_from_conflict = False

    while i < len(ours_changes) or j < len(theirs_changes):
        ours_change = ours_changes[i] if i < len(ours_changes) else None
        theirs_change = theirs_changes[j] if j < len(theirs_changes) else None

        if ours_change is None:
            merged_changes.append(theirs_change)
            j += 1
            continue

        if theirs_change is None:
            merged_changes.append(ours_change)
            i += 1
            continue

        ours_start = int(ours_change['start'])
        ours_end = int(ours_change['end'])
        theirs_start = int(theirs_change['start'])
        theirs_end = int(theirs_change['end'])

        if ours_end < theirs_start or (ours_end == theirs_start and not (_is_text_change_insert(ours_change) and _is_text_change_insert(theirs_change) and ours_start == theirs_start)):
            merged_changes.append(ours_change)
            i += 1
            continue

        if theirs_end < ours_start or (theirs_end == ours_start and not (_is_text_change_insert(ours_change) and _is_text_change_insert(theirs_change) and ours_start == theirs_start)):
            merged_changes.append(theirs_change)
            j += 1
            continue

        cluster_start = min(ours_start, theirs_start)
        cluster_end = max(ours_end, theirs_end)
        cluster_ours = [ours_change]
        cluster_theirs = [theirs_change]
        i += 1
        j += 1

        while i < len(ours_changes):
            next_change = ours_changes[i]
            if not _text_change_touches_cluster(next_change, cluster_start, cluster_end):
                break
            cluster_ours.append(next_change)
            cluster_start = min(cluster_start, int(next_change['start']))
            cluster_end = max(cluster_end, int(next_change['end']))
            i += 1

        while j < len(theirs_changes):
            next_change = theirs_changes[j]
            if not _text_change_touches_cluster(next_change, cluster_start, cluster_end):
                break
            cluster_theirs.append(next_change)
            cluster_start = min(cluster_start, int(next_change['start']))
            cluster_end = max(cluster_end, int(next_change['end']))
            j += 1

        if (
            len(cluster_ours) == len(cluster_theirs) and
            all(
                int(ours_item['start']) == int(theirs_item['start']) and
                int(ours_item['end']) == int(theirs_item['end']) and
                str(ours_item['replacement']) == str(theirs_item['replacement'])
                for ours_item, theirs_item in zip(cluster_ours, cluster_theirs)
            )
        ):
            merged_changes.extend(cluster_ours)
            continue

        if (
            all(_is_text_change_insert(change) for change in cluster_ours) and
            all(_is_text_change_insert(change) for change in cluster_theirs) and
            len({int(change['start']) for change in cluster_ours + cluster_theirs}) == 1
        ):
            replacement = ''.join(str(change['replacement']) for change in cluster_theirs)
            replacement += ''.join(str(change['replacement']) for change in cluster_ours)
            merged_changes.append({
                'start': cluster_start,
                'end': cluster_end,
                'replacement': replacement
            })
            merged_from_conflict = True
            continue

        if (
            len(cluster_ours) == len(cluster_theirs) == 1 and
            cluster_ours[0]['start'] == cluster_theirs[0]['start'] and
            cluster_ours[0]['end'] == cluster_theirs[0]['end'] and
            cluster_ours[0]['replacement'] == cluster_theirs[0]['replacement']
        ):
            merged_changes.append(cluster_ours[0])
            continue

        cluster_base = base_text[cluster_start:cluster_end]
        normalized_ours = [
            {
                'start': int(change['start']) - cluster_start,
                'end': int(change['end']) - cluster_start,
                'replacement': str(change['replacement'])
            }
            for change in cluster_ours
        ]
        merged_changes.append({
            'start': cluster_start,
            'end': cluster_end,
            'replacement': apply_text_change_spans(cluster_base, normalized_ours)
        })
        merged_from_conflict = True

    return apply_text_change_spans(base_text, merged_changes), merged_from_conflict

def encode_text_editor_content(content, encoding='utf-8', newline='\n'):
    """Docstring."""
    encoding = (encoding or 'utf-8').lower()
    newline = newline or '\n'

    if newline not in ('\n', '\r\n', '\r'):
        newline = '\n'

    normalized_content = str(content or '').replace('\r\n', '\n').replace('\r', '\n')
    if newline != '\n':
        normalized_content = normalized_content.replace('\n', newline)

    if encoding not in TEXT_EDIT_ENCODINGS:
        encoding = 'utf-8'

    try:
        encoded_content = normalized_content.encode(encoding)
    except UnicodeEncodeError:
        encoding = 'utf-8'
        encoded_content = normalized_content.encode(encoding)

    return encoded_content, encoding, newline, normalized_content

def write_text_content_to_file(filepath, content, encoding='utf-8', newline='\n'):
    """Docstring."""
    encoded_content, normalized_encoding, normalized_newline, _ = encode_text_editor_content(content, encoding, newline)

    if len(encoded_content) > MAX_EDITABLE_TEXT_FILE_SIZE:
        raise ValueError(
            f'文件内容超过 {get_file_size(MAX_EDITABLE_TEXT_FILE_SIZE)}，请改用本地编辑后再上传。'
        )

    temp_path = f"{filepath}.editing.{uuid.uuid4().hex}.tmp"
    try:
        with open(temp_path, 'wb') as f:
            f.write(encoded_content)
        os.replace(temp_path, filepath)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return {
        'bytes': encoded_content,
        'encoding': normalized_encoding,
        'newline': normalized_newline,
        'size': len(encoded_content),
        'mtime_ns': get_file_mtime_token(filepath)
    }

def _remember_text_realtime_snapshot_locked(state):
    """Docstring."""
    revision = int(state.get('revision', 0))
    snapshots = state.setdefault('snapshots', {})
    order = state.setdefault('snapshot_order', [])
    snapshots[revision] = state.get('content', '')
    if revision not in order:
        order.append(revision)

    while len(order) > TEXT_REALTIME_HISTORY_LIMIT:
        expired_revision = order.pop(0)
        snapshots.pop(expired_revision, None)

def _get_text_realtime_content_at_revision_locked(state, revision):
    """Docstring."""
    revision = int(revision)
    if revision == int(state.get('revision', 0)):
        return state.get('content', '')
    return state.get('snapshots', {}).get(revision)

def ensure_text_realtime_document_locked(filepath):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    current_token = get_file_mtime_token(filepath)
    state = text_realtime_state.get(file_key)

    if state and state.get('mtime_ns') == current_token:
        state['last_activity'] = time.time()
        return state

    preview = load_text_file_preview(filepath)
    content = preview['content']
    revision = int(state.get('revision', 0) + 1) if state else 0

    refreshed_state = {
        'filepath': filepath,
        'content': content,
        'encoding': preview['encoding'],
        'newline': preview['newline'],
        'mtime_ns': str(preview['mtime_ns']),
        'revision': revision,
        'snapshots': {},
        'snapshot_order': [],
        'clients': state.get('clients', {}) if state else {},
        'last_activity': time.time(),
        'file_size': preview['file_size']
    }
    text_realtime_state[file_key] = refreshed_state
    _remember_text_realtime_snapshot_locked(refreshed_state)
    return refreshed_state

def update_text_realtime_presence(filepath, client_id, username):
    """Docstring."""
    with text_realtime_lock:
        state = ensure_text_realtime_document_locked(filepath)
        state.setdefault('clients', {})[client_id] = {
            'client_id': client_id,
            'username': username or '匿名用户',
            'updated_at': time.time()
        }
        state['last_activity'] = time.time()
        return {
            'content': state.get('content', ''),
            'revision': int(state.get('revision', 0)),
            'encoding': state.get('encoding', 'utf-8'),
            'newline': state.get('newline', '\n'),
            'mtime_ns': state.get('mtime_ns', get_file_mtime_token(filepath))
        }

def apply_text_realtime_update(filepath, client_id, username, content, base_revision):
    """Docstring."""
    file_key = get_document_collaboration_file_key(filepath)
    with text_realtime_lock:
        state = ensure_text_realtime_document_locked(filepath)
        clients = state.setdefault('clients', {})
        clients[client_id] = {
            'client_id': client_id,
            'username': username or '匿名用户',
            'updated_at': time.time()
        }
        state['last_activity'] = time.time()

        try:
            base_revision = int(base_revision)
        except (TypeError, ValueError):
            base_revision = int(state.get('revision', 0))

        current_revision = int(state.get('revision', 0))
        current_content = str(state.get('content', ''))
        incoming_content = '' if content is None else str(content)

        if base_revision == current_revision:
            merged_content = incoming_content
            merged = False
        else:
            base_content = _get_text_realtime_content_at_revision_locked(state, base_revision)
            if base_content is None:
                return {
                    'success': False,
                    'refresh_required': True,
                    'message': '当前文本协作版本已过期，请先同步最新内容后再继续编辑。',
                    'content': current_content,
                    'revision': current_revision,
                    'encoding': state.get('encoding', 'utf-8'),
                    'newline': state.get('newline', '\n'),
                    'mtime_ns': state.get('mtime_ns', get_file_mtime_token(filepath))
                }
            merged_content, merged = merge_text_versions(base_content, incoming_content, current_content)
            merged = merged or merged_content != incoming_content or merged_content != current_content

        if merged_content == current_content:
            return {
                'success': True,
                'content': current_content,
                'revision': current_revision,
                'encoding': state.get('encoding', 'utf-8'),
                'newline': state.get('newline', '\n'),
                'mtime_ns': state.get('mtime_ns', get_file_mtime_token(filepath)),
                'size': get_file_size(int(state.get('file_size', 0) or 0)),
                'merged': merged,
                'saved': False
            }

        encoded_content, normalized_encoding, normalized_newline, normalized_content = encode_text_editor_content(
            merged_content,
            state.get('encoding', 'utf-8'),
            state.get('newline', '\n')
        )

        state['content'] = normalized_content
        state['revision'] = current_revision + 1
        state['encoding'] = normalized_encoding
        state['newline'] = normalized_newline
        state['file_size'] = len(encoded_content)
        state['last_activity'] = time.time()
        _remember_text_realtime_snapshot_locked(state)
        response_payload = {
            'success': True,
            'content': state['content'],
            'revision': int(state['revision']),
            'encoding': state['encoding'],
            'newline': state['newline'],
            'mtime_ns': state['mtime_ns'],
            'size': get_file_size(state['file_size']),
            'merged': merged,
            'saved': False
        }

    publish_document_realtime_event(
        filepath,
        'text_update',
        {
            'content': response_payload['content'],
            'revision': response_payload['revision'],
            'encoding': response_payload['encoding'],
            'newline': response_payload['newline'],
            'mtime_ns': response_payload['mtime_ns'],
            'author': username or '匿名用户'
        },
        exclude_client_id=client_id
    )
    return response_payload

def get_excel_collaboration_file_key(filepath):
    """将文件路径标准化为协同状态键"""
    return os.path.normcase(os.path.abspath(filepath))

def normalize_excel_collaboration_cell(cell_data):
    """Docstring."""
    if not isinstance(cell_data, dict):
        return None

    sheet_name = str(cell_data.get('sheet') or cell_data.get('sheet_name') or '').strip()
    if not sheet_name:
        return None

    try:
        row = int(cell_data.get('row'))
        col = int(cell_data.get('col'))
    except (TypeError, ValueError, AttributeError):
        return None

    if row < 1 or col < 1 or row > EXCEL_MAX_ROWS or col > EXCEL_MAX_COLS:
        return None

    return {
        'sheet': sheet_name,
        'row': row,
        'col': col,
        'cell': f'{get_column_letter(col)}{row}'
    }

def normalize_excel_collaboration_viewport(viewport):
    """Docstring."""
    if not isinstance(viewport, dict):
        return None

    try:
        start_row = int(viewport.get('start_row'))
        start_col = int(viewport.get('start_col'))
        end_row = int(viewport.get('end_row'))
        end_col = int(viewport.get('end_col'))
    except (TypeError, ValueError, AttributeError):
        return None

    if min(start_row, start_col, end_row, end_col) < 1:
        return None

    return {
        'start_row': start_row,
        'start_col': start_col,
        'end_row': max(start_row, end_row),
        'end_col': max(start_col, end_col)
    }

def cleanup_expired_excel_collaboration(now=None):
    """Docstring."""
    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)

def _cleanup_expired_excel_collaboration_locked(now=None):
    """在已持锁状态下清理过期的 Excel 协作数据。"""
    current_time = time.time() if now is None else float(now)
    file_keys_to_delete = []

    for file_key, file_state in excel_collaboration_state.items():
        sessions = file_state.get('sessions', {})
        locks = file_state.get('locks', {})

        expired_sessions = [
            client_id for client_id, session_data in sessions.items()
            if current_time - float(session_data.get('updated_at', 0)) > EXCEL_COLLAB_HEARTBEAT_TTL
        ]
        for client_id in expired_sessions:
            sessions.pop(client_id, None)

        expired_lock_keys = [
            lock_key for lock_key, lock_data in locks.items()
            if (
                current_time - float(lock_data.get('updated_at', 0)) > EXCEL_COLLAB_LOCK_TTL or
                lock_data.get('client_id') not in sessions
            )
        ]
        for lock_key in expired_lock_keys:
            locks.pop(lock_key, None)

        if not sessions and not locks:
            file_keys_to_delete.append(file_key)

    for file_key in file_keys_to_delete:
        excel_collaboration_state.pop(file_key, None)

def _release_excel_collaboration_locks_locked(file_state, client_id, except_lock_key=None):
    """Docstring."""
    locks = file_state.get('locks', {})
    owned_keys = [
        lock_key for lock_key, lock_data in locks.items()
        if lock_data.get('client_id') == client_id and lock_key != except_lock_key
    ]
    for lock_key in owned_keys:
        locks.pop(lock_key, None)

def build_excel_collaboration_snapshot(filepath, client_id):
    """Docstring."""
    file_key = get_excel_collaboration_file_key(filepath)
    now = time.time()

    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)
        file_state = excel_collaboration_state.setdefault(file_key, {'sessions': {}, 'locks': {}})

        editors = []
        for session_data in file_state.get('sessions', {}).values():
            active_cell = session_data.get('active_cell')
            editors.append({
                'client_id': session_data.get('client_id'),
                'username': session_data.get('username') or '匿名用户',
                'is_self': session_data.get('client_id') == client_id,
                'active_sheet': session_data.get('active_sheet') or '',
                'active_cell': active_cell,
                'active_cell_ref': active_cell.get('cell') if active_cell else '',
                'viewport': session_data.get('viewport'),
                'updated_at': session_data.get('updated_at', now)
            })

        editors.sort(key=lambda item: (not item['is_self'], item['username'], item['client_id'] or ''))

        locks = []
        own_lock = None
        for lock_data in file_state.get('locks', {}).values():
            payload = {
                'client_id': lock_data.get('client_id'),
                'username': lock_data.get('username') or '匿名用户',
                'sheet': lock_data.get('sheet'),
                'row': lock_data.get('row'),
                'col': lock_data.get('col'),
                'cell': lock_data.get('cell'),
                'updated_at': lock_data.get('updated_at', now),
                'is_self': lock_data.get('client_id') == client_id
            }
            locks.append(payload)
            if payload['is_self']:
                own_lock = payload

    return {
        'client_id': client_id,
        'mtime_ns': get_file_mtime_token(filepath),
        'editors': editors,
        'locks': locks,
        'own_lock': own_lock
    }

def sync_excel_collaboration_presence(filepath, username, client_id, ip_address, sheet_name='', active_cell=None, viewport=None, lock_cell=None, release_lock=False):
    """Docstring."""
    file_key = get_excel_collaboration_file_key(filepath)
    now = time.time()
    active_cell = normalize_excel_collaboration_cell(active_cell)
    lock_cell = normalize_excel_collaboration_cell(lock_cell)
    viewport = normalize_excel_collaboration_viewport(viewport)
    lock_denied = None

    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)
        file_state = excel_collaboration_state.setdefault(file_key, {'sessions': {}, 'locks': {}})
        sessions = file_state['sessions']
        session_data = sessions.get(client_id, {})
        desired_sheet = str(sheet_name or (active_cell or {}).get('sheet') or session_data.get('active_sheet') or '').strip()

        sessions[client_id] = {
            'client_id': client_id,
            'username': username or '匿名用户',
            'ip': ip_address or '',
            'active_sheet': desired_sheet,
            'active_cell': active_cell,
            'viewport': viewport,
            'updated_at': now
        }

        if release_lock or not lock_cell:
            _release_excel_collaboration_locks_locked(file_state, client_id)
        else:
            desired_lock_key = f"{lock_cell['sheet']}::{lock_cell['row']}::{lock_cell['col']}"
            _release_excel_collaboration_locks_locked(file_state, client_id, desired_lock_key)
            current_lock = file_state['locks'].get(desired_lock_key)

            if current_lock and current_lock.get('client_id') != client_id:
                lock_denied = {
                    'client_id': current_lock.get('client_id'),
                    'username': current_lock.get('username') or '匿名用户',
                    'sheet': current_lock.get('sheet'),
                    'row': current_lock.get('row'),
                    'col': current_lock.get('col'),
                    'cell': current_lock.get('cell')
                }
            else:
                file_state['locks'][desired_lock_key] = {
                    'client_id': client_id,
                    'username': username or '匿名用户',
                    'sheet': lock_cell['sheet'],
                    'row': lock_cell['row'],
                    'col': lock_cell['col'],
                    'cell': lock_cell['cell'],
                    'updated_at': now
                }

    snapshot = build_excel_collaboration_snapshot(filepath, client_id)
    snapshot['lock_denied'] = lock_denied
    snapshot['lock_acquired'] = lock_denied is None if lock_cell and not release_lock else True
    return snapshot

def release_excel_collaboration_presence(filepath, client_id, remove_session=False):
    """Docstring."""
    file_key = get_excel_collaboration_file_key(filepath)
    now = time.time()

    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)
        file_state = excel_collaboration_state.get(file_key)
        if not file_state:
            return

        _release_excel_collaboration_locks_locked(file_state, client_id)
        if remove_session:
            file_state.get('sessions', {}).pop(client_id, None)

        if not file_state.get('sessions') and not file_state.get('locks'):
            excel_collaboration_state.pop(file_key, None)

def get_excel_collaboration_other_locks(filepath, client_id):
    """返回当前文件中其他用户持有的单元格锁"""
    file_key = get_excel_collaboration_file_key(filepath)
    now = time.time()

    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)
        file_state = excel_collaboration_state.get(file_key)
        if not file_state:
            return []

        return [
            {
                'username': lock_data.get('username') or '匿名用户',
                'sheet': lock_data.get('sheet'),
                'row': lock_data.get('row'),
                'col': lock_data.get('col'),
                'cell': lock_data.get('cell'),
                'client_id': lock_data.get('client_id')
            }
            for lock_data in file_state.get('locks', {}).values()
            if lock_data.get('client_id') != client_id
        ]

def find_excel_locked_cells_for_changes(filepath, sheets, client_id):
    """检查即将保存的单元格是否被其他协作者锁定。"""
    file_key = get_excel_collaboration_file_key(filepath)
    now = time.time()
    requested_cells = set()

    for sheet_data in sheets or []:
        sheet_name = str(sheet_data.get('name') or '').strip()
        if not sheet_name:
            continue
        for cell in sheet_data.get('cells') or []:
            normalized_cell = normalize_excel_collaboration_cell({
                'sheet': sheet_name,
                'row': cell.get('row'),
                'col': cell.get('col')
            })
            if normalized_cell:
                requested_cells.add((normalized_cell['sheet'], normalized_cell['row'], normalized_cell['col']))

    if not requested_cells:
        return []

    with excel_collaboration_lock:
        _cleanup_expired_excel_collaboration_locked(now)
        file_state = excel_collaboration_state.get(file_key)
        if not file_state:
            return []

        blockers = []
        for lock_data in file_state.get('locks', {}).values():
            if lock_data.get('client_id') == client_id:
                continue
            lock_signature = (lock_data.get('sheet'), lock_data.get('row'), lock_data.get('col'))
            if lock_signature in requested_cells:
                blockers.append({
                    'username': lock_data.get('username') or '匿名用户',
                    'sheet': lock_data.get('sheet'),
                    'row': lock_data.get('row'),
                    'col': lock_data.get('col'),
                    'cell': lock_data.get('cell')
                })
        return blockers

def is_text_previewable_file(filename):
    """检查文件是否支持文本预览或编辑。"""
    if is_macos_metadata_file(filename):
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in TEXT_PREVIEW_EXTENSIONS

def is_word_editable_file(filename):
    """检查文件是否支持 DOCX 在线编辑。"""
    if is_macos_metadata_file(filename):
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in WORD_EDITABLE_EXTENSIONS

def is_excel_editable_file(filename):
    """检查文件是否支持 Excel 在线编辑。"""
    if is_macos_metadata_file(filename):
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return ext in EXCEL_EDITABLE_EXTENSIONS

def is_previewable_file(filename):
    """检查文件是否支持预览或在线编辑。"""
    if is_macos_metadata_file(filename):
        return False
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    return (
        ext in DIRECT_PREVIEW_EXTENSIONS or
        is_text_previewable_file(filename) or
        is_word_editable_file(filename) or
        is_excel_editable_file(filename)
    )

def guess_inline_mimetype(filename):
    """为音视频内联预览推断更合适的 MIME 类型"""
    ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
    overrides = {
        'wav': 'audio/wav',
        'mp3': 'audio/mpeg',
        'flac': 'audio/flac',
        'm4a': 'audio/mp4',
        'aac': 'audio/aac',
        'ogg': 'audio/ogg',
        'mp4': 'video/mp4',
        'webm': 'video/webm',
        'pdf': 'application/pdf'
    }
    mime = overrides.get(ext)
    if mime:
        return mime
    guessed, _ = mimetypes.guess_type(filename)
    return guessed or 'application/octet-stream'

def should_use_excel_com(filepath):
    """判断 Excel 文件是否需要走本机 Excel/WPS 兼容模式。"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        return True
    if ext in {'.xlsx', '.xlsm'} and zipfile.is_zipfile(filepath):
        return False
    return True

def can_use_excel_com():
    """检测当前环境是否可用 Excel/WPS COM 组件。"""
    global excel_com_available_cache
    if excel_com_available_cache is not None:
        return excel_com_available_cache

    pythoncom_module = None
    excel = None
    try:
        import pythoncom as imported_pythoncom
        import win32com.client
        pythoncom_module = imported_pythoncom
        pythoncom_module.CoInitialize()
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel_com_available_cache = True
    except Exception:
        excel_com_available_cache = False
    finally:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        if pythoncom_module is not None:
            try:
                pythoncom_module.CoUninitialize()
            except Exception:
                pass

    return excel_com_available_cache

def should_prefer_excel_com_for_save(filepath):
    """大文件保存时优先走本机 Excel/WPS 组件，减少 openpyxl 全量重写的耗时。"""
    if should_use_excel_com(filepath):
        return True

    ext = os.path.splitext(filepath)[1].lower()
    if ext not in {'.xlsx', '.xlsm'}:
        return False

    try:
        file_size = os.path.getsize(filepath)
    except OSError:
        return False

    return file_size > MAX_EDITABLE_EXCEL_FILE_SIZE and can_use_excel_com()

def format_excel_cell_value(value):
    """将 Excel 单元格值格式化为前端可编辑文本。"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d %H:%M:%S')
    return str(value)

def get_excel_window_limits(file_size, backend):
    """Docstring."""
    if backend == 'excel_com' or file_size > MAX_EDITABLE_EXCEL_FILE_SIZE:
        return HEAVY_EXCEL_PREVIEW_ROWS, HEAVY_EXCEL_PREVIEW_COLS
    return DEFAULT_EXCEL_PREVIEW_ROWS, DEFAULT_EXCEL_PREVIEW_COLS

def normalize_excel_rows(rows_iterable):
    """Docstring."""
    rows = []
    for row in rows_iterable:
        if isinstance(row, tuple):
            rows.append([format_excel_cell_value(value) for value in row])
        elif isinstance(row, list):
            rows.append([format_excel_cell_value(value) for value in row])
        else:
            rows.append([format_excel_cell_value(row)])
    return rows

def normalize_excel_range_values(raw_values, row_count, col_count):
    """Docstring."""
    if row_count <= 0 or col_count <= 0:
        return []

    if row_count == 1 and col_count == 1:
        return [[format_excel_cell_value(raw_values)]]

    if row_count == 1:
        if not isinstance(raw_values, (list, tuple)):
            raw_values = (raw_values,)
        return [[format_excel_cell_value(value) for value in list(raw_values)[:col_count]]]

    if col_count == 1:
        if not isinstance(raw_values, (list, tuple)):
            raw_values = (raw_values,)
        rows = []
        for item in list(raw_values)[:row_count]:
            cell_value = item[0] if isinstance(item, (list, tuple)) and item else item
            rows.append([format_excel_cell_value(cell_value)])
        return rows

    rows = []
    for row in list(raw_values)[:row_count]:
        if isinstance(row, (list, tuple)):
            rows.append([format_excel_cell_value(value) for value in list(row)[:col_count]])
        else:
            rows.append([format_excel_cell_value(row)])
    return rows

def _create_excel_com_app():
    """创建Excel COM实例"""
    try:
        import pythoncom
        import win32com.client
    except ImportError as e:
        raise RuntimeError('当前系统缺少 Excel 兼容组件，无法打开该表格。') from e

    pythoncom.CoInitialize()
    excel = None
    try:
        excel = win32com.client.DispatchEx('Excel.Application')
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False
        return pythoncom, excel
    except Exception:
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()
        raise

def _close_excel_com_app(pythoncom_module, excel):
    """关闭Excel COM实例"""
    try:
        if excel is not None:
            excel.Quit()
    finally:
        pythoncom_module.CoUninitialize()

def open_excel_com_workbook(excel, filepath, read_only=True):
    """Docstring."""
    try:
        excel.Calculation = -4135  # xlCalculationManual
        excel.CalculateBeforeSave = False
    except Exception:
        pass
    return excel.Workbooks.Open(
        os.path.abspath(filepath),
        UpdateLinks=0,
        ReadOnly=read_only,
        IgnoreReadOnlyRecommended=True,
        AddToMru=False,
        Notify=False
    )

def read_excel_range_via_com(worksheet, start_row, start_col, end_row, end_col):
    """使用 COM 整块读取单元格，避免逐格访问带来的性能问题。"""
    cell_range = worksheet.Range(worksheet.Cells(start_row, start_col), worksheet.Cells(end_row, end_col))
    return cell_range.Value

def open_excel_workbook(filepath, read_only=False):
    """Docstring."""
    keep_vba = filepath.lower().endswith('.xlsm')
    return load_workbook(filepath, keep_vba=keep_vba, read_only=read_only, data_only=False, keep_links=False)

def save_openpyxl_workbook_fast(workbook, target_path, source_file_size=0):
    """用更低压缩等级保存大表，减少 openpyxl 写盘时间。"""
    compresslevel = 1 if int(source_file_size or 0) > MAX_EDITABLE_EXCEL_FILE_SIZE else 6
    with zipfile.ZipFile(
        target_path,
        'w',
        compression=zipfile.ZIP_DEFLATED,
        allowZip64=True,
        compresslevel=compresslevel
    ) as archive:
        writer = ExcelWriter(workbook, archive)
        writer.save()

def detect_newline_style(raw_bytes):
    """Docstring."""
    if b'\r\n' in raw_bytes:
        return '\r\n'
    if b'\r' in raw_bytes:
        return '\r'
    return '\n'

def decode_text_bytes(raw_bytes):
    """Docstring."""
    for encoding in TEXT_EDIT_ENCODINGS:
        try:
            return raw_bytes.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode('utf-8', errors='replace'), 'utf-8'

def load_text_file_preview(filepath):
    """读取文本文件，用于浏览器预览和编辑。"""
    file_size = os.path.getsize(filepath)
    editable = file_size <= MAX_EDITABLE_TEXT_FILE_SIZE
    preview_size = file_size if editable else min(file_size, TEXT_PREVIEW_FALLBACK_BYTES)
    truncated = preview_size < file_size

    with open(filepath, 'rb') as f:
        raw_bytes = f.read(preview_size)

    content, encoding = decode_text_bytes(raw_bytes)
    return {
        'content': content,
        'encoding': encoding,
        'newline': detect_newline_style(raw_bytes),
        'editable': editable,
        'truncated': truncated,
        'file_size': file_size,
        'mtime_ns': os.stat(filepath).st_mtime_ns
    }

def iter_docx_body_blocks(document):
    """Docstring."""
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield 'paragraph', DocxParagraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield 'table', DocxTable(child, document)

def normalize_docx_text(value):
    """Docstring."""
    if value is None:
        return ''
    return str(value).replace('\r\n', '\n').replace('\r', '\n')

def docx_xml_local_name(element):
    """返回 lxml/docx XML 节点的本地标签名"""
    tag = getattr(element, 'tag', '')
    if not isinstance(tag, str):
        return ''
    return tag.split('}', 1)[-1] if '}' in tag else tag

def get_docx_style_name(item):
    """Docstring."""
    try:
        return item.style.name if item.style else ''
    except Exception:
        return ''

def build_docx_target_label(scope_label, block_label):
    """为协同提示生成更可读的块标签"""
    scope_label = str(scope_label or '').strip()
    block_label = str(block_label or '').strip()
    if scope_label and block_label:
        return f'{scope_label} / {block_label}'
    return block_label or scope_label

def build_docx_image_meta_target_key(block_id, field_name):
    """为 DOCX 图片说明字段生成稳定的目标键。"""
    field_name = str(field_name or '').strip().lower()
    if field_name not in {'title', 'description'}:
        field_name = 'description'
    return f'image_meta::{str(block_id or "").strip()}::{field_name}'

def get_docx_story_root_element(container, story_kind='body'):
    """杩斿洖姝ｆ枃/椤电湁/椤佃剼/鎵规敞绛夋晠浜嬪鍣ㄧ殑鏍?XML 鑺傜偣"""
    if story_kind == 'body':
        return container.element.body
    return getattr(container, '_element', None)

def iter_docx_story_blocks(container, story_kind='body'):
    """Docstring."""
    if story_kind == 'body':
        yield from iter_docx_body_blocks(container)
        return

    if hasattr(container, 'iter_inner_content'):
        for item in container.iter_inner_content():
            if isinstance(item, DocxParagraph):
                yield 'paragraph', item
            elif isinstance(item, DocxTable):
                yield 'table', item

def build_docx_story_sources(document):
    """Docstring."""
    stories = [{
        'story_id': 'body',
        'story_kind': 'body',
        'scope_label': '正文',
        'scope_meta': '',
        'container': document,
        'root_element': get_docx_story_root_element(document, 'body')
    }]

    header_seen = {}
    footer_seen = {}

    for section_index, section in enumerate(document.sections, start=1):
        header = section.header
        header_part_name = str(getattr(header.part, 'partname', '') or f'/word/header-section-{section_index}.xml')
        header_token = re.sub(r'[^0-9A-Za-z_-]+', '_', os.path.splitext(os.path.basename(header_part_name))[0]).strip('_') or f'header_{section_index}'
        if header_token not in header_seen:
            header_seen[header_token] = len(header_seen) + 1
            stories.append({
                'story_id': f'header-{header_token}',
                'story_kind': 'header',
                'scope_label': f'页眉 {header_seen[header_token]}',
                'scope_meta': f'第 {section_index} 节',
                'container': header,
                'root_element': get_docx_story_root_element(header, 'header')
            })

        footer = section.footer
        footer_part_name = str(getattr(footer.part, 'partname', '') or f'/word/footer-section-{section_index}.xml')
        footer_token = re.sub(r'[^0-9A-Za-z_-]+', '_', os.path.splitext(os.path.basename(footer_part_name))[0]).strip('_') or f'footer_{section_index}'
        if footer_token not in footer_seen:
            footer_seen[footer_token] = len(footer_seen) + 1
            stories.append({
                'story_id': f'footer-{footer_token}',
                'story_kind': 'footer',
                'scope_label': f'页脚 {footer_seen[footer_token]}',
                'scope_meta': f'第 {section_index} 节',
                'container': footer,
                'root_element': get_docx_story_root_element(footer, 'footer')
            })

    for comment_index, comment in enumerate(list(document.comments), start=1):
        author = str(comment.author or '').strip()
        scope_meta = author or ''
        stories.append({
            'story_id': f'comment-{int(comment.comment_id)}',
            'story_kind': 'comment',
            'scope_label': f'批注 {comment_index}',
            'scope_meta': scope_meta,
            'container': comment,
            'root_element': get_docx_story_root_element(comment, 'comment')
        })

    return stories

def build_docx_table_rows(table):
    """Docstring."""
    rows = []
    col_count = 0
    for row_index, row in enumerate(table.rows):
        row_cells = []
        for col_index, cell in enumerate(row.cells):
            row_cells.append({
                'row': row_index,
                'col': col_index,
                'label': f'{get_column_letter(col_index + 1)}{row_index + 1}',
                'text': cell.text or ''
            })
        col_count = max(col_count, len(row_cells))
        rows.append(row_cells)
    return rows, len(rows), col_count

def collect_docx_textbox_entries_for_story(story):
    """提取故事容器里的文本框段落和表格"""
    root_element = story.get('root_element')
    container = story.get('container')
    story_id = story.get('story_id')
    scope_label = story.get('scope_label')

    if root_element is None or container is None:
        return [], 0, 0, 0

    entries = []
    paragraph_total = 0
    table_total = 0
    textbox_count = 0

    for textbox_index, textbox in enumerate(root_element.xpath('.//*[local-name()="txbxContent"]'), start=1):
        textbox_count += 1
        textbox_scope = build_docx_target_label(scope_label, f'文本框 {textbox_index}')
        textbox_paragraph_index = 0
        textbox_table_index = 0

        for child in textbox.iterchildren():
            local_name = docx_xml_local_name(child)
            if local_name == 'p':
                paragraph = DocxParagraph(child, container)
                textbox_paragraph_index += 1
                paragraph_total += 1
                block_label = f'段落 {textbox_paragraph_index}'
                entries.append({
                    'id': f'{story_id}-tbx-{textbox_index}-p-{textbox_paragraph_index - 1}',
                    'label': block_label,
                    'target_label': build_docx_target_label(textbox_scope, block_label),
                    'scope_label': textbox_scope,
                    'scope_meta': story.get('scope_meta', ''),
                    'story_kind': 'textbox',
                    'type': 'paragraph',
                    'text': paragraph.text or '',
                    'style': get_docx_style_name(paragraph),
                    '_object': paragraph
                })
                continue

            if local_name != 'tbl':
                continue

            table = DocxTable(child, container)
            textbox_table_index += 1
            table_total += 1
            block_label = f'表格 {textbox_table_index}'
            rows, row_count, col_count = build_docx_table_rows(table)
            entries.append({
                'id': f'{story_id}-tbx-{textbox_index}-t-{textbox_table_index - 1}',
                'label': block_label,
                'target_label': build_docx_target_label(textbox_scope, block_label),
                'scope_label': textbox_scope,
                'scope_meta': story.get('scope_meta', ''),
                'story_kind': 'textbox',
                'type': 'table',
                'rows': rows,
                'row_count': row_count,
                'col_count': col_count,
                'style': get_docx_style_name(table),
                '_object': table
            })

    return entries, textbox_count, paragraph_total, table_total

def collect_docx_image_entries_for_story(story):
    """提取故事容器里的图片标题/说明字段"""
    root_element = story.get('root_element')
    if root_element is None:
        return [], 0

    entries = []
    image_count = 0

    for image_index, image_element in enumerate(root_element.xpath('.//*[local-name()="drawing" or local-name()="pict"]'), start=1):
        doc_prop_elements = list(image_element.xpath('.//*[local-name()="docPr"]'))
        cnv_prop_elements = list(image_element.xpath('.//*[local-name()="cNvPr"]'))
        attr_elements = [element for element in (doc_prop_elements + cnv_prop_elements) if element is not None]
        if not attr_elements:
            continue

        image_count += 1
        label = f'图片 {image_count}'
        name = ''
        title = ''
        description = ''
        for attr_element in attr_elements:
            if not name:
                name = str(attr_element.get('name') or '').strip()
            if not title:
                title = str(attr_element.get('title') or '').strip()
            if not description:
                description = str(attr_element.get('descr') or '').strip()

        entries.append({
            'id': f"{story.get('story_id')}-img-{image_index}",
            'label': label,
            'target_label': build_docx_target_label(story.get('scope_label'), label),
            'scope_label': story.get('scope_label'),
            'scope_meta': story.get('scope_meta', ''),
            'story_kind': 'image',
            'type': 'image_meta',
            'name': name or label,
            'title': title,
            'description': description,
            '_attr_elements': attr_elements
        })

    return entries, image_count

def collect_docx_editable_entries(document):
    """将 DOCX 中可在线编辑的内容统一整理为块列表。"""
    entries = []
    stats = {
        'paragraph_count': 0,
        'table_count': 0,
        'header_count': 0,
        'footer_count': 0,
        'comment_count': 0,
        'textbox_count': 0,
        'image_count': 0
    }

    for story in build_docx_story_sources(document):
        story_kind = story.get('story_kind')
        if story_kind == 'header':
            stats['header_count'] += 1
        elif story_kind == 'footer':
            stats['footer_count'] += 1
        elif story_kind == 'comment':
            stats['comment_count'] += 1

        paragraph_index = 0
        table_index = 0
        for block_type, block in iter_docx_story_blocks(story.get('container'), story_kind):
            if block_type == 'paragraph':
                paragraph_index += 1
                stats['paragraph_count'] += 1
                block_label = f'段落 {paragraph_index}'
                paragraph_text = normalize_docx_text(block.text or '')
                is_placeholder = story_kind in {'header', 'footer'} and not paragraph_text.strip()
                entries.append({
                    'id': f"{story.get('story_id')}-p-{paragraph_index - 1}",
                    'label': block_label,
                    'target_label': build_docx_target_label(story.get('scope_label'), block_label),
                    'scope_label': story.get('scope_label'),
                    'scope_meta': story.get('scope_meta', ''),
                    'story_kind': story_kind,
                    'type': 'paragraph',
                    'text': '' if is_placeholder else paragraph_text,
                    'style': get_docx_style_name(block),
                    'is_placeholder': is_placeholder,
                    'placeholder_text': (
                        f"{story.get('scope_label')} 当前为空，直接输入即可。"
                        if is_placeholder else ''
                    ),
                    '_object': block
                })
                continue

            table_index += 1
            stats['table_count'] += 1
            block_label = f'表格 {table_index}'
            rows, row_count, col_count = build_docx_table_rows(block)
            entries.append({
                'id': f"{story.get('story_id')}-t-{table_index - 1}",
                'label': block_label,
                'target_label': build_docx_target_label(story.get('scope_label'), block_label),
                'scope_label': story.get('scope_label'),
                'scope_meta': story.get('scope_meta', ''),
                'story_kind': story_kind,
                'type': 'table',
                'rows': rows,
                'row_count': row_count,
                'col_count': col_count,
                'style': get_docx_style_name(block),
                '_object': block
            })

        textbox_entries, textbox_count, textbox_paragraphs, textbox_tables = collect_docx_textbox_entries_for_story(story)
        entries.extend(textbox_entries)
        stats['textbox_count'] += textbox_count
        stats['paragraph_count'] += textbox_paragraphs
        stats['table_count'] += textbox_tables

        image_entries, image_count = collect_docx_image_entries_for_story(story)
        entries.extend(image_entries)
        stats['image_count'] += image_count

    return entries, stats

def apply_docx_image_meta_changes(entry, title=None, description=None):
    """Docstring."""
    for element in entry.get('_attr_elements', []):
        if title is not None:
            if title:
                element.set('title', title)
            elif 'title' in element.attrib:
                del element.attrib['title']

        if description is not None:
            if description:
                element.set('descr', description)
            elif 'descr' in element.attrib:
                del element.attrib['descr']

def load_docx_file_preview(filepath):
    """读取 DOCX 文档用于在线预览/编辑"""
    file_size = os.path.getsize(filepath)
    editable = file_size <= MAX_EDITABLE_DOCX_FILE_SIZE
    document = WordDocument(filepath)
    entries, stats = collect_docx_editable_entries(document)
    blocks = []
    for entry in entries:
        block = {key: value for key, value in entry.items() if not key.startswith('_')}
        blocks.append(block)

    return {
        'blocks': blocks,
        'editable': editable,
        'file_size': file_size,
        'mtime_ns': os.stat(filepath).st_mtime_ns,
        'paragraph_count': stats['paragraph_count'],
        'table_count': stats['table_count'],
        'header_count': stats['header_count'],
        'footer_count': stats['footer_count'],
        'comment_count': stats['comment_count'],
        'textbox_count': stats['textbox_count'],
        'image_count': stats['image_count'],
        'block_count': len(blocks)
    }

def save_docx_file_content(filepath, blocks):
    """Docstring."""
    document = WordDocument(filepath)
    payload_map = {}

    for block in blocks or []:
        if not isinstance(block, dict):
            continue
        block_id = str(block.get('id') or '').strip()
        if block_id:
            payload_map[block_id] = block

    entries, stats = collect_docx_editable_entries(document)
    changed_paragraphs = 0
    changed_cells = 0
    changed_image_notes = 0
    pending_changes = []
    conflicts = []

    for entry in entries:
        block_id = entry['id']
        payload = payload_map.get(block_id)
        if not payload:
            continue

        if entry['type'] == 'paragraph':
            incoming_text = normalize_docx_text(payload.get('text'))
            original_text = normalize_docx_text(payload.get('original_text'))
            current_text = normalize_docx_text(entry['_object'].text)
            if incoming_text == original_text:
                continue
            if current_text != original_text and current_text != incoming_text:
                conflicts.append({
                    'key': build_docx_paragraph_target_key(block_id),
                    'label': str(payload.get('label') or entry.get('target_label') or block_id)
                })
                continue

            if current_text != incoming_text:
                pending_changes.append(('paragraph', entry['_object'], incoming_text))
                changed_paragraphs += 1
            continue

        if entry['type'] == 'table':
            table_label = str(payload.get('label') or entry.get('target_label') or block_id)
            row_payloads = payload.get('rows') if isinstance(payload.get('rows'), list) else []
            for row_index, row in enumerate(entry['_object'].rows):
                payload_row = row_payloads[row_index] if row_index < len(row_payloads) and isinstance(row_payloads[row_index], list) else []
                for col_index, cell in enumerate(row.cells):
                    payload_cell = payload_row[col_index] if col_index < len(payload_row) and isinstance(payload_row[col_index], dict) else {}
                    incoming_text = normalize_docx_text(payload_cell.get('text'))
                    original_text = normalize_docx_text(payload_cell.get('original_text'))
                    current_text = normalize_docx_text(cell.text)
                    if incoming_text == original_text:
                        continue
                    if current_text != original_text and current_text != incoming_text:
                        cell_label = str(payload_cell.get('label') or f'{get_column_letter(col_index + 1)}{row_index + 1}')
                        conflicts.append({
                            'key': build_docx_table_cell_target_key(block_id, row_index, col_index),
                            'label': f'{table_label} / {cell_label}'
                        })
                        continue

                    if current_text != incoming_text:
                        pending_changes.append(('table_cell', cell, incoming_text))
                        changed_cells += 1
            continue

        if entry['type'] != 'image_meta':
            continue

        payload_label = str(payload.get('label') or entry.get('target_label') or block_id)
        current_title = normalize_docx_text(entry.get('title'))
        current_description = normalize_docx_text(entry.get('description'))
        incoming_title = normalize_docx_text(payload.get('title'))
        incoming_description = normalize_docx_text(payload.get('description'))
        original_title = normalize_docx_text(payload.get('original_title'))
        original_description = normalize_docx_text(payload.get('original_description'))

        title_conflict = (
            incoming_title != original_title and
            current_title != original_title and
            current_title != incoming_title
        )
        description_conflict = (
            incoming_description != original_description and
            current_description != original_description and
            current_description != incoming_description
        )
        if title_conflict:
            conflicts.append({
                'key': build_docx_image_meta_target_key(block_id, 'title'),
                'label': f'{payload_label} / 标题'
            })
        if description_conflict:
            conflicts.append({
                'key': build_docx_image_meta_target_key(block_id, 'description'),
                'label': f'{payload_label} / 说明'
            })
        if title_conflict or description_conflict:
            continue

        next_title = incoming_title if incoming_title != current_title else None
        next_description = incoming_description if incoming_description != current_description else None
        if next_title is not None or next_description is not None:
            pending_changes.append(('image_meta', entry, {
                'title': next_title,
                'description': next_description
            }))
            if next_title is not None:
                changed_image_notes += 1
            if next_description is not None:
                changed_image_notes += 1

    if conflicts:
        return {
            'changed_paragraphs': 0,
            'changed_cells': 0,
            'changed_image_notes': 0,
            'paragraph_count': stats['paragraph_count'],
            'table_count': stats['table_count'],
            'header_count': stats['header_count'],
            'footer_count': stats['footer_count'],
            'comment_count': stats['comment_count'],
            'textbox_count': stats['textbox_count'],
            'image_count': stats['image_count'],
            'conflicts': conflicts
        }

    if pending_changes:
        for change_type, target, new_value in pending_changes:
            if change_type == 'image_meta':
                apply_docx_image_meta_changes(
                    target,
                    title=new_value.get('title'),
                    description=new_value.get('description')
                )
                continue

            target.text = new_value

        temp_path = f"{filepath}.editing.{uuid.uuid4().hex}.tmp"
        try:
            document.save(temp_path)
            os.replace(temp_path, filepath)
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

    return {
        'changed_paragraphs': changed_paragraphs,
        'changed_cells': changed_cells,
        'changed_image_notes': changed_image_notes,
        'paragraph_count': stats['paragraph_count'],
        'table_count': stats['table_count'],
        'header_count': stats['header_count'],
        'footer_count': stats['footer_count'],
        'comment_count': stats['comment_count'],
        'textbox_count': stats['textbox_count'],
        'image_count': stats['image_count'],
        'conflicts': []
    }

def get_file_mtime_token(filepath):
    """返回适合前后端传递的文件版本令牌"""
    return str(os.stat(filepath).st_mtime_ns)

def parse_excel_input_value(value):
    """尽量将前端输入还原为常见Excel类型"""
    if value is None:
        return None

    if not isinstance(value, str):
        return value

    if value == '':
        return None

    stripped = value.strip()
    if stripped == '':
        return value

    if value.startswith('='):
        return value

    lowered = stripped.lower()
    if lowered == 'true':
        return True
    if lowered == 'false':
        return False

    if re.fullmatch(r'-?(0|[1-9]\d*)', stripped):
        try:
            return int(stripped)
        except ValueError:
            pass

    if re.fullmatch(r'-?(0|[1-9]\d*)\.\d+', stripped):
        try:
            return float(stripped)
        except ValueError:
            pass

    return value

def clamp_excel_structure_amount(value, default=1):
    """Docstring."""
    return clamp_excel_position(value, 1, MAX_EXCEL_STRUCTURE_AMOUNT, default)

def normalize_excel_sheet_name(name):
    """Docstring."""
    sheet_name = str(name or '').strip()
    if not sheet_name:
        raise ValueError('工作表名称不能为空。')
    if len(sheet_name) > 31:
        raise ValueError('工作表名称不能超过 31 个字符。')
    if INVALID_EXCEL_SHEET_NAME_RE.search(sheet_name):
        raise ValueError('工作表名称不能包含: \\ / ? * [ ]')
    return sheet_name

def build_unique_excel_sheet_name(existing_names, desired_name='Sheet'):
    """Docstring."""
    existing = {str(name) for name in existing_names}
    base_name = normalize_excel_sheet_name(desired_name)
    if base_name not in existing:
        return base_name

    index = 2
    while True:
        suffix = f'_{index}'
        candidate = f'{base_name[:31 - len(suffix)]}{suffix}'
        if candidate not in existing:
            return candidate
        index += 1

def choose_preview_sheet_names(all_sheet_names, preferred_sheet_name=None):
    """Docstring."""
    sheet_names = [str(name) for name in all_sheet_names]
    visible = sheet_names[:MAX_EXCEL_PREVIEW_SHEETS]
    preferred = str(preferred_sheet_name or '').strip()

    if (
        preferred and
        preferred in sheet_names and
        preferred not in visible and
        MAX_EXCEL_PREVIEW_SHEETS > 0
    ):
        visible = visible[:-1] + [preferred] if visible else [preferred]

    return visible

def clamp_excel_position(value, minimum, maximum, default):
    """限制Excel行列范围"""
    try:
        value = int(value)
    except (TypeError, ValueError):
        return default
    return max(minimum, min(value, maximum))

def build_excel_sheet_chunk(sheet_name, row_count, col_count, start_row, start_col, rows):
    """统一构建Excel分块响应"""
    visible_row_count = len(rows)
    visible_col_count = len(rows[0]) if rows else min(MAX_EXCEL_PREVIEW_COLS, max(1, col_count - start_col + 1))
    end_row = min(row_count, start_row + visible_row_count - 1) if visible_row_count else start_row
    end_col = min(col_count, start_col + visible_col_count - 1) if visible_col_count else start_col

    return {
        'name': sheet_name,
        'start_row': start_row,
        'start_col': start_col,
        'end_row': end_row,
        'end_col': end_col,
        'max_row': visible_row_count,
        'max_col': visible_col_count,
        'columns': [get_column_letter(col_idx) for col_idx in range(start_col, end_col + 1)] if visible_col_count else [],
        'rows': rows,
        'row_count': row_count,
        'col_count': col_count,
        'truncated': row_count > end_row or col_count > end_col,
        'has_prev_rows': start_row > 1,
        'has_next_rows': end_row < row_count,
        'has_prev_cols': start_col > 1,
        'has_next_cols': end_col < col_count
    }

def build_excel_sheet_chunk_meta(sheet_name, row_count, col_count, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """构建不含单元格内容的轻量工作表视窗元数据。"""
    row_count = max(1, row_count or 1)
    col_count = max(1, col_count or 1)
    start_row = clamp_excel_position(start_row, 1, row_count, 1)
    start_col = clamp_excel_position(start_col, 1, col_count, 1)
    row_limit = clamp_excel_position(row_limit, 1, MAX_EXCEL_PREVIEW_ROWS, MAX_EXCEL_PREVIEW_ROWS)
    col_limit = clamp_excel_position(col_limit, 1, MAX_EXCEL_PREVIEW_COLS, MAX_EXCEL_PREVIEW_COLS)
    visible_row_count = min(row_limit, max(1, row_count - start_row + 1))
    visible_col_count = min(col_limit, max(1, col_count - start_col + 1))
    end_row = min(row_count, start_row + visible_row_count - 1)
    end_col = min(col_count, start_col + visible_col_count - 1)

    return {
        'name': str(sheet_name),
        'start_row': start_row,
        'start_col': start_col,
        'end_row': end_row,
        'end_col': end_col,
        'max_row': visible_row_count,
        'max_col': visible_col_count,
        'row_count': row_count,
        'col_count': col_count,
        'truncated': row_count > end_row or col_count > end_col,
        'has_prev_rows': start_row > 1,
        'has_next_rows': end_row < row_count,
        'has_prev_cols': start_col > 1,
        'has_next_cols': end_col < col_count
    }

def build_excel_sheet_chunk_from_openpyxl_worksheet(worksheet, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    row_count = max(1, worksheet.max_row or 1)
    col_count = max(1, worksheet.max_column or 1)
    start_row = clamp_excel_position(start_row, 1, row_count, 1)
    start_col = clamp_excel_position(start_col, 1, col_count, 1)
    row_limit = clamp_excel_position(row_limit, 1, MAX_EXCEL_PREVIEW_ROWS, MAX_EXCEL_PREVIEW_ROWS)
    col_limit = clamp_excel_position(col_limit, 1, MAX_EXCEL_PREVIEW_COLS, MAX_EXCEL_PREVIEW_COLS)
    end_row = min(row_count, start_row + row_limit - 1)
    end_col = min(col_count, start_col + col_limit - 1)
    rows = normalize_excel_rows(
        worksheet.iter_rows(
            min_row=start_row,
            max_row=end_row,
            min_col=start_col,
            max_col=end_col,
            values_only=True
        )
    )
    return build_excel_sheet_chunk(str(worksheet.title), row_count, col_count, start_row, start_col, rows)

def build_excel_sheet_meta_from_openpyxl_worksheet(worksheet, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    row_count = max(1, worksheet.max_row or 1)
    col_count = max(1, worksheet.max_column or 1)
    return build_excel_sheet_chunk_meta(str(worksheet.title), row_count, col_count, start_row, start_col, row_limit, col_limit)

def build_excel_sheet_chunk_from_com_worksheet(worksheet, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    used_range = worksheet.UsedRange
    used_start_row = int(used_range.Row or 1)
    used_start_col = int(used_range.Column or 1)
    used_row_count = int(used_range.Rows.Count or 1)
    used_col_count = int(used_range.Columns.Count or 1)
    row_count = max(1, used_start_row + used_row_count - 1)
    col_count = max(1, used_start_col + used_col_count - 1)
    start_row = clamp_excel_position(start_row, 1, row_count, 1)
    start_col = clamp_excel_position(start_col, 1, col_count, 1)
    row_limit = clamp_excel_position(row_limit, 1, MAX_EXCEL_PREVIEW_ROWS, MAX_EXCEL_PREVIEW_ROWS)
    col_limit = clamp_excel_position(col_limit, 1, MAX_EXCEL_PREVIEW_COLS, MAX_EXCEL_PREVIEW_COLS)
    end_row = min(row_count, start_row + row_limit - 1)
    end_col = min(col_count, start_col + col_limit - 1)
    raw_values = read_excel_range_via_com(worksheet, start_row, start_col, end_row, end_col)
    rows = normalize_excel_range_values(raw_values, end_row - start_row + 1, end_col - start_col + 1)
    return build_excel_sheet_chunk(str(worksheet.Name), row_count, col_count, start_row, start_col, rows)

def build_excel_sheet_meta_from_com_worksheet(worksheet, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    used_range = worksheet.UsedRange
    used_start_row = int(used_range.Row or 1)
    used_start_col = int(used_range.Column or 1)
    used_row_count = int(used_range.Rows.Count or 1)
    used_col_count = int(used_range.Columns.Count or 1)
    row_count = max(1, used_start_row + used_row_count - 1)
    col_count = max(1, used_start_col + used_col_count - 1)
    return build_excel_sheet_chunk_meta(str(worksheet.Name), row_count, col_count, start_row, start_col, row_limit, col_limit)

def _load_excel_preview_via_openpyxl(filepath, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS, preferred_sheet_name=None):
    """使用openpyxl读取标准OOXML表格"""
    workbook = open_excel_workbook(filepath, read_only=True)
    total_sheet_count = len(workbook.sheetnames)
    sheet_names = choose_preview_sheet_names(workbook.sheetnames, preferred_sheet_name)
    sheets = []
    truncated = total_sheet_count > MAX_EXCEL_PREVIEW_SHEETS

    for sheet_name in sheet_names:
        ws = workbook[sheet_name]
        max_row = min(max(1, ws.max_row or 1), row_limit)
        max_col = min(max(1, ws.max_column or 1), col_limit)
        rows = normalize_excel_rows(
            ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)
        )

        sheets.append({
            'name': sheet_name,
            'start_row': 1,
            'start_col': 1,
            'end_row': max_row,
            'end_col': max_col,
            'max_row': max_row,
            'max_col': max_col,
            'columns': [get_column_letter(col_idx) for col_idx in range(1, max_col + 1)],
            'rows': rows,
            'row_count': ws.max_row or 0,
            'col_count': ws.max_column or 0,
            'truncated': (ws.max_row or 0) > row_limit or (ws.max_column or 0) > col_limit
        })

    workbook.close()
    return {
        'sheets': sheets,
        'sheet_count': total_sheet_count,
        'truncated': truncated,
        'backend': 'openpyxl'
    }

def _load_excel_sheet_chunk_via_openpyxl(filepath, sheet_name, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    workbook = open_excel_workbook(filepath, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f'工作表不存在: {sheet_name}')

        return build_excel_sheet_chunk_from_openpyxl_worksheet(
            workbook[sheet_name],
            start_row,
            start_col,
            row_limit,
            col_limit
        )
    finally:
        workbook.close()

def _load_excel_preview_via_com(filepath, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS, preferred_sheet_name=None):
    """使用本机Excel兼容模式读取表格"""
    pythoncom_module, excel = _create_excel_com_app()
    workbook = None
    try:
        workbook = open_excel_com_workbook(excel, filepath, read_only=True)
        total_sheet_count = workbook.Worksheets.Count
        all_sheet_names = [str(workbook.Worksheets(i).Name) for i in range(1, total_sheet_count + 1)]
        sheet_names = choose_preview_sheet_names(all_sheet_names, preferred_sheet_name)
        sheets = []

        for sheet_name in sheet_names:
            ws = workbook.Worksheets(sheet_name)
            used_range = ws.UsedRange

            used_start_row = int(used_range.Row or 1)
            used_start_col = int(used_range.Column or 1)
            used_row_count = int(used_range.Rows.Count or 1)
            used_col_count = int(used_range.Columns.Count or 1)

            actual_max_row = max(1, used_start_row + used_row_count - 1)
            actual_max_col = max(1, used_start_col + used_col_count - 1)
            max_row = min(actual_max_row, row_limit)
            max_col = min(actual_max_col, col_limit)
            raw_values = read_excel_range_via_com(ws, 1, 1, max_row, max_col)
            rows = normalize_excel_range_values(raw_values, max_row, max_col)

            sheets.append({
                'name': str(ws.Name),
                'start_row': 1,
                'start_col': 1,
                'end_row': max_row,
                'end_col': max_col,
                'max_row': max_row,
                'max_col': max_col,
                'columns': [get_column_letter(col_idx) for col_idx in range(1, max_col + 1)],
                'rows': rows,
                'row_count': actual_max_row,
                'col_count': actual_max_col,
                'truncated': actual_max_row > row_limit or actual_max_col > col_limit
            })

        return {
            'sheets': sheets,
            'sheet_count': total_sheet_count,
            'truncated': total_sheet_count > MAX_EXCEL_PREVIEW_SHEETS,
            'backend': 'excel_com'
        }
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        _close_excel_com_app(pythoncom_module, excel)

def _load_excel_sheet_chunk_via_com(filepath, sheet_name, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """Docstring."""
    pythoncom_module, excel = _create_excel_com_app()
    workbook = None
    try:
        workbook = open_excel_com_workbook(excel, filepath, read_only=True)
        worksheet = None
        for idx in range(1, workbook.Worksheets.Count + 1):
            candidate = workbook.Worksheets(idx)
            if str(candidate.Name) == sheet_name:
                worksheet = candidate
                break

        if worksheet is None:
            raise KeyError(f'工作表不存在: {sheet_name}')

        return build_excel_sheet_chunk_from_com_worksheet(
            worksheet,
            start_row,
            start_col,
            row_limit,
            col_limit
        )
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        _close_excel_com_app(pythoncom_module, excel)

def load_excel_file_preview(filepath, preferred_sheet_name=None):
    """Docstring."""
    file_size = os.path.getsize(filepath)
    editable = True
    backend = 'excel_com' if should_use_excel_com(filepath) else 'openpyxl'
    row_limit, col_limit = get_excel_window_limits(file_size, backend)
    preview_data = (
        _load_excel_preview_via_com(filepath, row_limit, col_limit, preferred_sheet_name)
        if backend == 'excel_com'
        else _load_excel_preview_via_openpyxl(filepath, row_limit, col_limit, preferred_sheet_name)
    )

    return {
        'editable': editable,
        'large_file': file_size > MAX_EDITABLE_EXCEL_FILE_SIZE,
        'truncated': preview_data['truncated'],
        'file_size': file_size,
        'mtime_ns': get_file_mtime_token(filepath),
        'sheets': preview_data['sheets'],
        'sheet_count': preview_data['sheet_count'],
        'backend': preview_data['backend'],
        'row_limit': row_limit,
        'col_limit': col_limit
    }

def load_excel_sheet_chunk(filepath, sheet_name, start_row=1, start_col=1, row_limit=MAX_EXCEL_PREVIEW_ROWS, col_limit=MAX_EXCEL_PREVIEW_COLS):
    """读取Excel工作表指定范围的数据"""
    if should_use_excel_com(filepath):
        return _load_excel_sheet_chunk_via_com(filepath, sheet_name, start_row, start_col, row_limit, col_limit), 'excel_com'
    return _load_excel_sheet_chunk_via_openpyxl(filepath, sheet_name, start_row, start_col, row_limit, col_limit), 'openpyxl'

def _search_excel_sheet_via_openpyxl(filepath, sheet_name, keyword, max_results=MAX_EXCEL_SEARCH_RESULTS):
    """在标准OOXML工作表中搜索内容"""
    workbook = open_excel_workbook(filepath, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise KeyError(f'工作表不存在: {sheet_name}')

        ws = workbook[sheet_name]
        row_count = max(1, ws.max_row or 1)
        col_count = max(1, ws.max_column or 1)
        needle = keyword.casefold()
        results = []
        truncated = False

        for row_idx, row_values in enumerate(
            ws.iter_rows(min_row=1, max_row=row_count, min_col=1, max_col=col_count, values_only=True),
            start=1
        ):
            for col_idx, cell_value in enumerate(row_values, start=1):
                display_value = format_excel_cell_value(cell_value)
                if needle and needle in display_value.casefold():
                    results.append({
                        'sheet': sheet_name,
                        'row': row_idx,
                        'col': col_idx,
                        'cell': f'{get_column_letter(col_idx)}{row_idx}',
                        'value': display_value[:200]
                    })
                    if len(results) >= max_results:
                        truncated = True
                        return {
                            'sheet_name': sheet_name,
                            'results': results,
                            'truncated': truncated,
                            'row_count': row_count,
                            'col_count': col_count
                        }

        return {
            'sheet_name': sheet_name,
            'results': results,
            'truncated': truncated,
            'row_count': row_count,
            'col_count': col_count
        }
    finally:
        workbook.close()

def _search_excel_sheet_via_com(filepath, sheet_name, keyword, max_results=MAX_EXCEL_SEARCH_RESULTS):
    """在Excel兼容模式工作表中搜索内容"""
    pythoncom_module, excel = _create_excel_com_app()
    workbook = None
    try:
        workbook = open_excel_com_workbook(excel, filepath, read_only=True)
        worksheet = None
        for idx in range(1, workbook.Worksheets.Count + 1):
            candidate = workbook.Worksheets(idx)
            if str(candidate.Name) == sheet_name:
                worksheet = candidate
                break

        if worksheet is None:
            raise KeyError(f'工作表不存在: {sheet_name}')

        used_range = worksheet.UsedRange
        used_start_row = int(used_range.Row or 1)
        used_start_col = int(used_range.Column or 1)
        used_row_count = int(used_range.Rows.Count or 1)
        used_col_count = int(used_range.Columns.Count or 1)
        row_count = max(1, used_start_row + used_row_count - 1)
        col_count = max(1, used_start_col + used_col_count - 1)
        needle = keyword.casefold()
        results = []
        truncated = False

        for block_start_row in range(used_start_row, row_count + 1, EXCEL_SEARCH_BLOCK_ROWS):
            block_end_row = min(row_count, block_start_row + EXCEL_SEARCH_BLOCK_ROWS - 1)
            raw_values = read_excel_range_via_com(worksheet, block_start_row, used_start_col, block_end_row, col_count)
            rows = normalize_excel_range_values(raw_values, block_end_row - block_start_row + 1, col_count - used_start_col + 1)

            for row_offset, row_values in enumerate(rows):
                actual_row = block_start_row + row_offset
                for col_offset, display_value in enumerate(row_values):
                    if needle and needle in display_value.casefold():
                        actual_col = used_start_col + col_offset
                        results.append({
                            'sheet': sheet_name,
                            'row': actual_row,
                            'col': actual_col,
                            'cell': f'{get_column_letter(actual_col)}{actual_row}',
                            'value': display_value[:200]
                        })
                        if len(results) >= max_results:
                            truncated = True
                            return {
                                'sheet_name': sheet_name,
                                'results': results,
                                'truncated': truncated,
                                'row_count': row_count,
                                'col_count': col_count
                            }

        return {
            'sheet_name': sheet_name,
            'results': results,
            'truncated': truncated,
            'row_count': row_count,
            'col_count': col_count
        }
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        _close_excel_com_app(pythoncom_module, excel)

def search_excel_sheet(filepath, sheet_name, keyword, max_results=MAX_EXCEL_SEARCH_RESULTS):
    """Docstring."""
    safe_limit = clamp_excel_position(max_results, 1, MAX_EXCEL_SEARCH_RESULTS, MAX_EXCEL_SEARCH_RESULTS)
    if should_use_excel_com(filepath):
        return _search_excel_sheet_via_com(filepath, sheet_name, keyword, safe_limit), 'excel_com'
    return _search_excel_sheet_via_openpyxl(filepath, sheet_name, keyword, safe_limit), 'openpyxl'

def apply_excel_structure_operation(filepath, action, payload):
    """执行 Excel 行列或工作表结构操作"""
    action = str(action or '').strip()
    if not action:
        raise ValueError('缺少操作类型')

    if should_use_excel_com(filepath):
        return _apply_excel_structure_operation_via_com(filepath, action, payload)
    return _apply_excel_structure_operation_via_openpyxl(filepath, action, payload)

def _apply_excel_structure_operation_via_openpyxl(filepath, action, payload):
    """Docstring."""
    workbook = open_excel_workbook(filepath)
    try:
        active_sheet = str(payload.get('sheet_name') or '').strip()
        if active_sheet and active_sheet not in workbook.sheetnames:
            raise KeyError(f'工作表不存在: {active_sheet}')

        message = ''
        updated_chunk = None
        updated_meta = None
        include_sheet_snapshot = bool(payload.get('include_sheet_snapshot', True))

        if action == 'insert_rows':
            row_index = clamp_excel_position(payload.get('row'), 1, EXCEL_MAX_ROWS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            workbook[active_sheet].insert_rows(row_index, amount)
            message = f'已在第 {row_index} 行插入 {amount} 行。'
        elif action == 'delete_rows':
            row_index = clamp_excel_position(payload.get('row'), 1, EXCEL_MAX_ROWS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            workbook[active_sheet].delete_rows(row_index, amount)
            message = f'已从第 {row_index} 行开始删除 {amount} 行。'
        elif action == 'insert_cols':
            col_index = clamp_excel_position(payload.get('col'), 1, EXCEL_MAX_COLS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            workbook[active_sheet].insert_cols(col_index, amount)
            message = f'已在第 {col_index} 列插入 {amount} 列。'
        elif action == 'delete_cols':
            col_index = clamp_excel_position(payload.get('col'), 1, EXCEL_MAX_COLS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            workbook[active_sheet].delete_cols(col_index, amount)
            message = f'已从第 {col_index} 列开始删除 {amount} 列。'
        elif action == 'add_sheet':
            desired_name = payload.get('new_name') or 'Sheet'
            new_name = build_unique_excel_sheet_name(workbook.sheetnames, desired_name)
            workbook.create_sheet(title=new_name)
            active_sheet = new_name
            message = f'已新增工作表：{new_name}'
        elif action == 'rename_sheet':
            new_name = normalize_excel_sheet_name(payload.get('new_name'))
            if new_name in workbook.sheetnames and new_name != active_sheet:
                raise ValueError('工作表名称已存在')
            workbook[active_sheet].title = new_name
            active_sheet = new_name
            message = f'已将工作表重命名为：{new_name}'
        elif action == 'delete_sheet':
            if len(workbook.sheetnames) <= 1:
                raise ValueError('至少保留一个工作表，不能删除最后一个工作表')
            removed_index = workbook.sheetnames.index(active_sheet)
            workbook.remove(workbook[active_sheet])
            remaining_names = workbook.sheetnames
            active_sheet = remaining_names[min(removed_index, len(remaining_names) - 1)]
            message = '已删除当前工作表'
        else:
            raise ValueError(f'不支持的操作: {action}')

        if action in {'insert_rows', 'delete_rows', 'insert_cols', 'delete_cols'}:
            refresh_sheet_name = str(payload.get('refresh_sheet_name') or active_sheet or '').strip()
            if refresh_sheet_name in workbook.sheetnames:
                refresh_sheet = workbook[refresh_sheet_name]
                updated_meta = build_excel_sheet_meta_from_openpyxl_worksheet(
                    refresh_sheet,
                    payload.get('start_row', 1),
                    payload.get('start_col', 1),
                    payload.get('row_limit', MAX_EXCEL_PREVIEW_ROWS),
                    payload.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
                )
                if include_sheet_snapshot:
                    updated_chunk = build_excel_sheet_chunk_from_openpyxl_worksheet(
                        refresh_sheet,
                        payload.get('start_row', 1),
                        payload.get('start_col', 1),
                        payload.get('row_limit', MAX_EXCEL_PREVIEW_ROWS),
                        payload.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
                    )

        temp_path = f"{filepath}.editing.{uuid.uuid4().hex}.tmp"
        try:
            save_openpyxl_workbook_fast(workbook, temp_path, os.path.getsize(filepath))
            workbook.close()
            os.replace(temp_path, filepath)
        finally:
            if os.path.exists(temp_path):
                os.remove(temp_path)

        return {
            'active_sheet': active_sheet,
            'message': message,
            'sheet': updated_chunk,
            'sheet_meta': updated_meta
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass

def _apply_excel_structure_operation_via_com(filepath, action, payload):
    """Docstring."""
    pythoncom_module, excel = _create_excel_com_app()
    workbook = None
    try:
        workbook = open_excel_com_workbook(excel, filepath, read_only=False)
        sheet_name = str(payload.get('sheet_name') or '').strip()
        if sheet_name:
            existing_names = [str(workbook.Worksheets(i).Name) for i in range(1, workbook.Worksheets.Count + 1)]
            if sheet_name not in existing_names:
                raise KeyError(f'工作表不存在: {sheet_name}')
        else:
            existing_names = [str(workbook.Worksheets(i).Name) for i in range(1, workbook.Worksheets.Count + 1)]

        active_sheet = sheet_name
        message = ''
        updated_chunk = None
        updated_meta = None
        include_sheet_snapshot = bool(payload.get('include_sheet_snapshot', True))

        if action in {'insert_rows', 'delete_rows', 'insert_cols', 'delete_cols', 'rename_sheet', 'delete_sheet'}:
            worksheet = workbook.Worksheets(sheet_name)

        if action == 'insert_rows':
            row_index = clamp_excel_position(payload.get('row'), 1, EXCEL_MAX_ROWS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            worksheet.Rows(f'{row_index}:{row_index + amount - 1}').Insert()
            message = f'已在第 {row_index} 行插入 {amount} 行。'
        elif action == 'delete_rows':
            row_index = clamp_excel_position(payload.get('row'), 1, EXCEL_MAX_ROWS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            worksheet.Rows(f'{row_index}:{row_index + amount - 1}').Delete()
            message = f'已从第 {row_index} 行开始删除 {amount} 行。'
        elif action == 'insert_cols':
            col_index = clamp_excel_position(payload.get('col'), 1, EXCEL_MAX_COLS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            start_col = get_column_letter(col_index)
            end_col = get_column_letter(col_index + amount - 1)
            worksheet.Columns(f'{start_col}:{end_col}').Insert()
            message = f'已在第 {col_index} 列插入 {amount} 列。'
        elif action == 'delete_cols':
            col_index = clamp_excel_position(payload.get('col'), 1, EXCEL_MAX_COLS, 1)
            amount = clamp_excel_structure_amount(payload.get('amount'), 1)
            start_col = get_column_letter(col_index)
            end_col = get_column_letter(col_index + amount - 1)
            worksheet.Columns(f'{start_col}:{end_col}').Delete()
            message = f'已从第 {col_index} 列开始删除 {amount} 列。'
        elif action == 'add_sheet':
            desired_name = payload.get('new_name') or 'Sheet'
            new_name = build_unique_excel_sheet_name(existing_names, desired_name)
            new_sheet = workbook.Worksheets.Add(After=workbook.Worksheets(workbook.Worksheets.Count))
            new_sheet.Name = new_name
            active_sheet = new_name
            message = f'已新增工作表：{new_name}'
        elif action == 'rename_sheet':
            new_name = normalize_excel_sheet_name(payload.get('new_name'))
            if new_name in existing_names and new_name != sheet_name:
                raise ValueError('工作表名称已存在')
            worksheet.Name = new_name
            active_sheet = new_name
            message = f'已将工作表重命名为：{new_name}'
        elif action == 'delete_sheet':
            if workbook.Worksheets.Count <= 1:
                raise ValueError('至少保留一个工作表，不能删除最后一个工作表')
            removed_index = int(worksheet.Index)
            worksheet.Delete()
            remaining_count = workbook.Worksheets.Count
            next_index = min(removed_index, remaining_count)
            active_sheet = str(workbook.Worksheets(next_index).Name)
            message = '已删除当前工作表'
        else:
            raise ValueError(f'不支持的操作: {action}')

        workbook.Save()
        if action in {'insert_rows', 'delete_rows', 'insert_cols', 'delete_cols'}:
            refresh_sheet_name = str(payload.get('refresh_sheet_name') or active_sheet or '').strip()
            if refresh_sheet_name:
                refresh_sheet = workbook.Worksheets(refresh_sheet_name)
                updated_meta = build_excel_sheet_meta_from_com_worksheet(
                    refresh_sheet,
                    payload.get('start_row', 1),
                    payload.get('start_col', 1),
                    payload.get('row_limit', MAX_EXCEL_PREVIEW_ROWS),
                    payload.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
                )
                if include_sheet_snapshot:
                    updated_chunk = build_excel_sheet_chunk_from_com_worksheet(
                        refresh_sheet,
                        payload.get('start_row', 1),
                        payload.get('start_col', 1),
                        payload.get('row_limit', MAX_EXCEL_PREVIEW_ROWS),
                        payload.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
                    )
        return {
            'active_sheet': active_sheet,
            'message': message,
            'sheet': updated_chunk,
            'sheet_meta': updated_meta
        }
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        _close_excel_com_app(pythoncom_module, excel)

def _apply_excel_cell_changes_via_openpyxl(filepath, sheets, expected_token):
    """在 openpyxl 模式下保存单元格修改，并在版本变化时按单元格尝试合并。"""
    workbook = open_excel_workbook(filepath)
    current_token = get_file_mtime_token(filepath)
    token_matches = str(expected_token or '') == current_token
    conflicts = []
    changed_count = 0

    try:
        for sheet_data in sheets:
            sheet_name = sheet_data.get('name')
            cells = sheet_data.get('cells') or []
            if sheet_name not in workbook.sheetnames or not isinstance(cells, list):
                continue

            ws = workbook[sheet_name]
            for cell in cells:
                try:
                    row = int(cell.get('row'))
                    col = int(cell.get('col'))
                except (TypeError, ValueError, AttributeError):
                    continue

                if row < 1 or col < 1 or row > EXCEL_MAX_ROWS or col > EXCEL_MAX_COLS:
                    continue

                new_value = cell.get('value', '')
                current_display = format_excel_cell_value(ws.cell(row=row, column=col).value)
                original_display = '' if cell.get('original') is None else str(cell.get('original'))
                new_display = '' if new_value is None else str(new_value)

                if not token_matches and current_display not in {original_display, new_display}:
                    conflicts.append({
                        'sheet': sheet_name,
                        'row': row,
                        'col': col,
                        'cell': f'{get_column_letter(col)}{row}',
                        'current': current_display
                    })
                    continue

                if current_display == new_display:
                    continue

                ws.cell(row=row, column=col).value = parse_excel_input_value(new_value)
                changed_count += 1

        if conflicts:
            return {
                'success': False,
                'mtime_ns': current_token,
                'conflicts': conflicts
            }

        if changed_count > 0:
            temp_path = f"{filepath}.editing.{uuid.uuid4().hex}.tmp"
            try:
                save_openpyxl_workbook_fast(workbook, temp_path, os.path.getsize(filepath))
                workbook.close()
                os.replace(temp_path, filepath)
            finally:
                if os.path.exists(temp_path):
                    os.remove(temp_path)
            new_token = get_file_mtime_token(filepath)
        else:
            new_token = current_token

        return {
            'success': True,
            'changed_count': changed_count,
            'merged': not token_matches,
            'mtime_ns': new_token
        }
    finally:
        try:
            workbook.close()
        except Exception:
            pass

def _apply_excel_cell_changes_via_com(filepath, sheets, expected_token):
    """在 Excel COM 模式下保存单元格修改，并在版本变化时按单元格尝试合并。"""
    pythoncom_module, excel = _create_excel_com_app()
    workbook = None
    current_token = get_file_mtime_token(filepath)
    token_matches = str(expected_token or '') == current_token
    conflicts = []
    changed_count = 0

    try:
        workbook = open_excel_com_workbook(excel, filepath, read_only=False)
        try:
            workbook.Application.Calculation = -4135  # xlCalculationManual
            workbook.Application.CalculateBeforeSave = False
        except Exception:
            pass
        valid_sheet_names = {str(workbook.Worksheets(i).Name) for i in range(1, workbook.Worksheets.Count + 1)}

        for sheet_data in sheets:
            sheet_name = sheet_data.get('name')
            cells = sheet_data.get('cells') or []
            if sheet_name not in valid_sheet_names or not isinstance(cells, list):
                continue

            ws = workbook.Worksheets(sheet_name)
            for cell in cells:
                try:
                    row = int(cell.get('row'))
                    col = int(cell.get('col'))
                except (TypeError, ValueError, AttributeError):
                    continue

                if row < 1 or col < 1 or row > EXCEL_MAX_ROWS or col > EXCEL_MAX_COLS:
                    continue

                new_value = cell.get('value', '')
                current_display = format_excel_cell_value(ws.Cells(row, col).Value)
                original_display = '' if cell.get('original') is None else str(cell.get('original'))
                new_display = '' if new_value is None else str(new_value)

                if not token_matches and current_display not in {original_display, new_display}:
                    conflicts.append({
                        'sheet': sheet_name,
                        'row': row,
                        'col': col,
                        'cell': f'{get_column_letter(col)}{row}',
                        'current': current_display
                    })
                    continue

                if current_display == new_display:
                    continue

                ws.Cells(row, col).Value = parse_excel_input_value(new_value)
                changed_count += 1

        if conflicts:
            return {
                'success': False,
                'mtime_ns': current_token,
                'conflicts': conflicts
            }

        if changed_count > 0:
            workbook.Save()
            new_token = get_file_mtime_token(filepath)
        else:
            new_token = current_token

        return {
            'success': True,
            'changed_count': changed_count,
            'merged': not token_matches,
            'mtime_ns': new_token
        }
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        _close_excel_com_app(pythoncom_module, excel)

# HTML模板
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>局域网文件共享</title>
    <style>
        * {
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }
        
        /* ========== 全局性能优化 ========== */
        /* 减少重绘并启用 GPU 加速 */
        .file-item, .header, .upload-section, .files-section, .btn-download, .btn-delete {
            will-change: transform;
            transform: translateZ(0);
            backface-visibility: hidden;
        }
        
        /* 平滑滚动 */
        html {
            scroll-behavior: smooth;
        }
        
        /* 减少动画卡顿 */
        @media (prefers-reduced-motion: reduce) {
            *, *::before, *::after {
                animation-duration: 0.01ms !important;
                transition-duration: 0.01ms !important;
            }
        }
        
        @keyframes gradientShift {
            0% { background-position: 0% 50%; }
            50% { background-position: 100% 50%; }
            100% { background-position: 0% 50%; }
        }
        
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', 'Microsoft YaHei', 'Helvetica Neue', Arial, sans-serif;
            background: url('/static/bg.jpg') no-repeat center center fixed;
            background-size: cover;
            min-height: 100vh;
            padding: 30px;
            position: relative;
        }
        
        body::before {
            content: '';
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.15);
            z-index: -1;
        }
        
        .container {
            max-width: 1400px;
            margin: 0 auto;
        }
        
        .header {
            background: rgba(255, 255, 255, 0.08);
            backdrop-filter: blur(8px);
            padding: 40px;
            border-radius: 24px;
            box-shadow: 0 4px 16px rgba(31, 38, 135, 0.15);
            margin-bottom: 30px;
            text-align: center;
            border: 1px solid rgba(255, 255, 255, 0.2);
            transition: transform 0.3s ease, box-shadow 0.3s ease, background 0.3s ease;
            animation: fadeInUp 0.8s ease-out;
        }
        
        .header:hover {
            background: rgba(255, 255, 255, 0.15);
            box-shadow: 0 12px 48px rgba(31, 38, 135, 0.35);
            transform: translateY(-5px);
        }
        
        .header h1 {
            color: white;
            font-size: 2.5em;
            font-weight: 700;
            margin-bottom: 15px;
            letter-spacing: -1px;
            text-shadow: 0 4px 12px rgba(0, 0, 0, 0.4);
            transition: all 0.3s ease;
        }
        
        .header:hover h1 {
            transform: scale(1.05);
            text-shadow: 0 6px 20px rgba(255, 255, 255, 0.5);
        }
        
        .header .info {
            color: rgba(255, 255, 255, 0.95);
            font-size: 15px;
            line-height: 1.6;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .header .info strong {
            color: white;
            font-weight: 600;
        }
        .upload-section {
            background: rgba(255, 255, 255, 0.05);
            backdrop-filter: blur(5px);
            padding: 35px;
            border-radius: 24px;
            box-shadow: 0 4px 16px rgba(31, 38, 135, 0.1);
            margin-bottom: 30px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .upload-section:hover {
            background: rgba(255, 255, 255, 0.15);
            backdrop-filter: blur(25px);
            box-shadow: 0 12px 40px rgba(31, 38, 135, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.35);
            transform: translateY(-3px);
        }
        
        .upload-section h2 {
            color: rgba(255, 255, 255, 0.95);
            margin-bottom: 25px;
            font-size: 1.5em;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 10px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        .upload-section:hover h2 {
            transform: translateX(5px);
        }
        
        .upload-section h2:before {
            content: '';
            width: 2.1em;
            height: 2.1em;
            display: inline-block;
            background: url('/static/cloud.png') no-repeat center center;
            background-size: contain;
            transition: transform 0.3s ease;
        }
        
        .upload-section:hover h2:before {
            transform: scale(1.2) rotate(-10deg);
        }
        
        /* 创建新文件夹不显示云图标 */
        .upload-section h2[style*="position: relative"]:before {
            display: none !important;
        }
        
        /* 创建新文件夹区域 - 与共享文件列表完全一致的样式 */
        .folder-section h2 {
            color: rgba(255, 255, 255, 0.95);
            margin-bottom: 25px;
            font-size: 1.5em;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 10px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        /* 覆盖upload-section的云图标，使用文件夹图标 */
        .upload-section.folder-section h2:before {
            content: '📁' !important;
            background: none !important;
            width: auto !important;
            height: auto !important;
            font-size: 1.2em;
            display: inline-block;
            transition: transform 0.3s ease;
        }
        
        .folder-section:hover h2 {
            transform: translateX(5px);
        }
        
        .folder-section:hover h2:before {
            transform: scale(1.2) rotate(10deg);
        }
        
        /* 拖拽上传样式 */
        .upload-section.drag-over {
            background: rgba(16, 185, 129, 0.25) !important;
            backdrop-filter: blur(30px) !important;
            border: 2px dashed rgba(16, 185, 129, 0.8) !important;
            transform: scale(1.02);
            box-shadow: 0 16px 48px rgba(16, 185, 129, 0.4) !important;
        }
        
        .drag-hint {
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            font-size: 24px;
            font-weight: 700;
            color: white;
            text-shadow: 0 4px 12px rgba(0, 0, 0, 0.5);
            background: rgba(16, 185, 129, 0.4);
            padding: 30px 50px;
            border-radius: 20px;
            border: 2px solid rgba(16, 185, 129, 0.8);
            z-index: 10;
            pointer-events: none;
            animation: bounce 0.6s ease-in-out infinite alternate;
        }
        
        @keyframes bounce {
            from {
                transform: translate(-50%, -50%) scale(1);
            }
            to {
                transform: translate(-50%, -55%) scale(1.05);
            }
        }
        
        .upload-section {
            position: relative;
        }
        
        .upload-form {
            display: flex;
            gap: 15px;
            align-items: center;
            flex-wrap: wrap;
        }
        .file-input-wrapper {
            position: relative;
            overflow: visible;
            display: inline-block;
            margin: 5px;
        }
        .file-input-wrapper input[type=file] {
            position: absolute;
            left: -9999px;
        }
        .file-input-label {
            display: inline-block;
            padding: 14px 32px;
            background: rgba(255, 255, 255, 0.08);
            backdrop-filter: blur(5px);
            color: white;
            border-radius: 12px;
            cursor: pointer;
            transition: transform 0.2s ease, background 0.2s ease, box-shadow 0.2s ease;
            font-weight: 600;
            box-shadow: 0 4px 15px rgba(255, 255, 255, 0.1);
            border: 1px solid rgba(255, 255, 255, 0.2);
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .file-input-label:hover {
            background: rgba(255, 255, 255, 0.2);
            transform: translateY(-2px);
            box-shadow: 0 6px 20px rgba(255, 255, 255, 0.2);
        }
        
        .file-input-label:active {
            transform: translateY(0px);
        }
        .file-name {
            color: rgba(255, 255, 255, 0.9);
            font-size: 14px;
            margin-left: 10px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        .btn {
            padding: 14px 32px;
            border: none;
            border-radius: 12px;
            cursor: pointer;
            font-size: 16px;
            transition: transform 0.15s ease, background 0.15s ease, box-shadow 0.15s ease;
            font-weight: 600;
        }
        .btn-primary {
            background: rgba(16, 185, 129, 0.15);
            backdrop-filter: blur(5px);
            color: white;
            font-weight: 600;
            box-shadow: 0 4px 15px rgba(16, 185, 129, 0.2);
            border: 1px solid rgba(16, 185, 129, 0.3);
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .btn-primary:hover {
            background: rgba(16, 185, 129, 0.5);
            backdrop-filter: blur(20px);
            transform: translateY(-4px) scale(1.03);
            box-shadow: 0 8px 28px rgba(16, 185, 129, 0.45);
            border-color: rgba(16, 185, 129, 0.7);
        }
        
        .btn-primary:active {
            transform: translateY(-2px) scale(1.01);
        }
        
        .btn-secondary {
            background: rgba(156, 163, 175, 0.25);
            backdrop-filter: blur(10px);
            color: white;
            font-weight: 600;
            box-shadow: 0 4px 15px rgba(156, 163, 175, 0.25);
            border: 1px solid rgba(156, 163, 175, 0.4);
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .btn-secondary:hover {
            background: rgba(156, 163, 175, 0.4);
            transform: translateY(-3px);
            box-shadow: 0 6px 20px rgba(156, 163, 175, 0.35);
            border-color: rgba(156, 163, 175, 0.6);
        }
        
        .files-section {
            background: rgba(255, 255, 255, 0.05);
            backdrop-filter: blur(5px);
            padding: 35px;
            border-radius: 24px;
            box-shadow: 0 4px 16px rgba(31, 38, 135, 0.1);
            border: 1px solid rgba(255, 255, 255, 0.15);
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .files-section:hover {
            background: rgba(255, 255, 255, 0.15);
            backdrop-filter: blur(25px);
            box-shadow: 0 12px 40px rgba(31, 38, 135, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.35);
            transform: translateY(-3px);
        }
        
        .files-section h2 {
            color: rgba(255, 255, 255, 0.95);
            margin-bottom: 25px;
            font-size: 1.5em;
            font-weight: 600;
            display: flex;
            align-items: center;
            gap: 10px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        .files-section:hover h2 {
            transform: translateX(5px);
        }
        
        .files-section h2:before {
            content: '📁';
            font-size: 1.2em;
            display: inline-block;
            transition: transform 0.3s ease;
        }
        
        .files-section:hover h2:before {
            transform: scale(1.2) rotate(10deg);
        }
        .file-list {
            list-style: none;
            display: grid;
            gap: 12px;
        }
        
        .file-item {
            display: flex;
            justify-content: space-between;
            align-items: center;
            padding: 16px 24px;
            background: rgba(255, 255, 255, 0.03);
            backdrop-filter: blur(3px);
            border: 1px solid rgba(255, 255, 255, 0.1);
            border-radius: 16px;
            transition: transform 0.2s ease, box-shadow 0.2s ease, background 0.2s ease;
            box-shadow: 0 2px 8px rgba(0, 0, 0, 0.05);
            margin-bottom: 0;
        }
        
        .file-item:hover {
            background: rgba(255, 255, 255, 0.12);
            transform: translateY(-3px);
            box-shadow: 0 8px 24px rgba(0, 0, 0, 0.15);
        }
        
        .file-item:active {
            transform: translateY(-1px);
        }
        
        .file-info {
            flex: 1;
            display: flex;
            align-items: center;
            gap: 12px;
            min-width: 0;
        }
        
        .file-icon {
            font-size: 32px;
            flex-shrink: 0;
            width: 48px;
            height: 48px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: rgba(255, 255, 255, 0.05);
            border-radius: 12px;
            border: 1px solid rgba(255, 255, 255, 0.1);
            transition: transform 0.2s ease;
        }
        
        .file-item:hover .file-icon {
            background: rgba(255, 255, 255, 0.25);
            backdrop-filter: blur(15px);
            transform: rotate(5deg) scale(1.1);
            box-shadow: 0 4px 16px rgba(255, 255, 255, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.4);
        }

        .file-details {
            flex: 1;
            min-width: 0;
        }
        
        .file-name-text {
            font-weight: 600;
            color: rgba(255, 255, 255, 0.95);
            margin-bottom: 4px;
            font-size: 15px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        .file-item:hover .file-name-text {
            color: white;
            text-shadow: 0 2px 8px rgba(0, 0, 0, 0.5);
            transform: translateX(3px);
        }
         .folder-group {
             background: rgba(255, 200, 100, 0.25);
             backdrop-filter: blur(10px);
             padding: 18px 24px;
             margin: 0 0 12px 0;
             border: 2px solid rgba(255, 200, 100, 0.5);
             font-weight: 600;
             color: rgba(255, 255, 255, 0.95);
             cursor: pointer;
             user-select: none;
             transition: all 0.3s ease;
             border-radius: 16px;
             box-shadow: 0 4px 16px rgba(255, 152, 0, 0.2);
             font-size: 16px;
             text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
         }
         
         .folder-group:hover {
             background: rgba(255, 200, 100, 0.35);
             transform: translateY(-2px);
             box-shadow: 0 8px 24px rgba(255, 152, 0, 0.3);
             border-color: rgba(255, 200, 100, 0.7);
         }
         .folder-group .toggle-icon {
             display: inline-block;
             margin-right: 8px;
             transition: transform 0.3s;
         }
         .folder-group.collapsed .toggle-icon {
             transform: rotate(-90deg);
         }
         .folder-files {
             display: block;
         }
         .folder-files.hidden {
             display: none;
         }
         .folder-group.hidden {
             display: none;
         }
        .file-meta {
            font-size: 13px;
            color: rgba(255, 255, 255, 0.8);
            display: flex;
            gap: 16px;
            align-items: center;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .file-meta span {
            display: inline-flex;
            align-items: center;
            gap: 4px;
        }
        .file-actions {
            display: flex;
            gap: 10px;
            opacity: 0.6;
            transition: opacity 0.3s ease;
        }
        
        .file-item:hover .file-actions {
            opacity: 1;
        }
        .btn-download {
            background: rgba(16, 185, 129, 0.15);
            backdrop-filter: blur(3px);
            color: white;
            padding: 8px 20px;
            text-decoration: none;
            border-radius: 10px;
            transition: transform 0.15s ease, background 0.15s ease, box-shadow 0.15s ease;
            font-weight: 500;
            font-size: 14px;
            box-shadow: 0 2px 8px rgba(16, 185, 129, 0.2);
            border: 1px solid rgba(16, 185, 129, 0.3);
            display: inline-flex;
            align-items: center;
            gap: 6px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .btn-download:hover {
            background: rgba(16, 185, 129, 0.5);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(16, 185, 129, 0.4);
        }
        
        .btn-download:active {
            transform: translateY(0);
        }
        
        .btn-delete {
            background: rgba(244, 63, 94, 0.15);
            backdrop-filter: blur(3px);
            color: white;
            padding: 8px 20px;
            border: 1px solid rgba(244, 63, 94, 0.3);
            border-radius: 10px;
            cursor: pointer;
            transition: transform 0.15s ease, background 0.15s ease, box-shadow 0.15s ease;
            font-weight: 500;
            font-size: 14px;
            box-shadow: 0 2px 8px rgba(244, 63, 94, 0.2);
            display: inline-flex;
            align-items: center;
            gap: 6px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .btn-delete:hover {
            background: rgba(244, 63, 94, 0.5);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(244, 63, 94, 0.4);
        }
        
        .btn-delete:active {
            transform: translateY(0);
        }
        .empty-message {
            text-align: center;
            padding: 60px 40px;
            background: rgba(255, 255, 255, 0.05);
            backdrop-filter: blur(5px);
            border-radius: 16px;
            border: 2px dashed rgba(255, 255, 255, 0.2);
        }
        
        .empty-message:hover {
            background: rgba(255, 255, 255, 0.1);
            border-color: rgba(255, 255, 255, 0.3);
        }
        
        .empty-message p {
            color: rgba(255, 255, 255, 0.9);
            font-size: 16px;
            margin: 0;
            font-weight: 500;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .empty-message .emoji {
            font-size: 48px;
            margin-bottom: 16px;
            display: block;
        }
        input::placeholder {
            color: rgba(255, 255, 255, 0.5);
        }
        
        .alert {
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 20px;
            backdrop-filter: blur(10px);
        }
        .alert-success {
            background: rgba(16, 185, 129, 0.25);
            color: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(16, 185, 129, 0.5);
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        .alert-danger {
            background: rgba(244, 63, 94, 0.25);
            color: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(244, 63, 94, 0.5);
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        .alert-warning {
            background: rgba(255, 193, 7, 0.25);
            color: rgba(255, 255, 255, 0.95);
            border: 1px solid rgba(255, 193, 7, 0.5);
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        /* 进度条样式 */
        .progress-container {
            position: fixed;
            top: 20px;
            right: 360px;  /* 避免与在线用户面板重叠 */
            background: rgba(255, 255, 255, 0.15);
            backdrop-filter: blur(20px);
            padding: 20px 30px;
            border-radius: 16px;
            box-shadow: 0 8px 32px rgba(31, 38, 135, 0.25);
            border: 1px solid rgba(255, 255, 255, 0.3);
            z-index: 10000;
            min-width: 350px;
            max-width: 400px;
            display: none;
        }
        
        .progress-title {
            color: white;
            font-size: 16px;
            font-weight: 600;
            margin-bottom: 12px;
            text-align: left;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .progress-bar-bg {
            width: 100%;
            height: 30px;
            background: rgba(255, 255, 255, 0.2);
            border-radius: 15px;
            overflow: hidden;
            position: relative;
            box-shadow: inset 0 2px 5px rgba(0, 0, 0, 0.2);
        }
        
        .progress-bar-fill {
            height: 100%;
            background: linear-gradient(90deg, #10b981, #059669);
            border-radius: 15px;
            transition: width 0.3s ease;
            box-shadow: 0 2px 10px rgba(16, 185, 129, 0.5);
            display: flex;
            align-items: center;
            justify-content: center;
            color: white;
            font-weight: 600;
            font-size: 14px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .progress-info {
            margin-top: 12px;
            text-align: left;
            color: rgba(255, 255, 255, 0.9);
            font-size: 13px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .progress-speed {
            margin-top: 6px;
            text-align: left;
            color: rgba(255, 255, 255, 0.85);
            font-size: 12px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
            font-weight: 500;
        }
        
        .progress-speed .speed-value {
            color: #10b981;
            font-weight: 600;
            font-size: 15px;
        }
        
        .progress-actions {
            display: flex;
            gap: 10px;
            margin-top: 15px;
            justify-content: flex-end;
        }
        
        .progress-btn {
            padding: 8px 16px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 13px;
            font-weight: 600;
            transition: all 0.3s;
            color: white;
            text-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);
        }
        
        .progress-btn-pause {
            background: rgba(255, 193, 7, 0.6);
        }
        
        .progress-btn-pause:hover {
            background: rgba(255, 193, 7, 0.8);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(255, 193, 7, 0.4);
        }
        
        .progress-btn-resume {
            background: rgba(16, 185, 129, 0.6);
        }
        
        .progress-btn-resume:hover {
            background: rgba(16, 185, 129, 0.8);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.4);
        }
        
        .progress-btn-cancel {
            background: rgba(244, 63, 94, 0.6);
        }
        
        .progress-btn-cancel:hover {
            background: rgba(244, 63, 94, 0.8);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(244, 63, 94, 0.4);
        }
        
        .progress-overlay {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.3);
            z-index: 9999;
            display: none;
            pointer-events: none;  /* 不阻止点击 */
        }
        
        /* 未完成任务通知样式 */
        .incomplete-tasks-notification {
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            background: rgba(255, 255, 255, 0.18);
            backdrop-filter: blur(30px);
            padding: 40px;
            border-radius: 24px;
            box-shadow: 0 8px 32px rgba(31, 38, 135, 0.35);
            border: 2px solid rgba(255, 255, 255, 0.3);
            z-index: 10001;
            max-width: 600px;
            width: 90%;
            animation: slideInDown 0.3s ease-out;
        }
        
        @keyframes slideInDown {
            from {
                opacity: 0;
                transform: translate(-50%, -60%);
            }
            to {
                opacity: 1;
                transform: translate(-50%, -50%);
            }
        }
        
        .notification-content {
            display: flex;
            gap: 20px;
            align-items: flex-start;
        }
        
        .notification-icon {
            font-size: 48px;
            flex-shrink: 0;
        }
        
        .notification-body {
            flex: 1;
        }
        
        .notification-title {
            color: white;
            font-size: 1.5em;
            margin: 0 0 15px 0;
            text-shadow: 0 2px 8px rgba(0, 0, 0, 0.3);
        }
        
        .notification-message {
            color: rgba(255, 255, 255, 0.9);
            margin-bottom: 20px;
            font-size: 14px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .incomplete-tasks-list {
            background: rgba(0, 0, 0, 0.15);
            padding: 15px;
            border-radius: 12px;
            margin-bottom: 20px;
            max-height: 300px;
            overflow-y: auto;
        }
        
        .incomplete-task-item {
            background: rgba(255, 255, 255, 0.12);
            padding: 15px;
            border-radius: 10px;
            margin-bottom: 10px;
            border: 1px solid rgba(255, 255, 255, 0.2);
        }
        
        .incomplete-task-item:last-child {
            margin-bottom: 0;
        }
        
        .incomplete-task-name {
            color: white;
            font-weight: 600;
            margin-bottom: 8px;
            font-size: 15px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .incomplete-task-progress {
            color: rgba(255, 255, 255, 0.85);
            font-size: 13px;
            margin-bottom: 10px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .incomplete-task-actions {
            display: flex;
            gap: 10px;
        }
        
        .task-btn {
            padding: 8px 16px;
            border: none;
            border-radius: 8px;
            cursor: pointer;
            font-size: 13px;
            font-weight: 600;
            transition: all 0.3s;
            color: white;
            text-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);
        }
        
        .task-btn-continue {
            background: rgba(16, 185, 129, 0.6);
        }
        
        .task-btn-continue:hover {
            background: rgba(16, 185, 129, 0.8);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(16, 185, 129, 0.4);
        }
        
        .task-btn-delete {
            background: rgba(244, 63, 94, 0.6);
        }
        
        .task-btn-delete:hover {
            background: rgba(244, 63, 94, 0.8);
            transform: translateY(-1px);
            box-shadow: 0 4px 12px rgba(244, 63, 94, 0.4);
        }
        
        .notification-actions {
            display: flex;
            gap: 10px;
            justify-content: flex-end;
            margin-top: 10px;
        }
        
        .notification-btn {
            padding: 12px 24px;
            border: none;
            border-radius: 10px;
            cursor: pointer;
            font-size: 14px;
            font-weight: 600;
            transition: all 0.3s;
            color: white;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .notification-btn-primary {
            background: rgba(16, 185, 129, 0.6);
        }
        
        .notification-btn-primary:hover {
            background: rgba(16, 185, 129, 0.8);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(16, 185, 129, 0.4);
        }
        
        /* 闈㈠寘灞戝鑸牱寮?*/
        .breadcrumb-link:hover {
            background: rgba(255, 255, 255, 0.18) !important;
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(255, 255, 255, 0.2);
        }
        
        /* 闈㈠寘灞戞嫋鎷界洰鏍囨牱寮?*/
        .breadcrumb-drop-target {
            transition: all 0.3s ease;
        }
        
        .breadcrumb-drop-target.drag-over-breadcrumb {
            background: rgba(16, 185, 129, 0.4) !important;
            transform: scale(1.1) translateY(-2px);
            box-shadow: 0 6px 20px rgba(16, 185, 129, 0.5) !important;
            border: 2px solid rgba(16, 185, 129, 0.8) !important;
        }
        
        /* 页面加载动画 */
        @keyframes fadeInUp {
            from {
                opacity: 0;
                transform: translateY(30px);
            }
            to {
                opacity: 1;
                transform: translateY(0);
            }
        }
        
        @keyframes fadeInRight {
            from {
                opacity: 0;
                transform: translateX(30px);
            }
            to {
                opacity: 1;
                transform: translateX(0);
            }
        }
        
        @keyframes scaleIn {
            from {
                opacity: 0;
                transform: scale(0.9);
            }
            to {
                opacity: 1;
                transform: scale(1);
            }
        }
        
        .upload-section {
            animation: fadeInUp 0.6s ease-out;
        }
        
        .files-section {
            animation: fadeInUp 0.6s ease-out 0.1s backwards;
        }
        
        .online-panel {
            animation: fadeInRight 0.6s ease-out 0.2s backwards;
        }
        
        .file-item {
            animation: scaleIn 0.4s ease-out backwards;
        }
        
        .file-item:nth-child(1) { animation-delay: 0.05s; }
        .file-item:nth-child(2) { animation-delay: 0.1s; }
        .file-item:nth-child(3) { animation-delay: 0.15s; }
        .file-item:nth-child(4) { animation-delay: 0.2s; }
        .file-item:nth-child(5) { animation-delay: 0.25s; }
        .file-item:nth-child(6) { animation-delay: 0.3s; }
        .file-item:nth-child(7) { animation-delay: 0.35s; }
        .file-item:nth-child(8) { animation-delay: 0.4s; }
        .file-item:nth-child(9) { animation-delay: 0.45s; }
        .file-item:nth-child(10) { animation-delay: 0.5s; }
        
        /* 鼠标光晕效果 */
        .mouse-glow {
            position: fixed;
            width: 300px;
            height: 300px;
            border-radius: 50%;
            background: radial-gradient(circle, rgba(255, 255, 255, 0.15) 0%, rgba(255, 255, 255, 0) 70%);
            pointer-events: none;
            z-index: 1;
            transform: translate(-50%, -50%);
            transition: opacity 0.3s ease;
            opacity: 0;
        }
        
        body:hover .mouse-glow {
            opacity: 1;
        }
        
        /* 登录弹窗 */
        .login-modal {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.8);
            display: flex;
            justify-content: center;
            align-items: center;
            z-index: 10000;
        }
        
        .login-box {
            background: rgba(255, 255, 255, 0.18);
            backdrop-filter: blur(30px);
            padding: 50px 60px;
            border-radius: 24px;
            box-shadow: 0 8px 32px rgba(31, 38, 135, 0.35);
            border: 2px solid rgba(255, 255, 255, 0.3);
            text-align: center;
            max-width: 450px;
            width: 90%;
        }
        
        .login-box h2 {
            color: white;
            font-size: 2em;
            margin-bottom: 15px;
            text-shadow: 0 2px 8px rgba(0, 0, 0, 0.3);
        }
        
        .login-box p {
            color: rgba(255, 255, 255, 0.9);
            margin-bottom: 30px;
            font-size: 14px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .login-box input {
            width: 100%;
            padding: 16px 24px;
            border: 2px solid rgba(255, 255, 255, 0.35);
            background: rgba(255, 255, 255, 0.12);
            backdrop-filter: blur(10px);
            border-radius: 12px;
            font-size: 16px;
            color: white;
            outline: none;
            transition: all 0.3s;
            margin-bottom: 16px;
            text-align: center;
        }
        
        .login-box input::placeholder {
            color: rgba(255, 255, 255, 0.6);
        }
        
        .login-box input:focus {
            border-color: rgba(255, 255, 255, 0.8);
            background: rgba(255, 255, 255, 0.3);
            box-shadow: 0 0 0 4px rgba(255, 255, 255, 0.1);
        }
        
        .login-box button {
            width: 100%;
            padding: 16px 32px;
            background: rgba(16, 185, 129, 0.4);
            backdrop-filter: blur(10px);
            color: white;
            border: 2px solid rgba(16, 185, 129, 0.6);
            border-radius: 12px;
            font-size: 16px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .login-box button:hover {
            background: rgba(16, 185, 129, 0.5);
            border-color: rgba(16, 185, 129, 0.8);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(16, 185, 129, 0.4);
        }

        .login-form-actions {
            display: grid;
            gap: 12px;
            margin-top: 8px;
        }

        .login-box button.secondary {
            background: rgba(59, 130, 246, 0.28);
            border-color: rgba(96, 165, 250, 0.5);
        }

        .login-box button.secondary:hover {
            background: rgba(59, 130, 246, 0.42);
            border-color: rgba(147, 197, 253, 0.8);
            box-shadow: 0 4px 16px rgba(59, 130, 246, 0.35);
        }

        .login-note {
            margin-top: 14px;
            color: rgba(255, 255, 255, 0.75);
            font-size: 12px;
            line-height: 1.6;
        }
        
        /* 在线用户面板 */
        .online-panel {
            position: fixed;
            top: 20px;
            right: 20px;
            width: 320px;
            background: rgba(255, 255, 255, 0.08);
            backdrop-filter: blur(10px);
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 4px 16px rgba(31, 38, 135, 0.15);
            border: 1px solid rgba(255, 255, 255, 0.15);
            z-index: 100;
            max-height: 80vh;
            overflow-y: auto;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
        }
        
        .online-panel:hover {
            background: rgba(255, 255, 255, 0.18);
            backdrop-filter: blur(25px);
            box-shadow: 0 12px 40px rgba(31, 38, 135, 0.3);
            border: 1px solid rgba(255, 255, 255, 0.35);
        }
        
        .online-panel h3 {
            color: white;
            font-size: 16px;
            margin-bottom: 12px;
            display: flex;
            align-items: center;
            gap: 8px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            cursor: pointer;
            user-select: none;
            transition: all 0.3s;
        }
        
        .online-panel h3:hover {
            opacity: 0.8;
        }
        
        .online-panel h3 .toggle-arrow {
            font-size: 12px;
            transition: transform 0.3s;
            display: inline-block;
        }
        
        .online-panel h3.collapsed .toggle-arrow {
            transform: rotate(0deg);
        }
        
        .online-panel h3:not(.collapsed) .toggle-arrow {
            transform: rotate(90deg);
        }
        
        .section-content {
            max-height: 500px;
            overflow: hidden;
            transition: max-height 0.3s ease-out, opacity 0.3s;
            opacity: 1;
        }
        
        .section-content.collapsed {
            max-height: 0;
            opacity: 0;
            overflow: hidden;
        }
        
        /* Activities section specific styling */
        #activities:not(.collapsed) {
            max-height: 300px;
            overflow-y: auto;
            overflow-x: hidden;
        }
        
        #activities {
            overflow-x: hidden;
            position: relative;
            z-index: 1;
        }
        
        @keyframes pulse {
            0% {
                opacity: 0.3;
                transform: scale(0.8);
            }
            50% {
                opacity: 1;
                transform: scale(1.2);
            }
            100% {
                opacity: 0.3;
                transform: scale(0.8);
            }
        }
        
        .online-count {
            background: rgba(16, 185, 129, 0.4);
            padding: 4px 12px;
            border-radius: 12px;
            font-size: 13px;
            font-weight: 600;
        }
        
        .user-item, .activity-item {
            background: rgba(255, 255, 255, 0.15);
            padding: 10px 12px;
            border-radius: 10px;
            margin-bottom: 8px;
            color: rgba(255, 255, 255, 0.95);
            font-size: 13px;
            text-shadow: 0 1px 2px rgba(0, 0, 0, 0.2);
            border: 1px solid rgba(255, 255, 255, 0.2);
            position: relative;
            z-index: 1;
        }
        
        .user-item strong {
            color: white;
            font-weight: 600;
        }
        
        .activity-item {
            font-size: 12px;
            line-height: 1.5;
        }
        
        .activity-time {
            color: rgba(255, 255, 255, 0.7);
            font-size: 11px;
            margin-top: 4px;
        }
        
        .active-task {
            animation: pulse 2s ease-in-out infinite;
        }
        
        @keyframes pulse {
            0%, 100% {
                box-shadow: 0 0 0 0 rgba(16, 185, 129, 0.4);
            }
            50% {
                box-shadow: 0 0 0 8px rgba(16, 185, 129, 0);
            }
        }
        
        .current-user {
            background: rgba(16, 185, 129, 0.2);
            padding: 12px;
            border-radius: 10px;
            margin-bottom: 15px;
            color: white;
            font-size: 14px;
            font-weight: 600;
            text-align: center;
            border: 1px solid rgba(16, 185, 129, 0.5);
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }

        .current-user-name {
            display: block;
            margin-bottom: 4px;
        }

        .request-admin-btn {
            margin-top: 10px;
            width: 100%;
            padding: 8px 12px;
            background: rgba(255, 215, 0, 0.15);
            border: 1px solid rgba(255, 215, 0, 0.3);
            border-radius: 8px;
            color: white;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            transition: all 0.4s cubic-bezier(0.4, 0, 0.2, 1);
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .request-admin-btn:hover {
            background: rgba(255, 215, 0, 0.4);
            border-color: rgba(255, 215, 0, 0.6);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(255, 215, 0, 0.4);
        }
        
        .request-admin-btn:disabled {
            opacity: 0.5;
            cursor: not-allowed;
        }

        .logout-btn {
            margin-top: 10px;
            width: 100%;
            padding: 8px 12px;
            background: rgba(244, 63, 94, 0.16);
            border: 1px solid rgba(244, 63, 94, 0.35);
            border-radius: 8px;
            color: white;
            cursor: pointer;
            font-size: 12px;
            font-weight: 600;
            transition: all 0.3s ease;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }

        .logout-btn:hover {
            background: rgba(244, 63, 94, 0.3);
            border-color: rgba(251, 113, 133, 0.6);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(244, 63, 94, 0.25);
        }
        
        /* 拖放样式 */
        .file-item.dragging {
            opacity: 0.5;
            cursor: move;
            border: 2px dashed rgba(255, 255, 255, 0.5);
            background: rgba(59, 130, 246, 0.2) !important;
        }
        
        .folder-group.drag-over {
            background: rgba(16, 185, 129, 0.35) !important;
            border-color: rgba(16, 185, 129, 0.7) !important;
            transform: scale(1.02);
        }
        
        .folder-item.drag-over {
            background: rgba(16, 185, 129, 0.35) !important;
            border: 2px solid rgba(16, 185, 129, 0.8) !important;
            transform: scale(1.02);
            box-shadow: 0 8px 24px rgba(16, 185, 129, 0.4) !important;
        }
        
        .folder-item {
            transition: all 0.3s ease;
        }
        
        .folder-item:hover {
            transform: translateY(-2px);
        }
        
        .file-item.draggable {
            cursor: move;
            position: relative;
        }
        
        .file-item.draggable:hover {
            cursor: move;
        }
        
        
        /* 鎷栨嫿鎻愮ず妗?*/
        #dragHintBox {
            position: fixed;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            background: rgba(59, 130, 246, 0.95);
            color: white;
            padding: 30px 50px;
            border-radius: 20px;
            font-size: 24px;
            font-weight: 600;
            box-shadow: 0 10px 40px rgba(0, 0, 0, 0.5);
            z-index: 9999;
            display: none;
            backdrop-filter: blur(10px);
            border: 3px solid rgba(255, 255, 255, 0.3);
            animation: pulse 1s infinite;
        }
        
        @keyframes pulse {
            0%, 100% { transform: translate(-50%, -50%) scale(1); }
            50% { transform: translate(-50%, -50%) scale(1.05); }
        }
        
        /* 批量操作样式 */
        .file-item {
            display: flex;
            align-items: center;
        }
        
        .file-checkbox {
            flex-shrink: 0;
        }
        
        .batch-btn {
            transition: all 0.3s ease;
        }
        
        .batch-btn:hover {
            transform: translateY(-2px);
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.3);
            opacity: 0.9;
        }
        
        .batch-btn:active {
            transform: translateY(0);
        }
        
        .file-item.selected {
            background: rgba(59, 130, 246, 0.2) !important;
            border-left: 4px solid rgba(59, 130, 246, 0.8);
        }
        
        .stats {
            display: flex;
            gap: 30px;
            margin-top: 25px;
            justify-content: center;
            flex-wrap: wrap;
        }
        
        .stat-item:nth-child(1) {
            animation-delay: 0.1s;
        }
        
        .stat-item:nth-child(2) {
            animation-delay: 0.2s;
        }
        
        .stat-item {
            background: rgba(255, 255, 255, 0.05);
            backdrop-filter: blur(5px);
            padding: 20px 35px;
            border-radius: 16px;
            box-shadow: 0 4px 16px rgba(31, 38, 135, 0.15);
            transition: all 0.5s cubic-bezier(0.4, 0, 0.2, 1);
            min-width: 180px;
            border: 1px solid rgba(255, 255, 255, 0.15);
            animation: scaleIn 0.6s ease-out backwards;
        }
        
        .stat-item:hover {
            transform: translateY(-6px) scale(1.05);
            box-shadow: 0 16px 48px rgba(255, 255, 255, 0.35);
            background: rgba(255, 255, 255, 0.3);
            backdrop-filter: blur(30px);
            border-color: rgba(255, 255, 255, 0.6);
        }
        
        .stat-label {
            font-size: 13px;
            color: rgba(255, 255, 255, 0.85);
            font-weight: 500;
            margin-bottom: 8px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        .stat-item:hover .stat-label {
            color: white;
            transform: translateY(-2px);
        }
        
        .stat-value {
            font-size: 32px;
            font-weight: 700;
            color: white;
            text-shadow: 0 2px 8px rgba(0, 0, 0, 0.3);
            transition: all 0.3s ease;
        }
        
        .stat-item:hover .stat-value {
            transform: scale(1.15);
            text-shadow: 0 4px 20px rgba(255, 255, 255, 0.7);
        }
        
        /* 任务管理面板 */
        .tasks-panel {
            position: fixed;
            bottom: 20px;
            right: 20px;
            width: 400px;
            max-height: 400px;
            background: rgba(255, 255, 255, 0.25);
            backdrop-filter: blur(20px);
            border-radius: 16px;
            padding: 20px;
            box-shadow: 0 8px 32px rgba(31, 38, 135, 0.37);
            border: 1px solid rgba(255, 255, 255, 0.4);
            z-index: 100;
            overflow-y: auto;
            display: none;
        }
        
        .tasks-panel h3 {
            color: white;
            font-size: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
            user-select: none;
        }
        
        .task-item {
            background: rgba(255, 255, 255, 0.15);
            padding: 12px;
            border-radius: 10px;
            margin-bottom: 10px;
            border: 1px solid rgba(255, 255, 255, 0.2);
        }
        
        .task-header {
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin-bottom: 8px;
        }
        
        .task-name {
            color: white;
            font-weight: 600;
            font-size: 13px;
            flex: 1;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
            text-shadow: 0 1px 2px rgba(0, 0, 0, 0.2);
        }
        
        .task-status {
            font-size: 11px;
            padding: 2px 8px;
            border-radius: 8px;
            font-weight: 600;
        }
        
        .task-status.running {
            background: rgba(16, 185, 129, 0.4);
            color: white;
        }
        
        .task-status.paused {
            background: rgba(255, 193, 7, 0.4);
            color: white;
        }
        
        .task-status.completed {
            background: rgba(16, 185, 129, 0.4);
            color: white;
        }
        
        .task-progress {
            width: 100%;
            height: 6px;
            background: rgba(255, 255, 255, 0.2);
            border-radius: 3px;
            overflow: hidden;
            margin-bottom: 8px;
        }
        
        .task-progress-bar {
            height: 100%;
            background: linear-gradient(90deg, #10b981, #059669);
            transition: width 0.3s ease;
            border-radius: 3px;
        }
        
        .task-actions {
            display: flex;
            gap: 6px;
        }
        
        .task-btn {
            padding: 4px 12px;
            border: none;
            border-radius: 6px;
            font-size: 11px;
            cursor: pointer;
            transition: all 0.3s;
            font-weight: 600;
            text-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);
        }
        
        .task-btn-pause {
            background: rgba(255, 193, 7, 0.4);
            color: white;
        }
        
        .task-btn-resume {
            background: rgba(16, 185, 129, 0.4);
            color: white;
        }
        
        .task-btn-delete {
            background: rgba(244, 63, 94, 0.4);
            color: white;
        }
        
        .task-btn:hover {
            opacity: 0.8;
            transform: translateY(-1px);
        }
        
        /* 任务管理面板已移除，样式保留但不显示 */
        .tasks-toggle {
            display: none !important;
        }
        
        .tasks-panel {
            display: none !important;
        }
        
        /* 自定义确认对话框 */
        .confirm-modal {
            position: fixed;
            top: 0;
            left: 0;
            right: 0;
            bottom: 0;
            background: rgba(0, 0, 0, 0.6);
            backdrop-filter: blur(5px);
            display: none;
            justify-content: center;
            align-items: center;
            z-index: 20000;
        }
        
        .confirm-modal.show {
            display: flex;
        }
        
        .confirm-dialog {
            background: rgba(255, 255, 255, 0.18);
            backdrop-filter: blur(30px);
            padding: 30px 40px;
            border-radius: 20px;
            box-shadow: 0 8px 32px rgba(31, 38, 135, 0.35);
            border: 2px solid rgba(255, 255, 255, 0.3);
            min-width: 400px;
            max-width: 500px;
            text-align: center;
        }
        
        .confirm-dialog h3 {
            color: white;
            font-size: 20px;
            font-weight: 600;
            margin-bottom: 20px;
            text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        }
        
        .confirm-dialog p {
            color: rgba(255, 255, 255, 0.95);
            font-size: 15px;
            line-height: 1.6;
            margin-bottom: 30px;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .confirm-buttons {
            display: flex;
            gap: 15px;
            justify-content: center;
        }
        
        .confirm-btn {
            padding: 12px 30px;
            border: none;
            border-radius: 12px;
            font-size: 15px;
            font-weight: 600;
            cursor: pointer;
            transition: all 0.3s;
            text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);
        }
        
        .confirm-btn-cancel {
            background: rgba(156, 163, 175, 0.4);
            backdrop-filter: blur(10px);
            color: white;
            border: 2px solid rgba(156, 163, 175, 0.6);
        }
        
        .confirm-btn-cancel:hover {
            background: rgba(156, 163, 175, 0.6);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(156, 163, 175, 0.4);
        }
        
        .confirm-btn-ok {
            background: rgba(244, 63, 94, 0.4);
            backdrop-filter: blur(10px);
            color: white;
            border: 2px solid rgba(244, 63, 94, 0.6);
        }
        
        .confirm-btn-ok:hover {
            background: rgba(244, 63, 94, 0.6);
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(244, 63, 94, 0.4);
        }
    </style>
</head>
<body>
    <!-- 鼠标光晕效果 -->
    <div class="mouse-glow" id="mouseGlow"></div>
    
    <!-- 鎷栨嫿鎻愮ず妗?-->
    <div id="dragHintBox"></div>
    
    <!-- 登录弹窗 -->
    <div class="login-modal" id="loginModal" style="display: {% if not session.get('username') %}flex{% else %}none{% endif %};">
        <div class="login-box">
            <h2>欢迎使用</h2>
            <p>请输入昵称和密码后继续。昵称必须是中文，四个字以内；新用户可以直接注册，之后可在任意电脑登录。</p>
            <form id="loginForm" method="post" action="{{ url_for('login') }}?redirect=1">
                <input type="hidden" id="loginModeInput" name="mode" value="login">
                <input type="text" id="usernameInput" name="username" placeholder="请输入中文昵称（1-4 个字）" maxlength="{{ nickname_max_length }}" required autocomplete="username">
                <input type="password" id="passwordInput" name="password" placeholder="请输入密码（至少 {{ password_min_length }} 个字符）" minlength="{{ password_min_length }}" required autocomplete="current-password">
                <div class="login-form-actions">
                    <button type="submit" onclick="document.getElementById('loginModeInput').value='login'">登录</button>
                    <button type="submit" class="secondary" onclick="document.getElementById('loginModeInput').value='register'">注册并进入</button>
                </div>
                <div class="login-note">昵称和密码都会保存在服务器本地。后续换设备时，只需要输入同一套昵称和密码即可。</div>
            </form>
        </div>
    </div>

    <!-- 在线用户面板 -->
    <div class="online-panel" id="onlinePanel" style="display: {% if session.get('username') %}block{% else %}none{% endif %};">
        <div class="current-user" id="currentUser">
            <span class="current-user-name"><span id="currentUserName">{{ session.get('username', '未登录') }}</span></span>
            <div id="adminBadge" style="display: none; margin-top: 5px; font-size: 12px; color: rgba(255, 215, 0, 0.9);">
                管理员
            </div>
            <button id="requestAdminBtn" 
                    class="request-admin-btn"
                    style="display: none;"
                    onclick="requestAdminPermission()">
                申请管理员
            </button>
            <button id="logoutBtn" class="logout-btn" onclick="logoutCurrentUser()">退出登录</button>
        </div>
        
        <!-- 管理员申请审批面板，仅主机可见 -->
        <div id="adminRequestsPanel" style="display: none; margin-bottom: 15px;">
            <h3 onclick="toggleSection('adminRequests', this)" class="collapsed">
                <span class="toggle-arrow">▶</span>
                管理员申请 <span class="online-count" id="adminRequestsCount">{{ (initial_admin_state.requests or {})|length }}</span>
            </h3>
            <div id="adminRequests" class="section-content collapsed"></div>
        </div>
        
        <h3 onclick="toggleSection('onlineUsers', this)" class="collapsed">
            <span class="toggle-arrow">▶</span>
            在线用户
            <span class="online-count" id="onlineCount">{{ initial_online_users|length }}</span>
        </h3>
        <div id="onlineUsers" class="section-content collapsed">
            {% if initial_online_users %}
                {% for online_user in initial_online_users %}
                <div class="user-item"><strong>{{ online_user.username }}</strong></div>
                {% endfor %}
            {% else %}
                <div style="color: rgba(255,255,255,0.7); padding: 10px; text-align: center; font-size: 12px;">暂无在线用户</div>
            {% endif %}
        </div>
        
        <h3 style="margin-top: 20px;" onclick="toggleSection('activities', this)" class="collapsed">
            <span class="toggle-arrow">▶</span>
            实时动态
        </h3>
        <div id="activities" class="section-content collapsed"></div>
    </div>

    <!-- 进度条遮罩层 -->
    <div class="progress-overlay" id="progressOverlay"></div>
    
    <!-- 进度条容器 -->
    <div class="progress-container" id="progressContainer">
        <div class="progress-title" id="progressTitle">上传中...</div>
        <div class="progress-bar-bg">
            <div class="progress-bar-fill" id="progressBar" style="width: 0%">0%</div>
        </div>
        <div class="progress-info" id="progressInfo">准备中...</div>
        <div class="progress-speed" id="progressSpeed"></div>
        <div class="progress-actions" id="progressActions" style="display: none;">
            <button class="progress-btn progress-btn-pause" id="progressPauseBtn" onclick="handleProgressPause()" style="display: none;">暂停</button>
            <button class="progress-btn progress-btn-resume" id="progressResumeBtn" onclick="handleProgressResume()" style="display: none;">继续</button>
            <button class="progress-btn progress-btn-cancel" id="progressCancelBtn" onclick="handleProgressCancel()">取消</button>
        </div>
    </div>
    
    <!-- 未完成任务通知区域 -->
    <div class="incomplete-tasks-notification" id="incompleteTasksNotification" style="display: none;">
        <div class="notification-content">
            <div class="notification-icon">!</div>
            <div class="notification-body">
                <h3 class="notification-title">检测到未完成的上传任务</h3>
                <div class="notification-message">
                    您有以下文件上传未完成，可以继续上传：<br>
                    <small style="color: rgba(255, 255, 255, 0.7); margin-top: 5px; display: block;">
                        由于浏览器安全限制，需要重新选择文件后才能继续上传，系统会自动从断点恢复。
                    </small>
                </div>
                <div class="incomplete-tasks-list" id="incompleteTasksList"></div>
                <div class="notification-actions">
                    <button class="notification-btn notification-btn-primary" onclick="closeIncompleteTasksNotification()">知道了</button>
                </div>
            </div>
        </div>
    </div>
    
    <!-- 自定义确认对话框 -->
    <div class="confirm-modal" id="confirmModal">
        <div class="confirm-dialog">
            <h3 id="confirmTitle">确认操作</h3>
            <p id="confirmMessage">确定要执行这个操作吗？</p>
            <div class="confirm-buttons">
                <button class="confirm-btn confirm-btn-cancel" id="confirmCancel">取消</button>
                <button class="confirm-btn confirm-btn-ok" id="confirmOk">确定</button>
            </div>
        </div>
    </div>
    
    <div class="container">
        <div class="header">
            <h1><img src="/static/client.png" style="width: 60px; height: 60px; vertical-align: middle; margin-right: 8px;">局域网文件共享中心</h1>
            <div class="info">
                <p>服务器地址: <strong>http://{{ ip }}:{{ port }}</strong></p>
                <p>局域网内其他设备可以通过上面的地址访问这个页面</p>
            </div>
            <div class="stats">
                <div class="stat-item">
                    <div class="stat-label">文件总数</div>
                    <div class="stat-value">{{ files|length }}</div>
                </div>
                <div class="stat-item">
                    <div class="stat-label">总大小</div>
                    <div class="stat-value">{{ total_size }}</div>
                </div>
            </div>
        </div>

        {% with messages = get_flashed_messages(with_categories=true) %}
            {% if messages %}
                {% for category, message in messages %}
                    <div class="alert alert-{{ category }}">{{ message }}</div>
                {% endfor %}
            {% endif %}
        {% endwith %}

        <div class="upload-section" id="uploadSection">
            <h2>上传文件</h2>
            <div class="drag-hint" id="dragHint" style="display: none;">
                松开鼠标即可上传文件
            </div>
            <form method="post" enctype="multipart/form-data" class="upload-form" id="uploadForm">
                <div class="file-input-wrapper">
                    <input type="file" name="files" id="fileInput" multiple>
                    <label for="fileInput" class="file-input-label">选择文件</label>
                </div>
                <div class="file-input-wrapper">
                    <input type="file" name="folder" id="folderInput" webkitdirectory directory>
                    <label for="folderInput" class="file-input-label">选择文件夹</label>
                </div>
                <select id="uploadPathSelect" name="upload_path" style="padding: 12px 20px; border: 2px solid rgba(255, 255, 255, 0.4); background: rgba(255, 255, 255, 0.15); backdrop-filter: blur(10px); border-radius: 10px; font-size: 14px; color: white; cursor: pointer; outline: none; transition: all 0.3s; min-width: 180px;">
                    <option value="{{ current_path or '' }}" style="background: #1a1a2e; color: white;">当前目录{% if current_path %} ({{ current_path }}){% endif %}</option>
                    {% for folder in all_folders %}
                        {% if folder.path != current_path %}
                        <option value="{{ folder.path }}" style="background: #1a1a2e; color: white;">{{ folder.name }}</option>
                        {% endif %}
                    {% endfor %}
                </select>
                <span class="file-name" id="fileName">未选择</span>
                <button type="button" class="btn btn-secondary" id="cancelBtn" style="display: none;">取消选择</button>
                <button type="submit" class="btn btn-primary">上传</button>
            </form>
            <div style="display: none; margin-top: 15px; padding: 12px; background: rgba(255, 255, 255, 0.15); backdrop-filter: blur(10px); border-radius: 10px; border-left: 4px solid rgba(255, 255, 255, 0.5);">
                <div style="color: rgba(255, 255, 255, 0.95); font-weight: 600; margin-bottom: 8px; text-shadow: 0 1px 3px rgba(0, 0, 0, 0.3);">提示：</div>
                <div style="color: rgba(255, 255, 255, 0.85); font-size: 13px; line-height: 1.6; text-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);">
                    <strong>选择文件</strong>：普通文件最大 100MB，ZIP 文件不受这个限制<br>
                    <strong>选择文件夹</strong>：最多 1000 个文件，并保留目录结构<br>
                    <strong>超大文件夹</strong>：建议先压缩成 ZIP 再上传
                </div>
            </div>
        </div>

        <div class="upload-section folder-section" style="margin-top: 20px;">
            <h2>创建新文件夹</h2>
            <form method="post" action="{{ url_for('create_folder') }}" style="display: flex; gap: 10px; align-items: center;">
                <input type="hidden" name="current_path" value="{{ current_path }}">
                <input type="text" 
                       name="folder_name" 
                       placeholder="输入文件夹名称" 
                       style="flex: 1; padding: 12px 20px; border: 2px solid rgba(255, 255, 255, 0.4); background: rgba(255, 255, 255, 0.15); backdrop-filter: blur(10px); border-radius: 10px; font-size: 14px; outline: none; transition: all 0.3s; color: white;"
                       onfocus="this.style.borderColor='rgba(255, 255, 255, 0.6)'; this.style.boxShadow='0 0 0 3px rgba(255,255,255,0.1)'; this.style.background='rgba(255, 255, 255, 0.2)'"
                       onblur="this.style.borderColor='rgba(255, 255, 255, 0.4)'; this.style.boxShadow='none'; this.style.background='rgba(255, 255, 255, 0.15)'"
                       required>
                <button type="submit" class="btn btn-primary" style="white-space: nowrap;">
                    创建文件夹
                </button>
            </form>
            <div style="display: none; margin-top: 10px; color: rgba(255, 255, 255, 0.75); font-size: 12px; text-shadow: 0 1px 2px rgba(0, 0, 0, 0.3);">
                提示：文件夹名称只能包含中文、字母、数字、下划线和连字符
            </div>
        </div>

        <div class="files-section">
            <h2>共享文件列表</h2>
            
            <!-- 鏂囦欢鎼滅储妗?-->
            <div style="margin-bottom: 20px;">
                <div style="display: flex; gap: 10px; align-items: center;">
                    <div style="flex: 1; position: relative;">
                        <input type="text" 
                               id="searchInput" 
                               placeholder="输入文件名或扩展名进行搜索..." 
                               style="width: 100%; 
                                      padding: 12px 45px 12px 15px; 
                                      border-radius: 12px; 
                                      border: 2px solid rgba(16, 185, 129, 0.3); 
                                      background: rgba(255, 255, 255, 0.1); 
                                      color: white; 
                                      font-size: 14px;
                                      backdrop-filter: blur(10px);
                                      box-sizing: border-box;">
                        <button id="clearSearchBtn" 
                                onclick="clearSearch()" 
                                style="position: absolute; 
                                       right: 10px; 
                                       top: 50%; 
                                       transform: translateY(-50%); 
                                       background: rgba(244, 63, 94, 0.6); 
                                       border: none; 
                                       color: white; 
                                       padding: 6px 12px; 
                                       border-radius: 8px; 
                                       cursor: pointer; 
                                       font-size: 12px;
                                       display: none;">
                            清除
                        </button>
                    </div>
                    <label style="display: flex; align-items: center; gap: 8px; color: rgba(255, 255, 255, 0.9); white-space: nowrap;">
                        <input type="checkbox" id="searchSubfolders" checked style="width: 18px; height: 18px; cursor: pointer;">
                        <span style="font-size: 14px;">搜索子文件夹</span>
                    </label>
                </div>
                <div id="searchResults" style="margin-top: 10px; display: none;">
                    <div style="padding: 10px 15px; background: rgba(16, 185, 129, 0.15); border-radius: 8px; border: 1px solid rgba(16, 185, 129, 0.3);">
                        <span style="color: rgba(255, 255, 255, 0.9); font-size: 13px;">
                            找到 <strong id="searchCount">0</strong> 个结果
                        </span>
                    </div>
                </div>
            </div>
            
            <!-- 面包屑导航 -->
            <div style="margin-bottom: 20px; padding: 15px 20px; background: rgba(255, 255, 255, 0.1); border-radius: 12px; display: flex; align-items: center; gap: 10px; flex-wrap: wrap;">
                <a href="{{ url_for('index') }}" class="breadcrumb-link breadcrumb-drop-target" data-breadcrumb-path="" data-breadcrumb-name="根目录" style="color: rgba(255, 255, 255, 0.9); text-decoration: none; padding: 5px 12px; background: rgba(255, 255, 255, 0.08); border-radius: 8px; transition: all 0.3s;">
                    根目录
                </a>
                {% for crumb in breadcrumbs %}
                    <span style="color: rgba(255, 255, 255, 0.5);">/</span>
                    {% if loop.last %}
                        <span class="breadcrumb-current" style="color: rgba(255, 255, 255, 0.9); padding: 5px 12px; background: rgba(16, 185, 129, 0.2); border-radius: 8px; font-weight: 600;">
                            {{ crumb.name }}
                        </span>
                    {% else %}
                        <a href="{{ url_for('index', subpath=crumb.path) }}" class="breadcrumb-link breadcrumb-drop-target" data-breadcrumb-path="{{ crumb.path }}" data-breadcrumb-name="{{ crumb.name }}" style="color: rgba(255, 255, 255, 0.9); text-decoration: none; padding: 5px 12px; background: rgba(255, 255, 255, 0.08); border-radius: 8px; transition: all 0.3s;">
                            {{ crumb.name }}
                        </a>
                    {% endif %}
                {% endfor %}
            </div>
            
            <!-- 批量操作工具栏 -->
            <div id="batchToolbar" style="display: none; margin-bottom: 15px; padding: 15px 20px; background: rgba(59, 130, 246, 0.15); border-radius: 12px; border: 2px solid rgba(59, 130, 246, 0.3);">
                <div style="display: flex; align-items: center; gap: 15px; flex-wrap: wrap;">
                    <span id="selectedCount" style="color: rgba(255, 255, 255, 0.9); font-weight: 600;">已选择 0 项</span>
                    <button onclick="batchDownload()" class="batch-btn" style="padding: 8px 16px; background: rgba(59, 130, 246, 0.8); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600;">
                        <img src="/static/下载.png" style="width: 28px; height: 28px; vertical-align: middle; margin-right: 4px;">批量下载
                    </button>
                    <button onclick="batchMoveTo()" class="batch-btn" style="padding: 8px 16px; background: rgba(16, 185, 129, 0.8); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600;">
                        <img src="/static/目录.png" style="width: 28px; height: 28px; vertical-align: middle; margin-right: 4px;">批量移动
                    </button>
                    <button onclick="batchDelete()" class="batch-btn" style="padding: 8px 16px; background: rgba(244, 63, 94, 0.8); color: white; border: none; border-radius: 8px; cursor: pointer; font-weight: 600;">
                        <img src="/static/删除.png" style="width: 28px; height: 28px; vertical-align: middle; margin-right: 4px;">批量删除
                    </button>
                    <button onclick="cancelBatchSelection()" style="padding: 8px 16px; background: rgba(107, 114, 128, 0.6); color: white; border: none; border-radius: 8px; cursor: pointer;">
                        取消选择
                    </button>
                </div>
            </div>
            
            {% if files %}
                <!-- 全选控制 -->
                <div style="margin-bottom: 10px; padding: 10px 15px; background: rgba(255, 255, 255, 0.05); border-radius: 8px;">
                    <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; color: rgba(255, 255, 255, 0.9);">
                        <input type="checkbox" id="selectAll" onchange="toggleSelectAll()" style="width: 18px; height: 18px; cursor: pointer;">
                        <span style="font-weight: 600;">全选</span>
                    </label>
                </div>
                
                <ul class="file-list">
                    {% for file in files %}
                        {% if file.is_folder %}
                            {# 文件夹，可点击进入 #}
                            {% set folder_display_name = file.display_name %}
                            {% set folder_path = file.relative_path %}
                            <li class="file-item folder-item" data-folder="{{ folder_path }}" data-filepath="{{ folder_path }}" data-is-folder="true">
                                <input type="checkbox" class="file-checkbox" data-path="{{ folder_path }}" onclick="event.stopPropagation(); updateBatchToolbar();" style="width: 20px; height: 20px; cursor: pointer; margin-right: 10px;">
                                <div class="file-info" style="cursor: pointer;" onclick="window.location.href='{{ url_for('index', subpath=folder_path) }}'">
                                    <div class="file-icon" style="font-size: 36px;">📁</div>
                                    <div class="file-details">
                                        <div class="file-name-text" title="{{ folder_display_name }}">{{ folder_display_name }}</div>
                                        <div class="file-meta">
                                            <span><img src="/static/目录.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">{{ file.size }}</span>
                                            <span>{{ file.time }}</span>
                                    </div>
                                </div>
                                </div>
                                <div class="file-actions">
                                    <button class="btn-download" 
                                            onclick="event.stopPropagation(); renameItem('{{ folder_path }}', '{{ folder_display_name }}', true);"
                                            style="padding: 8px 16px; font-size: 14px;">
                                        <img src="/static/钢笔.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">重命名
                                    </button>
                                    <a href="{{ url_for('download_folder', folder_name=folder_path) }}" 
                                       class="btn-download" 
                                       onclick="event.stopPropagation()">
                                        <img src="/static/下载.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">下载
                                    </a>
                                    <form method="post" 
                                          action="{{ url_for('delete_folder', foldername=folder_path) }}" 
                                          style="display: inline; margin: 0;"
                                          onclick="event.stopPropagation()">
                                        <button type="submit" 
                                                class="btn-delete" 
                                                onclick="event.stopPropagation(); return checkDeletePermission(event, '文件夹', '{{ folder_display_name }}', true);">
                                            <img src="/static/删除.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">删除
                                        </button>
                                    </form>
                                </div>
                            </li>
                        {% else %}
                            {# 文件 #}
                            {% set file_path = file.relative_path %}
                            <li class="file-item" data-filepath="{{ file_path }}" data-is-folder="false">
                            <input type="checkbox" class="file-checkbox" data-path="{{ file_path }}" onclick="event.stopPropagation(); updateBatchToolbar();" style="width: 20px; height: 20px; cursor: pointer; margin-right: 10px;">
                            <div class="file-info">
                                <div class="file-icon">{{ get_file_icon(file.name) }}</div>
                                <div class="file-details">
                                    <div class="file-name-text" title="{{ file_path }}">{{ file.name }}</div>
                                    <div class="file-meta">
                                        <span><img src="/static/目录.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">{{ file.size }}</span>
                                        <span>{{ file.time }}</span>
                                    </div>
                                </div>
                            </div>
                                <div class="file-actions">
                                    {% if is_previewable_file(file.name) %}
                                    <button class="btn-download" 
                                            onclick="openFilePreview('{{ file_path }}', '{{ 'excel' if is_excel_editable_file(file.name) else ('word' if is_word_editable_file(file.name) else ('text' if is_text_previewable_file(file.name) else 'preview')) }}');"
                                            style="padding: 8px 16px; font-size: 14px;">
                                        {% if is_excel_editable_file(file.name) or is_word_editable_file(file.name) or is_text_previewable_file(file.name) %}预览/编辑{% else %}预览{% endif %}
                                    </button>
                                    {% endif %}
                                    <button class="btn-download" 
                                            onclick="createShareLink('{{ file_path }}', '{{ file.name }}');"
                                            style="padding: 8px 16px; font-size: 14px;">
                                        分享
                                    </button>
                                    <button class="btn-download" 
                                            onclick="renameItem('{{ file_path }}', '{{ file.name }}', false);"
                                            style="padding: 8px 16px; font-size: 14px;">
                                        <img src="/static/钢笔.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">重命名
                                    </button>
                                    <a href="{{ url_for('download_file', filename=file_path) }}" class="btn-download"><img src="/static/下载.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">下载</a>
                                    <form method="post" action="{{ url_for('delete_file', filename=file_path) }}" style="display: inline; margin: 0;">
                                        <button type="submit" class="btn-delete" onclick="return checkDeletePermission(event, '文件', '{{ file.name }}');"><img src="/static/删除.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">删除</button>
                                    </form>
                                </div>
                        </li>
                        {% endif %}
                    {% endfor %}
                </ul>
            {% else %}
                <div class="empty-message">
                    <span class="emoji">...</span>
                    <p>暂无文件，请先上传文件</p>
                </div>
            {% endif %}
        </div>
    </div>

    <script>
        // ========== 根据文件名获取图标 ==========
        function getFileIcon(filename) {
            const ext = filename.includes('.') ? filename.split('.').pop().toLowerCase() : '';

            if (['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg', 'ico'].includes(ext)) {
                return '\\uD83D\\uDDBC\\uFE0F';
            }
            if (['mp4', 'avi', 'mkv', 'mov', 'wmv', 'flv', 'webm'].includes(ext)) {
                return '\\uD83C\\uDFAC';
            }
            if (['mp3', 'wav', 'flac', 'aac', 'ogg', 'm4a'].includes(ext)) {
                return '\\uD83C\\uDFB5';
            }
            if (['zip', 'rar', '7z', 'tar', 'gz'].includes(ext)) {
                return '\\uD83D\\uDCE6';
            }
            if (['doc', 'docx'].includes(ext)) {
                return '\\uD83D\\uDCC4';
            }
            if (['xls', 'xlsx', 'xlsm'].includes(ext)) {
                return '\\uD83D\\uDCCA';
            }
            if (['ppt', 'pptx'].includes(ext)) {
                return '\\uD83D\\uDCF9';
            }
            if (ext === 'pdf') {
                return '\\uD83D\\uDCD5';
            }
            if (['py', 'java', 'cpp', 'c', 'h', 'js', 'ts', 'html', 'css'].includes(ext)) {
                return '\\uD83D\\uDCBB';
            }
            if (['txt', 'md', 'json', 'xml', 'yaml', 'yml', 'log', 'csv'].includes(ext)) {
                return '\\uD83D\\uDCDD';
            }
            if (['exe', 'msi', 'apk', 'dmg', 'deb', 'rpm', 'pkg'].includes(ext)) {
                return '\\u2699\\uFE0F';
            }
            if (ext === 'iso') {
                return '\\uD83D\\uDCBF';
            }
            if (['sh', 'bat', 'cmd', 'ps1'].includes(ext)) {
                return '\\uD83E\\uDDF0';
            }
            if (['sql', 'db', 'sqlite', 'mdb'].includes(ext)) {
                return '\\uD83D\\uDDD3\\uFE0F';
            }
            return '\\uD83D\\uDCC1';
        }
        
        // ========== 性能优化工具函数 ==========
        // 防抖函数：延迟执行，适合搜索输入
        function debounce(func, wait) {
            let timeout;
            return function executedFunction(...args) {
                const later = () => {
                    clearTimeout(timeout);
                    func(...args);
                };
                clearTimeout(timeout);
                timeout = setTimeout(later, wait);
            };
        }
        
        // 节流函数：限制执行频率，适合滚动和拖拽
        function throttle(func, limit) {
            let inThrottle;
            return function(...args) {
                if (!inThrottle) {
                    func.apply(this, args);
                    inThrottle = true;
                    setTimeout(() => inThrottle = false, limit);
                }
            };
        }
        
        // 使用requestAnimationFrame优化动画
        function rafThrottle(func) {
            let rafId = null;
            return function(...args) {
                if (rafId) return;
                rafId = requestAnimationFrame(() => {
                    func.apply(this, args);
                    rafId = null;
                });
            };
        }
        
        // ========== 任务管理变量（提前定义） ==========
        let activeTasks = {}; // {task_id: {xhr, file, ...}}
        let localRealtimeUploads = {}; // {task_id: {filename, progress, status, ...}}
        let currentSessionUsername = {{ (session.get('username') or '') | tojson }};
        const passwordMinLength = {{ password_min_length | int }};
        const nicknameMaxLength = {{ nickname_max_length | int }};
        const nicknamePattern = /^[\u3400-\u4dbf\u4e00-\u9fff\uf900-\ufaff]{1,4}$/;
        const sessionLoggedIn = {{ 'true' if session.get('username') else 'false' }};
        const pageSessionId = (() => {
            try {
                const storageKey = 'lanfs_page_session_id';
                let value = sessionStorage.getItem(storageKey);
                if (!value) {
                    value = `page-${Date.now().toString(36)}-${Math.random().toString(36).slice(2, 10)}`;
                    sessionStorage.setItem(storageKey, value);
                }
                return value;
            } catch (error) {
                return `page-fallback-${Date.now().toString(36)}-${Math.random().toString(36).slice(2, 10)}`;
            }
        })();
        function setCurrentUserName(username) {
            currentSessionUsername = username || '';
            const currentUserNameEl = document.getElementById('currentUserName');
            if (currentUserNameEl) {
                currentUserNameEl.textContent = username || '未登录';
            }
        }
        let realtimeEventSource = null;
        let realtimeReconnectTimer = null;

        function escapeHtml(value) {
            return String(value ?? '').replace(/[&<>"']/g, function(char) {
                return {
                    '&': '&amp;',
                    '<': '&lt;',
                    '>': '&gt;',
                    '"': '&quot;',
                    "'": '&#39;'
                }[char];
            });
        }

        function buildUploadDisplayName(filename, uploadPath = '') {
            const cleanName = String(filename || '').replace(/^[/\\\\]+/, '');
            const cleanPath = String(uploadPath || '')
                .replace(/\\\\/g, '/')
                .replace(/^\/+|\/+$/g, '');
            return cleanPath ? `${cleanPath}/${cleanName}` : cleanName;
        }

        function createLocalUploadTaskId(prefix = 'upload') {
            return `${prefix}-${Date.now().toString(36)}-${Math.random().toString(36).slice(2, 8)}`;
        }

        function fetchWithPresence(url, options = {}) {
            const nextOptions = { ...options };
            const headers = new Headers(options.headers || {});
            headers.set('X-Page-Session-Id', pageSessionId);
            nextOptions.headers = headers;
            if (!Object.prototype.hasOwnProperty.call(nextOptions, 'credentials')) {
                nextOptions.credentials = 'same-origin';
            }
            return fetch(url, nextOptions);
        }

        function setLocalRealtimeUpload(taskId, patch) {
            const existing = localRealtimeUploads[taskId] || {};
            localRealtimeUploads[taskId] = {
                ...existing,
                ...patch,
                task_id: taskId
            };
        }

        function removeLocalRealtimeUpload(taskId) {
            if (taskId && localRealtimeUploads[taskId]) {
                delete localRealtimeUploads[taskId];
            }
        }

        function clampProgress(value) {
            const progress = Number(value);
            if (!Number.isFinite(progress)) {
                return 0;
            }
            return Math.max(0, Math.min(100, progress));
        }

        function formatRealtimeProgress(value) {
            return clampProgress(value).toFixed(1);
        }

        function mergeRealtimeActiveTasks(serverTasks) {
            const mergedTasks = [];
            const seenTaskIds = new Set();

            (serverTasks || []).forEach(serverTask => {
                const localTask = localRealtimeUploads[serverTask.task_id] || {};
                const mergedTask = {
                    ...serverTask,
                    filename: buildUploadDisplayName(
                        localTask.displayName || serverTask.filename || '',
                        localTask.displayName ? '' : (serverTask.upload_path || '')
                    )
                };

                if (localTask.username) {
                    mergedTask.username = localTask.username;
                }
                if (localTask.status) {
                    mergedTask.status = localTask.status;
                }
                if (Number.isFinite(Number(localTask.liveProgress))) {
                    mergedTask.progress = clampProgress(localTask.liveProgress);
                }
                if (Number.isFinite(Number(localTask.uploadedChunks))) {
                    mergedTask.uploaded_chunks = Math.max(Number(mergedTask.uploaded_chunks || 0), Number(localTask.uploadedChunks));
                }
                if (Number.isFinite(Number(localTask.totalChunks)) && Number(localTask.totalChunks) > 0) {
                    mergedTask.total_chunks = Number(localTask.totalChunks);
                }
                if (localTask.detailText) {
                    mergedTask.detail_text = localTask.detailText;
                }

                mergedTasks.push(mergedTask);
                seenTaskIds.add(serverTask.task_id);
            });

            Object.entries(localRealtimeUploads).forEach(([taskId, localTask]) => {
                if (seenTaskIds.has(taskId)) {
                    return;
                }
                if (!localTask || localTask.hidden || localTask.status === 'completed' || localTask.status === 'error') {
                    return;
                }

                mergedTasks.push({
                    task_id: taskId,
                    username: localTask.username || currentSessionUsername,
                    type: 'upload',
                    filename: localTask.displayName || localTask.filename || '未命名文件',
                    progress: clampProgress(
                        Number.isFinite(Number(localTask.liveProgress))
                            ? localTask.liveProgress
                            : localTask.progress
                    ),
                    status: localTask.status || 'running',
                    uploaded_chunks: Number.isFinite(Number(localTask.uploadedChunks)) ? Number(localTask.uploadedChunks) : 0,
                    total_chunks: Number.isFinite(Number(localTask.totalChunks)) ? Number(localTask.totalChunks) : 0,
                    detail_text: localTask.detailText || ''
                });
            });

            return mergedTasks;
        }
        
        // ========== 拖拽上传功能 ==========
        const uploadSection = document.getElementById('uploadSection');
        const dragHint = document.getElementById('dragHint');
        const fileInput = document.getElementById('fileInput');
        let dragCounter = 0;  // 防止拖拽事件冒泡导致的闪烁
        // 阻止浏览器默认的拖拽行为
        ['dragenter', 'dragover', 'dragleave', 'drop'].forEach(eventName => {
            document.body.addEventListener(eventName, preventDefaults, false);
            uploadSection.addEventListener(eventName, preventDefaults, false);
            document.documentElement.addEventListener(eventName, preventDefaults, false);
            window.addEventListener(eventName, preventDefaults, false);
        });
        
        function preventDefaults(e) {
            e.preventDefault();
            e.stopPropagation();
        }
        
        // 完全阻止浏览器的默认拖拽处理
        window.addEventListener('drop', function(e) {
            e.preventDefault();
            return false;
        }, true);
        
        window.addEventListener('dragover', function(e) {
            e.preventDefault();
            return false;
        }, true);
        
        // 拖拽进入上传区域
        uploadSection.addEventListener('dragenter', function(e) {
            dragCounter++;
            uploadSection.classList.add('drag-over');
            dragHint.style.display = 'block';
        });
        
        // 拖拽离开上传区域
        uploadSection.addEventListener('dragleave', function(e) {
            dragCounter--;
            if (dragCounter === 0) {
                uploadSection.classList.remove('drag-over');
                dragHint.style.display = 'none';
            }
        });
        
        // 拖拽悬停
        uploadSection.addEventListener('dragover', function(e) {
            e.dataTransfer.dropEffect = 'copy';
        });
        
        // 释放文件
        uploadSection.addEventListener('drop', function(e) {
            dragCounter = 0;
            uploadSection.classList.remove('drag-over');
            dragHint.style.display = 'none';
            
            const files = e.dataTransfer.files;
            if (files.length > 0) {
                handleDroppedFiles(files);
            }
        });
        
        // 处理拖入的文件
        async function handleDroppedFiles(files) {
            console.log('拖入项目数量:', files.length);
            
            const fileArray = Array.from(files);
            
            // 过滤掉文件夹，只保留文件
            const validFiles = fileArray.filter(file => {
                // 文件对象通常会带 size/type，文件夹则常见为 size=0 且 type 为空
                const isFile = file.size > 0 || file.type !== '';
                
                if (!isFile) {
                    console.warn('跳过文件夹:', file.name);
                }
                
                return isFile;
            });
            
            // 如果没有有效文件
            if (validFiles.length === 0) {
                await showAlert(
                    '无法拖拽上传文件夹',
                    '检测到您拖入的是文件夹。\\n\\n浏览器不支持直接拖拽上传整个文件夹。\\n\\n请使用下方的“选择文件夹”按钮上传，这样会保留完整目录结构。'
                );
                return;
            }
            
            // 如果过滤掉了一些项目，提示用户
            if (validFiles.length < fileArray.length) {
                const folderCount = fileArray.length - validFiles.length;
                await showAlert(
                    '⚠️ 提示',
                    `已过滤 ${folderCount} 个文件夹，仅上传 ${validFiles.length} 个文件。\\n\\n文件夹请使用“选择文件夹”按钮上传。`
                );
            }
            
            // 如果文件很多，再次确认
            if (validFiles.length > 20) {
                const confirmed = await showConfirm(
                    '文件数量较多',
                    `即将上传 ${validFiles.length} 个文件到当前目录。\\n\\n是否继续？`
                );
                
                if (!confirmed) {
                    return;
                }
            }
            
            // 直接上传文件，避免浏览器默认拖拽提示
            uploadFilesDirectly(validFiles);
        }
        
        // 直接上传文件列表
        function uploadFilesDirectly(fileArray) {
            fileArray.forEach(file => {
                if (file.size > 10 * 1024 * 1024) {
                    // 大文件使用分块上传
                    const taskId = Date.now().toString() + Math.random().toString(36).substr(2, 9);
                    uploadFileInChunks(file, taskId);
                } else {
                    // 小文件直接上传
                    uploadSmallFile(file);
                }
            });
        }
        
        // 上传文件夹，保留目录结构
        function uploadFolderFiles(fileArray) {
            console.log('上传文件夹，文件总数:', fileArray.length);
            
            showProgress('上传文件夹中...', null);
            updateProgress(0, `准备上传 ${fileArray.length} 个文件...`);
            
            // 使用 API 上传文件夹
            const formData = new FormData();
            
            // 获取用户选择的上传路径
            const uploadPathSelect = document.getElementById('uploadPathSelect');
            const selectedPath = uploadPathSelect ? uploadPathSelect.value : '{{ current_path }}';
            formData.append('upload_path', selectedPath || '');
            const folderName = (fileArray[0] && fileArray[0].webkitRelativePath)
                ? fileArray[0].webkitRelativePath.split('/')[0]
                : '文件夹上传';
            const localTaskId = createLocalUploadTaskId('folder-upload');
            setLocalRealtimeUpload(localTaskId, {
                type: 'upload',
                status: 'running',
                username: currentSessionUsername,
                filename: folderName,
                displayName: buildUploadDisplayName(folderName, selectedPath || ''),
                detailText: `准备上传 ${fileArray.length} 个文件`,
                liveProgress: 0
            });
            updateActivities();
            
            // 添加所有文件，保留相对路径
            fileArray.forEach(file => {
                // 使用 webkitRelativePath 保留目录结构
                const path = file.webkitRelativePath || file.name;
                formData.append('folder', file, path);
            });
            
            const xhr = new XMLHttpRequest();
            
            // 总体进度监听
            let uploadedBytes = 0;
            xhr.upload.addEventListener('progress', function(e) {
                if (e.lengthComputable) {
                    const percent = Math.round((e.loaded / e.total) * 100);
                    const uploadedMB = (e.loaded / 1024 / 1024).toFixed(2);
                    const totalMB = (e.total / 1024 / 1024).toFixed(2);
                    setLocalRealtimeUpload(localTaskId, {
                        liveProgress: percent,
                        detailText: `已上传 ${uploadedMB}MB / ${totalMB}MB`
                    });
                    updateProgress(percent, `已上传 ${uploadedMB}MB / ${totalMB}MB`);
                }
            });
            
            xhr.onload = function() {
                removeLocalRealtimeUpload(localTaskId);
                updateActivities();
                if (xhr.status === 200) {
                    updateProgress(100, '上传完成');
                    setTimeout(() => window.location.reload(), 500);
                } else {
                    hideProgress();
                    console.error('文件夹上传失败，状态码:', xhr.status);
                    showAlert('上传失败', '文件夹上传失败，请重试');
                }
            };
            
            xhr.onerror = function() {
                removeLocalRealtimeUpload(localTaskId);
                updateActivities();
                hideProgress();
                console.error('文件夹上传出错');
                showAlert('上传失败', '上传过程中发生错误');
            };
            
            // 使用当前页面路径作为上传地址
            const currentPath = '{{ current_path }}';
            const uploadUrl = currentPath ? `/browse/${currentPath}` : '/';
            
            xhr.open('POST', uploadUrl, true);
            xhr.send(formData);
        }
        
        // 上传小文件
        function uploadSmallFile(file) {
            // 检查当前是否有大文件正在上传
            const hasActiveChunkUpload = Object.keys(activeTasks).some(tid => {
                const t = activeTasks[tid];
                return t && t.type === 'upload' && !t.paused;
            });
            
            if (!hasActiveChunkUpload) {
                showProgress('上传中...', null);
            }
            
            const formData = new FormData();
            formData.append('files', file);
            const uploadPathSelect = document.getElementById('uploadPathSelect');
            const selectedPath = uploadPathSelect ? uploadPathSelect.value : '{{ current_path }}';
            const localTaskId = createLocalUploadTaskId('small-upload');
            setLocalRealtimeUpload(localTaskId, {
                type: 'upload',
                status: 'running',
                username: currentSessionUsername,
                filename: file.name,
                displayName: buildUploadDisplayName(file.name, selectedPath || ''),
                detailText: '准备上传...',
                liveProgress: 0
            });
            updateActivities();
            
            const xhr = new XMLHttpRequest();
            
            xhr.upload.addEventListener('progress', function(e) {
                if (hasActiveChunkUpload) return;
                if (e.lengthComputable) {
                    const percent = Math.round((e.loaded / e.total) * 100);
                    const uploadedMB = (e.loaded / 1024 / 1024).toFixed(2);
                    const totalMB = (e.total / 1024 / 1024).toFixed(2);
                    setLocalRealtimeUpload(localTaskId, {
                        liveProgress: percent,
                        detailText: `已上传 ${uploadedMB}MB / ${totalMB}MB`
                    });
                    updateProgress(percent, `上传中 ${file.name}`);
                }
            });
            
            xhr.onload = function() {
                console.log('上传响应状态:', xhr.status);
                removeLocalRealtimeUpload(localTaskId);
                updateActivities();
                
                if (xhr.status === 200) {
                    try {
                        const response = JSON.parse(xhr.responseText);
                        console.log('上传成功:', response);
                        
                        if (!hasActiveChunkUpload) {
                            updateProgress(100, '上传完成');
                        }
                        
                        // 如果有失败的文件，输出警告
                        if (response.failed && response.failed.length > 0) {
                            console.warn('部分文件上传失败:', response.failed);
                        }
                        
                        setTimeout(() => window.location.reload(), 500);
                    } catch (e) {
                        console.error('解析响应失败:', e);
                        setTimeout(() => window.location.reload(), 500);
                    }
                } else {
                    if (!hasActiveChunkUpload) {
                        hideProgress();
                    }
                    console.error('上传失败，状态码:', xhr.status, '响应:', xhr.responseText);
                    
                    try {
                        const response = JSON.parse(xhr.responseText);
                        showAlert('上传失败', response.message || '上传失败，请重试');
                    } catch (e) {
                        showAlert('上传失败', '上传失败，请重试 (状态码: ' + xhr.status + ')');
                    }
                }
            };
            
            xhr.onerror = function(e) {
                console.error('上传错误详情:', e);
                console.error('XHR״̬:', xhr.status, xhr.statusText);
                console.error('XHR响应:', xhr.responseText);
                removeLocalRealtimeUpload(localTaskId);
                updateActivities();
                if (!hasActiveChunkUpload) {
                    hideProgress();
                }
                showAlert('上传失败', '上传过程中发生错误，详情请查看控制台（F12）');
            };
            
            xhr.onabort = function() {
                console.log('上传已中止');
                removeLocalRealtimeUpload(localTaskId);
                updateActivities();
            };
            
            xhr.onloadstart = function() {
                console.log('开始上传文件:', file.name);
            };
            
            xhr.onloadend = function() {
                console.log('上传结束，状态:', xhr.status);
            };
            
            // 使用 API 上传，返回 JSON，不做页面跳转
            const uploadUrl = selectedPath ? `/api/upload_files/${selectedPath}` : '/api/upload_files';
            console.log('上传到:', uploadUrl, '文件:', file.name);
            
            xhr.open('POST', uploadUrl, true);
            xhr.send(formData);
        }
        
        // 也可以拖拽到整个页面
        document.body.addEventListener('dragenter', function(e) {
            if (e.target === document.body || e.target === document.documentElement) {
                uploadSection.classList.add('drag-over');
                dragHint.style.display = 'block';
            }
        });
        
        // ========== 管理员权限系统 ==========
        let isAdmin = false;
        let isHost = false;

        function applyAdminState(data = {}) {
            isAdmin = Boolean(data.is_admin);
            isHost = Boolean(data.is_host);

            const adminBadge = document.getElementById('adminBadge');
            const requestAdminBtn = document.getElementById('requestAdminBtn');
            const adminRequestsPanel = document.getElementById('adminRequestsPanel');

            if (adminBadge) {
                adminBadge.style.display = isAdmin ? 'block' : 'none';
            }
            if (requestAdminBtn) {
                requestAdminBtn.style.display = (!isAdmin && !isHost) ? 'block' : 'none';
            }
            if (adminRequestsPanel) {
                adminRequestsPanel.style.display = isHost ? 'block' : 'none';
            }
        }

        function renderAdminRequests(requests) {
            const safeRequests = (requests && typeof requests === 'object') ? requests : {};
            const count = Object.keys(safeRequests).length;
            const countEl = document.getElementById('adminRequestsCount');
            const container = document.getElementById('adminRequests');

            if (countEl) {
                countEl.textContent = count;
            }
            if (!container) {
                return;
            }

            container.innerHTML = '';
            if (count === 0) {
                container.innerHTML = '<div style="color: rgba(255,255,255,0.7); padding: 10px; text-align: center; font-size: 12px;">暂无申请</div>';
                return;
            }

            for (const [requestId, req] of Object.entries(safeRequests)) {
                const div = document.createElement('div');
                div.style.cssText = `
                    padding: 10px;
                    margin-bottom: 8px;
                    background: rgba(255, 215, 0, 0.15);
                    border-radius: 8px;
                    border: 1px solid rgba(255, 215, 0, 0.3);
                `;

                const timestamp = req && req.timestamp ? new Date(req.timestamp).toLocaleString('zh-CN') : '未知时间';
                const requesterName = escapeHtml(req && req.username ? req.username : '未知用户');
                const requesterIp = escapeHtml(req && req.ip ? req.ip : '未知');

                div.innerHTML = `
                    <div style="color: white; font-weight: 600; margin-bottom: 5px;">
                        用户 ${requesterName}
                    </div>
                    <div style="color: rgba(255,255,255,0.7); font-size: 11px; margin-bottom: 8px;">
                        IP: ${requesterIp}<br>
                        时间: ${escapeHtml(timestamp)}
                    </div>
                    <div style="display: flex; gap: 5px;">
                        <button onclick="approveAdminRequest('${requestId}')"
                                style="flex: 1;
                                       padding: 5px 10px;
                                       background: rgba(16, 185, 129, 0.3);
                                       border: 1px solid rgba(16, 185, 129, 0.5);
                                       color: white;
                                       border-radius: 6px;
                                       cursor: pointer;
                                       font-size: 11px;
                                       transition: all 0.3s;">
                            同意
                        </button>
                        <button onclick="rejectAdminRequest('${requestId}')"
                                style="flex: 1;
                                       padding: 5px 10px;
                                       background: rgba(244, 63, 94, 0.3);
                                       border: 1px solid rgba(244, 63, 94, 0.5);
                                       color: white;
                                       border-radius: 6px;
                                       cursor: pointer;
                                       font-size: 11px;
                                       transition: all 0.3s;">
                            拒绝
                        </button>
                    </div>
                `;

                container.appendChild(div);
            }
        }
        
        // 检查管理员权限
        async function checkAdminPermission(forceRefresh = false) {
            if (!forceRefresh && realtimeEventSource && realtimeEventSource.readyState === EventSource.OPEN) {
                return;
            }
            try {
                const response = await fetchWithPresence('/api/check_admin');
                const data = await response.json();
                applyAdminState(data);
                renderAdminRequests(Boolean(data.is_host) ? (data.requests || {}) : {});
                
                // 控制台输出权限检查结果，便于排查权限问题
                console.log('权限检查结果:', {
                    '是否管理员': Boolean(data.is_admin),
                    '是否主机': Boolean(data.is_host),
                    '昵称': data.username,
                    '客户端IP': data.client_ip,
                    '本机IP': data.local_ip
                });
            } catch (error) {
                console.error('检查管理员权限失败:', error);
            }
        }
        
        // 申请管理员权限
        async function requestAdminPermission() {
            const confirmed = await showConfirm(
                '申请管理员',
                '申请成为管理员后，您将获得删除文件和文件夹的权限。\\n\\n申请会发送给主机审批，是否继续？'
            );
            
            if (!confirmed) return;
            
            try {
                const response = await fetchWithPresence('/api/request_admin', { method: 'POST' });
                const data = await response.json();
                
                if (data.success) {
                    await showAlert('申请已提交', data.message);
                    document.getElementById('requestAdminBtn').disabled = true;
                    document.getElementById('requestAdminBtn').textContent = '等待审批...';
                } else {
                    await showAlert('申请失败', data.message);
                }
            } catch (error) {
                console.error('申请管理员失败:', error);
                await showAlert('错误', '申请失败: ' + error.message);
            }
        }
        
        // 更新管理员申请列表，仅主机可见
        async function updateAdminRequests() {
            if (!isHost) {
                renderAdminRequests({});
                return;
            }
            
            try {
                const response = await fetchWithPresence('/api/admin_requests');
                const data = await response.json();
                renderAdminRequests(data && data.success ? data.requests : {});
            } catch (error) {
                console.error('获取管理员申请失败:', error);
                renderAdminRequests({});
            }
        }
        
        // 批准管理员申请
        async function approveAdminRequest(requestId) {
            try {
                const response = await fetchWithPresence(`/api/admin_approve/${requestId}`, { method: 'POST' });
                const data = await response.json();
                
                if (data.success) {
                    await showAlert('已批准', data.message);
                    checkAdminPermission(true);
                } else {
                    await showAlert('失败', data.message);
                }
            } catch (error) {
                console.error('批准失败:', error);
                await showAlert('错误', '批准失败');
            }
        }
        
        // 拒绝管理员申请
        async function rejectAdminRequest(requestId) {
            const confirmed = await showConfirm('拒绝申请', '确定要拒绝这个管理员申请吗？');
            if (!confirmed) return;
            
            try {
                const response = await fetchWithPresence(`/api/admin_reject/${requestId}`, { method: 'POST' });
                const data = await response.json();
                
                if (data.success) {
                    await showAlert('已拒绝', data.message);
                    checkAdminPermission(true);
                } else {
                    await showAlert('失败', data.message);
                }
            } catch (error) {
                console.error('拒绝失败:', error);
                await showAlert('错误', '拒绝失败');
            }
        }
        
        // 检查删除权限
        async function checkDeletePermission(event, itemType, itemName, isFolder = false) {
            event.preventDefault();
            
            // 检查当前用户是否为管理员
            if (!isAdmin) {
                await showAlert(
                    '权限不足',
                    `只有管理员才能删除${itemType}。\\n\\n请点击右上角的“申请管理员”按钮申请权限。`
                );
                return false;
            }
            
            // 管理员删除前仍然进行二次确认
            const message = isFolder 
                ? `确定要删除文件夹“${itemName}”及其全部内容吗？此操作不可恢复。`
                : `确定要删除文件“${itemName}”吗？`;
            
            const result = await showConfirm(`删除${itemType}`, message);
            
            if (result) {
                // 用户确认后提交删除表单
                event.target.closest('form').submit();
            }
            
            return false;
        }
        
        // 页面加载后先检查一次管理员权限
        checkAdminPermission();
        
        // 定期刷新管理员申请和权限状态
        setInterval(() => {
            if (isHost) {
                updateAdminRequests();
            }
            checkAdminPermission();
        }, 5000);
        
        // ========== 鼠标光晕跟随效果 ==========
        const mouseGlow = document.getElementById('mouseGlow');
        let mouseX = 0;
        let mouseY = 0;
        
        document.addEventListener('mousemove', (e) => {
            mouseX = e.clientX;
            mouseY = e.clientY;
            
            if (mouseGlow) {
                mouseGlow.style.left = mouseX + 'px';
                mouseGlow.style.top = mouseY + 'px';
            }
        });
        
        // 自定义提示对话框函数（只有确定按钮）
        function showAlert(title, message) {
            return new Promise((resolve) => {
                const modal = document.getElementById('confirmModal');
                const titleEl = document.getElementById('confirmTitle');
                const messageEl = document.getElementById('confirmMessage');
                const cancelBtn = document.getElementById('confirmCancel');
                const okBtn = document.getElementById('confirmOk');
                
                titleEl.textContent = title || '提示';
                messageEl.textContent = message || '';
                
                // 隐藏取消按钮
                cancelBtn.style.display = 'none';
                
                modal.classList.add('show');
                
                const handleOk = () => {
                    modal.classList.remove('show');
                    okBtn.removeEventListener('click', handleOk);
                    // 恢复取消按钮显示
                    cancelBtn.style.display = 'inline-block';
                    resolve(true);
                };
                
                okBtn.addEventListener('click', handleOk);
                
                // 点击背景关闭
                modal.addEventListener('click', function(e) {
                    if (e.target === modal) {
                        handleOk();
                    }
                });
            });
        }
        
        // 自定义确认对话框函数（有确定和取消按钮）
        function showConfirm(title, message) {
            return new Promise((resolve) => {
                const modal = document.getElementById('confirmModal');
                const titleEl = document.getElementById('confirmTitle');
                const messageEl = document.getElementById('confirmMessage');
                const cancelBtn = document.getElementById('confirmCancel');
                const okBtn = document.getElementById('confirmOk');
                
                titleEl.textContent = title || '确认';
                messageEl.textContent = message || '';
                
                // 显示取消按钮
                cancelBtn.style.display = 'inline-block';
                
                modal.classList.add('show');
                
                const handleOk = () => {
                    cleanup();
                    resolve(true);
                };
                
                const handleCancel = () => {
                    cleanup();
                    resolve(false);
                };
                
                const cleanup = () => {
                    modal.classList.remove('show');
                    okBtn.removeEventListener('click', handleOk);
                    cancelBtn.removeEventListener('click', handleCancel);
                };
                
                okBtn.addEventListener('click', handleOk);
                cancelBtn.addEventListener('click', handleCancel);
                
                // 点击背景关闭（视为取消）
                modal.addEventListener('click', function(e) {
                    if (e.target === modal) {
                        handleCancel();
                    }
                });
            });
        }
        
        // 页面加载后清理已完成任务，并初始化登录表单状态
        window.addEventListener('DOMContentLoaded', function() {
            setCurrentUserName(currentSessionUsername);
            const usernameInput = document.getElementById('usernameInput');
            if (!sessionLoggedIn && usernameInput) {
                usernameInput.focus();
            }
            
            // 页面加载后自动清理已完成任务
            setTimeout(function() {
                fetch('/api/all_tasks')
                    .then(response => response.json())
                    .then(data => {
                        if (data.success && data.tasks) {
                            Object.keys(data.tasks).forEach(async (taskId) => {
                                const task = data.tasks[taskId];
                                // 仅清理已经完成的任务，避免刷新后残留
                                if (task.status === 'completed') {
                                    console.log('清理已完成任务:', taskId);
                                    await deleteTask(taskId, false);
                                }
                            });
                        }
                    })
                    .catch(error => console.error('清理任务失败:', error));
            }, 1000);
        });
        
        // 切换面板区域折叠/展开
        function toggleSection(sectionId, headerElement) {
            const section = document.getElementById(sectionId);
            const isCollapsed = section.classList.contains('collapsed');
            
            if (isCollapsed) {
                section.classList.remove('collapsed');
                headerElement.classList.remove('collapsed');
            } else {
                section.classList.add('collapsed');
                headerElement.classList.add('collapsed');
            }
        }
        
        // 登录处理
        const loginForm = document.getElementById('loginForm');
        if (loginForm) {
            loginForm.addEventListener('submit', function(e) {
                e.preventDefault();
                const username = document.getElementById('usernameInput').value.trim();
                const password = document.getElementById('passwordInput').value;
                const mode = (document.getElementById('loginModeInput').value || 'login').trim().toLowerCase();
                
                if (!nicknamePattern.test(username) || username.length > nicknameMaxLength) {
                    showAlert('输入错误', `昵称必须是 1 到 ${nicknameMaxLength} 个中文字符。`);
                    return;
                }

                if (!password || password.length < passwordMinLength) {
                    showAlert('输入错误', `密码长度不能少于 ${passwordMinLength} 个字符。`);
                    return;
                }
                
                const formData = new FormData(loginForm);
                formData.set('username', username);
                formData.set('password', password);
                formData.set('mode', mode === 'register' ? 'register' : 'login');
                
                fetchWithPresence('/login', {
                    method: 'POST',
                    body: formData
                })
                .then(async response => {
                    try {
                        return await response.json();
                    } catch (error) {
                        console.warn('登录响应不是 JSON，改走原生表单提交。', error);
                        HTMLFormElement.prototype.submit.call(loginForm);
                        return null;
                    }
                })
                .then(data => {
                    if (!data) {
                        return;
                    }
                    if (data.success) {
                        setCurrentUserName(data.username || username);
                        window.location.reload();
                    } else {
                        showAlert(mode === 'register' ? '注册失败' : '登录失败', data.message);
                    }
                })
                .catch(error => {
                    console.error('登录失败，改走原生表单提交。', error);
                    HTMLFormElement.prototype.submit.call(loginForm);
                });
            });
        }

        async function logoutCurrentUser() {
            const confirmed = await showConfirm('退出登录', '确定要退出当前登录吗？');
            if (!confirmed) {
                return;
            }

            try {
                const response = await fetchWithPresence('/logout', { method: 'POST' });
                const data = await response.json();
                if (!data.success) {
                    await showAlert('退出失败', data.message || '退出登录失败。');
                    return;
                }

                if (realtimeEventSource) {
                    realtimeEventSource.close();
                    realtimeEventSource = null;
                }
                window.location.reload();
            } catch (error) {
                console.error('退出登录失败:', error);
                await showAlert('退出失败', '网络异常或服务器不可用。');
            }
        }
        
        // 更新在线用户列表
        function updateOnlineUsers() {
            fetchWithPresence('/get_online_users')
            .then(response => response.json())
            .then(data => renderOnlineUsers(data.users || []))
            .catch(error => console.error('获取在线用户失败:', error));
        }
        
        // 更新实时活动
        function updateActivities() {
            fetchWithPresence('/get_activities')
            .then(response => response.json())
            .then(data => renderActivities(data.activities || [], data.active_tasks || []))
            .catch(error => {
                console.error('获取活动失败:', error);
                // 显示兜底提示，但不影响下次刷新
                const container = document.getElementById('activities');
                if (container && !container.querySelector('.activity-item')) {
                    container.innerHTML = '<div style="color: rgba(255,255,255,0.7); padding: 10px; text-align: center; font-size: 12px;">暂无动态</div>';
                }
            });
        }

        function renderOnlineUsers(users) {
            const safeUsers = Array.isArray(users) ? users : [];
            document.getElementById('onlineCount').textContent = safeUsers.length;
            const container = document.getElementById('onlineUsers');
            if (!container) {
                return;
            }
            container.innerHTML = '';
            safeUsers.forEach(user => {
                const div = document.createElement('div');
                div.className = 'user-item';
                div.innerHTML = `<strong>${escapeHtml(user && user.username ? user.username : '匿名用户')}</strong>`;
                container.appendChild(div);
            });
            if (safeUsers.length === 0) {
                container.innerHTML = '<div style="color: rgba(255,255,255,0.7); padding: 10px; text-align: center; font-size: 12px;">暂无在线用户</div>';
            }
        }

        function renderActivities(activities, activeTasks) {
            const container = document.getElementById('activities');
            if (!container) {
                return;
            }

            const refreshing = container.querySelector('.refreshing-indicator');
            if (!refreshing) {
                const indicator = document.createElement('div');
                indicator.className = 'refreshing-indicator';
                indicator.style.cssText = `
                    position: absolute;
                    top: 5px;
                    right: 5px;
                    width: 8px;
                    height: 8px;
                    background: rgba(16, 185, 129, 0.8);
                    border-radius: 50%;
                    animation: pulse 0.5s ease-in-out;
                `;
                container.style.position = 'relative';
                container.appendChild(indicator);
            }

            container.innerHTML = '';
            let renderedItems = 0;

            const updateIndicator = document.createElement('div');
            updateIndicator.style.cssText = `
                font-size: 10px;
                color: rgba(255,255,255,0.5);
                text-align: right;
                margin-bottom: 8px;
                padding: 2px 4px;
            `;
            updateIndicator.textContent = `最后更新 ${new Date().toLocaleTimeString()}`;
            container.appendChild(updateIndicator);

            const mergedActiveTasks = mergeRealtimeActiveTasks(activeTasks || []);
            if (mergedActiveTasks.length > 0) {
                mergedActiveTasks.forEach(task => {
                    const div = document.createElement('div');
                    div.className = 'activity-item active-task';
                    div.style.cssText = `
                        background: rgba(16, 185, 129, 0.15);
                        border: 1px solid rgba(16, 185, 129, 0.3);
                        padding: 12px;
                        border-radius: 8px;
                        margin-bottom: 8px;
                        word-wrap: break-word;
                        overflow-wrap: break-word;
                        overflow: hidden;
                        max-width: 100%;
                        box-sizing: border-box;
                    `;

                    let icon = '上传中';
                    let actionText = '正在上传';
                    if (task.status === 'paused') {
                        icon = '已暂停';
                        actionText = '已暂停';
                    }

                    const progressText = formatRealtimeProgress(task.progress);
                    const progressDetail = task.detail_text
                        || ((Number(task.total_chunks) > 0)
                            ? `${Number(task.uploaded_chunks || 0)} / ${Number(task.total_chunks)} 分块`
                            : '传输中');

                    const progressHtml = `
                        <div style="margin-top: 6px; width: 100%; box-sizing: border-box;">
                            <div style="display: flex; justify-content: space-between; margin-bottom: 4px; width: 100%; box-sizing: border-box;">
                                <span style="font-size: 11px; color: rgba(255,255,255,0.8); flex-shrink: 0;">
                                    ${escapeHtml(progressDetail)}
                                </span>
                                <span style="font-size: 11px; color: rgba(255,255,255,0.8); flex-shrink: 0;">
                                    ${progressText}%
                                </span>
                            </div>
                            <div style="width: 100%; height: 6px; background: rgba(0,0,0,0.2); border-radius: 3px; overflow: hidden; box-sizing: border-box;">
                                <div style="width: ${clampProgress(task.progress)}%; height: 100%; background: linear-gradient(90deg, #10b981, #059669); border-radius: 3px; transition: width 0.3s ease; max-width: 100%;"></div>
                            </div>
                        </div>
                    `;

                    div.innerHTML = `
                        <div style="display: flex; justify-content: space-between; align-items: flex-start; width: 100%; box-sizing: border-box;">
                            <div style="flex: 1; min-width: 0; margin-right: 8px;">
                                <div style="margin-bottom: 4px; word-wrap: break-word; overflow-wrap: break-word;">
                                    ${icon} <strong>${escapeHtml(task.username || currentSessionUsername)}</strong> ${escapeHtml(actionText)}
                                </div>
                                <div style="font-size: 12px; color: rgba(255,255,255,0.9); word-break: break-all; overflow: hidden; text-overflow: ellipsis;">
                                    文件 ${escapeHtml(task.filename || '未命名文件')}
                                </div>
                            </div>
                        </div>
                        ${progressHtml}
                    `;

                    container.appendChild(div);
                    renderedItems += 1;
                });

                if (Array.isArray(activities) && activities.length > 0) {
                    const divider = document.createElement('div');
                    divider.style.cssText = `
                        height: 1px;
                        background: rgba(255, 255, 255, 0.2);
                        margin: 12px 0;
                    `;
                    container.appendChild(divider);
                }
            }

            if (Array.isArray(activities)) {
                activities.slice(-10).reverse().forEach(activity => {
                    const div = document.createElement('div');
                    div.className = 'activity-item';
                    let icon = '动态';
                    if (activity.action === '上传') icon = '上传';
                    else if (activity.action === '下载') icon = '下载';
                    else if (activity.action === '上线') icon = '在线';
                    else if (activity.action === '删除') icon = '<img src="/static/删除.png" style="width: 28px; height: 28px; vertical-align: middle;">';
                    else if (activity.action === '重命名') icon = '<img src="/static/钢笔.png" style="width: 28px; height: 28px; vertical-align: middle;">';

                    const timeAgo = formatTimeAgo(activity.timestamp);
                    const fileInfo = activity.filename ? ` ${escapeHtml(activity.filename)}` : '';
                    div.innerHTML = `${icon} <strong>${escapeHtml(activity.username || '匿名用户')}</strong> ${escapeHtml(activity.action || '动态')}${fileInfo}<div class="activity-time">${escapeHtml(timeAgo)}</div>`;
                    container.appendChild(div);
                    renderedItems += 1;
                });
            }

            if (renderedItems === 0) {
                container.innerHTML = '<div style="color: rgba(255,255,255,0.7); padding: 10px; text-align: center; font-size: 12px;">暂无动态</div>';
            }
        }
        
        // 格式化时间
        function formatTimeAgo(timestamp) {
            const now = Date.now() / 1000;
            const diff = now - timestamp;
            if (diff < 60) return '刚刚';
            if (diff < 3600) return Math.floor(diff / 60) + ' 分钟前';
            if (diff < 86400) return Math.floor(diff / 3600) + ' Сʱǰ';
            return Math.floor(diff / 86400) + ' 天前';
        }

        let pollingStarted = false;
        let onlineUsersPollTimer = null;
        let activitiesPollTimer = null;
        let realtimeFallbackStarted = false;
        let realtimeFallbackOnlineTimer = null;
        let realtimeFallbackActivitiesTimer = null;
        let realtimeFallbackAdminTimer = null;

        function handleRealtimeSnapshot(payload) {
            if (!payload || typeof payload !== 'object') {
                return;
            }
            if (payload.current_username) {
                setCurrentUserName(payload.current_username);
            }
            if (payload.admin) {
                applyAdminState(payload.admin);
                renderAdminRequests(Boolean(payload.admin.is_host) ? (payload.admin.requests || {}) : {});
            }
            renderOnlineUsers(payload.online_users || []);
            renderActivities(payload.activities || [], payload.active_tasks || []);
        }

        function clearRealtimeFallback() {
            if (realtimeFallbackOnlineTimer) {
                clearInterval(realtimeFallbackOnlineTimer);
                realtimeFallbackOnlineTimer = null;
            }
            if (realtimeFallbackActivitiesTimer) {
                clearInterval(realtimeFallbackActivitiesTimer);
                realtimeFallbackActivitiesTimer = null;
            }
            if (realtimeFallbackAdminTimer) {
                clearInterval(realtimeFallbackAdminTimer);
                realtimeFallbackAdminTimer = null;
            }
            realtimeFallbackStarted = false;
        }

        function startPollingFallback() {
            if (realtimeFallbackStarted) {
                return;
            }
            realtimeFallbackStarted = true;
            updateOnlineUsers();
            updateActivities();
            checkAdminPermission(true);
            realtimeFallbackOnlineTimer = setInterval(updateOnlineUsers, 4000);
            realtimeFallbackActivitiesTimer = setInterval(updateActivities, 2500);
            realtimeFallbackAdminTimer = setInterval(() => checkAdminPermission(true), 12000);
        }

        function scheduleRealtimeReconnect() {
            if (realtimeReconnectTimer) {
                return;
            }
            realtimeReconnectTimer = setTimeout(() => {
                realtimeReconnectTimer = null;
                if (document.hidden) {
                    return;
                }
                startRealtimeUpdates();
            }, 3000);
        }

        function stopRealtimeStream() {
            if (realtimeEventSource) {
                realtimeEventSource.close();
                realtimeEventSource = null;
            }
            if (realtimeReconnectTimer) {
                clearTimeout(realtimeReconnectTimer);
                realtimeReconnectTimer = null;
            }
        }

        function startRealtimeStream() {
            if (document.hidden) {
                return;
            }
            if (!window.EventSource) {
                startPollingFallback();
                return;
            }
            if (realtimeEventSource) {
                return;
            }

            clearRealtimeFallback();

            const streamUrl = `/api/realtime_stream?page_session_id=${encodeURIComponent(pageSessionId)}`;
            realtimeEventSource = new EventSource(streamUrl);

            realtimeEventSource.onopen = function() {
                if (realtimeReconnectTimer) {
                    clearTimeout(realtimeReconnectTimer);
                    realtimeReconnectTimer = null;
                }
            };

            realtimeEventSource.addEventListener('snapshot', function(event) {
                try {
                    handleRealtimeSnapshot(JSON.parse(event.data || '{}'));
                } catch (error) {
                    console.error('实时动态快照解析失败:', error);
                }
            });

            realtimeEventSource.onerror = function() {
                if (realtimeEventSource) {
                    realtimeEventSource.close();
                    realtimeEventSource = null;
                }
                startPollingFallback();
                scheduleRealtimeReconnect();
            };
        }

        function startRealtimeUpdates() {
            updateOnlineUsers();
            updateActivities();
            checkAdminPermission(true);
            if (window.EventSource) {
                startRealtimeStream();
                return;
            }
            startPollingFallback();
        }
        
        // 启动轮询
        function startPolling() {
            if (pollingStarted) {
                return;
            }
            pollingStarted = true;
            updateOnlineUsers();
            updateActivities();
            onlineUsersPollTimer = setInterval(updateOnlineUsers, 3000);  // 每 3 秒刷新一次在线用户
            activitiesPollTimer = setInterval(updateActivities, 3000);    // SSE 失败后再用低频轮询兜底
        }
        
        // 如果已登录，启动轮询
        {% if session.get('username') %}
        startRealtimeUpdates();
        
        // 监听浏览器关闭事件，发送离线通知
        window.addEventListener('beforeunload', function(e) {
            // 使用 navigator.sendBeacon 发送离线通知，即使页面关闭也尽量送达
            const formData = new FormData();
            formData.append('page_session_id', pageSessionId);
            navigator.sendBeacon('/offline', formData);
        });
        
        // 页面隐藏时也发送离线通知，例如切换标签页
        document.addEventListener('visibilitychange', function() {
            if (document.hidden) {
                stopRealtimeStream();
                clearRealtimeFallback();
            } else {
                startRealtimeUpdates();
            }
        });
        {% endif %}
        
        // 文件夹折叠/展开功能
        function toggleFolder(element) {
            const folderGroup = element;
            // 使用 data-folder 获取准确路径
            const folderName = folderGroup.getAttribute('data-folder') || '';
            const isCollapsed = folderGroup.classList.contains('collapsed');
            
            // 切换折叠状态
            folderGroup.classList.toggle('collapsed');
            
            // 该文件夹和其所有子层级文件
            const allFiles = document.querySelectorAll('.folder-files');
            allFiles.forEach(fileItem => {
                const itemFolder = fileItem.getAttribute('data-folder') || '';
                if (itemFolder === folderName || itemFolder.startsWith(folderName + '/')) {
                    if (isCollapsed) {
                        fileItem.classList.remove('hidden');
                    } else {
                        fileItem.classList.add('hidden');
                    }
                }
            });

            // 同时处理所有子文件夹的显示/隐藏
            const allGroups = document.querySelectorAll('.folder-group');
            allGroups.forEach(group => {
                if (group === folderGroup) return;
                const groupName = group.getAttribute('data-folder') || '';
                if (groupName.startsWith(folderName + '/')) {
                    if (isCollapsed) {
                        group.classList.remove('hidden');
                    } else {
                        group.classList.add('hidden');
                        // 折叠父级时，强制子级也折叠
                        group.classList.add('collapsed');
                    }
                }
            });
        }
        
        // 显示/隐藏取消按钮
        function updateCancelButton() {
            const fileInput = document.getElementById('fileInput');
            const folderInput = document.getElementById('folderInput');
            const cancelBtn = document.getElementById('cancelBtn');
            
            if (fileInput.files.length > 0 || folderInput.files.length > 0) {
                cancelBtn.style.display = 'inline-block';
            } else {
                cancelBtn.style.display = 'none';
            }
        }

        function resetUploadSelectionLabel() {
            const fileNameEl = document.getElementById('fileName');
            if (fileNameEl) {
                fileNameEl.textContent = '未选择';
            }
        }

        // 显示选中的文件名
        document.getElementById('fileInput').addEventListener('change', function(e) {
            const fileCount = e.target.files.length;
            if (fileCount > 0) {
                if (fileCount === 1) {
                    const file = e.target.files[0];
                    const sizeMB = (file.size / 1024 / 1024).toFixed(1);
                    document.getElementById('fileName').textContent = `${file.name} (${sizeMB}MB)`;
                } else {
                    document.getElementById('fileName').textContent = `已选择 ${fileCount} 个文件`;
                }
                // 清空文件夹选择
                document.getElementById('folderInput').value = '';
            } else {
                resetUploadSelectionLabel();
            }
            updateCancelButton();
        });

        // 显示选中的文件夹
        document.getElementById('folderInput').addEventListener('change', function(e) {
            const fileCount = e.target.files.length;
            if (fileCount > 0) {
                const folderName = e.target.files[0].webkitRelativePath.split('/')[0];
                if (fileCount > 1000) {
                    document.getElementById('fileName').textContent = `文件夹 ${folderName} (${fileCount} 个文件 - 超过限制)`;
                    showAlert('警告', `文件夹包含 ${fileCount} 个文件，超过 1000 个限制！\\n\\n建议：\\n1. 先压缩成 ZIP 后再上传（不受此限制）\\n2. 或分批上传`);
                } else {
                    document.getElementById('fileName').textContent = `文件夹 ${folderName} (${fileCount} 个文件)`;
                }
                // 清空文件选择
                document.getElementById('fileInput').value = '';
            } else {
                resetUploadSelectionLabel();
            }
            updateCancelButton();
        });

        // 取消选择按钮
        document.getElementById('cancelBtn').addEventListener('click', function() {
            document.getElementById('fileInput').value = '';
            document.getElementById('folderInput').value = '';
            resetUploadSelectionLabel();
            updateCancelButton();
        });
        
        // 当前显示在顶部进度条的任务ID（用于分块上传）
        let currentProgressTaskId = null;
        
        // 显示进度条（不阻止点击）
        function showProgress(title, taskId = null) {
            // 不显示遮罩层，允许继续点击其他按钮
            document.getElementById('progressOverlay').style.display = 'none';
            document.getElementById('progressContainer').style.display = 'block';
            document.getElementById('progressTitle').textContent = title;
            document.getElementById('progressBar').style.width = '0%';
            document.getElementById('progressBar').textContent = '0%';
            document.getElementById('progressInfo').textContent = '准备中...';
            // 记录当前显示的任务ID
            if (taskId) {
                currentProgressTaskId = taskId;
                // 显示操作按钮，仅分块上传任务使用
                document.getElementById('progressActions').style.display = 'flex';
                updateProgressButtons();
            } else {
                // 小文件上传不显示操作按钮；如果当前已有大文件上传，则不覆盖它
                if (currentProgressTaskId !== null) {
                    // 已有大文件上传时，不覆盖进度条
                    return;
                }
                document.getElementById('progressActions').style.display = 'none';
            }
        }
        
        // 更新进度条按钮状态
        function updateProgressButtons() {
            const taskId = currentProgressTaskId;
            const pauseBtn = document.getElementById('progressPauseBtn');
            const resumeBtn = document.getElementById('progressResumeBtn');
            const cancelBtn = document.getElementById('progressCancelBtn');
            
            if (!taskId || !activeTasks[taskId]) {
                pauseBtn.style.display = 'none';
                resumeBtn.style.display = 'none';
                return;
            }
            
            const task = activeTasks[taskId];
            const isPaused = task.paused || task.pauseRequested;
            
            if (isPaused) {
                pauseBtn.style.display = 'none';
                resumeBtn.style.display = 'block';
            } else {
                pauseBtn.style.display = 'block';
                resumeBtn.style.display = 'none';
            }
        }
        
        // 处理进度条暂停
        function handleProgressPause() {
            if (currentProgressTaskId) {
                pauseTask(currentProgressTaskId);
            }
        }
        
        // 处理进度条继续
        function handleProgressResume() {
            if (currentProgressTaskId) {
                resumeTask(currentProgressTaskId);
            }
        }
        
        // 处理进度条取消
        async function handleProgressCancel() {
            if (!currentProgressTaskId) {
                hideProgress();
                return;
            }
            
            const result = await showConfirm('取消任务', '确定要取消当前上传任务吗？已上传的部分将会丢失。');
            if (result) {
                // 删除任务
                await deleteTask(currentProgressTaskId, false);
                hideProgress();
            }
        }
        
        // 隐藏进度条
        function hideProgress() {
            document.getElementById('progressOverlay').style.display = 'none';
            document.getElementById('progressContainer').style.display = 'none';
            currentProgressTaskId = null;
        }
        
        // 更新进度条
        function updateProgress(percent, info, speed, taskId = null) {
            // 如果指定了 taskId，只更新当前正在显示的那个任务
            if (taskId !== null && currentProgressTaskId !== null && taskId !== currentProgressTaskId) {
                // 不是当前显示的任务，不更新进度条
                return;
            }
            
            // 小文件上传未指定 taskId 时，如果当前有大文件任务正在显示，就不要覆盖进度条
            if (taskId === null && currentProgressTaskId !== null) {
                return;
            }
            
            const progressBar = document.getElementById('progressBar');
            progressBar.style.width = percent + '%';
            progressBar.textContent = percent + '%';
            if (info !== undefined && info !== null) {
                document.getElementById('progressInfo').textContent = info;
            }
            if (speed !== undefined && speed !== null) {
                document.getElementById('progressSpeed').innerHTML = `速度: <span class="speed-value">${speed}</span>`;
            } else if (speed === null) {
                // speed 为 null 表示清空速度显示
                document.getElementById('progressSpeed').innerHTML = '';
            }
            // speed 为 undefined 表示保持当前速度显示不变
        }
        
        // 格式化速度显示
        function formatSpeed(bytesPerSecond) {
            if (!bytesPerSecond || bytesPerSecond <= 0) {
                return '0 B/s';
            }
            if (bytesPerSecond < 1024) {
                return bytesPerSecond.toFixed(0) + ' B/s';
            } else if (bytesPerSecond < 1024 * 1024) {
                return (bytesPerSecond / 1024).toFixed(2) + ' KB/s';
            } else if (bytesPerSecond < 1024 * 1024 * 1024) {
                return (bytesPerSecond / 1024 / 1024).toFixed(2) + ' MB/s';
            } else {
                return (bytesPerSecond / 1024 / 1024 / 1024).toFixed(2) + ' GB/s';
            }
        }
        
        // 计算剩余时间
        function formatTimeRemaining(seconds) {
            if (seconds < 60) {
                return Math.ceil(seconds) + '秒';
            } else if (seconds < 3600) {
                const minutes = Math.floor(seconds / 60);
                const secs = Math.ceil(seconds % 60);
                return minutes + '分' + secs + '秒';
            } else {
                const hours = Math.floor(seconds / 3600);
                const minutes = Math.floor((seconds % 3600) / 60);
                return hours + '小时' + minutes + '分钟';
            }
        }
        
        // 拖放功能：把文件或文件夹移动到目标文件夹
        let draggedElement = null;
        let draggedFilename = null;
        let draggedIsFolder = false;
        
        // 为所有文件项和文件夹添加拖动功能
        function enableFileDrag() {
            document.querySelectorAll('.file-item').forEach(function(fileItem) {
                const nameEl = fileItem.querySelector('.file-name-text');
                if (nameEl) {
                    fileItem.setAttribute('draggable', 'true');
                    fileItem.classList.add('draggable');
                    
                    fileItem.addEventListener('dragstart', function(e) {
                        draggedElement = this;
                        draggedIsFolder = this.classList.contains('folder-item');
                        
                        // 文件夹使用 data-folder，普通文件优先使用 title 中的完整路径
                        if (draggedIsFolder) {
                            draggedFilename = this.getAttribute('data-folder') || '';
                        } else {
                            const titlePath = nameEl.getAttribute('title');
                            draggedFilename = (titlePath || nameEl.textContent).trim();
                        }
                        
                        this.classList.add('dragging');
                        e.dataTransfer.effectAllowed = 'move';
                        e.dataTransfer.setData('text/html', this.innerHTML);
                        
                        // 显示拖拽提示框
                        const itemType = draggedIsFolder ? '文件夹' : '文件';
                        const itemName = nameEl.textContent.trim();
                        const dragHintBox = document.getElementById('dragHintBox');
                        dragHintBox.innerHTML = `拖拽 ${itemType} “${itemName}”<br>到目标文件夹`;
                        dragHintBox.style.display = 'block';
                        
                        console.log(`开始拖拽${itemType}: ${itemName}`);
                    });
                    
                    fileItem.addEventListener('dragend', function(e) {
                        this.classList.remove('dragging');
                        // 隐藏拖拽提示框
                        const dragHintBox = document.getElementById('dragHintBox');
                        dragHintBox.style.display = 'none';
                    });
                }
            });
        }
        
        // 为所有文件夹添加拖放目标能力
        function enableFolderDrop() {
            // 文件夹目标：移动到指定文件夹
            document.querySelectorAll('.folder-item').forEach(function(folder) {
                folder.addEventListener('dragover', function(e) {
                    if (draggedElement && !this.classList.contains('dragging')) {
                        e.preventDefault();
                        e.dataTransfer.dropEffect = 'move';
                        this.classList.add('drag-over');
                    }
                });
                
                folder.addEventListener('dragleave', function(e) {
                    this.classList.remove('drag-over');
                });
                
                folder.addEventListener('drop', function(e) {
                    e.preventDefault();
                    this.classList.remove('drag-over');
                    
                    if (!draggedElement || !draggedFilename) return;
                    
                    // 不能把项目拖到自己身上
                    if (this === draggedElement) {
                        showAlert('提示', '不能将项目移动到自身上');
                        return;
                    }
                    
                    // 直接从 data-folder 读取准确的文件夹名
                    let targetFolder = (this.getAttribute('data-folder') || '').trim();
                    
                    if (!targetFolder) {
                        showAlert('错误', '无法识别目标文件夹');
                        return;
                    }
                    
                    // 获取被拖拽项目和目标文件夹名称
                    const draggedItemName = draggedElement.querySelector('.file-name-text').textContent.trim();
                    const targetFolderName = this.querySelector('.file-name-text').textContent.trim();
                    const itemType = draggedIsFolder ? '文件夹' : '文件';
                    
                    // 显示确认提示
                    const confirmMsg = `确定要将${itemType}“${draggedItemName}”移动到“${targetFolderName}”吗？`;
                    showConfirm('确认移动', confirmMsg).then(confirmed => {
                        if (!confirmed) {
                            return;
                        }
                        
                        // 发送移动请求
                        showProgress('移动中...', null);
                        updateProgress(50, `正在移动${itemType}...`);
                        
                        fetch('/move_file', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({
                                filename: draggedFilename,
                                target_folder: targetFolder
                            })
                        })
                        .then(response => response.json())
                        .then(data => {
                            updateProgress(100, '移动完成');
                            setTimeout(function() {
                                hideProgress();
                                if (data.success) {
                                    showAlert('成功', data.message).then(() => window.location.reload());
                                } else {
                                    showAlert('失败', data.message);
                                }
                            }, 500);
                        })
                        .catch(error => {
                            hideProgress();
                            console.error('Error:', error);
                            showAlert('错误', '移动文件时发生错误');
                        });
                        
                        draggedElement = null;
                        draggedFilename = null;
                    });
                });
            });

            // 禁用拖到空白区域的默认行为（现在使用面包屑和文件夹拖放）
            const filesSection = document.querySelector('.files-section');
            if (filesSection) {
                filesSection.addEventListener('dragover', function(e) {
                    if (!draggedElement) return;
                    // 如果悬停在文件夹或面包屑上，交由它们处理
                    if (e.target && e.target.closest) {
                        if (e.target.closest('.folder-item') || e.target.closest('.breadcrumb-drop-target')) {
                            return;
                        }
                    }
                    // 空白区域不接受拖放，这里不阻止默认行为
                });
            }
        }
        
        // 为面包屑导航添加拖放目标功能
        function enableBreadcrumbDrop() {
            document.querySelectorAll('.breadcrumb-drop-target').forEach(function(breadcrumb) {
                breadcrumb.addEventListener('dragover', function(e) {
                    if (draggedElement) {
                        e.preventDefault();
                        e.dataTransfer.dropEffect = 'move';
                        this.classList.add('drag-over-breadcrumb');
                    }
                });
                
                breadcrumb.addEventListener('dragleave', function(e) {
                    this.classList.remove('drag-over-breadcrumb');
                });
                
                breadcrumb.addEventListener('drop', function(e) {
                    e.preventDefault();
                    this.classList.remove('drag-over-breadcrumb');
                    
                    if (!draggedElement || !draggedFilename) return;
                    
                    const targetPath = this.getAttribute('data-breadcrumb-path') || '';
                    const targetName = this.getAttribute('data-breadcrumb-name') || '根目录';
                    
                    // 获取被拖拽项目名称
                    const draggedItemName = draggedElement.querySelector('.file-name-text').textContent.trim();
                    const itemType = draggedIsFolder ? '文件夹' : '文件';
                    
                    // 显示移动确认对话框
                    const confirmMsg = `确定要将${itemType}“${draggedItemName}”移动到“${targetName}”吗？`;
                    showConfirm('确认移动', confirmMsg).then(confirmed => {
                        if (!confirmed) {
                            return;
                        }
                        
                        // 发送移动请求
                        showProgress('移动中...', null);
                        updateProgress(50, `正在移动${itemType}到${targetName}...`);
                        
                        fetch('/move_file', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json',
                            },
                            body: JSON.stringify({
                                filename: draggedFilename,
                                target_folder: targetPath
                            })
                        })
                        .then(response => response.json())
                        .then(data => {
                            updateProgress(100, '移动完成');
                            setTimeout(function() {
                                hideProgress();
                                if (data.success) {
                                    showAlert('成功', data.message).then(() => window.location.reload());
                                } else {
                                    showAlert('失败', data.message);
                                }
                            }, 500);
                        })
                        .catch(error => {
                            hideProgress();
                            console.error('Error:', error);
                            showAlert('错误', '移动文件时发生错误');
                        });
                        
                        draggedElement = null;
                        draggedFilename = null;
                    });
                });
            });
        }
        
        // 初始化拖放功能
        enableFileDrag();
        enableFolderDrop();
        enableBreadcrumbDrop();
        
        // ========== 批量操作功能 ==========
        // 全选 / 取消全选
        function toggleSelectAll() {
            const selectAllCheckbox = document.getElementById('selectAll');
            const checkboxes = document.querySelectorAll('.file-checkbox');
            checkboxes.forEach(cb => {
                cb.checked = selectAllCheckbox.checked;
                if (cb.checked) {
                    cb.closest('.file-item').classList.add('selected');
                } else {
                    cb.closest('.file-item').classList.remove('selected');
                }
            });
            updateBatchToolbar();
        }
        
        // 更新批量操作工具栏
        function updateBatchToolbar() {
            const checkboxes = document.querySelectorAll('.file-checkbox:checked');
            const toolbar = document.getElementById('batchToolbar');
            const countSpan = document.getElementById('selectedCount');
            const selectAllCheckbox = document.getElementById('selectAll');
            
            if (checkboxes.length > 0) {
                toolbar.style.display = 'block';
                countSpan.textContent = `已选择 ${checkboxes.length} 项`;
                
                // 更新选中项的样式
                document.querySelectorAll('.file-checkbox').forEach(cb => {
                    if (cb.checked) {
                        cb.closest('.file-item').classList.add('selected');
                    } else {
                        cb.closest('.file-item').classList.remove('selected');
                    }
                });
                
                // 更新全选复选框状态
                const allCheckboxes = document.querySelectorAll('.file-checkbox');
                selectAllCheckbox.checked = (checkboxes.length === allCheckboxes.length);
            } else {
                toolbar.style.display = 'none';
                selectAllCheckbox.checked = false;
                document.querySelectorAll('.file-item').forEach(item => {
                    item.classList.remove('selected');
                });
            }
        }
        
        // 获取选中的文件路径
        function getSelectedPaths() {
            const checkboxes = document.querySelectorAll('.file-checkbox:checked');
            return Array.from(checkboxes).map(cb => cb.getAttribute('data-path'));
        }
        
        // 取消选择
        function cancelBatchSelection() {
            document.querySelectorAll('.file-checkbox').forEach(cb => {
                cb.checked = false;
            });
            updateBatchToolbar();
        }
        
        // 批量删除
        function batchDelete() {
            const paths = getSelectedPaths();
            if (paths.length === 0) {
                showAlert('⚠️ 提示', '请先选择要删除的项目');
                return;
            }
            
            const confirmMsg = `确定要删除选中的 ${paths.length} 项吗？此操作不可撤销。`;
            showConfirm('<img src="/static/删除.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">批量删除', confirmMsg).then(confirmed => {
                if (!confirmed) return;
                
                showProgress('<img src="/static/删除.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">删除中...');
                updateProgress(30, '正在删除文件...');
                
                fetch('/api/batch_delete', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ paths: paths })
                })
                .then(r => r.json())
                .then(data => {
                    updateProgress(100, '删除完成！');
                    setTimeout(() => {
                        hideProgress();
                        if (data.success) {
                            let message = data.message;
                            if (data.failed_items && data.failed_items.length > 0) {
                                message += '\\n\\n失败项：\\n' + data.failed_items.join('\\n');
                            }
                            showAlert('✅ 完成', message).then(() => window.location.reload());
                        } else {
                            showAlert('❌ 失败', data.message);
                        }
                    }, 500);
                })
                .catch(err => {
                    hideProgress();
                    console.error(err);
                    showAlert('❌ 错误', '批量删除时发生错误');
                });
            });
        }
        
        // 批量移动
        function batchMoveTo() {
            const paths = getSelectedPaths();
            if (paths.length === 0) {
                showAlert('⚠️ 提示', '请先选择要移动的项目');
                return;
            }
            
            // 创建目标文件夹选择对话框
            const folders = [];
            folders.push({ path: '', name: '根目录' });
            
            // 获取所有文件夹
            document.querySelectorAll('.folder-item').forEach(folder => {
                const path = folder.getAttribute('data-folder');
                const name = folder.querySelector('.file-name-text').textContent.trim();
                if (path && !paths.includes(path)) {  // 排除选中的文件夹
                    folders.push({ path: path, name: `📁 ${name}` });
                }
            });
            
            // 构建对话框HTML
            let optionsHTML = folders.map(f => 
                `<option value="${f.path}">${f.name}</option>`
            ).join('');
            
            const modal = document.getElementById('confirmModal');
            const titleEl = document.getElementById('confirmTitle');
            const messageEl = document.getElementById('confirmMessage');
            const cancelBtn = document.getElementById('confirmCancel');
            const okBtn = document.getElementById('confirmOk');
            
            titleEl.innerHTML = '<img src="/static/下载.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">批量移动';
            messageEl.innerHTML = `
                <div style="margin-bottom: 15px;">已选择 ${paths.length} 项，请选择目标文件夹：</div>
                <select id="batchMoveTarget" style="width: 100%; padding: 10px; border-radius: 8px; border: 2px solid rgba(255, 255, 255, 0.3); background: rgba(255, 255, 255, 0.1); color: white; font-size: 16px;">
                    ${optionsHTML}
                </select>
            `;
            
            cancelBtn.style.display = 'inline-block';
            modal.classList.add('show');
            
            const handleOk = () => {
                const targetFolder = document.getElementById('batchMoveTarget').value;
                cleanup();
                
                showProgress('<img src="/static/下载.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">移动中...');
                updateProgress(30, '正在移动文件...');
                
                fetch('/api/batch_move', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({ paths: paths, target_folder: targetFolder })
                })
                .then(r => r.json())
                .then(data => {
                    updateProgress(100, '移动完成！');
                    setTimeout(() => {
                        hideProgress();
                        if (data.success) {
                            let message = data.message;
                            if (data.failed_items && data.failed_items.length > 0) {
                                message += '\\n\\n失败项：\\n' + data.failed_items.join('\\n');
                            }
                            showAlert('✅ 完成', message).then(() => window.location.reload());
                        } else {
                            showAlert('❌ 失败', data.message);
                        }
                    }, 500);
                })
                .catch(err => {
                    hideProgress();
                    console.error(err);
                    showAlert('❌ 错误', '批量移动时发生错误');
                });
            };
            
            const handleCancel = () => {
                cleanup();
            };
            
            const cleanup = () => {
                modal.classList.remove('show');
                okBtn.removeEventListener('click', handleOk);
                cancelBtn.removeEventListener('click', handleCancel);
            };
            
            okBtn.addEventListener('click', handleOk);
            cancelBtn.addEventListener('click', handleCancel);
        }
        
        // 批量下载（优化：使用隐藏表单，避免页面跳转中断下载）
        function batchDownload() {
            const paths = getSelectedPaths();
            if (paths.length === 0) {
                showAlert('⚠️ 提示', '请先选择要下载的项目');
                return;
            }
            
            // 显示提示
            const pathsCount = paths.length;
            showProgress('📦 正在打包文件...');
            updateProgress(10, `正在准备打包 ${pathsCount} 个项目，请勿关闭或刷新页面...`);
            
            // 模拟进度增长
            let currentProgress = 10;
            const progressInterval = setInterval(() => {
                currentProgress += 5;
                if (currentProgress >= 85) {
                    clearInterval(progressInterval);
                    updateProgress(85, '正在压缩大文件，请耐心等待...');
                } else {
                    updateProgress(currentProgress, '正在打包文件，请勿切换页面...');
                }
            }, 800);
            
            // 使用 XMLHttpRequest 以便更好地控制下载
            const xhr = new XMLHttpRequest();
            xhr.open('POST', '/api/batch_download', true);
            xhr.setRequestHeader('Content-Type', 'application/json');
            xhr.responseType = 'blob';
            
            xhr.onload = function() {
                clearInterval(progressInterval);
                
                if (xhr.status === 200) {
                    updateProgress(95, '准备下载...');
                    
                    const blob = xhr.response;
                    const url = window.URL.createObjectURL(blob);
                    const a = document.createElement('a');
                    a.href = url;
                    a.download = `批量下载_${new Date().getTime()}.zip`;
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                    window.URL.revokeObjectURL(url);
                    
                    updateProgress(100, '下载完成！');
                    setTimeout(() => {
                        hideProgress();
                        cancelBatchSelection();
                    }, 1000);
                } else {
                    hideProgress();
                    showAlert('❌ 错误', '批量下载失败');
                }
            };
            
            xhr.onerror = function() {
                clearInterval(progressInterval);
                hideProgress();
                showAlert('❌ 错误', '批量下载时发生网络错误，请检查网络连接');
            };
            
            xhr.onabort = function() {
                clearInterval(progressInterval);
                hideProgress();
                showAlert('⚠️ 已取消', '批量下载已被取消');
            };
            
            // 发送请求
            xhr.send(JSON.stringify({ paths: paths }));
            
            // 阻止页面跳转（警告用户）
            const beforeUnloadHandler = (e) => {
                e.preventDefault();
                e.returnValue = '正在下载文件，确定要离开吗？';
                return e.returnValue;
            };
            
            window.addEventListener('beforeunload', beforeUnloadHandler);
            
            // 下载完成后移除警告
            xhr.addEventListener('loadend', () => {
                setTimeout(() => {
                    window.removeEventListener('beforeunload', beforeUnloadHandler);
                }, 2000);
            });
        }
        
        // 文件预览
        function openFilePreview(filePath, mode = 'preview') {
            let basePath = '/preview/';
            if (mode === 'text') {
                basePath = '/edit/';
            } else if (mode === 'word') {
                basePath = '/edit_docx/';
            } else if (mode === 'excel') {
                basePath = '/edit_excel/';
            }
            window.open(basePath + encodeURIComponent(filePath), '_blank');
        }
        
        function previewFile(filePath) {
            openFilePreview(filePath, 'preview');
        }
        
        // 创建分享链接
        function createShareLink(filePath, fileName) {
            const modal = document.getElementById('confirmModal');
            const titleEl = document.getElementById('confirmTitle');
            const messageEl = document.getElementById('confirmMessage');
            const cancelBtn = document.getElementById('confirmCancel');
            const okBtn = document.getElementById('confirmOk');
            
            titleEl.innerHTML = '🔗 创建分享链接';
            messageEl.innerHTML = `
                <div style="text-align: left; margin-bottom: 15px;">
                    <div style="margin-bottom: 10px; font-weight: 600; color: rgba(255, 255, 255, 0.9);">文件：${fileName}</div>
                    <label style="display: block; margin-bottom: 10px; color: rgba(255, 255, 255, 0.8);">
                        过期时间（小时）：
                        <input type="number" id="shareExpireHours" value="24" min="1" max="168" 
                               style="width: 100%; padding: 8px; border-radius: 6px; border: 2px solid rgba(255, 255, 255, 0.3); background: rgba(255, 255, 255, 0.1); color: white; margin-top: 5px;">
                    </label>
                    <label style="display: block; margin-bottom: 10px; color: rgba(255, 255, 255, 0.8);">
                        密码（可选）：
                        <input type="text" id="sharePassword" placeholder="留空则无需密码" 
                               style="width: 100%; padding: 8px; border-radius: 6px; border: 2px solid rgba(255, 255, 255, 0.3); background: rgba(255, 255, 255, 0.1); color: white; margin-top: 5px;">
                    </label>
                    <label style="display: block; color: rgba(255, 255, 255, 0.8);">
                        最大下载次数（可选）：
                        <input type="number" id="shareMaxDownloads" placeholder="留空则不限制" min="1" 
                               style="width: 100%; padding: 8px; border-radius: 6px; border: 2px solid rgba(255, 255, 255, 0.3); background: rgba(255, 255, 255, 0.1); color: white; margin-top: 5px;">
                    </label>
                </div>
            `;
            
            cancelBtn.style.display = 'inline-block';
            modal.classList.add('show');
            
            const handleOk = () => {
                const expireHours = document.getElementById('shareExpireHours').value;
                const password = document.getElementById('sharePassword').value;
                const maxDownloads = document.getElementById('shareMaxDownloads').value;
                
                cleanup();
                showProgress('🔗 生成分享链接...');
                
                fetch('/api/create_share', {
                    method: 'POST',
                    headers: { 'Content-Type': 'application/json' },
                    body: JSON.stringify({
                        file_path: filePath,
                        expire_hours: expireHours,
                        password: password,
                        max_downloads: maxDownloads
                    })
                })
                .then(r => r.json())
                .then(data => {
                    hideProgress();
                    if (data.success) {
                        const shareInfo = `
                            <div style="text-align: left;">
                                <div style="margin-bottom: 15px; padding: 12px; background: rgba(16, 185, 129, 0.15); border-radius: 8px; border: 2px solid rgba(16, 185, 129, 0.3);">
                                    <div style="color: rgba(255, 255, 255, 0.8); margin-bottom: 8px; font-weight: 600;">分享链接：</div>
                                    <div style="color: rgba(255, 255, 255, 0.95); word-break: break-all; font-family: monospace; font-size: 14px;">${data.share_url}</div>
                                </div>
                                ${password ? `<div style="margin-bottom: 10px; color: rgba(255, 255, 255, 0.9);">🔐 密码：<strong>${password}</strong></div>` : ''}
                                <div style="color: rgba(255, 255, 255, 0.8);">⏰ ${data.expires_in} 小时后过期</div>
                                ${maxDownloads ? `<div style="color: rgba(255, 255, 255, 0.8);">📥 最多下载 ${maxDownloads} 次</div>` : ''}
                            </div>
                        `;
                        
                        showAlert('✅ 分享链接已创建', shareInfo).then(() => {
                            // 复制链接到剪贴板
                            navigator.clipboard.writeText(data.share_url).catch(() => {});
                        });
                    } else {
                        showAlert('❌ 失败', data.message);
                    }
                })
                .catch(err => {
                    hideProgress();
                    console.error(err);
                    showAlert('❌ 错误', '创建分享链接失败');
                });
            };
            
            const handleCancel = () => {
                cleanup();
            };
            
            const cleanup = () => {
                modal.classList.remove('show');
                okBtn.removeEventListener('click', handleOk);
                cancelBtn.removeEventListener('click', handleCancel);
            };
            
            okBtn.addEventListener('click', handleOk);
            cancelBtn.addEventListener('click', handleCancel);
        }
        
        // 为所有下载按钮添加进度条
        document.querySelectorAll('.btn-download').forEach(function(btn) {
            btn.addEventListener('click', function(e) {
                // 如果是文件夹下载，显示特殊提示
                if (this.href.includes('/download_folder/')) {
                    e.preventDefault();
                    showProgress('<img src="/static/下载.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">打包中...');
                    updateProgress(30, '正在压缩文件夹，请稍候...');
                    
                    // 创建隐藏的 iframe 来下载
                    const iframe = document.createElement('iframe');
                    iframe.style.display = 'none';
                    iframe.src = this.href;
                    document.body.appendChild(iframe);
                    
                    // 模拟进度
                    let progress = 30;
                    const interval = setInterval(function() {
                        progress += 10;
                        if (progress >= 90) {
                            clearInterval(interval);
                            updateProgress(90, '即将完成...');
                        } else {
                            updateProgress(progress, '正在压缩文件夹，请稍候...');
                        }
                    }, 500);
                    
                    // 5秒后隐藏进度条（假设下载已开始）
                    setTimeout(function() {
                        clearInterval(interval);
                        updateProgress(100, '下载已开始！');
                        setTimeout(hideProgress, 1000);
                        document.body.removeChild(iframe);
                    }, 5000);
                } else {
                    // 普通文件下载不显示进度条，直接开始下载
                    // 让浏览器直接处理下载
                }
            });
        });
        
        // ========== 任务管理功能 ==========
        // activeTasks 已在前面定义
        
        // 切换任务面板显示
        // 任务管理面板已移除
        function toggleTasksPanel() {
            // 任务管理面板已移除，保留空函数避免旧按钮报错
            return;
        }
        
        // 更新任务列表
        function updateTasksList() {
            fetch('/api/tasks')
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        renderTasks(data.tasks);
                        // 如果任务列表为空，保留面板显示，由用户手动关闭
                        if (Object.keys(data.tasks).length === 0) {
                            // no-op
                        }
                    }
                })
                .catch(error => console.error('获取任务列表失败:', error));
        }
        
        // 渲染任务列表
        function renderTasks(tasks) {
            const container = document.getElementById('tasksList');
            container.innerHTML = '';
            
            if (Object.keys(tasks).length === 0) {
                container.innerHTML = '<div style="color: rgba(255,255,255,0.7); text-align: center; padding: 20px;">暂无任务</div>';
                return;
            }
            
            Object.entries(tasks).forEach(([taskId, task]) => {
                const taskDiv = document.createElement('div');
                taskDiv.className = 'task-item';
                taskDiv.id = `task-${taskId}`;
                
                // 计算进度，优先使用服务端返回的 progress，其次根据分块数计算
                let progress = task.progress || 0;
                if (task.total_chunks && task.uploaded_chunks !== undefined) {
                    const calculatedProgress = (task.uploaded_chunks / task.total_chunks) * 100;
                    if (calculatedProgress > progress || (progress === 0 && task.uploaded_chunks > 0)) {
                        progress = calculatedProgress;
                    }
                }

                const statusText = task.status === 'running' ? '运行中' : task.status === 'paused' ? '已暂停' : '已完成';
                const isActive = !!activeTasks[taskId];

                // 如果前端仍持有活动任务，用前端的实时进度覆盖服务端快照
                if (isActive && activeTasks[taskId] && activeTasks[taskId].uploadedChunks !== undefined && activeTasks[taskId].totalChunks) {
                    progress = (activeTasks[taskId].uploadedChunks / activeTasks[taskId].totalChunks) * 100;
                }
                
                taskDiv.innerHTML = `
                    <div class="task-header">
                        <div class="task-name" title="${task.filename}">${task.type === 'upload' ? '⬆️' : '⬇️'} ${task.filename}</div>
                        <span class="task-status ${task.status}">${statusText}${!isActive && task.status === 'running' ? ' (已断开)' : ''}</span>
                    </div>
                    <div class="task-progress">
                        <div class="task-progress-bar" style="width: ${progress}%"></div>
                    </div>
                    <div style="color: rgba(255,255,255,0.8); font-size: 11px; margin-bottom: 8px;">
                        ${progress.toFixed(1)}% (${isActive && activeTasks[taskId] ? activeTasks[taskId].uploadedChunks : (task.uploaded_chunks || 0)}/${task.total_chunks || 1})
                    </div>
                    <div class="task-actions">
                        ${task.status === 'running' && isActive ? 
                            `<button class="task-btn task-btn-pause" onclick="pauseTask('${taskId}')">⏸️ 暂停</button>` :
                            task.status === 'paused' || (!isActive && task.status === 'running') ?
                            `<button class="task-btn task-btn-resume" onclick="resumeTask('${taskId}')">▶️ 继续</button>` : ''
                        }
                        <button class="task-btn task-btn-delete" onclick="deleteTask('${taskId}')"><img src="/static/删除.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">删除</button>
                    </div>
                `;
                
                container.appendChild(taskDiv);
            });
        }
        
        // 暂停任务
        async function pauseTask(taskId) {
            const task = activeTasks[taskId];

            console.log('暂停任务:', taskId);
            console.log('任务对象:', task);

            if (task) {
                task.paused = true;
                task.pauseRequested = true;
                console.log('已设置暂停标志，文件对象是否存在:', !!task.file);

                const uploadedBytes = Math.min(task.file?.size || 0, (task.uploadedChunks || 0) * 10 * 1024 * 1024);
                setLocalRealtimeUpload(taskId, {
                    status: 'paused',
                    totalChunks: task.totalChunks || 0,
                    uploadedChunks: task.uploadedChunks || 0,
                    liveProgress: task.totalChunks ? ((task.uploadedChunks || 0) / task.totalChunks) * 100 : 0,
                    detailText: task.file
                        ? `已暂停 · 已上传 ${(uploadedBytes / 1024 / 1024).toFixed(2)}MB / ${(task.file.size / 1024 / 1024).toFixed(2)}MB`
                        : '已暂停'
                });
                updateActivities();
            }

            if (task && task.xhr) {
                try {
                    task.xhr.abort();
                    task.xhr = null;
                    console.log('已中止 xhr 请求');
                } catch (e) {
                    console.warn('中止请求时出错:', e);
                }
            }

            // 不要删除 activeTasks[taskId]，保留文件对象以便恢复
            if (task && task.type === 'upload' && currentProgressTaskId === taskId) {
                const progress = task.uploadedChunks ? ((task.uploadedChunks / task.totalChunks) * 100) : 0;
                const uploadedMB = (task.uploadedChunks * 10).toFixed(2);
                const totalMB = ((task.file?.size || 0) / 1024 / 1024).toFixed(2);
                console.log('暂停任务，保留文件对象，已上传分块:', task.uploadedChunks);

                updateProgress(
                    Math.round(progress),
                    `已暂停 · 已上传 ${uploadedMB}MB / ${totalMB}MB`,
                    '',
                    taskId
                );
                updateProgressButtons();

                setTimeout(() => {
                    const activeUploadTask = Object.keys(activeTasks).find(tid => {
                        const t = activeTasks[tid];
                        return t && t.type === 'upload' && !t.paused && tid !== taskId;
                    });

                    if (activeUploadTask) {
                        const t = activeTasks[activeUploadTask];
                        updateTopProgress(activeUploadTask, t.uploadedChunks || 0, t.totalChunks || 1, t.file?.size || 0);
                    } else {
                        updateProgressButtons();
                    }
                }, 100);
            }

            try {
                const response = await fetch(`/api/tasks/${taskId}/pause`, { method: 'POST' });
                const data = await response.json();

                if (data.success) {
                    updateTasksList();
                    console.log('任务已暂停:', taskId);
                    updateProgressButtons();
                } else {
                    console.error('暂停失败:', data.message);
                    if (task) {
                        task.paused = false;
                        task.pauseRequested = false;
                    }
                    updateTasksList();
                    updateProgressButtons();
                    await showAlert('错误', data.message || '暂停任务失败');
                }
            } catch (error) {
                console.error('暂停请求失败:', error);
                if (task) {
                    task.paused = false;
                    task.pauseRequested = false;
                }
                updateTasksList();
                updateProgressButtons();
                await showAlert('错误', '暂停任务失败: ' + error.message);
            }
        }
        
        // 恢复任务
        async function resumeTask(taskId) {
            try {
                const tasksResponse = await fetch('/api/tasks');
                const tasksData = await tasksResponse.json();

                if (!tasksData.success) {
                    await showAlert('错误', '无法获取任务信息');
                    return;
                }

                const serverTask = tasksData.tasks[taskId];
                if (!serverTask) {
                    await showAlert('错误', '任务不存在或已完成');
                    return;
                }

                if (serverTask.status !== 'paused') {
                    await showAlert('提示', '该任务当前不是暂停状态，无法继续');
                    return;
                }

                const resumeResponse = await fetch(`/api/tasks/${taskId}/resume`, { method: 'POST' });
                const resumeData = await resumeResponse.json();

                if (!resumeData.success) {
                    await showAlert('错误', '恢复失败: ' + resumeData.message);
                    return;
                }

                if (serverTask.type === 'upload') {
                    await resumeUploadTask(taskId, serverTask);
                } else if (serverTask.type === 'download') {
                    await resumeDownloadTask(taskId, serverTask);
                }

                updateProgressButtons();
                updateTasksList();
            } catch (error) {
                console.error('恢复任务失败:', error);
                await showAlert('错误', '恢复任务失败: ' + error.message);
            }
        }
        
        // 恢复上传任务
        async function resumeUploadTask(taskId, serverTask) {
            const frontendTask = activeTasks[taskId];

            console.log('恢复上传任务:', taskId);
            console.log('前端任务:', frontendTask);
            console.log('前端任务是否有文件对象:', frontendTask ? !!frontendTask.file : false);

            if (!frontendTask || !frontendTask.file) {
                console.log('文件对象已丢失，需要重新选择');

                const result = await showConfirm(
                    '📂 需要重新选择文件',
                    `文件对象已丢失。\n\n文件名：${serverTask.filename}\n已上传：${serverTask.uploaded_chunks || 0}/${serverTask.total_chunks || 0} 分块\n\n是否重新选择文件继续上传？`
                );

                if (!result) {
                    return;
                }

                const fileInput = document.createElement('input');
                fileInput.type = 'file';
                fileInput.accept = '*/*';
                fileInput.onchange = async function(e) {
                    const file = e.target.files[0];
                    if (!file) return;

                    if (file.name !== serverTask.filename) {
                        const confirmed = await showConfirm(
                            '⚠️ 文件名不匹配',
                            `选择的文件名“${file.name}”与原文件名“${serverTask.filename}”不一致。\n\n继续上传可能导致文件损坏，是否继续？`
                        );
                        if (!confirmed) return;
                    }

                    const CHUNK_SIZE = 10 * 1024 * 1024;
                    const expectedChunks = serverTask.total_chunks || 0;
                    const actualChunks = Math.ceil(file.size / CHUNK_SIZE);

                    if (expectedChunks && actualChunks !== expectedChunks) {
                        await showAlert('错误', `文件大小不匹配。\n原文件分块数：${expectedChunks}\n当前文件分块数：${actualChunks}`);
                        return;
                    }

                    activeTasks[taskId] = {
                        type: 'upload',
                        file: file,
                        totalChunks: expectedChunks || actualChunks,
                        uploadedChunks: serverTask.uploaded_chunks || 0,
                        paused: false,
                        pauseRequested: false
                    };

                    showProgress('⬆️ 上传中...', taskId);
                    startUploadFromChunk(taskId, file, serverTask.uploaded_chunks || 0);
                };
                fileInput.click();
            } else {
                console.log('恢复上传，文件对象存在');
                console.log('文件名:', frontendTask.file.name);
                console.log('文件大小:', frontendTask.file.size);
                console.log('从分块继续:', serverTask.uploaded_chunks || 0);

                frontendTask.paused = false;
                frontendTask.pauseRequested = false;

                showProgress('⬆️ 上传中...', taskId);
                startUploadFromChunk(taskId, frontendTask.file, serverTask.uploaded_chunks || 0);
            }
        }
        
        // 恢复下载任务
        async function resumeDownloadTask(taskId, serverTask) {
            if (activeTasks[taskId]) {
                activeTasks[taskId].paused = false;
                activeTasks[taskId].pauseRequested = false;
            } else {
                activeTasks[taskId] = {
                    type: 'download',
                    filename: serverTask.filename,
                    paused: false
                };
            }

            downloadFileWithResume(serverTask.filename, taskId);
        }
        
        // 删除任务
        async function deleteTask(taskId, showConfirmDialog = true) {
            if (showConfirmDialog) {
                const result = await showConfirm('🗑️ 删除任务', '确定要删除这个任务吗？');
                if (!result) return;
            }

            console.log('开始删除任务:', taskId);

            if (activeTasks[taskId]) {
                if (activeTasks[taskId].xhr) {
                    activeTasks[taskId].xhr.abort();
                }
                delete activeTasks[taskId];
            }
            removeLocalRealtimeUpload(taskId);

            try {
                const response = await fetch(`/api/tasks/${taskId}`, { method: 'DELETE' });
                const data = await response.json();

                if (data.success) {
                    console.log('任务删除成功:', taskId);
                    updateTasksList();
                    updateActivities();
                } else {
                    console.error('删除任务失败:', data.message);
                    await showAlert('错误', data.message || '删除任务失败');
                }
            } catch (error) {
                console.error('删除任务请求失败:', error);
                await showAlert('错误', '删除任务失败: ' + error.message);
            }
        }

        // 清理所有任务
        async function clearAllTasks() {
            console.log('开始清理所有任务...');

            try {
                const response = await fetch('/api/all_tasks');
                const data = await response.json();

                if (data.success && data.tasks) {
                    const taskIds = Object.keys(data.tasks);
                    console.log('找到任务数量:', taskIds.length);

                    for (const taskId of taskIds) {
                        console.log('删除任务:', taskId);
                        await deleteTask(taskId, false);
                    }

                    updateTasksList();
                    await showAlert('成功', '已清理所有任务');
                }
            } catch (error) {
                console.error('清理任务失败:', error);
                await showAlert('错误', '清理任务失败');
            }
        }

        // 分块上传文件
        function uploadFileInChunks(file, taskId) {
            const CHUNK_SIZE = 10 * 1024 * 1024; // 10MB per chunk (提高上传速度)
            const totalChunks = Math.ceil(file.size / CHUNK_SIZE);
            const uploadPathSelect = document.getElementById('uploadPathSelect');
            const selectedPath = uploadPathSelect ? uploadPathSelect.value : '{{ current_path }}';
            
            // 创建任务对象
            activeTasks[taskId] = {
                type: 'upload',
                file: file,
                totalChunks: totalChunks,
                uploadedChunks: 0,
                paused: false,
                taskId: taskId // 保存taskId用于更新进度
            };
            setLocalRealtimeUpload(taskId, {
                type: 'upload',
                status: 'running',
                username: currentSessionUsername,
                filename: file.name,
                displayName: buildUploadDisplayName(file.name, selectedPath || ''),
                totalChunks: totalChunks,
                uploadedChunks: 0,
                detailText: `0 / ${totalChunks} 分块`,
                liveProgress: 0
            });
            updateActivities();
            
            // 显示顶部进度条（关联任务 ID）
            showProgress('⬆️ 上传中...', taskId);
            
            // 先检查服务器上已上传的分块数，实现断点续传
            fetch('/api/tasks')
                .then(response => response.json())
                .then(data => {
                    if (data.success && data.tasks[taskId]) {
                        const uploadedChunks = data.tasks[taskId].uploaded_chunks || 0;
                        activeTasks[taskId].uploadedChunks = uploadedChunks;
                        updateTopProgress(taskId, uploadedChunks, totalChunks, file.size);
                        uploadChunkRecursive(taskId, file, uploadedChunks, totalChunks);
                    } else {
                        updateTopProgress(taskId, 0, totalChunks, file.size);
                        uploadChunkRecursive(taskId, file, 0, totalChunks);
                    }
                })
                .catch(() => {
                    updateTopProgress(taskId, 0, totalChunks, file.size);
                    uploadChunkRecursive(taskId, file, 0, totalChunks);
                });
        }
        
        // 更新顶部进度条
        function updateTopProgress(taskId, uploadedChunks, totalChunks, totalSize) {
            // 检查任务是否存在
            if (!activeTasks[taskId]) {
                return;
            }
            
            // 检查任务是否已经暂停
            if (activeTasks[taskId].paused || activeTasks[taskId].pauseRequested) {
                // 如果暂停的是当前显示的任务，更新按钮状态并尝试切到其他活动任务
                if (currentProgressTaskId === taskId) {
                    updateProgressButtons();
                    // 查找其他活动的上传任务来显示
                    const activeUploadTask = Object.keys(activeTasks).find(tid => {
                        const t = activeTasks[tid];
                        return t && t.type === 'upload' && !t.paused && !t.pauseRequested && tid !== taskId;
                    });
                    
                    if (activeUploadTask) {
                        const t = activeTasks[activeUploadTask];
                        updateTopProgress(activeUploadTask, t.uploadedChunks || 0, t.totalChunks || 1, t.file?.size || 0);
                    }
                }
                return;
            }
            
            // 如果当前没有显示的任务，或者这就是当前显示的任务，则更新顶部进度
            if (currentProgressTaskId === null || currentProgressTaskId === taskId) {
                currentProgressTaskId = taskId;

                const progress = (uploadedChunks / totalChunks) * 100;
                const uploadedBytes = Math.min(totalSize, uploadedChunks * 10 * 1024 * 1024);
                const uploadedMB = (uploadedBytes / 1024 / 1024).toFixed(2);
                const totalMB = (totalSize / 1024 / 1024).toFixed(2);
                setLocalRealtimeUpload(taskId, {
                    status: activeTasks[taskId].paused ? 'paused' : 'running',
                    totalChunks: totalChunks,
                    uploadedChunks: uploadedChunks,
                    liveProgress: progress,
                    detailText: `已上传 ${uploadedMB}MB / ${totalMB}MB`
                });

                updateProgress(
                    Math.round(progress),
                    `已上传 ${uploadedMB}MB / ${totalMB}MB`,
                    '',
                    taskId
                );

                updateProgressButtons();
            }
        }
        
        // 从指定分块开始上传
        function startUploadFromChunk(taskId, file, startChunkIndex) {
            const CHUNK_SIZE = 10 * 1024 * 1024; // 10MB per chunk (提高上传速度)
            const totalChunks = Math.ceil(file.size / CHUNK_SIZE);
            let uploadedChunks = startChunkIndex;
            const uploadPathSelect = document.getElementById('uploadPathSelect');
            const selectedPath = uploadPathSelect ? uploadPathSelect.value : '{{ current_path }}';
            
            console.log('startUploadFromChunk - taskId:', taskId, '文件:', file.name, '从分块:', startChunkIndex, '开始');
            
            // 显示进度条
            showProgress('⬆️ 继续上传...', taskId);
            
            // 确保任务对象存在
            if (!activeTasks[taskId]) {
                console.log('创建新的任务对象');
                activeTasks[taskId] = {
                    type: 'upload',
                    file: file,
                    totalChunks: totalChunks,
                    uploadedChunks: startChunkIndex,
                    paused: false,
                    pauseRequested: false,
                    taskId: taskId
                };
            } else {
                console.log('更新现有任务对象');
                activeTasks[taskId].file = file;
                activeTasks[taskId].totalChunks = totalChunks;
                activeTasks[taskId].uploadedChunks = startChunkIndex;
                activeTasks[taskId].paused = false;
                activeTasks[taskId].pauseRequested = false;
                activeTasks[taskId].taskId = taskId;
            }
            setLocalRealtimeUpload(taskId, {
                type: 'upload',
                status: 'running',
                username: currentSessionUsername,
                filename: file.name,
                displayName: buildUploadDisplayName(file.name, selectedPath || ''),
                totalChunks: totalChunks,
                uploadedChunks: startChunkIndex,
                liveProgress: totalChunks ? (startChunkIndex / totalChunks) * 100 : 0,
                detailText: `${startChunkIndex} / ${totalChunks} 分块`
            });
            updateActivities();
            
            // 先通知服务器恢复任务状态，再继续上传
            fetch(`/api/tasks/${taskId}/resume`, { method: 'POST' })
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        console.log('服务器任务状态已恢复');
                    }
                    
                    updateTopProgress(taskId, startChunkIndex, totalChunks, file.size);
                    console.log('开始递归上传分块');
                    uploadChunkRecursive(taskId, file, startChunkIndex, totalChunks);
                })
                .catch(error => {
                    console.warn('恢复服务器任务状态失败，仍尝试继续上传:', error);
                    
                    // 即使恢复失败也尝试上传（可能任务不存在，会自动创建）
                    updateTopProgress(taskId, startChunkIndex, totalChunks, file.size);
                    uploadChunkRecursive(taskId, file, startChunkIndex, totalChunks);
                });
        }
        
        // 递归上传分块，并在每轮检查暂停状态
        function uploadChunkRecursive(taskId, file, chunkIndex, totalChunks) {
            const CHUNK_SIZE = 10 * 1024 * 1024; // 10MB per chunk (提高上传速度)
            const task = activeTasks[taskId];
            
            console.log('uploadChunkRecursive - 分块:', chunkIndex, '/', totalChunks);
            
            // 严格检查任务是否存在、是否暂停、是否已被删除
            if (!task || task.paused || task.pauseRequested) {
                console.log('任务已暂停或删除，停止上传。任务存在?', !!task, '暂停标记:', task?.paused, task?.pauseRequested);
                return;
            }
            
            // 检查是否已完成
            if (chunkIndex >= totalChunks) {
                // 所有分块上传完成
                    removeLocalRealtimeUpload(taskId);
                    delete activeTasks[taskId];
                    fetch(`/api/tasks/${taskId}`, { method: 'DELETE' })
                        .then(() => {
                            updateTasksList();
                            updateActivities();
                            setTimeout(() => window.location.reload(), 500);
                        })
                        .catch(() => {
                            updateTasksList();
                            updateActivities();
                            setTimeout(() => window.location.reload(), 500);
                        });
                    return;
                }
                
            // 准备分块数据
            const start = chunkIndex * CHUNK_SIZE;
                const end = Math.min(start + CHUNK_SIZE, file.size);
                const chunk = file.slice(start, end);
                
                const formData = new FormData();
                formData.append('task_id', taskId);
            formData.append('chunk_index', chunkIndex);
                formData.append('total_chunks', totalChunks);
                formData.append('filename', file.name);
                
                // 获取用户选择的上传路径
                const uploadPathSelect = document.getElementById('uploadPathSelect');
                const selectedPath = uploadPathSelect ? uploadPathSelect.value : '{{ current_path }}';
                formData.append('upload_path', selectedPath || '');
                
                formData.append('chunk', chunk);
                
            // 创建新的xhr请求
                const xhr = new XMLHttpRequest();
                task.xhr = xhr; // 保存引用以便暂停时中止

                // 监听单个分块的上传进度
                xhr.upload.addEventListener('progress', function(e) {
                if (e.lengthComputable && activeTasks[taskId] && !activeTasks[taskId].paused) {
                    const chunkProgress = e.loaded / e.total;
                    const totalProgress = (chunkIndex + chunkProgress) / totalChunks * 100;
                    const totalUploadedMB = ((chunkIndex * CHUNK_SIZE + e.loaded) / 1024 / 1024).toFixed(2);
                    const totalMB = (file.size / 1024 / 1024).toFixed(2);
                    
                    // 计算上传速度
                    const now = Date.now();
                    if (!task.lastProgressTime) {
                        task.lastProgressTime = now;
                        task.lastProgressBytes = 0;
                    }
                    
                    const timeDiff = (now - task.lastProgressTime) / 1000;
                    if (timeDiff >= 0.5) { // 每 0.5 秒更新一次速度
                        const bytesDiff = (chunkIndex * CHUNK_SIZE + e.loaded) - task.lastProgressBytes;
                        const speed = bytesDiff / timeDiff;
                        const speedText = formatSpeed(speed);
                        setLocalRealtimeUpload(taskId, {
                            status: 'running',
                            totalChunks: totalChunks,
                            uploadedChunks: chunkIndex,
                            liveProgress: totalProgress,
                            detailText: `已上传 ${totalUploadedMB}MB / ${totalMB}MB`
                        });
                        
                        updateProgress(
                            Math.round(totalProgress),
                            `已上传 ${totalUploadedMB}MB / ${totalMB}MB`,
                            speedText,
                            taskId
                        );
                        
                        task.lastProgressTime = now;
                        task.lastProgressBytes = chunkIndex * CHUNK_SIZE + e.loaded;
                    } else {
                        // 更新进度但不更新速度
                        setLocalRealtimeUpload(taskId, {
                            status: 'running',
                            totalChunks: totalChunks,
                            uploadedChunks: chunkIndex,
                            liveProgress: totalProgress,
                            detailText: `已上传 ${totalUploadedMB}MB / ${totalMB}MB`
                        });
                        updateProgress(
                            Math.round(totalProgress),
                            `已上传 ${totalUploadedMB}MB / ${totalMB}MB`,
                            undefined,
                            taskId
                        );
                    }
                }
            });
                
                xhr.onload = function() {
                // 再次严格检查暂停状态，防止异步回调时状态已变化
                if (!activeTasks[taskId] || activeTasks[taskId].paused || activeTasks[taskId].pauseRequested) {
                    console.log('任务已暂停，忽略响应');
                    return;
                }
                
                if (xhr.status === 200) {
                    try {
                        const response = JSON.parse(xhr.responseText);
                        
                        if (response.paused) {
                            console.log('服务器返回暂停状态');
                            return;
                        }
                        
                        if (response.completed) {
                            removeLocalRealtimeUpload(taskId);
                            delete activeTasks[taskId];
                            fetch(`/api/tasks/${taskId}`, { method: 'DELETE' })
                                .then(() => {
                                    updateTasksList();
                                    updateActivities();
                                    setTimeout(() => window.location.reload(), 500);
                                })
                                .catch(() => {
                                    updateTasksList();
                                    updateActivities();
                                    setTimeout(() => window.location.reload(), 500);
                                });
                            return;
                        }
                        
                        // 更新进度（chunkIndex + 1 表示已完成的分块数）
                        const completedChunks = chunkIndex + 1;
                        if (activeTasks[taskId]) {
                            activeTasks[taskId].uploadedChunks = completedChunks;
                            setLocalRealtimeUpload(taskId, {
                                status: 'running',
                                totalChunks: totalChunks,
                                uploadedChunks: completedChunks,
                                liveProgress: totalChunks ? (completedChunks / totalChunks) * 100 : 0,
                                detailText: `${completedChunks} / ${totalChunks} 分块`
                            });
                            updateTopProgress(taskId, completedChunks, totalChunks, file.size);
                        }
                        updateTasksList();
                        
                        // 继续上传下一个分块
                        uploadChunkRecursive(taskId, file, chunkIndex + 1, totalChunks);
                    } catch (e) {
                        console.error('解析响应失败:', e);
                        if (activeTasks[taskId]) {
                            activeTasks[taskId].error = '响应解析失败';
                            setLocalRealtimeUpload(taskId, {
                                status: 'error',
                                detailText: '响应解析失败'
                            });
                            updateTasksList();
                        }
                    }
                } else {
                        console.error('上传分块失败:', xhr.statusText);
                        if (activeTasks[taskId]) {
                            activeTasks[taskId].error = xhr.statusText;
                            setLocalRealtimeUpload(taskId, {
                                status: 'error',
                                detailText: xhr.statusText || '上传失败'
                            });
                            updateTasksList();
                        }
                    }
                };
                
                xhr.onerror = function() {
                    console.error('上传分块出错');
                    if (activeTasks[taskId] && !activeTasks[taskId].paused) {
                        activeTasks[taskId].error = '网络错误';
                        setLocalRealtimeUpload(taskId, {
                            status: 'error',
                            detailText: '网络错误'
                        });
                        updateTasksList();
                    }
                };
                
                xhr.onabort = function() {
                    console.log('上传请求已中止');
                };
                
                // 发送请求
                xhr.open('POST', '/api/upload_chunk', true);
                xhr.send(formData);
        }
        
        // 断点续传下载
        function downloadFileWithResume(filename, taskId) {
            const filepath = `/download/${encodeURIComponent(filename)}`;
            let downloadedBytes = 0;
            
            // 确保任务对象存在
            if (!activeTasks[taskId]) {
            activeTasks[taskId] = {
                type: 'download',
                    filename: filename,
                    paused: false
            };
            } else {
                activeTasks[taskId].paused = false;
                activeTasks[taskId].pauseRequested = false;
            }
            
            function downloadChunk() {
                const task = activeTasks[taskId];
                
                // 严格检查：任务是否存在、是否暂停、是否已被删除
                if (!task || task.paused || task.pauseRequested) {
                    console.log('任务已暂停或删除，停止下载');
                    return;
                }
                
                const xhr = new XMLHttpRequest();
                task.xhr = xhr; // 保存引用以便暂停时中止
                xhr.open('GET', filepath, true);
                xhr.setRequestHeader('Range', `bytes=${downloadedBytes}-`);
                xhr.responseType = 'blob';
                
                xhr.onload = function() {
                    // 再次严格检查暂停状态，防止异步回调时状态已变化
                    if (!activeTasks[taskId] || activeTasks[taskId].paused || activeTasks[taskId].pauseRequested) {
                        console.log('任务已暂停，忽略响应');
                        return;
                    }
                    
                    if (xhr.status === 206 || xhr.status === 200) {
                        const blob = xhr.response;
                        // 创建下载链接
                        const url = window.URL.createObjectURL(blob);
                        const a = document.createElement('a');
                        a.href = url;
                        a.download = filename;
                        a.click();
                        window.URL.revokeObjectURL(url);
                        
                        // 下载完成后立即清理任务
                        delete activeTasks[taskId];
                        // 立即从服务器删除任务
                        fetch(`/api/tasks/${taskId}`, { method: 'DELETE' })
                            .then(() => updateTasksList())
                            .catch(() => updateTasksList());
                    }
                };
                
                xhr.onerror = function() {
                    console.error('下载出错');
                    if (activeTasks[taskId] && !activeTasks[taskId].paused) {
                        activeTasks[taskId].error = '网络错误';
                        updateTasksList();
                    }
                };
                
                xhr.onabort = function() {
                    console.log('下载请求已中止');
                };
                
                xhr.send();
            }
            
            downloadChunk();
        }
        
        // 恢复下载
        function resumeDownload(taskId, task) {
            downloadFileWithResume(task.filename, taskId);
        }
        
        // 定期更新任务列表（后台清理已完成的任务）
        setInterval(function() {
            // 后台清理，不显示面板
                    }, 1000);
        
        // 后台定期清理已完成的任务（即使面板关闭）
        setInterval(function() {
            fetch('/api/tasks')
                .then(response => response.json())
                .then(data => {
                    if (data.success) {
                        // 检查所有任务，包括已完成的任务
                        fetch('/api/all_tasks')
                            .then(r => r.json())
                            .then(allData => {
                                if (allData.success) {
                                    Object.keys(allData.tasks || {}).forEach(taskId => {
                                        const task = allData.tasks[taskId];
                                        if (task.status === 'completed') {
                                            // 自动删除已完成超过 2 秒的任务
                                            const completedAt = new Date(task.completed_at);
                                            const now = new Date();
                                            if (now - completedAt > 2000) {
                                                fetch(`/api/tasks/${taskId}`, { method: 'DELETE' })
                                                    .catch(() => {});
                                            }
                                        }
                                    });
                                }
                            })
                            .catch(() => {});
                    }
                })
                .catch(() => {});
        }, 500);
        
        // 统一上传表单提交流程，避免浏览器默认行为打断自定义上传
        document.getElementById('uploadForm').addEventListener('submit', function(e) {
            e.preventDefault();
            
            const fileInput = document.getElementById('fileInput');
            const folderInput = document.getElementById('folderInput');
            
            if (!fileInput.files.length && !folderInput.files.length) {
                showAlert('⚠️ 提示', '请先选择文件或文件夹');
                return false;
            }
            
            // 获取所有文件（普通文件或文件夹中的文件）
            const files = fileInput.files.length ? Array.from(fileInput.files) : Array.from(folderInput.files);
            const isFolder = folderInput.files.length > 0;
            
            console.log('表单提交，文件数量:', files.length, '是否为文件夹:', isFolder);
            
            // 如果是文件夹上传，需要特殊处理以保留结构
            if (isFolder) {
                uploadFolderFiles(files);
            } else {
                // 普通文件上传
                uploadFilesDirectly(files);
            }
            
            return false;
        });
        
        // ========== 检查未完成任务（页面加载时） ==========
        function checkIncompleteTasks() {
            fetch('/api/tasks')
                .then(response => response.json())
                .then(data => {
                    if (data.success && data.tasks && Object.keys(data.tasks).length > 0) {
                        // 有未完成的上传任务
                        showIncompleteTasksNotification(data.tasks);
                    }
                })
                .catch(error => console.error('检查未完成任务失败:', error));
        }
        
        // 显示未完成任务通知
        function showIncompleteTasksNotification(tasks) {
            const notification = document.getElementById('incompleteTasksNotification');
            const tasksList = document.getElementById('incompleteTasksList');
            
            // 清空任务列表
            tasksList.innerHTML = '';
            
            // 只显示上传类型的任务
            let hasUploadTasks = false;
            Object.entries(tasks).forEach(([taskId, task]) => {
                if (task.type === 'upload' && task.status !== 'completed') {
                    hasUploadTasks = true;
                    const taskItem = document.createElement('div');
                    taskItem.className = 'incomplete-task-item';
                    
                    const progress = task.uploaded_chunks && task.total_chunks 
                        ? Math.round((task.uploaded_chunks / task.total_chunks) * 100) 
                        : 0;
                    
                    taskItem.innerHTML = `
                        <div class="incomplete-task-name">📄 ${task.filename || '未知文件'}</div>
                        <div class="incomplete-task-progress">
                            进度: ${progress}% (${task.uploaded_chunks || 0} / ${task.total_chunks || 0} 分块)
                        </div>
                        <div class="incomplete-task-actions">
                            <button class="task-btn task-btn-continue" onclick="continueUpload('${taskId}', '${task.filename}', ${task.total_chunks})">
                                ▶️ 继续上传
                            </button>
                            <button class="task-btn task-btn-delete" onclick="deleteIncompleteTask('${taskId}')">
                                <img src="/static/删除.png" style="width: 25px; height: 25px; vertical-align: middle; margin-right: 4px;">删除任务
                            </button>
                        </div>
                    `;
                    
                    tasksList.appendChild(taskItem);
                }
            });
            
            // 只有存在上传任务时才显示通知
            if (hasUploadTasks) {
                notification.style.display = 'block';
            }
        }
        
        // 继续上传
        async function continueUpload(taskId, filename, totalChunks) {
            console.log('继续上传:', taskId, filename, totalChunks);

            const proceed = await showConfirm(
                '📂 选择文件继续上传',
                `请重新选择文件“${filename}”以继续上传。\n\n系统会从上次中断的位置继续，已上传的分块不会重复上传。`
            );

            if (!proceed) {
                return;
            }

            const fileInput = document.createElement('input');
            fileInput.type = 'file';
            fileInput.accept = '*/*';

            fileInput.onchange = async function(e) {
                const file = e.target.files[0];
                if (!file) {
                    return;
                }

                const originalName = filename;
                const selectedName = file.name;

                if (selectedName !== originalName) {
                    const confirmed = await showConfirm(
                        '⚠️ 文件名不匹配',
                        `您选择的文件名是“${selectedName}”，而原文件名是“${originalName}”。\n\n确定要继续吗？这可能导致上传后的文件不完整。`
                    );
                    if (!confirmed) {
                        return;
                    }
                }

                closeIncompleteTasksNotification();

                fetch('/api/tasks')
                    .then(response => response.json())
                    .then(data => {
                        if (data.success && data.tasks[taskId]) {
                            const uploadedChunks = data.tasks[taskId].uploaded_chunks || 0;
                            console.log('从分块继续上传:', uploadedChunks);
                            startUploadFromChunk(taskId, file, uploadedChunks);
                        } else {
                            showAlert('错误', '无法获取任务信息，请删除该任务后重新上传。');
                        }
                    })
                    .catch(error => {
                        console.error('获取任务信息失败:', error);
                        showAlert('错误', '获取任务信息失败');
                    });
            };

            fileInput.click();
        }
        
        // 删除未完成任务
        async function deleteIncompleteTask(taskId) {
            const confirmed = await showConfirm(
                '<img src="/static/删除.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">确认删除',
                '确定要删除这个未完成任务吗？已上传的分块也会一并清除。'
            );

            if (confirmed) {
                try {
                    const response = await fetch(`/api/tasks/${taskId}`, { method: 'DELETE' });
                    const data = await response.json();

                    if (data.success) {
                        showAlert('成功', '任务已删除');
                        setTimeout(checkIncompleteTasks, 500);
                    } else {
                        showAlert('错误', '删除任务失败: ' + data.message);
                    }
                } catch (error) {
                    console.error('删除任务失败:', error);
                    showAlert('错误', '删除任务失败');
                }
            }
        }

        // 关闭未完成任务通知
        function closeIncompleteTasksNotification() {
            const notification = document.getElementById('incompleteTasksNotification');
            notification.style.display = 'none';
        }

        // 页面加载后延迟检查未完成任务，避免首屏初始化互相打断
        setTimeout(checkIncompleteTasks, 3000);

        // ========== 重命名功能 ==========
        async function renameItem(itemPath, currentName, isFolder) {
            const itemType = isFolder ? '文件夹' : '文件';

            const newName = await showPrompt(
                `<img src="/static/钢笔.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">重命名${itemType}`,
                `请输入新的${itemType}名称：`,
                currentName
            );

            if (!newName || newName === currentName) {
                return;
            }

            showProgress('<img src="/static/钢笔.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">重命名中...', null);
            updateProgress(50, `正在重命名${itemType}...`);

            try {
                const response = await fetch('/api/rename', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        old_path: itemPath,
                        new_name: newName,
                        current_path: '{{ current_path }}'
                    })
                });

                const data = await response.json();
                hideProgress();

                if (data.success) {
                    await showAlert('成功', data.message);
                    window.location.reload();
                } else {
                    await showAlert('失败', data.message);
                }
            } catch (error) {
                hideProgress();
                console.error('重命名失败:', error);
                await showAlert('错误', '重命名失败: ' + error.message);
            }
        }

        // 自定义输入对话框
        function showPrompt(title, message, defaultValue = '') {
            return new Promise((resolve) => {
                const modal = document.getElementById('confirmModal');
                const titleEl = document.getElementById('confirmTitle');
                const messageEl = document.getElementById('confirmMessage');
                const cancelBtn = document.getElementById('confirmCancel');
                const okBtn = document.getElementById('confirmOk');
                
                titleEl.innerHTML = title || '<img src="/static/钢笔.png" style="width: 35px; height: 35px; vertical-align: middle; margin-right: 8px;">输入';
                
                // 创建输入框
                const inputHtml = `
                    <div style="margin-bottom: 10px; color: rgba(255, 255, 255, 0.9);">${message}</div>
                    <input type="text" 
                           id="promptInput" 
                           value="${defaultValue}" 
                           style="width: 100%; 
                                  padding: 12px; 
                                  border: 2px solid rgba(255, 255, 255, 0.3); 
                                  background: rgba(255, 255, 255, 0.1); 
                                  color: white; 
                                  border-radius: 8px; 
                                  font-size: 14px;
                                  outline: none;"
                           placeholder="请输入新名称">
                `;
                messageEl.innerHTML = inputHtml;
                
                // 显示两个按钮
                cancelBtn.style.display = 'inline-block';
                
                modal.classList.add('show');
                
                // 聚焦并选中输入框
                setTimeout(() => {
                    const input = document.getElementById('promptInput');
                    input.focus();
                    input.select();
                }, 100);
                
                const handleCancel = () => {
                    modal.classList.remove('show');
                    cancelBtn.removeEventListener('click', handleCancel);
                    okBtn.removeEventListener('click', handleOk);
                    messageEl.textContent = '';  // 清理HTML
                    resolve(null);
                };
                
                const handleOk = () => {
                    const input = document.getElementById('promptInput');
                    const value = input ? input.value.trim() : '';
                    modal.classList.remove('show');
                    cancelBtn.removeEventListener('click', handleCancel);
                    okBtn.removeEventListener('click', handleOk);
                    messageEl.textContent = '';  // 清理HTML
                    resolve(value);
                };
                
                cancelBtn.addEventListener('click', handleCancel);
                okBtn.addEventListener('click', handleOk);
                
                // 按回车确认
                const input = document.getElementById('promptInput');
                if (input) {
                    input.addEventListener('keypress', (e) => {
                        if (e.key === 'Enter') {
                            handleOk();
                        }
                    });
                }
            });
        }
        
        // ========== 自动刷新功能 ==========
        {% if auto_refresh %}
        let currentHash = null;  // 当前文件列表哈希
        let refreshInterval = {{ refresh_interval * 1000 }};  // 刷新间隔（毫秒）
        let currentPath = '{{ current_path }}';  // 当前路径
        let isRefreshing = false;  // 防止重复刷新
        
        // 检查文件列表是否有更新
        async function checkForUpdates() {
            if (isRefreshing) return;
            
            try {
                const url = currentPath ? `/api/check_updates/${currentPath}` : '/api/check_updates';
                const response = await fetch(url);
                const data = await response.json();
                
                if (data.success) {
                    if (currentHash === null) {
                        // 首次加载时记录当前哈希
                        currentHash = data.hash;
                    } else if (currentHash !== data.hash) {
                        // 文件列表有变化，显示提示并刷新
                        console.log('检测到文件变化，准备自动刷新...');
                        showUpdateNotification();
                        currentHash = data.hash;
                        
                        // 延迟刷新，让用户看到提示
                        setTimeout(() => {
                            window.location.reload();
                        }, 1000);
                    }
                }
            } catch (error) {
                console.error('检查更新失败:', error);
            }
        }
        
        // 显示更新提示
        function showUpdateNotification() {
            // 创建提示元素
            const notification = document.createElement('div');
            notification.style.cssText = `
                position: fixed;
                top: 100px;
                right: 20px;
                background: rgba(16, 185, 129, 0.9);
                color: white;
                padding: 15px 25px;
                border-radius: 12px;
                box-shadow: 0 4px 20px rgba(16, 185, 129, 0.4);
                z-index: 10000;
                font-weight: 600;
                animation: slideInRight 0.3s ease-out;
                backdrop-filter: blur(10px);
            `;
            notification.innerHTML = '检测到新文件，正在刷新...';
            document.body.appendChild(notification);
            
            // 添加动画
            const style = document.createElement('style');
            style.textContent = `
                @keyframes slideInRight {
                    from {
                        opacity: 0;
                        transform: translateX(100px);
                    }
                    to {
                        opacity: 1;
                        transform: translateX(0);
                    }
                }
            `;
            document.head.appendChild(style);
            
            // 1.5秒后移除提示
            setTimeout(() => {
                notification.style.opacity = '0';
                notification.style.transform = 'translateX(100px)';
                notification.style.transition = 'all 0.3s ease-out';
                setTimeout(() => notification.remove(), 300);
            }, 1500);
        }
        
        // 启动定时检查
        console.log(`自动刷新已启用，每 ${refreshInterval / 1000} 秒检查一次文件变化`);
        setInterval(checkForUpdates, refreshInterval);
        
        // 首次检查延迟 3 秒，避免与页面加载冲突
        setTimeout(checkForUpdates, 3000);
        {% else %}
        console.log('自动刷新已禁用');
        {% endif %}

    </script>
    <script>
        // ========== 文件搜索功能 ==========
        let searchTimeout = null;
        let originalFileList = null; // 保存原始文件列表
        let isSearching = false;
        
        // 搜索输入框事件监听
        document.getElementById('searchInput').addEventListener('input', function(e) {
            const query = e.target.value.trim();
            const clearBtn = document.getElementById('clearSearchBtn');
            
            // 显示或隐藏清除按钮
            if (query) {
                clearBtn.style.display = 'block';
            } else {
                clearBtn.style.display = 'none';
                clearSearch();
                return;
            }
            
            // 防抖处理
            clearTimeout(searchTimeout);
            searchTimeout = setTimeout(() => {
                performSearch(query);
            }, 300);
        });
        
        // 执行搜索
        async function performSearch(query) {
            if (!query) {
                clearSearch();
                return;
            }
            
            try {
                const searchSubfolders = document.getElementById('searchSubfolders').checked;
                const currentPath = '{{ current_path }}';
                
                const response = await fetch('/api/search', {
                    method: 'POST',
                    headers: {
                        'Content-Type': 'application/json'
                    },
                    body: JSON.stringify({
                        query: query,
                        current_path: currentPath,
                        search_subfolders: searchSubfolders
                    })
                });
                
                const data = await response.json();
                
                if (data.success) {
                    displaySearchResults(data.results, data.count);
                    // 如果结果被截断，显示警告
                    if (data.is_truncated && data.message) {
                        showMessage(data.message, 'warning');
                    }
                } else {
                    showMessage(data.message, 'danger');
                }
            } catch (error) {
                console.error('搜索失败:', error);
                showMessage('搜索失败，请重试', 'danger');
            }
        }
        
        // 显示搜索结果
        function displaySearchResults(results, count) {
            const fileList = document.querySelector('.file-list');
            const noFilesMsg = document.querySelector('.files-section > p');
            const searchResults = document.getElementById('searchResults');
            const searchCount = document.getElementById('searchCount');
            
            // 仅首次进入搜索状态时保存原始文件列表
            if (!isSearching && fileList) {
                originalFileList = fileList.cloneNode(true);
                isSearching = true;
            }
            
            // 显示搜索结果统计
            searchResults.style.display = 'block';
            searchCount.textContent = count;
            
            if (results.length === 0) {
                if (fileList) fileList.style.display = 'none';
                if (noFilesMsg) noFilesMsg.style.display = 'none';
                
                // 显示无结果提示
                let noResultsMsg = document.getElementById('noSearchResults');
                if (!noResultsMsg) {
                    noResultsMsg = document.createElement('p');
                    noResultsMsg.id = 'noSearchResults';
                    noResultsMsg.style.cssText = 'text-align: center; color: rgba(255, 255, 255, 0.7); margin-top: 20px; font-size: 14px;';
                    searchResults.parentNode.insertBefore(noResultsMsg, searchResults.nextSibling);
                }
                noResultsMsg.textContent = '😔 未找到匹配的文件或文件夹';
                noResultsMsg.style.display = 'block';
                return;
            }
            
            // 隐藏无结果提示
            const noResultsMsg = document.getElementById('noSearchResults');
            if (noResultsMsg) noResultsMsg.style.display = 'none';
            
            // 创建搜索结果列表，使用 DocumentFragment 优化 DOM 操作
            if (fileList) {
                fileList.style.display = 'block';
                
                // 使用DocumentFragment减少重绘
                const fragment = document.createDocumentFragment();
                
                results.forEach(item => {
                    const li = document.createElement('li');
                    li.className = 'file-item' + (item.is_folder ? ' folder-item' : '');
                    li.setAttribute('data-filepath', item.relative_path);
                    li.setAttribute('data-is-folder', item.is_folder);
                    
                    // 使用 innerHTML 一次性创建结构，减少重排
                    const folderPath = item.relative_path.includes('/') ? item.relative_path.substring(0, item.relative_path.lastIndexOf('/')) : '根目录';
                    const icon = item.is_folder ? '📁' : getFileIcon(item.name);
                    const actionHref = (item.is_folder ? '/browse/' : '/download/') + encodeURIComponent(item.relative_path);
                    const actionText = item.is_folder ? '📂 打开' : '⬇️ 下载';
                    const actionStyle = item.is_folder ? 'padding: 8px 16px; font-size: 14px;' : '';
                    
                    li.innerHTML = `
                        <input type="checkbox" class="file-checkbox" data-path="${item.relative_path}" style="width: 20px; height: 20px; cursor: pointer; margin-right: 10px;">
                        <div class="file-info" ${item.is_folder ? 'style="cursor: pointer;"' : ''}>
                            <div class="file-icon" style="font-size: 36px;">${icon}</div>
                            <div class="file-details">
                                <div class="file-name-text" title="${item.relative_path}">${item.name}</div>
                                <div class="file-meta">
                                    <span>📂 ${folderPath}</span>
                                    <span>📦 ${item.size}</span>
                                    <span>🕒 ${item.time}</span>
                                </div>
                            </div>
                        </div>
                        <div class="file-actions">
                            <a class="btn-download" href="${actionHref}" style="${actionStyle}">${actionText}</a>
                        </div>
                    `;
                    
                    // 添加事件监听器
                    const checkbox = li.querySelector('.file-checkbox');
                    checkbox.onclick = function(e) {
                        e.stopPropagation();
                        updateBatchToolbar();
                    };
                    
                    if (item.is_folder) {
                        const fileInfo = li.querySelector('.file-info');
                        fileInfo.onclick = function() {
                            window.location.href = '/browse/' + encodeURIComponent(item.relative_path);
                        };
                    }
                    
                    fragment.appendChild(li);
                });
                
                // 涓€娆℃€ф洿鏂癉OM
                fileList.innerHTML = '';
                fileList.appendChild(fragment);
            }
            
            if (noFilesMsg) noFilesMsg.style.display = 'none';
        }
        
        // 清除搜索
        function clearSearch() {
            const searchInput = document.getElementById('searchInput');
            const clearBtn = document.getElementById('clearSearchBtn');
            const searchResults = document.getElementById('searchResults');
            const fileList = document.querySelector('.file-list');
            const noResultsMsg = document.getElementById('noSearchResults');
            
            searchInput.value = '';
            clearBtn.style.display = 'none';
            searchResults.style.display = 'none';
            
            if (noResultsMsg) {
                noResultsMsg.style.display = 'none';
            }
            
            // 恢复原始文件列表
            if (isSearching && originalFileList && fileList) {
                fileList.parentNode.replaceChild(originalFileList.cloneNode(true), fileList);
                isSearching = false;
            }
        }
        
        // 搜索子文件夹选项变化时，如果当前有查询就重新搜索
        document.getElementById('searchSubfolders').addEventListener('change', function() {
            const searchInput = document.getElementById('searchInput');
            const query = searchInput.value.trim();
            if (query) {
                performSearch(query);
            }
        });
    </script>
</body>
</html>
"""

@app.route('/', methods=['GET', 'POST'])
@app.route('/browse', methods=['GET', 'POST'])
@app.route('/browse/<path:subpath>', methods=['GET', 'POST'])
def index(subpath=''):
    """Docstring."""
    success, current_path, error = safe_join_path(UPLOAD_FOLDER, subpath)
    if not success:
        flash(f'路径访问错误: {error}', 'danger')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        uploaded_files = []
        failed_files = []
        
        upload_path = request.form.get('upload_path', subpath)
        success, target_upload_path, error = safe_join_path(UPLOAD_FOLDER, upload_path)
        if not success:
            flash(f'上传路径错误: {error}', 'danger')
            return redirect(url_for('index', subpath=subpath))
        
        # 确保目标路径存在
        if not os.path.exists(target_upload_path):
            try:
                os.makedirs(target_upload_path, exist_ok=True)
            except Exception as e:
                flash(f'创建上传目录失败: {str(e)}', 'danger')
                return redirect(url_for('index', subpath=subpath))
        
        # 处理单个文件上传
        if 'files' in request.files:
            files = request.files.getlist('files')
            for file in files:
                if file and file.filename:
                    try:
                        filename = secure_filename(file.filename)
                        
                        # 检查文件扩展名
                        if not allowed_file(file.filename):
                            failed_files.append(f"{file.filename}: 不支持的文件类型")
                            continue
                        
                        if not filename.lower().endswith('.zip'):
                            file.seek(0, 2)  # 移到文件末尾
                            file_size = file.tell()
                            file.seek(0)  # 重置到开头
                            if file_size > MAX_FILE_SIZE:
                                size_mb = file_size / 1024 / 1024
                                limit_mb = MAX_FILE_SIZE / 1024 / 1024
                                failed_files.append(f"{file.filename}: 文件过大 ({size_mb:.1f}MB > {limit_mb}MB限制)")
                                continue
                        
                        filepath = os.path.join(target_upload_path, filename)
                        success, error_msg = save_file_safely(file, filepath)
                        
                        if success:
                            uploaded_files.append(filename)
                        else:
                            failed_files.append(f"{file.filename}: {error_msg}")
                    except Exception as e:
                        failed_files.append(f"{file.filename}: {str(e)}")
        
        if 'folder' in request.files:
            files = request.files.getlist('folder')
            
            # 检查文件夹中的文件数量
            if len(files) > MAX_FOLDER_FILES:
                flash(
                    f'文件夹包含 {len(files)} 个文件，超过限制（最多 {MAX_FOLDER_FILES} 个），建议压缩为 ZIP 后上传。',
                    'danger'
                )
            else:
                for file in files:
                    if file and file.filename:
                        try:
                            # 保留文件夹结构（使用完整的相对路径）
                            filename = file.filename.replace('\\', '/')  # 统一路径分隔符
                            # 检查文件扩展名
                            if not allowed_file(file.filename):
                                failed_files.append(f"{file.filename}: 不支持的文件类型")
                                continue
                            
                            # 保存到用户选择的路径，并保留完整文件夹结构
                            filepath = os.path.join(target_upload_path, filename)
                            success, error_msg = save_file_safely(file, filepath)
                            
                            if success:
                                uploaded_files.append(filename)
                            else:
                                failed_files.append(f"{file.filename}: {error_msg}")
                        except Exception as e:
                            failed_files.append(f"{file.filename}: {str(e)}")
        
        if uploaded_files:
            username = get_current_username()
            upload_path_display = upload_path if upload_path else '根目录'
            if len(uploaded_files) == 1:
                flash(f'文件 "{uploaded_files[0]}" 已上传到 {upload_path_display}。', 'success')
                add_activity(username, 'upload', uploaded_files[0])
            else:
                flash(f'成功上传 {len(uploaded_files)} 个文件到 {upload_path_display}。', 'success')
                add_activity(username, 'upload', f'{len(uploaded_files)} files')
        
        if failed_files:
            preview_failed = '; '.join(failed_files[:5])
            suffix = '...' if len(failed_files) > 5 else ''
            flash(f'部分文件上传失败：{preview_failed}{suffix}', 'danger')
        
        if not uploaded_files and not failed_files:
            flash('没有选择文件', 'danger')
        
        # 重定向回当前路径
        if subpath:
            return redirect(url_for('index', subpath=subpath))
        return redirect(url_for('index'))
    
    files = []
    total_size = 0
    total_size_recursive = 0
    breadcrumbs = []
    all_folders = []
    if subpath:
        parts = subpath.split('/')
        cumulative_path = ''
        for part in parts:
            cumulative_path = os.path.join(cumulative_path, part) if cumulative_path else part
            breadcrumbs.append({
                'name': part,
                'path': cumulative_path
            })
    
    if os.path.exists(current_path):
        try:
            items = os.listdir(current_path)
            
            for item in items:
                try:
                    # 检查该项目是否应被隐藏
                    if should_hide_shared_item(item):
                        continue
                    
                    item_path = os.path.join(current_path, item)
                    
                    if not os.access(item_path, os.R_OK):
                        continue
                    
                    stat = os.stat(item_path)
                    mtime = datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                    
                    if subpath:
                        relative_path = os.path.join(subpath, item).replace('\\', '/')
                    else:
                        relative_path = item
                    
                    if os.path.isdir(item_path):
                        # 文件夹，添加斜杠标记
                        files.append({
                            'name': item + '/',
                            'display_name': item,
                            'relative_path': relative_path,
                            'size': '文件夹',
                            'time': mtime,
                            'is_folder': True,
                            'is_empty_folder': False
                        })
                    else:
                        # 文件
                        size = stat.st_size
                        total_size += size
                        files.append({
                            'name': item,
                            'display_name': item,
                            'relative_path': relative_path,
                            'size': get_file_size(size),
                            'time': mtime,
                            'is_folder': False,
                            'is_empty_folder': False
                        })
                except (OSError, PermissionError) as e:
                    print(f"⚠️  无法访问: {item_path} - {e}")
                    continue
                    
        except (OSError, PermissionError) as e:
            print(f"⚠️  无法读取目录: {current_path} - {e}")
            flash(f'无法访问该目录：{e}', 'danger')
    
    def sort_key(item):
        is_folder = item.get('is_folder', False)
        name = item['name'].lower()
        return (not is_folder, name)
    files.sort(key=sort_key)
    
    try:
        for root, dirs, filenames in os.walk(current_path):
            if HIDE_SYSTEM_FOLDERS:
                dirs[:] = [d for d in dirs if d.lower() not in HIDDEN_ITEMS]
            
            for filename in filenames:
                try:
                    if should_hide_shared_item(filename):
                        continue
                    file_path = os.path.join(root, filename)
                    if os.access(file_path, os.R_OK):
                        total_size_recursive += os.path.getsize(file_path)
                except (OSError, PermissionError):
                    continue
    except (OSError, PermissionError):
        pass
    
    try:
        all_folders.append({'path': '', 'name': '根目录'})
        for root, dirs, _ in os.walk(UPLOAD_FOLDER):
            if HIDE_SYSTEM_FOLDERS:
                dirs[:] = [d for d in dirs if d.lower() not in HIDDEN_ITEMS]
            
            for dir_name in dirs:
                dir_path = os.path.join(root, dir_name)
                relative_path = os.path.relpath(dir_path, UPLOAD_FOLDER).replace('\\', '/')
                # 计算文件夹深度，用纯文本树形前缀避免浏览器/字体导致的乱码缩进
                depth = relative_path.count('/')
                indent = '|  ' * depth
                if depth > 0:
                    indent += '|- '
                all_folders.append({
                    'path': relative_path,
                    'name': f'{indent}{dir_name}'
                })
    except (OSError, PermissionError):
        pass
    
    current_session_username = normalize_username(session.get('username'))
    if session.get('username') and not current_session_username:
        stale_presence_id = get_page_session_id()
        session.pop('username', None)
        with user_lock:
            online_users.pop(stale_presence_id, None)

    initial_online_users = []
    initial_admin_state = {'requests': {}}
    if current_session_username:
        sync_current_online_presence()
        initial_online_users = get_online_users_snapshot()
        initial_admin_state = get_admin_state_snapshot()

    return render_template_string(
        HTML_TEMPLATE, 
        files=files, 
        ip=get_local_ip(), 
        port=SERVER_PORT,
        total_size=get_file_size(total_size_recursive),
        current_path=subpath,
        breadcrumbs=breadcrumbs,
        all_folders=all_folders,
        auto_refresh=AUTO_REFRESH,
        refresh_interval=REFRESH_INTERVAL,
        get_file_icon=get_file_icon,
        is_text_previewable_file=is_text_previewable_file,
        is_previewable_file=is_previewable_file,
        is_word_editable_file=is_word_editable_file,
        is_excel_editable_file=is_excel_editable_file,
        initial_online_users=initial_online_users,
        initial_admin_state=initial_admin_state,
        password_min_length=PASSWORD_MIN_LENGTH,
        nickname_max_length=NICKNAME_MAX_LENGTH
    )

def _handle_file_upload(current_path):
    uploaded_files = []
    failed_files = []
    
    if 'files' in request.files:
        files = request.files.getlist('files')
        for file in files:
            if file and file.filename:
                try:
                    filename = secure_filename(file.filename)
                    
                    # 检查文件扩展名
                    if not allowed_file(file.filename):
                        failed_files.append(f"{file.filename}: 不支持的文件类型")
                        continue
                    
                    # 保存到当前浏览的路径
                    filepath = os.path.join(current_path, filename)
                    success, error_msg = save_file_safely(file, filepath)
                    
                    if success:
                        uploaded_files.append(filename)
                        add_activity(get_current_username(), 'upload', filename)
                    else:
                        failed_files.append(f"{file.filename}: {error_msg}")
                except Exception as e:
                    failed_files.append(f"{file.filename}: {str(e)}")
    
    # 返回JSON结果
    if uploaded_files:
        return jsonify({
            'success': True,
            'message': f'成功上传 {len(uploaded_files)} 个文件。',
            'uploaded': uploaded_files,
            'failed': failed_files
        })
    elif failed_files:
        return jsonify({
            'success': False,
            'message': '所有文件上传失败。',
            'failed': failed_files
        }), 400
    else:
        return jsonify({
            'success': False,
            'message': '没有选择文件'
        }), 400

@app.route('/api/upload_files', methods=['POST'])
@app.route('/api/upload_files/<path:subpath>', methods=['POST'])
def api_upload_files(subpath=''):
    """API方式上传文件（返回JSON，不重定向）"""
    try:
        success, current_path, error = safe_join_path(UPLOAD_FOLDER, subpath)
        if not success:
            return jsonify({'success': False, 'message': error}), 403
        
        if not os.path.exists(current_path):
            return jsonify({'success': False, 'message': '目标路径不存在。'}), 404
        
        # 调用上传处理函数
        return _handle_file_upload(current_path)
    
    except Exception as e:
        print(f"[upload-api] failed: {str(e)}")
        return jsonify({'success': False, 'message': f'上传失败: {str(e)}'}), 500

@app.route('/api/check_updates')
@app.route('/api/check_updates/<path:subpath>')
def check_updates(subpath=''):
    """检查文件列表是否有更新。"""
    try:
        success, current_path, error = safe_join_path(UPLOAD_FOLDER, subpath)
        if not success:
            return jsonify({'success': False, 'message': error}), 403
        
        if not os.path.exists(current_path):
            return jsonify({'success': False, 'message': '路径不存在。'}), 404
        
        # 获取文件列表信息
        items = []
        try:
            for item in os.listdir(current_path):
                if should_hide_shared_item(item):
                    continue
                
                item_path = os.path.join(current_path, item)
                if not os.access(item_path, os.R_OK):
                    continue
                
                try:
                    stat = os.stat(item_path)
                    items.append({
                        'name': item,
                        'is_dir': os.path.isdir(item_path),
                        'mtime': stat.st_mtime,
                        'size': stat.st_size if not os.path.isdir(item_path) else 0
                    })
                except:
                    continue
        except:
            pass
        
        # 计算文件列表哈希，用于检测变化
        import hashlib
        items_str = '|'.join(sorted([f"{item['name']}_{item['mtime']}_{item['size']}" for item in items]))
        hash_value = hashlib.md5(items_str.encode()).hexdigest()
        
        return jsonify({
            'success': True,
            'hash': hash_value,
            'count': len(items),
            'items': items
        })
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/search', methods=['POST'])
def search_files():
    """搜索文件和文件夹"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据。'}), 400
        
        query = data.get('query', '').strip().lower()
        current_path = data.get('current_path', '')
        search_subfolders = data.get('search_subfolders', True)
        
        if not query:
            return jsonify({'success': False, 'message': '搜索关键词不能为空。'}), 400
        
        success, base_path, error = safe_join_path(UPLOAD_FOLDER, current_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 403
        
        if not os.path.exists(base_path):
            return jsonify({'success': False, 'message': '路径不存在。'}), 404
        
        results = []
        MAX_RESULTS = 500  # 最大返回结果数
        MAX_DEPTH = 20     # 最大搜索深度
        def search_in_directory(directory, relative_base='', depth=0):
            if depth > MAX_DEPTH:
                return
            
            if len(results) >= MAX_RESULTS:
                return
            
            try:
                for item in os.listdir(directory):
                    if len(results) >= MAX_RESULTS:
                        return
                    
                    if should_hide_shared_item(item):
                        continue
                    
                    item_path = os.path.join(directory, item)
                    
                    if not os.access(item_path, os.R_OK):
                        continue
                    
                    # 计算相对路径
                    if relative_base:
                        relative_path = os.path.join(relative_base, item).replace('\\', '/')
                    else:
                        relative_path = item
                    
                    # 检查是否匹配搜索关键词
                    if query in item.lower():
                        try:
                            stat = os.stat(item_path)
                            is_folder = os.path.isdir(item_path)
                            
                            result = {
                                'name': item,
                                'relative_path': relative_path,
                                'is_folder': is_folder,
                                'time': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                            }
                            
                            if is_folder:
                                result['size'] = '文件夹'
                            else:
                                result['size'] = get_file_size(stat.st_size)
                            
                            results.append(result)
                        except (OSError, PermissionError):
                            continue
                    
                    if search_subfolders and os.path.isdir(item_path) and len(results) < MAX_RESULTS:
                        search_in_directory(item_path, relative_path, depth + 1)
                        
            except (OSError, PermissionError):
                pass
        
        search_in_directory(base_path)
        
        results.sort(key=lambda x: (not x['is_folder'], x['name'].lower()))
        
        is_truncated = len(results) >= MAX_RESULTS
        
        return jsonify({
            'success': True,
            'results': results,
            'count': len(results),
            'is_truncated': is_truncated,
            'message': f'结果已达到上限（{MAX_RESULTS}），请使用更具体的关键词。' if is_truncated else None
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/download/<path:filename>')
def download_file(filename):
    """Download a single file with range request support."""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath):
            flash('文件不存在。', 'danger')
            return redirect(url_for('index'))
        
        file_size = os.path.getsize(filepath)
        download_name = os.path.basename(filename)
        range_header = request.headers.get('Range', None)
        
        if range_header:
            byte_start = 0
            byte_end = file_size - 1
            
            match = re.search(r'bytes=(\d+)-(\d*)', range_header)
            if match:
                byte_start = int(match.group(1))
                if match.group(2):
                    byte_end = int(match.group(2))
            
            def generate_chunk():
                chunk_size = 1024 * 1024  # 1MB chunks for faster transfer
                with open(filepath, 'rb') as f:
                    f.seek(byte_start)
                    remaining = byte_end - byte_start + 1
                    while remaining > 0:
                        read_size = min(chunk_size, remaining)
                        data = f.read(read_size)
                        if not data:
                            break
                        remaining -= len(data)
                        yield data
            
            response = Response(
                generate_chunk(),
                206,
                mimetype='application/octet-stream',
                direct_passthrough=True
            )
            response.headers.add('Content-Range', f'bytes {byte_start}-{byte_end}/{file_size}')
            response.headers.add('Accept-Ranges', 'bytes')
            response.headers.add('Content-Length', str(byte_end - byte_start + 1))
            response.headers.add('Content-Disposition', f'attachment; filename="{download_name}"')
            
            if byte_start == 0:
                add_activity(get_current_username(), 'download', download_name)
            
            return response
        else:
            response = send_file(
                filepath,
                as_attachment=True,
                download_name=download_name,
                mimetype='application/octet-stream',
                conditional=True,
                max_age=0
            )
            response.headers.add('Accept-Ranges', 'bytes')
            
        add_activity(get_current_username(), 'download', download_name)
             
        return response
             
    except Exception as e:
        flash(f'下载失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/stream/<path:filename>')
def stream_file(filename):
    """Stream an inline-playable media file."""
    try:
        actual_target = resolve_macos_metadata_target(filename)
        if actual_target:
            return redirect(url_for('stream_file', filename=actual_target))
        if is_macos_metadata_file(filename):
            flash('这个以 ._ 开头的文件是 macOS 生成的元数据文件，不是真正的音视频文件。', 'warning')
            parent_path = os.path.dirname(filename).replace('\\', '/')
            return redirect(url_for('index', subpath=parent_path) if parent_path else url_for('index'))

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], filename)
        if not success:
            flash(f'路径错误: {error}', 'danger')
            return redirect(url_for('index'))

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            flash('文件不存在。', 'danger')
            return redirect(url_for('index'))

        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        if ext not in INLINE_STREAM_EXTENSIONS:
            flash('该文件类型不支持内联播放', 'warning')
            return redirect(url_for('index'))

        response = send_file(
            filepath,
            mimetype=guess_inline_mimetype(filename),
            as_attachment=False,
            conditional=True,
            download_name=os.path.basename(filename),
            max_age=0
        )
        response.headers['Content-Disposition'] = f'inline; filename="{os.path.basename(filename)}"'
        response.headers['Accept-Ranges'] = 'bytes'
        return response
    except Exception as e:
        flash(f'播放失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/download_folder/<path:folder_name>')
def download_folder(folder_name):
    """下载整个文件夹，打包为 ZIP 并尽量流式压缩。"""
    try:
        success, folder_path, error = safe_join_path(app.config['UPLOAD_FOLDER'], folder_name)
        if not success:
            flash(f'路径错误: {error}', 'danger')
            return redirect(url_for('index'))
        
        if not os.path.exists(folder_path) or not os.path.isdir(folder_path):
            flash('文件夹不存在', 'danger')
            return redirect(url_for('index'))
        
        file_count = 0
        total_size = 0
        for root, dirs, files in os.walk(folder_path):
            file_count += len(files)
            for file in files:
                file_path = os.path.join(root, file)
                try:
                    total_size += os.path.getsize(file_path)
                except:
                    pass
        
        print(f"\n开始打包文件夹: {folder_name}")
        print(f"   文件数量: {file_count}")
        print(f"   总大小: {total_size / 1024 / 1024:.2f} MB")
        
        # 使用临时文件进行流式压缩（避免内存溢出）
        import tempfile
        temp_zip = tempfile.NamedTemporaryFile(delete=False, suffix='.zip')
        temp_zip.close()
        
        try:
            with zipfile.ZipFile(temp_zip.name, 'w', zipfile.ZIP_DEFLATED, compresslevel=1) as zipf:
                processed = 0
                no_compress_ext = {'.zip', '.rar', '.7z', '.gz', '.jpg', '.jpeg', '.png', '.gif', 
                                  '.mp4', '.mp3', '.avi', '.mkv', '.mov', '.flac', '.wav'}
                
                for root, dirs, files in os.walk(folder_path):
                    # 首先添加目录结构（确保空文件夹也被包含）
                    for dir_name in dirs:
                        dir_path = os.path.join(root, dir_name)
                        arcname = os.path.relpath(dir_path, folder_path) + '/'
                        zipf.writestr(arcname, '')
                    
                    # 添加文件
                    for file in files:
                        file_path = os.path.join(root, file)
                        arcname = os.path.relpath(file_path, folder_path)
                        
                        # 对已压缩文件使用ZIP_STORED（不压缩）以加快速度
                        file_ext = os.path.splitext(file)[1].lower()
                        if file_ext in no_compress_ext:
                            zipf.write(file_path, arcname, compress_type=zipfile.ZIP_STORED)
                        else:
                            zipf.write(file_path, arcname, compress_type=zipfile.ZIP_DEFLATED)
                        
                        processed += 1
                        if processed % 100 == 0:
                            print(f"   进度: {processed}/{file_count} ({processed*100//file_count}%)")
            
            print(f"打包完成: {folder_name}.zip")
            
            # 记录下载活动
            add_activity(get_current_username(), '下载', f'文件夹 {folder_name}')
            
            # 返回文件（优化：固定mimetype，支持条件请求）
            zip_filename = os.path.basename(folder_name) + '.zip'
            return send_file(
                temp_zip.name,
                mimetype='application/zip',
                as_attachment=True,
                download_name=zip_filename,
                conditional=True,  # 支持条件请求
                max_age=0  # 禁用缓存
            )
        except Exception as e:
            # 清理临时文件
            try:
                os.unlink(temp_zip.name)
            except:
                pass
            raise e
            
    except Exception as e:
        print(f"下载文件夹失败: {str(e)}")
        flash(f'下载文件夹失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.errorhandler(413)
def request_entity_too_large(error):
    """处理文件过大错误"""
    try:
        import werkzeug
        # 尝试获取版本号（不同版本属性名可能不同）
        werkzeug_version = getattr(werkzeug, '__version__', 
                                   getattr(werkzeug, 'version', 
                                          getattr(werkzeug, '__version_info__', '未知')))
    except:
        werkzeug_version = '未知'
    
    max_size = app.config.get('MAX_CONTENT_LENGTH', 'unknown')
    max_size_gb = max_size / (1024**3) if isinstance(max_size, int) else 0
    
    error_msg = (
        '上传失败：服务器返回 413（Request Entity Too Large）。\n'
        f'当前配置：MAX_CONTENT_LENGTH = {max_size_gb:.1f}GB\n'
        f'Werkzeug 版本：{werkzeug_version}\n\n'
        '请重启服务后重试；如果仍失败，请检查文件大小和上传配置。'
    )
    flash(error_msg, 'danger')
    print("\n[upload] 413 error details:")
    print(f"   MAX_CONTENT_LENGTH: {max_size}")
    print(f"   Werkzeug version: {werkzeug_version}")
    return redirect(url_for('index'))

# 分块上传和任务管理API
@app.route('/api/upload_chunk', methods=['POST'])
def upload_chunk():
    """分块上传文件"""
    try:
        task_id = request.form.get('task_id')
        chunk_index = int(request.form.get('chunk_index', 0))
        total_chunks = int(request.form.get('total_chunks', 1))
        filename = request.form.get('filename', '')
        upload_path = request.form.get('upload_path', '')  # 获取上传路径
        chunk_data = request.files.get('chunk')
        
        if not task_id or not chunk_data:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400
        
        # 安全化文件名
        filename = secure_filename(filename)
        
        if upload_path:
            success, target_dir, error = safe_join_path(app.config['UPLOAD_FOLDER'], upload_path)
            if not success:
                return jsonify({'success': False, 'message': error}), 400
            # 确保目标目录存在
            os.makedirs(target_dir, exist_ok=True)
            filepath = os.path.join(target_dir, filename)
        else:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        # 使用临时文件夹存储分块文件，避免在上传文件夹中产生小文件
        temp_filepath = os.path.join(TEMP_FOLDER, f'{task_id}_{filename}.tmp')
        
        with tasks_lock:
            if task_id not in tasks:
                username = get_current_username()
                
                tasks[task_id] = {
                    'type': 'upload',
                    'status': 'running',
                    'filename': filename,
                    'upload_path': (upload_path or '').replace('\\', '/').strip('/'),
                    'total_chunks': total_chunks,
                    'uploaded_chunks': 0,
                    'created_at': datetime.now().isoformat(),
                    'ip': request.remote_addr,
                    'username': username
                }
            else:
                uploaded_count = sum(1 for i in range(total_chunks) 
                                    if os.path.exists(temp_filepath + f'.chunk{i}'))
                tasks[task_id]['uploaded_chunks'] = uploaded_count
                if upload_path is not None:
                    tasks[task_id]['upload_path'] = (upload_path or '').replace('\\', '/').strip('/')
                # 如果任务状态不是 running，也不要强制覆盖，允许保持 paused
            
            task = tasks[task_id]
            
            if task['status'] == 'paused':
                return jsonify({'success': False, 'message': '任务已暂停。', 'paused': True}), 200
            
            if task['status'] != 'running':
                task['status'] = 'running'
            
            # 检查分块是否已存在，用于断点续传
            chunk_file = temp_filepath + f'.chunk{chunk_index}'
            if os.path.exists(chunk_file):
                # 分块已存在，跳过保存
                pass
            else:
                # 保存分块数据
                chunk_data.save(chunk_file)
            
            # 更新已上传的分块数量
            uploaded_count = sum(1 for i in range(total_chunks) 
                                if os.path.exists(temp_filepath + f'.chunk{i}'))
            task['uploaded_chunks'] = uploaded_count
            
            if uploaded_count >= total_chunks:
                try:
                    with open(temp_filepath, 'wb') as outfile:
                        for i in range(total_chunks):
                            chunk_file = temp_filepath + f'.chunk{i}'
                            if os.path.exists(chunk_file):
                                with open(chunk_file, 'rb') as infile:
                                    shutil.copyfileobj(infile, outfile)
                    
                    if not os.path.exists(temp_filepath):
                        raise Exception('合并后的临时文件不存在。')
                    
                    # 原子性地替换目标文件
                    # 使用临时文件名，然后原子性重命名
                    if os.path.exists(filepath):
                        backup_path = filepath + '.backup'
                        try:
                            shutil.move(filepath, backup_path)
                            shutil.move(temp_filepath, filepath)
                            if os.path.exists(backup_path):
                                os.remove(backup_path)
                        except Exception as e:
                            # 恢复备份
                            if os.path.exists(backup_path):
                                shutil.move(backup_path, filepath)
                            raise e
                    else:
                        shutil.move(temp_filepath, filepath)
                    
                    for i in range(total_chunks):
                        chunk_file = temp_filepath + f'.chunk{i}'
                        if os.path.exists(chunk_file):
                            try:
                                os.remove(chunk_file)
                            except:
                                pass  # 忽略清理失败
                    
                except Exception as merge_error:
                    # 合并失败，保留分块文件供重试
                    print(f"[upload-chunk] merge failed: {merge_error}")
                    raise merge_error
                
                task['status'] = 'completed'
                task['completed_at'] = datetime.now().isoformat()
                task['updated_at'] = datetime.now().isoformat()
                
                # 记录活动
                add_activity(get_current_username(), 'upload', filename)
            else:
                task['progress'] = round((task['uploaded_chunks'] / total_chunks) * 100, 1)
                # 如果进度大于 99 但小于 100，则显示为 99.9，避免提前显示 100%
                if task['progress'] > 99 and task['progress'] < 100:
                    task['progress'] = 99.9
            
            save_tasks()
        
        return jsonify({
            'success': True,
            'progress': task.get('progress', 0),
            'uploaded_chunks': task['uploaded_chunks'],
            'total_chunks': total_chunks,
            'completed': task['status'] == 'completed'
        })
        
    except Exception as e:
        print(f"[upload-chunk] failed: {str(e)}")
        with tasks_lock:
            if task_id in tasks:
                tasks[task_id]['status'] = 'error'
                tasks[task_id]['error'] = str(e)
                tasks[task_id]['updated_at'] = datetime.now().isoformat()
                save_tasks()
        
        try:
            temp_filepath = os.path.join(TEMP_FOLDER, f'{task_id}_{filename}.tmp')
            if os.path.exists(temp_filepath):
                os.remove(temp_filepath)
            for i in range(total_chunks):
                chunk_file = temp_filepath + f'.chunk{i}'
                if os.path.exists(chunk_file):
                    os.remove(chunk_file)
        except Exception as cleanup_error:
            print(f"[upload-chunk] cleanup failed: {cleanup_error}")
        
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/api/tasks', methods=['GET'])
def get_tasks():
    """Docstring."""
    with tasks_lock:
        # 只返回当前IP的任务，排除已完成的任务
        client_id = request.remote_addr
        user_tasks = {}
        tasks_to_delete = []
        
        for tid, task in tasks.items():
            if task.get('ip') == client_id and task['status'] != 'completed':
                # 检查文件是否已上传完成，通过检查目标文件是否存在
                if task['type'] == 'upload':
                    filename = task.get('filename', '')
                    if filename:
                        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
                        if os.path.exists(filepath):
                            tasks_to_delete.append(tid)
                            continue
                
                user_tasks[tid] = task
        
        # 删除已完成的任务
        for tid in tasks_to_delete:
            del tasks[tid]
        
        if tasks_to_delete:
            save_tasks()
        
        return jsonify({'success': True, 'tasks': user_tasks})

@app.route('/api/all_tasks', methods=['GET'])
def get_all_tasks():
    """获取所有任务列表，包括已完成任务，供清理使用。"""
    with tasks_lock:
        # 返回当前IP的所有任务，包括已完成的
        client_id = request.remote_addr
        user_tasks = {
            tid: task for tid, task in tasks.items()
            if task.get('ip') == client_id
        }
        return jsonify({'success': True, 'tasks': user_tasks})

@app.route('/api/tasks/<task_id>/pause', methods=['POST'])
def pause_task(task_id):
    """Pause a background task."""
    with tasks_lock:
        if task_id in tasks:
            tasks[task_id]['status'] = 'paused'
            tasks[task_id]['paused_at'] = datetime.now().isoformat()
            save_tasks()
            return jsonify({'success': True, 'message': '任务已暂停。'})
        return jsonify({'success': False, 'message': '任务不存在。'}), 404

@app.route('/api/tasks/<task_id>/resume', methods=['POST'])
def resume_task(task_id):
    """Resume a paused background task."""
    with tasks_lock:
        if task_id in tasks:
            tasks[task_id]['status'] = 'running'
            if 'resumed_at' not in tasks[task_id]:
                tasks[task_id]['resumed_at'] = []
            tasks[task_id]['resumed_at'].append(datetime.now().isoformat())
            save_tasks()
            return jsonify({'success': True, 'message': '任务已恢复。'})
        return jsonify({'success': False, 'message': '任务不存在。'}), 404

@app.route('/api/tasks/<task_id>', methods=['DELETE'])
def delete_task(task_id):
    """Delete a background task and its temporary files."""
    print(f"[tasks] delete requested: {task_id}")
    
    deleted_count = 0
    try:
        # 遍历临时文件夹，删除所有以 task_id 开头的文件
        for filename in os.listdir(TEMP_FOLDER):
            if filename.startswith(task_id + '_'):
                filepath = os.path.join(TEMP_FOLDER, filename)
                try:
                    os.remove(filepath)
                    deleted_count += 1
                    print(f"   removed temp file: {filename}")
                except Exception as e:
                    print(f"   failed removing temp file {filename}: {e}")
    except Exception as e:
        print(f"   failed cleaning temp folder: {e}")
    
    print(f"   removed {deleted_count} temp files")
    
    with tasks_lock:
        if task_id in tasks:
            task = tasks[task_id]
            print(f"   task info: {task.get('filename')} - status: {task.get('status')}")
            
            # 直接删除任务
            del tasks[task_id]
            save_tasks()
            print(f"[tasks] deleted {task_id}")
            return jsonify({'success': True, 'message': '任务已删除。', 'deleted_files': deleted_count})
        else:
            print("[tasks] task missing from registry, temp files already cleaned")
            return jsonify({'success': True, 'message': '任务已删除。', 'deleted_files': deleted_count})

@app.route('/api/rename', methods=['POST'])
def rename_item():
    """Rename a file or folder inside the shared directory."""
    try:
        data = request.get_json()
        old_path = data.get('old_path', '')
        new_name = data.get('new_name', '').strip()
        current_path = data.get('current_path', '')
        
        if not old_path or not new_name:
            return jsonify({'success': False, 'message': '缺少必要参数。'}), 400
        
        # 安全化新名称
        new_name = secure_filename(new_name)
        if not new_name:
            return jsonify({'success': False, 'message': '新名称无效。'}), 400
        
        # 构建完整路径
        old_full_path = os.path.join(UPLOAD_FOLDER, old_path)
        
        old_full_path = os.path.abspath(old_full_path)
        if not old_full_path.startswith(os.path.abspath(UPLOAD_FOLDER)):
            return jsonify({'success': False, 'message': '非法路径'}), 403
        
        if not os.path.exists(old_full_path):
            return jsonify({'success': False, 'message': '文件或文件夹不存在。'}), 404
        
        # 构建新路径，保持在同一目录下
        parent_dir = os.path.dirname(old_full_path)
        new_full_path = os.path.join(parent_dir, new_name)
        
        # 二次安全检查，确保新路径也在安全范围内
        new_full_path = os.path.abspath(new_full_path)
        if not new_full_path.startswith(os.path.abspath(UPLOAD_FOLDER)):
            return jsonify({'success': False, 'message': '非法目标路径'}), 403
        
        if os.path.exists(new_full_path):
            return jsonify({'success': False, 'message': f'"{new_name}" 已存在。'}), 400
        
        # 执行重命名（使用shutil.move支持跨磁盘移动）
        shutil.move(old_full_path, new_full_path)
        
        # 记录活动
        username = get_current_username()
        old_name = os.path.basename(old_path)
        item_type = '文件夹' if os.path.isdir(new_full_path) else '文件'
        add_activity(username, 'rename', f'{old_name} -> {new_name}')
        
        print(f"[rename] {old_name} -> {new_name}")
        
        return jsonify({
            'success': True, 
            'message': f'{item_type}已重命名为 "{new_name}"',
            'new_name': new_name
        })
        
    except Exception as e:
        print(f"[rename] failed: {str(e)}")
        return jsonify({'success': False, 'message': f'重命名失败: {str(e)}'}), 500

@app.route('/delete/<path:filename>', methods=['POST'])
def delete_file(filename):
    """删除文件，需要管理员权限。"""
    # 检查管理员权限
    username = normalize_username(session.get('username'))
    if not is_user_admin(username):
        flash('权限不足：只有管理员才能删除文件。', 'danger')
        return redirect(url_for('index'))
    
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if os.path.exists(filepath):
            os.remove(filepath)
            # 删除空文件夹
            try:
                parent_dir = os.path.dirname(filepath)
                while parent_dir != app.config['UPLOAD_FOLDER']:
                    if not os.listdir(parent_dir):
                        os.rmdir(parent_dir)
                        parent_dir = os.path.dirname(parent_dir)
                    else:
                        break
            except:
                pass
            flash(f'文件 "{os.path.basename(filename)}" 已删除。', 'success')
            # 记录活动
            add_activity(username, 'delete', os.path.basename(filename))
        else:
            flash('文件不存在。', 'danger')
    except Exception as e:
        flash(f'删除失败: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/delete_folder/<path:foldername>', methods=['POST'])
def delete_folder(foldername):
    """删除文件夹及其全部内容，需要管理员权限。"""
    # 检查管理员权限
    username = normalize_username(session.get('username'))
    if not is_user_admin(username):
        flash('权限不足：只有管理员才能删除文件夹。', 'danger')
        return redirect(url_for('index'))
    
    try:
        folder_path = os.path.join(app.config['UPLOAD_FOLDER'], foldername)
        
        if not os.path.exists(folder_path):
            flash('文件夹不存在。', 'danger')
            return redirect(url_for('index'))
        
        if not os.path.isdir(folder_path):
            flash('目标不是文件夹。', 'danger')
            return redirect(url_for('index'))
        
        # 统计文件数量
        file_count = sum(len(files) for _, _, files in os.walk(folder_path))
        
        shutil.rmtree(folder_path)
        
        print(f"[delete-folder] removed {foldername} ({file_count} files)")
        flash(f'文件夹 "{foldername}" 及其 {file_count} 个文件已删除。', 'success')
        
    except Exception as e:
        print(f"[delete-folder] failed: {str(e)}")
        flash(f'删除文件夹失败: {str(e)}', 'danger')
    
    return redirect(url_for('index'))

@app.route('/move_file', methods=['POST'])
def move_file():
    """移动文件或文件夹"""
    try:
        data = request.get_json()
        if not data:
            return {'success': False, 'message': '无效的请求数据。'}, 400
            
        filename = (data.get('filename') or '').strip()
        target_folder = (data.get('target_folder') or '').strip()
        
        if not filename:
            return {'success': False, 'message': '名称不能为空。'}, 400
        
        source_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(source_path):
            return {'success': False, 'message': '源项目不存在。'}, 404
        
        # 判断是文件还是文件夹
        is_folder = os.path.isdir(source_path)
        item_type = '文件夹' if is_folder else '文件'
        
        # 如果目标文件夹为空，表示移动到根目录
        if not target_folder:
            # 移动到根目录
            file_basename = os.path.basename(filename)
            target_path = os.path.join(app.config['UPLOAD_FOLDER'], file_basename)
            
            normalized_filename = filename.replace('\\', '/')
            if '/' not in normalized_filename:
                return {'success': False, 'message': f'{item_type}已在根目录。'}, 400
        else:
            target_folder_path = os.path.join(app.config['UPLOAD_FOLDER'], target_folder)
            
            if not os.path.exists(target_folder_path):
                return {'success': False, 'message': '目标文件夹不存在。'}, 404
            
            if not os.path.isdir(target_folder_path):
                return {'success': False, 'message': '目标不是文件夹。'}, 400
            
            # 如果移动的是文件夹，检查不能移动到自身子目录
            if is_folder:
                source_normalized = os.path.normpath(source_path)
                target_normalized = os.path.normpath(target_folder_path)
                
                if target_normalized.startswith(source_normalized + os.sep) or target_normalized == source_normalized:
                    return {'success': False, 'message': '不能将文件夹移动到自身或其子目录中。'}, 400
            
            file_basename = os.path.basename(filename)
            target_path = os.path.join(target_folder_path, file_basename)
            
            if os.path.normpath(source_path) == os.path.normpath(target_path):
                return {'success': False, 'message': f'{item_type}已在该位置。'}, 400
        
        # 检查目标位置是否已存在同名项目
        if os.path.exists(target_path) and os.path.normpath(source_path) != os.path.normpath(target_path):
            return {'success': False, 'message': f'目标位置已存在同名{item_type}。'}, 400
        
        # 移动文件或文件夹
        shutil.move(source_path, target_path)
        
        target_label = target_folder if target_folder else '根目录'
        print(f"[move] {item_type}: {filename} -> {target_label}")
        return {'success': True, 'message': f'{item_type}移动成功。'}, 200
        
    except Exception as e:
        print(f"[move] failed: {str(e)}")
        return {'success': False, 'message': f'移动失败: {str(e)}'}, 500

@app.route('/api/batch_delete', methods=['POST'])
def batch_delete():
    """Docstring."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据。'}), 400
            
        paths = data.get('paths', [])
        if not paths:
            return jsonify({'success': False, 'message': '请选择要删除的项目。'}), 400
        
        success_count = 0
        failed_items = []
        
        for path in paths:
            try:
                success, full_path, error = safe_join_path(app.config['UPLOAD_FOLDER'], path)
                if not success:
                    failed_items.append(f'{path}: {error}')
                    continue
                
                if not os.path.exists(full_path):
                    failed_items.append(f'{path}: 不存在')
                    continue
                
                # 删除文件或文件夹
                if os.path.isdir(full_path):
                    shutil.rmtree(full_path)
                    print(f"[batch-delete] folder: {path}")
                else:
                    os.unlink(full_path)
                    print(f"[batch-delete] file: {path}")
                
                success_count += 1
            except Exception as e:
                failed_items.append(f'{path}: {str(e)}')
                print(f"[batch-delete] failed {path}: {str(e)}")
        
        message = f'成功删除 {success_count} 项'
        if failed_items:
            message += f'，失败 {len(failed_items)} 项'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'failed_items': failed_items
        }), 200
        
    except Exception as e:
        print(f"[batch-delete] failed: {str(e)}")
        return jsonify({'success': False, 'message': f'批量删除失败: {str(e)}'}), 500

@app.route('/api/batch_move', methods=['POST'])
def batch_move():
    """Docstring."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据。'}), 400
            
        paths = data.get('paths', [])
        target_folder = (data.get('target_folder') or '').strip()
        
        if not paths:
            return jsonify({'success': False, 'message': '请选择要移动的项目。'}), 400
        
        if target_folder:
            target_folder_path = os.path.join(app.config['UPLOAD_FOLDER'], target_folder)
            if not os.path.exists(target_folder_path):
                return jsonify({'success': False, 'message': '目标文件夹不存在。'}), 404
            if not os.path.isdir(target_folder_path):
                return jsonify({'success': False, 'message': '目标不是文件夹。'}), 400
        else:
            target_folder_path = app.config['UPLOAD_FOLDER']
        
        success_count = 0
        failed_items = []
        
        for path in paths:
            try:
                success, source_path, error = safe_join_path(app.config['UPLOAD_FOLDER'], path)
                if not success:
                    failed_items.append(f'{path}: {error}')
                    continue
                
                if not os.path.exists(source_path):
                    failed_items.append(f'{path}: 不存在')
                    continue
                
                # 检查是否移动到自己或子目录
                if os.path.isdir(source_path):
                    source_normalized = os.path.normpath(source_path)
                    target_normalized = os.path.normpath(target_folder_path)
                    if target_normalized.startswith(source_normalized + os.sep) or target_normalized == source_normalized:
                        failed_items.append(f'{path}: 不能移动到自己或子文件夹')
                        continue
                
                # 构建目标路径
                file_basename = os.path.basename(path)
                target_path = os.path.join(target_folder_path, file_basename)
                
                if os.path.normpath(source_path) == os.path.normpath(target_path):
                    failed_items.append(f'{path}: 已在目标位置')
                    continue
                
                if os.path.exists(target_path):
                    failed_items.append(f'{path}: 目标位置存在同名项目')
                    continue
                
                # 移动文件
                shutil.move(source_path, target_path)
                target_label = target_folder if target_folder else '根目录'
                print(f"[batch-move] {path} -> {target_label}")
                success_count += 1
                
            except Exception as e:
                failed_items.append(f'{path}: {str(e)}')
                print(f"[batch-move] failed {path}: {str(e)}")
        
        message = f'成功移动 {success_count} 项'
        if failed_items:
            message += f'，失败 {len(failed_items)} 项'
        
        return jsonify({
            'success': True,
            'message': message,
            'success_count': success_count,
            'failed_items': failed_items
        }), 200
        
    except Exception as e:
        print(f"[batch-move] failed: {str(e)}")
        return jsonify({'success': False, 'message': f'批量移动失败: {str(e)}'}), 500

@app.route('/create_folder', methods=['POST'])
def create_folder():
    """创建新文件夹"""
    try:
        folder_name = request.form.get('folder_name', '').strip()
        current_path = request.form.get('current_path', '').strip()
        
        if not folder_name:
            flash('请输入文件夹名称。', 'danger')
            if current_path:
                return redirect(url_for('index', subpath=current_path))
            else:
                return redirect(url_for('index'))
        
        folder_name = secure_filename(folder_name)
        
        if not folder_name:
            flash('文件夹名称无效。', 'danger')
            if current_path:
                return redirect(url_for('index', subpath=current_path))
            else:
                return redirect(url_for('index'))
        
        # 构建完整路径（在当前目录下创建）
        if current_path:
            folder_path = os.path.join(app.config['UPLOAD_FOLDER'], current_path, folder_name)
        else:
            folder_path = os.path.join(app.config['UPLOAD_FOLDER'], folder_name)
        
        folder_path = os.path.abspath(folder_path)
        if not folder_path.startswith(os.path.abspath(app.config['UPLOAD_FOLDER'])):
            flash('非法路径', 'danger')
            if current_path:
                return redirect(url_for('index', subpath=current_path))
            else:
                return redirect(url_for('index'))
        
        if os.path.exists(folder_path):
            flash(f'文件夹 "{folder_name}" 已存在。', 'warning')
            if current_path:
                return redirect(url_for('index', subpath=current_path))
            else:
                return redirect(url_for('index'))
        
        os.makedirs(folder_path, exist_ok=True)
        
        location_desc = f'在 {current_path}' if current_path else '在根目录'
        print(f"[create-folder] {folder_name} {location_desc}")
        flash(f'文件夹 "{folder_name}" 创建成功。', 'success')
        
    except Exception as e:
        print(f"[create-folder] failed: {str(e)}")
        flash(f'创建文件夹失败: {str(e)}', 'danger')
    
    # 返回到当前目录（如果有的话）
    if current_path:
        return redirect(url_for('index', subpath=current_path))
    else:
        return redirect(url_for('index'))

# ==================== 分享链接功能 ====================
@app.route('/api/create_share', methods=['POST'])
def create_share():
    """创建文件分享链接"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据。'}), 400
        
        file_path = data.get('file_path', '').strip()
        password = data.get('password', '').strip() or None
        expire_hours = int(data.get('expire_hours', 24))
        max_downloads = data.get('max_downloads') or None
        if max_downloads:
            max_downloads = int(max_downloads)
        
        if not file_path:
            return jsonify({'success': False, 'message': '文件路径不能为空。'}), 400
        
        # 验证文件存在
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], file_path)
        if not os.path.exists(full_path) or not os.path.isfile(full_path):
            return jsonify({'success': False, 'message': '文件不存在。'}), 404
        
        # 生成分享链接ID
        link_id = str(uuid.uuid4())[:8]
        expires = time.time() + (expire_hours * 3600)
        
        with share_links_lock:
            share_links[link_id] = {
                'file_path': file_path,
                'password': password,
                'expires': expires,
                'downloads': 0,
                'max_downloads': max_downloads,
                'created_by': get_current_username(),
                'created_at': datetime.now().isoformat()
            }
            save_share_links()
        
        share_url = f"http://{request.host}/share/{link_id}"
        
        print(f"[share] created link for {file_path}: {link_id}")
        return jsonify({
            'success': True,
            'link_id': link_id,
            'share_url': share_url,
            'expires_in': expire_hours,
            'message': '分享链接创建成功。'
        })
    
    except Exception as e:
        print(f"[share] create failed: {str(e)}")
        return jsonify({'success': False, 'message': f'创建失败: {str(e)}'}), 500

@app.route('/share/<link_id>')
def access_share(link_id):
    """访问分享链接"""
    with share_links_lock:
        if link_id not in share_links:
            flash('分享链接不存在或已失效。', 'danger')
            return redirect(url_for('index'))
        
        link_info = share_links[link_id]
        
        if time.time() > link_info['expires']:
            flash('分享链接已过期。', 'danger')
            del share_links[link_id]
            save_share_links()
            return redirect(url_for('index'))
        
        if link_info['max_downloads'] and link_info['downloads'] >= link_info['max_downloads']:
            flash('分享链接下载次数已达上限', 'danger')
            del share_links[link_id]
            save_share_links()
            return redirect(url_for('index'))
        
        if link_info['password']:
            password = request.args.get('pwd', '')
            if password != link_info['password']:
                # 返回密码输入页面
                return render_template_string('''
                <!DOCTYPE html>
                <html>
                <head>
                    <title>请输入密码</title>
                    <style>
                        body {
                            font-family: Arial, sans-serif;
                            display: flex;
                            justify-content: center;
                            align-items: center;
                            min-height: 100vh;
                            margin: 0;
                            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        }
                        .password-box {
                            background: white;
                            padding: 40px;
                            border-radius: 15px;
                            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
                            text-align: center;
                            max-width: 400px;
                        }
                        h2 { color: #333; margin-bottom: 20px; }
                        input {
                            width: 100%;
                            padding: 12px;
                            border: 2px solid #ddd;
                            border-radius: 8px;
                            font-size: 16px;
                            margin-bottom: 15px;
                            box-sizing: border-box;
                        }
                        button {
                            width: 100%;
                            padding: 12px;
                            background: #667eea;
                            color: white;
                            border: none;
                            border-radius: 8px;
                            font-size: 16px;
                            cursor: pointer;
                            transition: background 0.3s;
                        }
                        button:hover { background: #764ba2; }
                        .error { color: red; margin-bottom: 10px; }
                    </style>
                </head>
                <body>
                    <div class="password-box">
                        <h2>此分享链接需要密码</h2>
                        {% if request.args.get('pwd') %}
                        <div class="error">密码错误，请重试</div>
                        {% endif %}
                        <form method="get">
                            <input type="password" name="pwd" placeholder="请输入密码" required autofocus>
                            <button type="submit">访问</button>
                        </form>
                    </div>
                </body>
                </html>
                ''')
        
        # 增加下载次数
        link_info['downloads'] += 1
        save_share_links()
        
        # 下载文件
        file_path = link_info['file_path']
        full_path = os.path.join(app.config['UPLOAD_FOLDER'], file_path)
        
        if not os.path.exists(full_path):
            flash('文件不存在。', 'danger')
            return redirect(url_for('index'))
        
        filename = os.path.basename(file_path)
        print(f"[share] download via link: {filename} ({link_id})")
        
        return send_file(full_path, as_attachment=True, download_name=filename)

# ==================== 批量下载功能 ====================
@app.route('/api/batch_download', methods=['POST'])
def batch_download():
    """批量下载选中的文件，打包为 ZIP 并尽量流式压缩。"""
    try:
        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'message': '无效的请求数据。'}), 400
        
        paths = data.get('paths', [])
        if not paths:
            return jsonify({'success': False, 'message': '请选择要下载的文件。'}), 400
        
        # 使用内存流进行ZIP压缩（避免大文件夹卡住）
        memory_file = io.BytesIO()
        
        print(f"[batch-download] start: {len(paths)} items")
        
        # 流式压缩
        with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            file_count = 0
            for path in paths:
                success, full_path, error = safe_join_path(app.config['UPLOAD_FOLDER'], path)
                if not success:
                    print(f"[batch-download] skip invalid path: {path}")
                    continue
                
                if not os.path.exists(full_path):
                    print(f"[batch-download] skip missing path: {path}")
                    continue
                
                # 如果是文件，直接添加
                if os.path.isfile(full_path):
                    try:
                        zipf.write(full_path, arcname=os.path.basename(path))
                        file_count += 1
                        print(f"  added file: {os.path.basename(path)}")
                    except Exception as e:
                        print(f"  failed adding file: {path}, error: {e}")
                
                # 如果是文件夹，则递归添加
                elif os.path.isdir(full_path):
                    print(f"  compressing folder: {os.path.basename(path)}")
                    folder_file_count = 0
                    for root, dirs, files in os.walk(full_path):
                        for file in files:
                            try:
                                file_path = os.path.join(root, file)
                                arcname = os.path.join(
                                    os.path.basename(path),
                                    os.path.relpath(file_path, full_path)
                                )
                                zipf.write(file_path, arcname=arcname)
                                folder_file_count += 1
                                file_count += 1
                                
                                if folder_file_count % 100 == 0:
                                    print(f"    compressed {folder_file_count} files...")
                            except Exception as e:
                                print(f"    skipped file: {file}, error: {e}")
                    
                    print(f"  finished folder: {os.path.basename(path)} ({folder_file_count} files)")
        
        print(f"[batch-download] completed: {file_count} files")
        
        memory_file.seek(0)
        
        # 返回ZIP文件
        zip_filename = f'批量下载_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip'
        
        return send_file(
            memory_file,
            mimetype='application/zip',
            as_attachment=True,
            download_name=zip_filename
        )
    
    except Exception as e:
        print(f"[batch-download] failed: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'批量下载失败: {str(e)}'}), 500

# ==================== 文件预览功能 ====================
def is_supported_document_editor(file_path, editor_type):
    """Docstring."""
    editor_type = str(editor_type or '').strip().lower()
    if editor_type == 'text':
        return is_text_previewable_file(file_path)
    if editor_type == 'docx':
        return is_word_editable_file(file_path)
    return False

@app.route('/api/document_collaboration_sync', methods=['POST'])
def document_collaboration_sync():
    """同步文档协作状态。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        editor_type = (data.get('editor_type') or '').strip().lower()
        active_target = data.get('active_target')
        lock_target = data.get('lock_target')
        release_lock = bool(data.get('release_lock'))

        if not file_path or not editor_type:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_supported_document_editor(file_path, editor_type):
            return jsonify({'success': False, 'message': '该文件类型暂不支持当前协作编辑器'}), 400

        snapshot = sync_document_collaboration_presence(
            filepath,
            get_current_username(),
            get_document_editor_client_id(),
            request.remote_addr,
            editor_type,
            active_target=active_target,
            lock_target=lock_target,
            release_lock=release_lock
        )
        snapshot['success'] = True
        return jsonify(snapshot)
    except Exception as e:
        print(f"文档协作同步失败: {str(e)}")
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'}), 500

@app.route('/api/document_collaboration_release', methods=['POST'])
def document_collaboration_release():
    """释放文档协作状态。"""
    try:
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            data = request.form.to_dict()

        file_path = (data.get('file_path') or '').strip()
        remove_session = str(data.get('remove_session', '')).lower() in {'1', 'true', 'yes', 'on'}
        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        client_id = get_document_editor_client_id()
        release_document_collaboration_presence(filepath, client_id, remove_session=remove_session)
        with text_realtime_lock:
            live_state = text_realtime_state.get(get_document_collaboration_file_key(filepath))
            if live_state:
                live_state.get('clients', {}).pop(client_id, None)
                live_state['last_activity'] = time.time()
        return jsonify({'success': True})
    except Exception as e:
        print(f"释放文档协作状态失败: {str(e)}")
        return jsonify({'success': False, 'message': f'释放失败: {str(e)}'}), 500

@app.route('/api/document_realtime_stream')
def document_realtime_stream():
    """为文本或 DOCX 编辑页提供实时事件流。"""
    file_path = (request.args.get('file_path') or '').strip()
    editor_type = (request.args.get('editor_type') or '').strip().lower()

    if not file_path or not editor_type:
        return jsonify({'success': False, 'message': '缺少必要参数'}), 400

    success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
    if not success:
        return jsonify({'success': False, 'message': error}), 400

    if not os.path.exists(filepath) or not os.path.isfile(filepath):
        return jsonify({'success': False, 'message': '文件不存在'}), 404

    if not is_supported_document_editor(file_path, editor_type):
        return jsonify({'success': False, 'message': '该文件类型暂不支持当前实时协作'}), 400

    client_id = get_document_editor_client_id()
    stream_id, stream_queue = register_document_realtime_stream(filepath, client_id, editor_type)

    def generate():
        try:
            if editor_type == 'text':
                snapshot = update_text_realtime_presence(filepath, client_id, get_current_username())
                yield (
                    "event: text_init\n"
                    f"data: {json.dumps(snapshot, ensure_ascii=False)}\n\n"
                )

            while True:
                try:
                    event_data = stream_queue.get(timeout=TEXT_REALTIME_STREAM_HEARTBEAT)
                    yield (
                        f"event: {event_data.get('type', 'message')}\n"
                        f"data: {json.dumps(event_data, ensure_ascii=False)}\n\n"
                    )
                except queue.Empty:
                    yield ": heartbeat\n\n"
        finally:
            unregister_document_realtime_stream(filepath, stream_id)

    response = Response(stream_with_context(generate()), mimetype='text/event-stream')
    response.headers['Cache-Control'] = 'no-cache'
    response.headers['X-Accel-Buffering'] = 'no'
    return response

@app.route('/api/text_realtime_update', methods=['POST'])
def text_realtime_update():
    """接收文本编辑器的实时内容更新。"""
    if not user_can_edit_files():
        return jsonify({'success': False, 'message': '只有管理员才能实时编辑文本文件'}), 403

    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        content = data.get('content')
        base_revision = data.get('base_revision')

        if not file_path or content is None:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_text_previewable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型不支持在线编辑'}), 400

        result = apply_text_realtime_update(
            filepath,
            get_document_editor_client_id(),
            get_current_username(),
            content,
            base_revision
        )
        if not result.get('success'):
            return jsonify(result), 409 if result.get('refresh_required') else 400
        return jsonify(result)
    except Exception as e:
        print(f"应用文本实时更新失败: {str(e)}")
        return jsonify({'success': False, 'message': f'实时更新失败: {str(e)}'}), 500

@app.route('/api/text_file_snapshot', methods=['POST'])
def text_file_snapshot():
    """获取文本文件最新快照，用于多人协作刷新。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()

        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_text_previewable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型不支持在线编辑'}), 400

        current_token = get_file_mtime_token(filepath)
        with text_realtime_lock:
            live_state = text_realtime_state.get(get_document_collaboration_file_key(filepath))
            if live_state and live_state.get('mtime_ns') != current_token:
                live_state = ensure_text_realtime_document_locked(filepath)
            if live_state:
                payload = {
                    'success': True,
                    'content': live_state.get('content', ''),
                    'encoding': live_state.get('encoding', 'utf-8'),
                    'newline': live_state.get('newline', '\n'),
                    'editable': True,
                    'mtime_ns': str(live_state.get('mtime_ns', current_token)),
                    'file_size': live_state.get('file_size', os.path.getsize(filepath)),
                    'revision': int(live_state.get('revision', 0))
                }
            else:
                payload = None

        if payload is not None:
            return jsonify(payload)

        preview = load_text_file_preview(filepath)
        return jsonify({
            'success': True,
            'content': preview['content'],
            'encoding': preview['encoding'],
            'newline': preview['newline'],
            'editable': preview['editable'],
            'mtime_ns': str(preview['mtime_ns']),
            'file_size': preview['file_size'],
            'revision': 0
        })
    except Exception as e:
        print(f"获取文本文件快照失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取快照失败: {str(e)}'}), 500

@app.route('/api/docx_file_snapshot', methods=['POST'])
def docx_file_snapshot():
    """获取 DOCX 文件最新快照，用于多人协作刷新。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()

        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_word_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        preview = load_docx_file_preview(filepath)
        return jsonify({
            'success': True,
            'editable': preview['editable'],
            'mtime_ns': str(preview['mtime_ns']),
            'file_size': preview['file_size'],
            'paragraph_count': preview['paragraph_count'],
            'table_count': preview['table_count'],
            'block_count': preview['block_count'],
            'blocks': preview['blocks']
        })
    except Exception as e:
        print(f"获取 DOCX 快照失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取快照失败: {str(e)}'}), 500

@app.route('/api/save_text_file', methods=['POST'])
def save_text_file():
    """保存在线编辑的文本文件。"""
    if not user_can_edit_files():
        return jsonify({'success': False, 'message': '只有管理员才能保存文件修改'}), 403

    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        content = data.get('content')
        encoding = (data.get('encoding') or 'utf-8').lower()
        newline = data.get('newline') or '\n'
        mtime_ns = data.get('mtime_ns')
        realtime_revision = data.get('realtime_revision')

        if not file_path or content is None:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_text_previewable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型不支持在线编辑'}), 400

        client_id = get_document_editor_client_id()
        with text_realtime_lock:
            live_state = text_realtime_state.get(get_document_collaboration_file_key(filepath))

        if live_state:
            result = apply_text_realtime_update(
                filepath,
                client_id,
                get_current_username(),
                content,
                realtime_revision if realtime_revision is not None else live_state.get('revision', 0)
            )
            if not result.get('success'):
                return jsonify(result), 409 if result.get('refresh_required') else 400

            write_result = write_text_content_to_file(
                filepath,
                result.get('content', ''),
                result.get('encoding', encoding),
                result.get('newline', newline)
            )
            with text_realtime_lock:
                refreshed_state = text_realtime_state.get(get_document_collaboration_file_key(filepath))
                if refreshed_state:
                    refreshed_state['content'] = result.get('content', '')
                    refreshed_state['encoding'] = write_result['encoding']
                    refreshed_state['newline'] = write_result['newline']
                    refreshed_state['mtime_ns'] = write_result['mtime_ns']
                    refreshed_state['file_size'] = write_result['size']
                    refreshed_state['last_activity'] = time.time()
                    _remember_text_realtime_snapshot_locked(refreshed_state)

            username = get_current_username()
            add_activity(username, '编辑', os.path.basename(file_path))
            print(f"已保存文本文件: {file_path} (用户: {username})")
            return jsonify({
                'success': True,
                'message': '保存成功',
                'mtime_ns': write_result['mtime_ns'],
                'encoding': write_result['encoding'],
                'size': get_file_size(write_result['size']),
                'realtime_revision': result['revision'],
                'merged': result.get('merged', False),
                'content': result.get('content'),
                'saved': True
            })

        try:
            mtime_ns = int(mtime_ns)
        except (TypeError, ValueError):
            return jsonify({'success': False, 'message': '文件版本信息无效，请刷新页面后重试'}), 400

        current_mtime_ns = os.stat(filepath).st_mtime_ns
        if current_mtime_ns != mtime_ns:
            return jsonify({
                'success': False,
                'message': '文件已被其他人修改，请刷新页面后再保存',
                'mtime_ns': str(current_mtime_ns)
            }), 409

        write_result = write_text_content_to_file(filepath, content, encoding, newline)
        username = get_current_username()
        add_activity(username, '编辑', os.path.basename(file_path))
        print(f"已保存文本文件: {file_path} (用户: {username})")

        return jsonify({
            'success': True,
            'message': '保存成功',
            'mtime_ns': write_result['mtime_ns'],
            'encoding': write_result['encoding'],
            'size': get_file_size(write_result['size']),
            'realtime_revision': 0
        })
    except Exception as e:
        print(f"在线保存文本文件失败: {str(e)}")
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

@app.route('/api/save_docx_file', methods=['POST'])
def save_docx_file():
    """保存在线编辑的 DOCX 文件。"""
    if not user_can_edit_files():
        return jsonify({'success': False, 'message': '只有管理员才能保存文件修改'}), 403

    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        blocks = data.get('blocks')
        mtime_ns = data.get('mtime_ns')

        if not file_path or not isinstance(blocks, list):
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_word_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        client_id = get_document_editor_client_id()
        changed_target_keys = set()
        for block in blocks:
            if not isinstance(block, dict):
                continue
            block_id = str(block.get('id') or '').strip()
            block_type = str(block.get('type') or '').strip()
            if not block_id or not block_type:
                continue
            if block_type == 'paragraph':
                incoming_text = normalize_docx_text(block.get('text'))
                original_text = normalize_docx_text(block.get('original_text'))
                if incoming_text != original_text:
                    changed_target_keys.add(build_docx_paragraph_target_key(block_id))
                continue

            if block_type == 'image_meta':
                for field_name in ('title', 'description'):
                    incoming_text = normalize_docx_text(block.get(field_name))
                    original_text = normalize_docx_text(block.get(f'original_{field_name}'))
                    if incoming_text != original_text:
                        changed_target_keys.add(build_docx_image_meta_target_key(block_id, field_name))
                continue

            if block_type != 'table':
                continue

            for row in block.get('rows') or []:
                if not isinstance(row, list):
                    continue
                for cell in row:
                    if not isinstance(cell, dict):
                        continue
                    incoming_text = normalize_docx_text(cell.get('text'))
                    original_text = normalize_docx_text(cell.get('original_text'))
                    if incoming_text == original_text:
                        continue
                    try:
                        row_index = int(cell.get('row'))
                        col_index = int(cell.get('col'))
                    except (TypeError, ValueError):
                        continue
                    changed_target_keys.add(build_docx_table_cell_target_key(block_id, row_index, col_index))

        locked_targets = find_document_locked_targets_for_changes(filepath, changed_target_keys, client_id)
        if locked_targets:
            lock_summary = '、'.join(
                f"{item.get('label') or '内容块'}（{item.get('username') or '未知用户'}）"
                for item in locked_targets[:5]
            )
            return jsonify({
                'success': False,
                'message': f'这些内容正被其他人编辑：{lock_summary}，请稍后再保存。',
                'mtime_ns': get_file_mtime_token(filepath),
                'locks': locked_targets
            }), 409

        save_summary = save_docx_file_content(filepath, blocks)
        if save_summary.get('conflicts'):
            conflict_summary = '、'.join(
                item.get('label', '内容块')
                for item in save_summary['conflicts'][:5]
            )
            return jsonify({
                'success': False,
                'message': f'这些内容已经被别人修改：{conflict_summary}。请先同步最新内容再保存。',
                'mtime_ns': get_file_mtime_token(filepath),
                'conflicts': save_summary['conflicts']
            }), 409

        new_mtime_ns = get_file_mtime_token(filepath)
        username = get_current_username()
        add_activity(username, '编辑', os.path.basename(file_path))
        print(f"已保存 DOCX 文件: {file_path} (用户: {username})")

        return jsonify({
            'success': True,
            'message': '保存成功',
            'mtime_ns': new_mtime_ns,
            'size': get_file_size(os.path.getsize(filepath)),
            **save_summary
        })
    except Exception as e:
        print(f"保存 DOCX 文件失败: {str(e)}")
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

@app.route('/api/save_excel_file', methods=['POST'])
def save_excel_file():
    """保存在线编辑的 Excel 文件。"""
    if not user_can_edit_files():
        return jsonify({'success': False, 'message': '只有管理员才能保存文件修改'}), 403

    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        sheets = data.get('sheets') or []
        mtime_ns = data.get('mtime_ns')

        if not file_path or not isinstance(sheets, list):
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        client_id = get_excel_editor_client_id()
        locked_cells = find_excel_locked_cells_for_changes(filepath, sheets, client_id)
        if locked_cells:
            lock_summary = '、'.join(
                f"{item.get('cell', '未知单元格')}（{item.get('username') or '未知用户'}）"
                for item in locked_cells[:5]
            )
            return jsonify({
                'success': False,
                'message': f'以下单元格正被其他人编辑：{lock_summary}，请稍后再试',
                'mtime_ns': get_file_mtime_token(filepath),
                'locked_cells': locked_cells
            }), 409

        result = (
            _apply_excel_cell_changes_via_com(filepath, sheets, mtime_ns)
            if should_prefer_excel_com_for_save(filepath)
            else _apply_excel_cell_changes_via_openpyxl(filepath, sheets, mtime_ns)
        )

        if not result['success']:
            conflict_cells = '、'.join(
                conflict.get('cell', '未知单元格')
                for conflict in result.get('conflicts', [])[:5]
            )
            return jsonify({
                'success': False,
                'message': f'文件已有新的修改，以下单元格存在冲突：{conflict_cells}，请刷新页面后处理。',
                'mtime_ns': result['mtime_ns'],
                'conflicts': result['conflicts']
            }), 409

        changed_count = result['changed_count']

        username = get_current_username()
        add_activity(username, '编辑', os.path.basename(file_path))
        new_token = result['mtime_ns']
        print(f"已保存 Excel 文件: {file_path} (用户: {username}, 修改单元格: {changed_count})")

        return jsonify({
            'success': True,
            'message': '已合并并保存' if result.get('merged') else '保存成功',
            'mtime_ns': new_token,
            'changed_cells': changed_count,
            'merged': result.get('merged', False)
        })
    except Exception as e:
        print(f"在线保存 Excel 失败: {str(e)}")
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'}), 500

@app.route('/api/excel_sheet_data', methods=['POST'])
def excel_sheet_data():
    """读取 Excel 工作表分块数据。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        sheet_name = (data.get('sheet_name') or '').strip()
        start_row = data.get('start_row', 1)
        start_col = data.get('start_col', 1)
        row_limit = data.get('row_limit', MAX_EXCEL_PREVIEW_ROWS)
        col_limit = data.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
        safe_row_limit = clamp_excel_position(row_limit, 1, MAX_EXCEL_PREVIEW_ROWS, MAX_EXCEL_PREVIEW_ROWS)
        safe_col_limit = clamp_excel_position(col_limit, 1, MAX_EXCEL_PREVIEW_COLS, MAX_EXCEL_PREVIEW_COLS)

        if not file_path or not sheet_name:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        chunk, backend = load_excel_sheet_chunk(filepath, sheet_name, start_row, start_col, safe_row_limit, safe_col_limit)
        return jsonify({
            'success': True,
            'sheet': chunk,
            'backend': backend,
            'row_limit': safe_row_limit,
            'col_limit': safe_col_limit,
            'mtime_ns': get_file_mtime_token(filepath)
        })
    except KeyError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except Exception as e:
        print(f"读取 Excel 分块失败: {str(e)}")
        return jsonify({'success': False, 'message': f'读取失败: {str(e)}'}), 500

@app.route('/api/excel_sheet_search', methods=['POST'])
def excel_sheet_search():
    """搜索 Excel 工作表内容。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        sheet_name = (data.get('sheet_name') or '').strip()
        keyword = (data.get('keyword') or '').strip()
        max_results = data.get('max_results', MAX_EXCEL_SEARCH_RESULTS)

        if not file_path or not sheet_name or not keyword:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        search_data, backend = search_excel_sheet(filepath, sheet_name, keyword, max_results)
        return jsonify({
            'success': True,
            'backend': backend,
            'sheet_name': search_data['sheet_name'],
            'results': search_data['results'],
            'truncated': search_data['truncated'],
            'row_count': search_data['row_count'],
            'col_count': search_data['col_count'],
            'keyword': keyword,
            'mtime_ns': get_file_mtime_token(filepath)
        })
    except KeyError as e:
        return jsonify({'success': False, 'message': str(e)}), 404
    except Exception as e:
        print(f"搜索 Excel 工作表失败: {str(e)}")
        return jsonify({'success': False, 'message': f'搜索失败: {str(e)}'}), 500

@app.route('/api/excel_file_status', methods=['POST'])
def excel_file_status():
    """获取 Excel 文件状态。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()

        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        return jsonify({
            'success': True,
            'mtime_ns': get_file_mtime_token(filepath),
            'file_size': os.path.getsize(filepath)
        })
    except Exception as e:
        print(f"获取 Excel 文件状态失败: {str(e)}")
        return jsonify({'success': False, 'message': f'获取状态失败: {str(e)}'}), 500

@app.route('/api/excel_collaboration_sync', methods=['POST'])
def excel_collaboration_sync():
    """同步 Excel 协作状态。"""
    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        sheet_name = (data.get('sheet_name') or '').strip()
        active_cell = data.get('active_cell')
        viewport = data.get('viewport')
        lock_cell = data.get('lock_cell')
        release_lock = bool(data.get('release_lock'))

        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        snapshot = sync_excel_collaboration_presence(
            filepath,
            get_current_username(),
            get_excel_editor_client_id(),
            request.remote_addr,
            sheet_name=sheet_name,
            active_cell=active_cell,
            viewport=viewport,
            lock_cell=lock_cell,
            release_lock=release_lock
        )
        snapshot['success'] = True
        return jsonify(snapshot)
    except Exception as e:
        print(f"同步 Excel 协作状态失败: {str(e)}")
        return jsonify({'success': False, 'message': f'同步失败: {str(e)}'}), 500

@app.route('/api/excel_collaboration_release', methods=['POST'])
def excel_collaboration_release():
    """释放 Excel 协作状态。"""
    try:
        data = request.get_json(silent=True)
        if not isinstance(data, dict):
            data = request.form.to_dict()

        file_path = (data.get('file_path') or '').strip()
        remove_session = str(data.get('remove_session', '')).lower() in {'1', 'true', 'yes', 'on'}
        if not file_path:
            return jsonify({'success': False, 'message': '缺少文件路径'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        release_excel_collaboration_presence(filepath, get_excel_editor_client_id(), remove_session=remove_session)
        return jsonify({'success': True})
    except Exception as e:
        print(f"释放 Excel 协作状态失败: {str(e)}")
        return jsonify({'success': False, 'message': f'释放失败: {str(e)}'}), 500

@app.route('/api/excel_structure_operation', methods=['POST'])
def excel_structure_operation():
    """执行 Excel 行列或工作表结构操作。"""
    if not user_can_edit_files():
        return jsonify({'success': False, 'message': '只有管理员才能修改表结构'}), 403

    try:
        data = request.get_json(silent=True) or {}
        file_path = (data.get('file_path') or '').strip()
        action = (data.get('action') or '').strip()
        sheet_name = (data.get('sheet_name') or '').strip()
        new_name = data.get('new_name')
        row = data.get('row')
        col = data.get('col')
        amount = data.get('amount', 1)
        mtime_ns = data.get('mtime_ns')
        start_row = data.get('start_row', 1)
        start_col = data.get('start_col', 1)
        row_limit = data.get('row_limit', MAX_EXCEL_PREVIEW_ROWS)
        col_limit = data.get('col_limit', MAX_EXCEL_PREVIEW_COLS)
        include_sheet_snapshot = str(data.get('include_sheet_snapshot', 'true')).lower() in {'1', 'true', 'yes', 'on'}

        if not file_path or not action:
            return jsonify({'success': False, 'message': '缺少必要参数'}), 400

        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], file_path)
        if not success:
            return jsonify({'success': False, 'message': error}), 400

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            return jsonify({'success': False, 'message': '文件不存在'}), 404

        if not is_excel_editable_file(file_path):
            return jsonify({'success': False, 'message': '该文件类型暂不支持在线编辑'}), 400

        client_id = get_excel_editor_client_id()
        other_locks = get_excel_collaboration_other_locks(filepath, client_id)
        if other_locks:
            lock_summary = '、'.join(
                f"{item.get('cell', '未知单元格')}（{item.get('username') or '未知用户'}）"
                for item in other_locks[:5]
            )
            return jsonify({
                'success': False,
                'message': f'其他人正在编辑这些单元格：{lock_summary}，请稍后再调整表结构',
                'mtime_ns': get_file_mtime_token(filepath),
                'locks': other_locks
            }), 409

        current_token = get_file_mtime_token(filepath)
        if str(mtime_ns or '') != current_token:
            return jsonify({
                'success': False,
                'message': '文件已被其他人修改，请刷新页面后再操作',
                'mtime_ns': current_token
            }), 409

        operation_result = apply_excel_structure_operation(filepath, action, {
            'sheet_name': sheet_name,
            'new_name': new_name,
            'row': row,
            'col': col,
            'amount': amount,
            'refresh_sheet_name': sheet_name,
            'start_row': start_row,
            'start_col': start_col,
            'row_limit': row_limit,
            'col_limit': col_limit,
            'include_sheet_snapshot': include_sheet_snapshot
        })

        username = get_current_username()
        add_activity(username, '编辑', os.path.basename(file_path))
        new_token = get_file_mtime_token(filepath)
        print(f"已修改 Excel 结构: {file_path} (用户: {username}, 操作: {action})")

        return jsonify({
            'success': True,
            'message': operation_result['message'],
            'active_sheet': operation_result['active_sheet'],
            'mtime_ns': new_token,
            'sheet': operation_result.get('sheet'),
            'sheet_meta': operation_result.get('sheet_meta')
        })
    except (ValueError, KeyError) as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except Exception as e:
        print(f"修改 Excel 结构失败: {str(e)}")
        return jsonify({'success': False, 'message': f'修改失败: {str(e)}'}), 500

@app.route('/edit_excel/<path:filename>')
def edit_excel_file(filename):
    """在线编辑 Excel 文件。"""
    try:
        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], filename)
        if not success:
            flash(f'路径错误: {error}', 'danger')
            return redirect(url_for('index'))

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            flash('文件不存在', 'danger')
            return redirect(url_for('index'))

        if not is_excel_editable_file(filename):
            return redirect(url_for('preview_file', filename=filename))

        requested_sheet = (request.args.get('sheet') or '').strip()
        preview = load_excel_file_preview(filepath, requested_sheet or None)
        can_edit = user_can_edit_files()
        parent_path = os.path.dirname(filename).replace('\\', '/')
        back_url = url_for('index', subpath=parent_path) if parent_path else url_for('index')
        active_sheet_name = requested_sheet if any(sheet['name'] == requested_sheet for sheet in preview['sheets']) else (preview['sheets'][0]['name'] if preview['sheets'] else '')

        warning_parts = []
        if preview.get('large_file'):
            warning_parts.append(
                f'当前文件较大（{get_file_size(preview["file_size"])}），为保证稳定性，页面会按块加载；'
                '你可以用“上一页/下一页”或输入起始行列继续查看整张表，保存时会写回你修改过的单元格。'
            )
        if preview['truncated']:
            warning_parts.append(f'工作表过多，当前仅显示前 {MAX_EXCEL_PREVIEW_SHEETS} 个工作表。')
        truncated_sheets = [sheet['name'] for sheet in preview['sheets'] if sheet['truncated']]
        if truncated_sheets:
            warning_parts.append(
                f'以下工作表首次加载前 {preview["row_limit"]} 行和 {preview["col_limit"]} 列，可继续翻页或指定起始行列查看完整内容：'
                + '、'.join(truncated_sheets)
            )
        if preview['editable'] and not can_edit:
            warning_parts.append('当前昵称可以预览此文件，但只有管理员才能保存修改。')
        warning_message = ' '.join(warning_parts) if warning_parts else None

        initial_status = '可直接编辑单元格并保存到共享目录，快捷键：Ctrl+S。'
        if not can_edit:
            initial_status = '当前为只读模式，只有管理员才能保存修改。'
        elif preview.get('large_file'):
            initial_status = (
                f'当前为大文件模式：每次加载约 {preview["row_limit"]} 行和 {preview["col_limit"]} 列，'
                '可以继续翻页并直接编辑保存。'
            )

        return render_template_string('''
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>表格编辑: {{ filename }}</title>
            <style>
                :root { color-scheme: dark; }
                body {
                    margin: 0;
                    min-height: 100vh;
                    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
                    background:
                        radial-gradient(circle at top left, rgba(34, 197, 94, 0.16), transparent 30%),
                        radial-gradient(circle at top right, rgba(59, 130, 246, 0.14), transparent 35%),
                        #0f172a;
                    color: #e2e8f0;
                }
                .page {
                    max-width: 1500px;
                    margin: 0 auto;
                    padding: 24px;
                }
                .header, .card {
                    background: rgba(15, 23, 42, 0.82);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    backdrop-filter: blur(14px);
                    border-radius: 20px;
                    box-shadow: 0 20px 45px rgba(2, 6, 23, 0.28);
                }
                .header {
                    padding: 22px 24px;
                    margin-bottom: 18px;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    gap: 16px;
                    flex-wrap: wrap;
                }
                .title h1 {
                    margin: 0 0 8px;
                    font-size: 28px;
                    color: #f8fafc;
                }
                .title p {
                    margin: 0;
                    color: #94a3b8;
                    word-break: break-all;
                }
                .actions {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                }
                .btn {
                    border: none;
                    border-radius: 12px;
                    padding: 12px 18px;
                    font-size: 14px;
                    font-weight: 600;
                    cursor: pointer;
                    text-decoration: none;
                    transition: transform 0.2s ease, opacity 0.2s ease;
                }
                .btn:hover { transform: translateY(-1px); }
                .btn:disabled { cursor: not-allowed; opacity: 0.55; transform: none; }
                .btn-primary { background: linear-gradient(135deg, #16a34a, #2563eb); color: white; }
                .btn-secondary {
                    background: rgba(30, 41, 59, 0.95);
                    color: #e2e8f0;
                    border: 1px solid rgba(148, 163, 184, 0.18);
                }
                .card { padding: 20px; }
                .meta {
                    display: flex;
                    gap: 16px;
                    flex-wrap: wrap;
                    margin-bottom: 14px;
                    font-size: 13px;
                    color: #94a3b8;
                }
                .notice {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(245, 158, 11, 0.16);
                    border: 1px solid rgba(245, 158, 11, 0.28);
                    color: #fde68a;
                    line-height: 1.6;
                }
                .guide {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(14, 165, 233, 0.12);
                    border: 1px solid rgba(14, 165, 233, 0.24);
                    color: #bae6fd;
                    line-height: 1.6;
                }
                .collab-panel {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    justify-content: space-between;
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(34, 197, 94, 0.1);
                    border: 1px solid rgba(34, 197, 94, 0.22);
                    color: #dcfce7;
                }
                .collab-users {
                    display: flex;
                    gap: 8px;
                    flex-wrap: wrap;
                    align-items: center;
                }
                .collab-user {
                    display: inline-flex;
                    align-items: center;
                    gap: 6px;
                    padding: 6px 10px;
                    border-radius: 999px;
                    background: rgba(15, 23, 42, 0.72);
                    border: 1px solid rgba(34, 197, 94, 0.18);
                    font-size: 12px;
                    color: #f0fdf4;
                }
                .collab-user.self {
                    border-color: rgba(59, 130, 246, 0.4);
                    background: rgba(30, 41, 59, 0.95);
                }
                .collab-hint {
                    font-size: 12px;
                    color: #bbf7d0;
                }
                .sheet-toolbar {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    margin-bottom: 14px;
                }
                .structure-toolbar {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    margin-bottom: 14px;
                    padding: 12px;
                    border-radius: 14px;
                    background: rgba(15, 23, 42, 0.68);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                }
                .toolbar-grow {
                    flex: 1 1 320px;
                }
                .toolbar-input {
                    min-width: 110px;
                    padding: 10px 12px;
                    border-radius: 10px;
                    border: 1px solid rgba(148, 163, 184, 0.2);
                    background: rgba(15, 23, 42, 0.9);
                    color: #e2e8f0;
                    outline: none;
                    box-sizing: border-box;
                }
                .toolbar-wide {
                    width: 100%;
                }
                .cell-badge {
                    display: inline-flex;
                    align-items: center;
                    min-height: 40px;
                    padding: 0 12px;
                    border-radius: 10px;
                    background: rgba(30, 41, 59, 0.9);
                    color: #cbd5e1;
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    font-size: 13px;
                    white-space: nowrap;
                }
                .search-results {
                    margin-bottom: 14px;
                    padding: 12px;
                    border-radius: 14px;
                    background: rgba(15, 23, 42, 0.68);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                }
                .search-results[hidden] {
                    display: none;
                }
                .search-results-title {
                    margin-bottom: 10px;
                    color: #94a3b8;
                    font-size: 13px;
                }
                .search-result-list {
                    display: flex;
                    flex-wrap: wrap;
                    gap: 8px;
                }
                .search-result-item {
                    display: inline-flex;
                    align-items: center;
                    gap: 8px;
                    padding: 8px 12px;
                    border-radius: 10px;
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    background: rgba(30, 41, 59, 0.9);
                    color: #e2e8f0;
                    cursor: pointer;
                }
                .search-result-item:hover {
                    background: rgba(51, 65, 85, 0.95);
                }
                .search-result-cell {
                    color: #7dd3fc;
                    font-weight: 700;
                }
                .search-empty {
                    color: #94a3b8;
                    font-size: 13px;
                }
                .tabs {
                    display: flex;
                    gap: 10px;
                    flex-wrap: wrap;
                    margin-bottom: 14px;
                }
                .tab-btn {
                    border: 1px solid rgba(148, 163, 184, 0.2);
                    background: rgba(30, 41, 59, 0.85);
                    color: #cbd5e1;
                    padding: 10px 14px;
                    border-radius: 12px;
                    cursor: pointer;
                    font-weight: 600;
                }
                .tab-btn.active {
                    background: linear-gradient(135deg, rgba(34, 197, 94, 0.9), rgba(37, 99, 235, 0.9));
                    color: white;
                }
                .sheet-panel { display: none; }
                .sheet-panel.active { display: block; }
                .sheet-meta {
                    color: #94a3b8;
                    font-size: 13px;
                    margin-bottom: 10px;
                }
                .sheet-controls {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    margin-bottom: 12px;
                }
                .control-group {
                    display: flex;
                    gap: 8px;
                    flex-wrap: wrap;
                    align-items: center;
                }
                .btn-small {
                    padding: 8px 12px;
                    font-size: 12px;
                    border-radius: 10px;
                }
                .range-label {
                    display: flex;
                    align-items: center;
                    gap: 6px;
                    color: #cbd5e1;
                    font-size: 12px;
                }
                .range-input {
                    width: 88px;
                    padding: 8px 10px;
                    border-radius: 10px;
                    border: 1px solid rgba(148, 163, 184, 0.2);
                    background: rgba(15, 23, 42, 0.9);
                    color: #e2e8f0;
                    outline: none;
                }
                .table-wrap {
                    overflow: auto;
                    border-radius: 16px;
                    border: 1px solid rgba(51, 65, 85, 0.95);
                    background: #020617;
                    max-height: calc(100vh - 290px);
                }
                table {
                    border-collapse: collapse;
                    min-width: 100%;
                }
                th, td {
                    border: 1px solid rgba(51, 65, 85, 0.95);
                    padding: 0;
                }
                th {
                    position: sticky;
                    top: 0;
                    z-index: 2;
                    background: #0f172a;
                    color: #cbd5e1;
                    font-size: 12px;
                    font-weight: 700;
                    text-align: center;
                    min-width: 110px;
                    height: 38px;
                }
                th.row-head {
                    left: 0;
                    z-index: 3;
                    min-width: 54px;
                }
                td.row-head {
                    position: sticky;
                    left: 0;
                    z-index: 1;
                    background: #0f172a;
                    color: #cbd5e1;
                    text-align: center;
                    font-size: 12px;
                    min-width: 54px;
                }
                .cell-input {
                    width: 100%;
                    min-width: 110px;
                    padding: 10px 12px;
                    box-sizing: border-box;
                    background: transparent;
                    border: none;
                    color: #e2e8f0;
                    outline: none;
                    font-size: 13px;
                }
                .cell-input[readonly] {
                    cursor: default;
                    color: #94a3b8;
                }
                .cell-input.changed {
                    background: rgba(34, 197, 94, 0.16);
                }
                .cell-input.active {
                    background: rgba(14, 165, 233, 0.18);
                    box-shadow: inset 0 0 0 1px rgba(56, 189, 248, 0.9);
                }
                .cell-input.locked {
                    background: rgba(248, 113, 113, 0.12);
                    box-shadow: inset 0 0 0 1px rgba(248, 113, 113, 0.4);
                    color: #fecaca;
                }
                .status {
                    min-height: 22px;
                    margin-top: 12px;
                    color: #cbd5e1;
                    font-size: 14px;
                }
                .status.success { color: #86efac; }
                .status.error { color: #fca5a5; }
                .status.info { color: #cbd5e1; }
                @media (max-width: 768px) {
                    .page { padding: 14px; }
                    .header, .card { padding: 16px; border-radius: 16px; }
                    .title h1 { font-size: 22px; }
                    .table-wrap { max-height: calc(100vh - 250px); }
                }
            </style>
        </head>
        <body>
            <div class="page">
                <div class="header">
                    <div class="title">
                        <h1>{{ filename }}</h1>
                        <p>{{ file_path }}</p>
                    </div>
                    <div class="actions">
                        <a href="{{ back_url }}" id="backToListBtn" class="btn btn-secondary">返回列表</a>
                        <a href="{{ url_for('download_file', filename=file_path) }}" class="btn btn-secondary">下载原文件</a>
                        <button id="saveBtn" class="btn btn-primary" {% if not editable or not can_edit %}disabled{% endif %}>保存修改</button>
                    </div>
                </div>
                <div class="card">
                    <div class="meta">
                        <span>大小: {{ file_size }}</span>
                        <span>工作表: {{ sheet_count }}</span>
                        <span>快捷键: Ctrl+S</span>
                    </div>
                    {% if warning_message %}
                    <div class="notice">{{ warning_message }}</div>
                    {% endif %}
                    <div class="guide">整张表支持继续查看和编辑：当前每次加载约 {{ excel_row_limit }} 行和 {{ excel_col_limit }} 列，可用“上一页 / 下一页”切换当前区域，或直接输入“起始行 / 起始列”后点击“加载范围”。</div>
                    <div class="collab-panel">
                        <div class="collab-users" id="collabUsers">协同连接中...</div>
                        <div class="collab-hint" id="collabHint">正在同步在线成员和单元格锁。</div>
                    </div>
                    <div class="sheet-toolbar">
                        <div class="control-group">
                            <span class="cell-badge" id="activeCellLabel">当前单元格 -</span>
                            <input id="cellRefInput" class="toolbar-input" type="text" placeholder="输入 A1 定位">
                            <button id="jumpCellBtn" class="btn btn-secondary btn-small" type="button">定位单元格</button>
                            <button id="autoSaveBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>自动保存: 开</button>
                        </div>
                        <div class="control-group toolbar-grow">
                            <input id="formulaBar" class="toolbar-input toolbar-wide" type="text" placeholder="当前单元格内容 / 公式（例如 =SUM(A1:A10)）" {% if not editable or not can_edit %}readonly{% endif %}>
                        </div>
                        <div class="control-group">
                            <input id="searchInput" class="toolbar-input" type="text" placeholder="搜索当前工作表">
                            <button id="searchBtn" class="btn btn-secondary btn-small" type="button">搜索</button>
                        </div>
                    </div>
                    <div class="structure-toolbar">
                        <div class="control-group">
                            <span class="cell-badge">行操作</span>
                            <input id="rowOpIndexInput" class="toolbar-input" type="number" min="1" value="1" placeholder="行号">
                            <input id="rowOpAmountInput" class="toolbar-input" type="number" min="1" value="1" placeholder="数量">
                            <button id="insertRowsBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>插入行</button>
                            <button id="deleteRowsBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>删除行</button>
                        </div>
                        <div class="control-group">
                            <span class="cell-badge">列操作</span>
                            <input id="colOpIndexInput" class="toolbar-input" type="text" value="A" placeholder="列号或字母">
                            <input id="colOpAmountInput" class="toolbar-input" type="number" min="1" value="1" placeholder="数量">
                            <button id="insertColsBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>插入列</button>
                            <button id="deleteColsBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>删除列</button>
                        </div>
                        <div class="control-group toolbar-grow">
                            <span class="cell-badge">工作表</span>
                            <input id="sheetNameInput" class="toolbar-input toolbar-wide" type="text" placeholder="输入工作表名称">
                            <button id="addSheetBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>新增工作表</button>
                            <button id="renameSheetBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>重命名当前表</button>
                            <button id="deleteSheetBtn" class="btn btn-secondary btn-small" type="button" {% if not editable or not can_edit %}disabled{% endif %}>删除当前表</button>
                        </div>
                    </div>
                    <div class="search-results" id="searchResults" hidden></div>
                    <div class="tabs">
                        {% for sheet in sheets %}
                        <button class="tab-btn {% if loop.first %}active{% endif %}" type="button" data-sheet-target="{{ sheet.name }}">{{ sheet.name }}</button>
                        {% endfor %}
                    </div>
                    {% for sheet in sheets %}
                    <div class="sheet-panel {% if loop.first %}active{% endif %}"
                         data-sheet-panel="{{ sheet.name }}"
                         data-row-count="{{ sheet.row_count }}"
                         data-col-count="{{ sheet.col_count }}"
                         data-start-row="{{ sheet.start_row }}"
                         data-start-col="{{ sheet.start_col }}"
                         data-end-row="{{ sheet.end_row }}"
                         data-end-col="{{ sheet.end_col }}">
                        <div class="sheet-controls">
                            <div class="control-group">
                                <button class="btn btn-secondary btn-small nav-rows-prev" type="button" data-sheet="{{ sheet.name }}">上一页行</button>
                                <button class="btn btn-secondary btn-small nav-rows-next" type="button" data-sheet="{{ sheet.name }}">下一页行</button>
                                <button class="btn btn-secondary btn-small nav-cols-prev" type="button" data-sheet="{{ sheet.name }}">上一页列</button>
                                <button class="btn btn-secondary btn-small nav-cols-next" type="button" data-sheet="{{ sheet.name }}">下一页列</button>
                            </div>
                            <div class="control-group">
                                <label class="range-label">起始行
                                    <input class="range-input range-row" type="number" min="1" max="{{ sheet.row_count }}" value="{{ sheet.start_row }}">
                                </label>
                                <label class="range-label">起始列
                                    <input class="range-input range-col" type="number" min="1" max="{{ sheet.col_count }}" value="{{ sheet.start_col }}">
                                </label>
                                <button class="btn btn-secondary btn-small load-range-btn" type="button" data-sheet="{{ sheet.name }}">加载范围</button>
                            </div>
                        </div>
                        <div class="sheet-meta" data-sheet-meta="{{ sheet.name }}">显示第 {{ sheet.start_row }}-{{ sheet.end_row }} 行 / 第 {{ sheet.start_col }}-{{ sheet.end_col }} 列，共 {{ sheet.row_count }} 行 / {{ sheet.col_count }} 列{% if sheet.truncated %}，当前工作表还有更多内容可继续加载{% endif %}</div>
                        <div class="table-wrap" data-table-wrap="{{ sheet.name }}">
                            <table>
                                <thead>
                                    <tr>
                                        <th class="row-head">#</th>
                                        {% for col_name in sheet.columns %}
                                        <th>{{ col_name }}</th>
                                        {% endfor %}
                                    </tr>
                                </thead>
                                <tbody>
                                    {% for row in sheet.rows %}
                                    {% set row_index = sheet.start_row + loop.index - 1 %}
                                    <tr>
                                        <td class="row-head">{{ row_index }}</td>
                                        {% for value in row %}
                                        <td>
                                            <input
                                                class="cell-input"
                                                type="text"
                                                value="{{ value }}"
                                                spellcheck="false"
                                                autocomplete="off"
                                                autocapitalize="off"
                                                data-original="{{ value }}"
                                                data-sheet="{{ sheet.name }}"
                                                data-row="{{ row_index }}"
                                                data-col="{{ sheet.start_col + loop.index - 1 }}"
                                                {% if not editable or not can_edit %}readonly{% endif %}
                                            >
                                        </td>
                                        {% endfor %}
                                    </tr>
                                    {% endfor %}
                                </tbody>
                            </table>
                        </div>
                    </div>
                    {% endfor %}
                    <div class="status info" id="status">{{ initial_status }}</div>
                </div>
            </div>
            <script>
                const state = {
                    filePath: {{ file_path|tojson }},
                    mtimeNs: {{ mtime_ns|tojson }},
                    editable: {{ 'true' if editable else 'false' }},
                    canEdit: {{ 'true' if can_edit else 'false' }},
                    sheetNames: {{ sheet_names|tojson }},
                    rowLimit: {{ excel_row_limit }},
                    colLimit: {{ excel_col_limit }},
                    backend: {{ backend|tojson }},
                    activeSheet: {{ active_sheet|tojson }},
                    activeCell: null,
                    autoSaveEnabled: {{ 'true' if editable and can_edit and not excel_large_file else 'false' }},
                    autoSaveDelay: {{ 1800 if excel_large_file else 1200 }},
                    pendingFocus: null,
                    isBulkEditing: false,
                    collabClientId: {{ collab_client_id|tojson }},
                    collabUsername: {{ collab_username|tojson }},
                    collabPollInterval: {{ collab_poll_interval_ms }},
                    collabEditors: [],
                    collabLockMap: new Map(),
                    collabLockTarget: null,
                    collabOwnLockKey: '',
                    collabSyncInFlight: false
                };

                const saveBtn = document.getElementById('saveBtn');
                const backToListBtn = document.getElementById('backToListBtn');
                const statusEl = document.getElementById('status');
                const formulaBar = document.getElementById('formulaBar');
                const searchInput = document.getElementById('searchInput');
                const searchBtn = document.getElementById('searchBtn');
                const searchResultsEl = document.getElementById('searchResults');
                const cellRefInput = document.getElementById('cellRefInput');
                const jumpCellBtn = document.getElementById('jumpCellBtn');
                const activeCellLabel = document.getElementById('activeCellLabel');
                const autoSaveBtn = document.getElementById('autoSaveBtn');
                const rowOpIndexInput = document.getElementById('rowOpIndexInput');
                const rowOpAmountInput = document.getElementById('rowOpAmountInput');
                const colOpIndexInput = document.getElementById('colOpIndexInput');
                const colOpAmountInput = document.getElementById('colOpAmountInput');
                const sheetNameInput = document.getElementById('sheetNameInput');
                const insertRowsBtn = document.getElementById('insertRowsBtn');
                const deleteRowsBtn = document.getElementById('deleteRowsBtn');
                const insertColsBtn = document.getElementById('insertColsBtn');
                const deleteColsBtn = document.getElementById('deleteColsBtn');
                const addSheetBtn = document.getElementById('addSheetBtn');
                const renameSheetBtn = document.getElementById('renameSheetBtn');
                const deleteSheetBtn = document.getElementById('deleteSheetBtn');
                const collabUsersEl = document.getElementById('collabUsers');
                const collabHintEl = document.getElementById('collabHint');
                const tabButtons = Array.from(document.querySelectorAll('.tab-btn'));
                const sheetPanels = Array.from(document.querySelectorAll('.sheet-panel'));
                const pendingChanges = new Map();
                let isSaving = false;
                let autoSaveTimer = null;
                let autoSaveRetryPending = false;
                let remoteCheckTimer = null;
                let isSyncingFormulaBar = false;
                let suppressBeforeUnload = false;
                const newlineChar = String.fromCharCode(10);
                const carriageReturnChar = String.fromCharCode(13);
                const tabChar = String.fromCharCode(9);

                function setStatus(message, type = 'info') {
                    statusEl.textContent = message;
                    statusEl.className = `status ${type}`;
                }

                function escapeHtml(value) {
                    return String(value ?? '')
                        .replace(/&/g, '&amp;')
                        .replace(/</g, '&lt;')
                        .replace(/>/g, '&gt;')
                        .replace(/"/g, '&quot;')
                        .replace(/'/g, '&#39;');
                }

                function getSheetPanel(sheetName) {
                    return sheetPanels.find(panel => panel.dataset.sheetPanel === sheetName) || null;
                }

                function getActiveSheetName() {
                    return state.activeSheet || state.sheetNames[0] || '';
                }

                function getFirstVisibleInput(sheetName) {
                    const panel = getSheetPanel(sheetName);
                    return panel ? panel.querySelector('.cell-input') : null;
                }

                function activateSheet(sheetName) {
                    state.activeSheet = sheetName;
                    tabButtons.forEach(btn => btn.classList.toggle('active', btn.dataset.sheetTarget === sheetName));
                    sheetPanels.forEach(panel => panel.classList.toggle('active', panel.dataset.sheetPanel === sheetName));
                    if (sheetNameInput) {
                        sheetNameInput.placeholder = `当前工作表：${sheetName}`;
                    }

                    if (state.activeCell && state.activeCell.sheet === sheetName) {
                        setActiveCell(sheetName, state.activeCell.row, state.activeCell.col);
                    } else {
                        const firstInput = getFirstVisibleInput(sheetName);
                        if (firstInput) {
                            setActiveCell(sheetName, Number(firstInput.dataset.row), Number(firstInput.dataset.col));
                        } else {
                            updateActiveCellDisplay();
                        }
                    }

                    syncCollaborationState({ quiet: true });
                }

                function buildCellKey(sheetName, row, col) {
                    return `${sheetName}::${row}::${col}`;
                }

                function getCellLockInfo(sheetName, row, col) {
                    return state.collabLockMap.get(buildCellKey(sheetName, row, col)) || null;
                }

                function isCellLockedByOther(sheetName, row, col) {
                    const lockInfo = getCellLockInfo(sheetName, row, col);
                    return Boolean(lockInfo && lockInfo.client_id !== state.collabClientId);
                }

                function setCollabHint(message) {
                    if (collabHintEl) {
                        collabHintEl.textContent = message;
                    }
                }

                function renderCollaborationUsers() {
                    if (!collabUsersEl) {
                        return;
                    }

                    const editors = Array.isArray(state.collabEditors) ? state.collabEditors : [];
                    if (editors.length === 0) {
                        collabUsersEl.innerHTML = '<span class="collab-user self">只有你</span>';
                        setCollabHint('当前没有其他在线协同编辑者。');
                        return;
                    }

                    collabUsersEl.innerHTML = editors.map(editor => {
                        const activeLabel = editor.active_sheet
                            ? `${escapeHtml(editor.active_sheet)}${editor.active_cell_ref ? ` / ${escapeHtml(editor.active_cell_ref)}` : ''}`
                            : '正在查看';
                        return `
                            <span class="collab-user${editor.is_self ? ' self' : ''}">
                                <strong>${escapeHtml(editor.username || '匿名用户')}</strong>
                                <span>${activeLabel}</span>
                            </span>
                        `;
                    }).join('');

                    const ownLock = state.collabOwnLockKey ? state.collabLockMap.get(state.collabOwnLockKey) : null;
                    if (ownLock) {
                        setCollabHint(`你当前锁定了 ${ownLock.sheet} / ${ownLock.cell}，其他人会看到该单元格为只读。`);
                    } else if (editors.length > 1) {
                        setCollabHint('其他人进入或保存后会自动同步到当前页面，已锁定的单元格会直接显示出来。');
                    } else {
                        setCollabHint('当前没有其他在线协同编辑者。');
                    }
                }

                function applyCollaborationLocks() {
                    const forceReadonly = !state.editable || !state.canEdit;
                    document.querySelectorAll('.cell-input').forEach(input => {
                        const lockInfo = getCellLockInfo(input.dataset.sheet, Number(input.dataset.row), Number(input.dataset.col));
                        const lockedByOther = Boolean(lockInfo && lockInfo.client_id !== state.collabClientId);
                        input.readOnly = forceReadonly || lockedByOther;
                        input.classList.toggle('locked', lockedByOther);
                        if (lockedByOther) {
                            input.title = `${lockInfo.username} 正在编辑 ${lockInfo.cell}`;
                        } else if (!forceReadonly) {
                            input.removeAttribute('title');
                        }
                    });

                    if (formulaBar) {
                        const activeLock = state.activeCell
                            ? getCellLockInfo(state.activeCell.sheet, state.activeCell.row, state.activeCell.col)
                            : null;
                        formulaBar.readOnly = forceReadonly || Boolean(activeLock && activeLock.client_id !== state.collabClientId);
                    }
                }

                function getCurrentViewport(sheetName = getActiveSheetName()) {
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return null;
                    }
                    return {
                        start_row: Number(panel.dataset.startRow || 1),
                        start_col: Number(panel.dataset.startCol || 1),
                        end_row: Number(panel.dataset.endRow || panel.dataset.startRow || 1),
                        end_col: Number(panel.dataset.endCol || panel.dataset.startCol || 1)
                    };
                }

                function updateCollabLockTarget(lockTarget) {
                    const currentKey = state.collabLockTarget
                        ? buildCellKey(state.collabLockTarget.sheet, state.collabLockTarget.row, state.collabLockTarget.col)
                        : '';
                    const nextKey = lockTarget
                        ? buildCellKey(lockTarget.sheet, lockTarget.row, lockTarget.col)
                        : '';

                    if (currentKey === nextKey) {
                        return;
                    }

                    state.collabLockTarget = lockTarget
                        ? { sheet: lockTarget.sheet, row: Number(lockTarget.row), col: Number(lockTarget.col) }
                        : null;
                    syncCollaborationState({ quiet: true });
                }

                async function syncCollaborationState(options = {}) {
                    const { quiet = true } = options;
                    if (!state.filePath || state.collabSyncInFlight) {
                        return;
                    }

                    const activeSheetName = getActiveSheetName();
                    state.collabSyncInFlight = true;

                    try {
                        const response = await fetch('/api/excel_collaboration_sync', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                sheet_name: activeSheetName,
                                active_cell: state.activeCell ? {
                                    sheet: state.activeCell.sheet,
                                    row: state.activeCell.row,
                                    col: state.activeCell.col
                                } : null,
                                viewport: getCurrentViewport(activeSheetName),
                                lock_cell: state.collabLockTarget ? {
                                    sheet: state.collabLockTarget.sheet,
                                    row: state.collabLockTarget.row,
                                    col: state.collabLockTarget.col
                                } : null,
                                release_lock: !state.collabLockTarget
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (!quiet) {
                                setStatus(data.message || '协同状态同步失败，请稍后重试。', 'error');
                            }
                            return;
                        }

                        if (data.client_id) {
                            state.collabClientId = data.client_id;
                        }

                        state.collabEditors = Array.isArray(data.editors) ? data.editors : [];
                        state.collabLockMap = new Map();
                        (data.locks || []).forEach(lock => {
                            state.collabLockMap.set(buildCellKey(lock.sheet, lock.row, lock.col), lock);
                        });
                        state.collabOwnLockKey = data.own_lock
                            ? buildCellKey(data.own_lock.sheet, data.own_lock.row, data.own_lock.col)
                            : '';

                        renderCollaborationUsers();
                        applyCollaborationLocks();

                        if (data.lock_denied) {
                            state.collabLockTarget = null;
                            applyCollaborationLocks();
                            setStatus(`${data.lock_denied.username} 正在编辑 ${data.lock_denied.sheet} / ${data.lock_denied.cell}，当前单元格暂时不可修改。`, 'error');
                        }

                        if (data.mtime_ns && data.mtime_ns !== state.mtimeNs) {
                            const activePanel = getSheetPanel(activeSheetName);
                            if (pendingChanges.size > 0) {
                                state.mtimeNs = data.mtime_ns;
                                setStatus('检测到其他人已保存新内容，你当前还有未保存修改，请先保存或刷新。', 'error');
                            } else if (activePanel) {
                                state.mtimeNs = data.mtime_ns;
                                await loadSheetRange(
                                    activeSheetName,
                                    Number(activePanel.dataset.startRow || 1),
                                    Number(activePanel.dataset.startCol || 1),
                                    { quiet: true }
                                );
                                setStatus('检测到其他人已保存，当前区域已自动刷新。', 'info');
                            }
                        }
                    } catch (error) {
                        if (!quiet) {
                            setStatus('协同状态同步失败：网络异常或服务器不可用。', 'error');
                        }
                    } finally {
                        state.collabSyncInFlight = false;
                    }
                }

                function releaseCollaborationState(removeSession = false) {
                    if (!state.filePath) {
                        return;
                    }

                    const payload = JSON.stringify({
                        file_path: state.filePath,
                        remove_session: removeSession
                    });

                    if (navigator.sendBeacon) {
                        const blob = new Blob([payload], { type: 'application/json' });
                        navigator.sendBeacon('/api/excel_collaboration_release', blob);
                        return;
                    }

                    fetch('/api/excel_collaboration_release', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: payload,
                        keepalive: true
                    }).catch(() => {});
                }

                function returnToList(event) {
                    if (!backToListBtn) {
                        return;
                    }

                    const backUrl = backToListBtn.href;
                    if (!window.opener || window.opener.closed) {
                        return;
                    }

                    event.preventDefault();
                    suppressBeforeUnload = true;
                    try {
                        window.opener.location.href = backUrl;
                        if (typeof window.opener.focus === 'function') {
                            window.opener.focus();
                        }
                    } catch (error) {
                    }

                    releaseCollaborationState(true);
                    window.close();
                    window.setTimeout(function() {
                        window.location.href = backUrl;
                    }, 120);
                }

                function columnNumberToName(colNumber) {
                    let value = Number(colNumber);
                    if (!Number.isFinite(value) || value < 1) {
                        return '';
                    }

                    let name = '';
                    while (value > 0) {
                        const remainder = (value - 1) % 26;
                        name = String.fromCharCode(65 + remainder) + name;
                        value = Math.floor((value - 1) / 26);
                    }
                    return name;
                }

                function columnNameToNumber(columnName) {
                    const text = String(columnName || '').trim().toUpperCase();
                    if (!/^[A-Z]+$/.test(text)) {
                        return null;
                    }

                    let value = 0;
                    for (const char of text) {
                        value = value * 26 + (char.charCodeAt(0) - 64);
                    }
                    return value;
                }

                function toA1Ref(row, col) {
                    return `${columnNumberToName(col)}${row}`;
                }

                function parseCellReference(value) {
                    const text = String(value || '').trim().toUpperCase();
                    if (!text) {
                        return null;
                    }

                    const a1Match = text.match(/^([A-Z]+)(\d+)$/);
                    if (a1Match) {
                        return {
                            row: Number(a1Match[2]),
                            col: columnNameToNumber(a1Match[1])
                        };
                    }

                    const rcMatch = text.match(/^(\d+)\s*[,，]\s*(\d+)$/);
                    if (rcMatch) {
                        return {
                            row: Number(rcMatch[1]),
                            col: Number(rcMatch[2])
                        };
                    }

                    return null;
                }

                function findCellInput(sheetName, row, col) {
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return null;
                    }

                    const targetRow = Number(row);
                    const targetCol = Number(col);
                    return Array.from(panel.querySelectorAll('.cell-input')).find(input => (
                        Number(input.dataset.row) === targetRow && Number(input.dataset.col) === targetCol
                    )) || null;
                }

                function getCellCurrentValue(sheetName, row, col) {
                    const input = findCellInput(sheetName, row, col);
                    if (input) {
                        return input.value ?? '';
                    }

                    const key = buildCellKey(sheetName, row, col);
                    return pendingChanges.has(key) ? (pendingChanges.get(key).value ?? '') : '';
                }

                function updateAutoSaveButton() {
                    if (!autoSaveBtn) {
                        return;
                    }
                    autoSaveBtn.textContent = `自动保存: ${state.autoSaveEnabled ? '开' : '关'}`;
                }

                function updateActiveCellDisplay() {
                    if (!state.activeCell) {
                        activeCellLabel.textContent = '当前单元格 -';
                        if (document.activeElement !== cellRefInput) {
                            cellRefInput.value = '';
                        }
                        isSyncingFormulaBar = true;
                        formulaBar.value = '';
                        isSyncingFormulaBar = false;
                        return;
                    }

                    activeCellLabel.textContent = `当前单元格: ${state.activeCell.sheet} / ${toA1Ref(state.activeCell.row, state.activeCell.col)}`;
                    if (document.activeElement !== cellRefInput) {
                        cellRefInput.value = toA1Ref(state.activeCell.row, state.activeCell.col);
                    }

                    isSyncingFormulaBar = true;
                    formulaBar.value = getCellCurrentValue(state.activeCell.sheet, state.activeCell.row, state.activeCell.col);
                    isSyncingFormulaBar = false;
                    applyCollaborationLocks();
                }

                function setActiveCell(sheetName, row, col, options = {}) {
                    const { focus = false, scroll = false } = options;
                    document.querySelectorAll('.cell-input.active').forEach(input => input.classList.remove('active'));
                    state.activeCell = { sheet: sheetName, row: Number(row), col: Number(col) };

                    const input = findCellInput(sheetName, row, col);
                    if (input) {
                        input.classList.add('active');
                        if (scroll) {
                            input.scrollIntoView({ block: 'nearest', inline: 'nearest' });
                        }
                        if (focus) {
                            input.focus();
                        }
                    }

                    updateActiveCellDisplay();
                }

                function updateSheetMeta(panel, sheet) {
                    panel.dataset.rowCount = sheet.row_count;
                    panel.dataset.colCount = sheet.col_count;
                    panel.dataset.startRow = sheet.start_row;
                    panel.dataset.startCol = sheet.start_col;
                    panel.dataset.endRow = sheet.end_row;
                    panel.dataset.endCol = sheet.end_col;

                    const meta = panel.querySelector('[data-sheet-meta]');
                    meta.textContent = `显示第 ${sheet.start_row}-${sheet.end_row} 行 / 第 ${sheet.start_col}-${sheet.end_col} 列，共 ${sheet.row_count} 行 / ${sheet.col_count} 列`;
                    if (sheet.truncated) {
                        meta.textContent += '，当前工作表还有更多内容可继续加载';
                    }

                    const rowInput = panel.querySelector('.range-row');
                    const colInput = panel.querySelector('.range-col');
                    rowInput.max = sheet.row_count;
                    colInput.max = sheet.col_count;
                    rowInput.value = sheet.start_row;
                    colInput.value = sheet.start_col;
                }

                function buildTableHtml(sheet) {
                    const readonlyAttr = (!state.editable || !state.canEdit) ? ' readonly' : '';
                    const headerCells = sheet.columns.map(colName => `<th>${escapeHtml(colName)}</th>`).join('');
                    const bodyRows = sheet.rows.map((rowValues, rowOffset) => {
                        const rowNumber = sheet.start_row + rowOffset;
                        const cells = rowValues.map((serverValue, colOffset) => {
                            const colNumber = sheet.start_col + colOffset;
                            const key = buildCellKey(sheet.name, rowNumber, colNumber);
                            const currentValue = pendingChanges.has(key) ? pendingChanges.get(key).value : serverValue;
                            const changedClass = pendingChanges.has(key) ? ' changed' : '';
                            return `
                                <td>
                                    <input
                                        class="cell-input${changedClass}"
                                        type="text"
                                        value="${escapeHtml(currentValue)}"
                                        spellcheck="false"
                                        autocomplete="off"
                                        autocapitalize="off"
                                        data-original="${escapeHtml(serverValue)}"
                                        data-sheet="${escapeHtml(sheet.name)}"
                                        data-row="${rowNumber}"
                                        data-col="${colNumber}"
                                        data-a1="${toA1Ref(rowNumber, colNumber)}"${readonlyAttr}
                                    >
                                </td>
                            `;
                        }).join('');
                        return `<tr><td class="row-head">${rowNumber}</td>${cells}</tr>`;
                    }).join('');

                    return `
                        <table>
                            <thead>
                                <tr>
                                    <th class="row-head">#</th>
                                    ${headerCells}
                                </tr>
                            </thead>
                            <tbody>${bodyRows}</tbody>
                        </table>
                    `;
                }

                function createBlankSheetRows(rowCount, colCount) {
                    return Array.from({ length: Math.max(0, rowCount) }, () => (
                        Array.from({ length: Math.max(0, colCount) }, () => '')
                    ));
                }

                function getRenderedSheetChunk(sheetName) {
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return null;
                    }

                    const startRow = Number(panel.dataset.startRow || 1);
                    const startCol = Number(panel.dataset.startCol || 1);
                    const endRow = Number(panel.dataset.endRow || startRow);
                    const endCol = Number(panel.dataset.endCol || startCol);
                    const rowCount = Number(panel.dataset.rowCount || endRow || 1);
                    const colCount = Number(panel.dataset.colCount || endCol || 1);
                    const rows = Array.from(panel.querySelectorAll('tbody tr')).map(row => (
                        Array.from(row.querySelectorAll('.cell-input')).map(input => input.value ?? '')
                    ));

                    return {
                        'name': sheetName,
                        'start_row': startRow,
                        'start_col': startCol,
                        'end_row': endRow,
                        'end_col': endCol,
                        'row_count': rowCount,
                        'col_count': colCount,
                        'rows': rows,
                        'columns': Array.from(
                            { length: Math.max(0, endCol - startCol + 1) },
                            (_, index) => columnNumberToName(startCol + index)
                        ),
                        'truncated': rowCount > endRow || colCount > endCol,
                        'has_prev_rows': startRow > 1,
                        'has_next_rows': endRow < rowCount,
                        'has_prev_cols': startCol > 1,
                        'has_next_cols': endCol < colCount
                    };
                }

                function normalizeLocalSheetChunk(sheet) {
                    const visibleRowCount = Math.max(0, (sheet.end_row || 0) - (sheet.start_row || 0) + 1);
                    const visibleColCount = Math.max(0, (sheet.end_col || 0) - (sheet.start_col || 0) + 1);
                    const sourceRows = Array.isArray(sheet.rows) ? sheet.rows : [];

                    sheet.rows = Array.from({ length: visibleRowCount }, (_, rowIndex) => {
                        const sourceRow = Array.isArray(sourceRows[rowIndex]) ? sourceRows[rowIndex] : [];
                        return Array.from({ length: visibleColCount }, (_, colIndex) => String(sourceRow[colIndex] ?? ''));
                    });
                    sheet.columns = Array.from(
                        { length: visibleColCount },
                        (_, index) => columnNumberToName((sheet.start_col || 1) + index)
                    );
                    sheet.max_row = visibleRowCount;
                    sheet.max_col = visibleColCount;
                    sheet.truncated = Boolean(sheet.row_count > sheet.end_row || sheet.col_count > sheet.end_col);
                    sheet.has_prev_rows = Boolean(sheet.start_row > 1);
                    sheet.has_next_rows = Boolean(sheet.end_row < sheet.row_count);
                    sheet.has_prev_cols = Boolean(sheet.start_col > 1);
                    sheet.has_next_cols = Boolean(sheet.end_col < sheet.col_count);
                    return sheet;
                }

                function applyOptimisticStructureSheetUpdate(sheetName, action, payload, sheetMeta) {
                    const currentSheet = getRenderedSheetChunk(sheetName);
                    if (!currentSheet || !sheetMeta) {
                        return { applied: false, needsReload: true };
                    }

                    const nextSheet = {
                        ...currentSheet,
                        start_row: Number(sheetMeta.start_row || currentSheet.start_row || 1),
                        start_col: Number(sheetMeta.start_col || currentSheet.start_col || 1),
                        end_row: Number(sheetMeta.end_row || currentSheet.end_row || currentSheet.start_row || 1),
                        end_col: Number(sheetMeta.end_col || currentSheet.end_col || currentSheet.start_col || 1),
                        row_count: Number(sheetMeta.row_count || currentSheet.row_count || 1),
                        col_count: Number(sheetMeta.col_count || currentSheet.col_count || 1),
                        rows: currentSheet.rows.map(row => row.slice())
                    };

                    const amount = Math.max(1, Math.floor(Number(payload.amount || 1) || 1));
                    const currentVisibleRowCount = Math.max(0, currentSheet.end_row - currentSheet.start_row + 1);
                    const currentVisibleColCount = Math.max(0, currentSheet.end_col - currentSheet.start_col + 1);
                    let needsReload = false;

                    if (action === 'insert_rows') {
                        const rowIndex = Math.max(1, Math.floor(Number(payload.row || 1) || 1));
                        if (rowIndex < currentSheet.start_row) {
                            needsReload = true;
                        } else if (rowIndex <= currentSheet.end_row) {
                            const insertOffset = Math.max(0, Math.min(nextSheet.rows.length, rowIndex - currentSheet.start_row));
                            nextSheet.rows.splice(insertOffset, 0, ...createBlankSheetRows(amount, currentVisibleColCount));
                        }
                    } else if (action === 'delete_rows') {
                        const rowIndex = Math.max(1, Math.floor(Number(payload.row || 1) || 1));
                        if (rowIndex <= currentSheet.end_row) {
                            needsReload = true;
                            if (rowIndex >= currentSheet.start_row) {
                                const removeOffset = Math.max(0, Math.min(nextSheet.rows.length, rowIndex - currentSheet.start_row));
                                nextSheet.rows.splice(removeOffset, amount);
                            }
                        }
                    } else if (action === 'insert_cols') {
                        const colIndex = Math.max(1, Math.floor(Number(payload.col || 1) || 1));
                        if (colIndex < currentSheet.start_col) {
                            needsReload = true;
                        } else if (colIndex <= currentSheet.end_col) {
                            const insertOffset = Math.max(0, Math.min(currentVisibleColCount, colIndex - currentSheet.start_col));
                            nextSheet.rows = nextSheet.rows.map(row => {
                                const copy = row.slice();
                                copy.splice(insertOffset, 0, ...Array.from({ length: amount }, () => ''));
                                return copy;
                            });
                        }
                    } else if (action === 'delete_cols') {
                        const colIndex = Math.max(1, Math.floor(Number(payload.col || 1) || 1));
                        if (colIndex <= currentSheet.end_col) {
                            needsReload = true;
                            if (colIndex >= currentSheet.start_col) {
                                const removeOffset = Math.max(0, Math.min(currentVisibleColCount, colIndex - currentSheet.start_col));
                                nextSheet.rows = nextSheet.rows.map(row => {
                                    const copy = row.slice();
                                    copy.splice(removeOffset, amount);
                                    return copy;
                                });
                            }
                        }
                    } else {
                        return { applied: false, needsReload: true };
                    }

                    normalizeLocalSheetChunk(nextSheet);
                    renderSheetChunk(nextSheet);
                    return { applied: true, needsReload };
                }

                function renderSheetChunk(sheet) {
                    const panel = getSheetPanel(sheet.name);
                    if (!panel) {
                        return;
                    }
                    updateSheetMeta(panel, sheet);
                    const tableWrap = panel.querySelector('.table-wrap');
                    tableWrap.innerHTML = buildTableHtml(sheet);
                    applyCollaborationLocks();

                    if (state.pendingFocus && state.pendingFocus.sheet === sheet.name) {
                        const { row, col, focus = true } = state.pendingFocus;
                        state.pendingFocus = null;
                        setActiveCell(sheet.name, row, col, { focus, scroll: true });
                        return;
                    }

                    if (state.activeCell && state.activeCell.sheet === sheet.name) {
                        const withinCurrentRange = (
                            state.activeCell.row >= sheet.start_row &&
                            state.activeCell.row <= sheet.end_row &&
                            state.activeCell.col >= sheet.start_col &&
                            state.activeCell.col <= sheet.end_col
                        );
                        if (withinCurrentRange) {
                            setActiveCell(sheet.name, state.activeCell.row, state.activeCell.col);
                            return;
                        }
                    }

                    if (getActiveSheetName() === sheet.name) {
                        const firstInput = getFirstVisibleInput(sheet.name);
                        if (firstInput) {
                            setActiveCell(sheet.name, Number(firstInput.dataset.row), Number(firstInput.dataset.col));
                        }
                    }
                }

                function collectChanges() {
                    const grouped = new Map();
                    for (const change of pendingChanges.values()) {
                        const sheetName = change.sheet;
                        if (!grouped.has(sheetName)) {
                            grouped.set(sheetName, []);
                        }
                        grouped.get(sheetName).push(change);
                    }
                    return Array.from(grouped.entries()).map(([name, cells]) => ({ name, cells }));
                }

                function isEditingCellOrFormula() {
                    const activeElement = document.activeElement;
                    return Boolean(
                        activeElement &&
                        (
                            activeElement.classList?.contains('cell-input') ||
                            activeElement === formulaBar
                        )
                    );
                }

                function scheduleAutoSave(force = false) {
                    clearTimeout(autoSaveTimer);
                    if (!state.autoSaveEnabled || !state.editable || !state.canEdit || pendingChanges.size === 0) {
                        autoSaveRetryPending = false;
                        return;
                    }
                    const delay = force ? Math.min(state.autoSaveDelay, 250) : state.autoSaveDelay;
                    autoSaveTimer = window.setTimeout(() => saveExcelFile({ auto: true }), delay);
                }

                function cancelAutoSave() {
                    clearTimeout(autoSaveTimer);
                    autoSaveTimer = null;
                    autoSaveRetryPending = false;
                }

                function updatePendingChangeFromInput(input, options = {}) {
                    const { deferAutoSave = false } = options;
                    const original = input.dataset.original ?? '';
                    const current = input.value ?? '';
                    const sheetName = input.dataset.sheet;
                    const row = Number(input.dataset.row);
                    const col = Number(input.dataset.col);
                    const key = buildCellKey(sheetName, row, col);

                    if (current === original) {
                        pendingChanges.delete(key);
                        input.classList.remove('changed');
                    } else {
                        pendingChanges.set(key, { sheet: sheetName, row, col, value: current, original });
                        input.classList.add('changed');
                    }

                    if (!deferAutoSave) {
                        scheduleAutoSave();
                    }

                    if (state.activeCell && state.activeCell.sheet === sheetName && state.activeCell.row === row && state.activeCell.col === col) {
                        updateActiveCellDisplay();
                    }
                }

                function applyValueToCell(sheetName, row, col, value, options = {}) {
                    const { deferAutoSave = false } = options;
                    if (isCellLockedByOther(sheetName, row, col)) {
                        return false;
                    }

                    const input = findCellInput(sheetName, row, col);
                    if (input) {
                        input.value = value;
                        updatePendingChangeFromInput(input, { deferAutoSave });
                        return true;
                    }

                    const key = buildCellKey(sheetName, row, col);
                    pendingChanges.set(key, { sheet: sheetName, row, col, value, original: '' });
                    if (!deferAutoSave) {
                        scheduleAutoSave();
                    }
                    if (state.activeCell && state.activeCell.sheet === sheetName && state.activeCell.row === row && state.activeCell.col === col) {
                        updateActiveCellDisplay();
                    }
                    return true;
                }

                async function loadSheetRange(sheetName, startRow = null, startCol = null, options = {}) {
                    const { quiet = false } = options;
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return null;
                    }

                    const rowInput = panel.querySelector('.range-row');
                    const colInput = panel.querySelector('.range-col');
                    const rowCount = Number(panel.dataset.rowCount || rowInput.max || 1);
                    const colCount = Number(panel.dataset.colCount || colInput.max || 1);
                    const nextStartRow = Math.max(1, Math.min(Number(startRow ?? rowInput.value ?? 1), rowCount));
                    const nextStartCol = Math.max(1, Math.min(Number(startCol ?? colInput.value ?? 1), colCount));

                    if (!quiet) {
                        setStatus(`正在加载 ${sheetName} 的更多内容...`, 'info');
                    }

                    try {
                        const response = await fetch('/api/excel_sheet_data', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                sheet_name: sheetName,
                                start_row: nextStartRow,
                                start_col: nextStartCol,
                                row_limit: state.rowLimit,
                                col_limit: state.colLimit
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (!quiet) {
                                setStatus(data.message || '加载失败，请稍后重试。', 'error');
                            }
                            return null;
                        }

                        if (data.mtime_ns) {
                            state.mtimeNs = data.mtime_ns;
                        }
                        renderSheetChunk(data.sheet);
                        syncCollaborationState({ quiet: true });
                        if (!quiet) {
                            setStatus(`已加载 ${sheetName}：第 ${data.sheet.start_row}-${data.sheet.end_row} 行 / 第 ${data.sheet.start_col}-${data.sheet.end_col} 列`, 'info');
                        }
                        return data.sheet;
                    } catch (error) {
                        if (!quiet) {
                            setStatus('加载失败：网络异常或服务器不可用。', 'error');
                        }
                        return null;
                    }
                }

                function shiftSheetRange(sheetName, rowDelta = 0, colDelta = 0) {
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return;
                    }
                    const rowCount = Number(panel.dataset.rowCount || 1);
                    const colCount = Number(panel.dataset.colCount || 1);
                    const currentStartRow = Number(panel.dataset.startRow || 1);
                    const currentStartCol = Number(panel.dataset.startCol || 1);
                    const nextStartRow = Math.max(1, Math.min(currentStartRow + rowDelta, rowCount));
                    const nextStartCol = Math.max(1, Math.min(currentStartCol + colDelta, colCount));
                    loadSheetRange(sheetName, nextStartRow, nextStartCol);
                }

                async function jumpToCell(sheetName, row, col, options = {}) {
                    const { focus = true } = options;
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return;
                    }

                    const rowCount = Number(panel.dataset.rowCount || 1);
                    const colCount = Number(panel.dataset.colCount || 1);
                    const safeRow = Math.max(1, Math.min(Number(row || 1), rowCount));
                    const safeCol = Math.max(1, Math.min(Number(col || 1), colCount));
                    activateSheet(sheetName);

                    const startRow = Number(panel.dataset.startRow || 1);
                    const startCol = Number(panel.dataset.startCol || 1);
                    const endRow = Number(panel.dataset.endRow || 1);
                    const endCol = Number(panel.dataset.endCol || 1);
                    if (safeRow >= startRow && safeRow <= endRow && safeCol >= startCol && safeCol <= endCol) {
                        setActiveCell(sheetName, safeRow, safeCol, { focus, scroll: true });
                        return;
                    }

                    const nextStartRow = Math.max(1, Math.min(safeRow - Math.floor(state.rowLimit / 2), rowCount));
                    const nextStartCol = Math.max(1, Math.min(safeCol - Math.floor(state.colLimit / 2), colCount));
                    state.pendingFocus = { sheet: sheetName, row: safeRow, col: safeCol, focus };
                    await loadSheetRange(sheetName, nextStartRow, nextStartCol);
                }

                function parsePositiveInputValue(input, defaultValue = 1) {
                    const value = Number(input?.value ?? defaultValue);
                    if (!Number.isFinite(value) || value < 1) {
                        return defaultValue;
                    }
                    return Math.max(1, Math.floor(value));
                }

                function parseColumnInputValue(value) {
                    const text = String(value || '').trim();
                    if (!text) {
                        return 1;
                    }
                    if (/^\d+$/.test(text)) {
                        return parsePositiveInputValue({ value: text }, 1);
                    }
                    const columnNumber = columnNameToNumber(text);
                    return columnNumber || 1;
                }

                function buildExcelEditorUrl(activeSheetName) {
                    const url = new URL(window.location.href);
                    if (activeSheetName) {
                        url.searchParams.set('sheet', activeSheetName);
                    } else {
                        url.searchParams.delete('sheet');
                    }
                    return url.toString();
                }

                async function runStructureOperation(action, extraPayload = {}, options = {}) {
                    const { confirmMessage = '' } = options;
                    if (!state.canEdit || !state.editable) {
                        setStatus('只有管理员才能修改表结构。', 'error');
                        return;
                    }

                    if (pendingChanges.size > 0) {
                        setStatus('请先保存当前单元格修改，再进行插入、删除行列或工作表操作。', 'error');
                        return;
                    }

                    if (confirmMessage && !window.confirm(confirmMessage)) {
                        return;
                    }

                    setStatus('正在修改表结构...', 'info');
                    try {
                        const activeSheetName = getActiveSheetName();
                        const activePanel = getSheetPanel(activeSheetName);
                        const currentStartRow = activePanel ? Number(activePanel.dataset.startRow || 1) : 1;
                        const currentStartCol = activePanel ? Number(activePanel.dataset.startCol || 1) : 1;
                        const isRowOrColOperation = ['insert_rows', 'delete_rows', 'insert_cols', 'delete_cols'].includes(action);
                        const response = await fetch('/api/excel_structure_operation', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                mtime_ns: state.mtimeNs,
                                action,
                                sheet_name: activeSheetName,
                                start_row: currentStartRow,
                                start_col: currentStartCol,
                                row_limit: state.rowLimit,
                                col_limit: state.colLimit,
                                include_sheet_snapshot: !isRowOrColOperation,
                                ...extraPayload
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (data.mtime_ns) {
                                state.mtimeNs = data.mtime_ns;
                            }
                            setStatus(data.message || '表结构修改失败，请稍后重试。', 'error');
                            return;
                        }

                        state.mtimeNs = data.mtime_ns || state.mtimeNs;
                        const nextActiveSheet = data.active_sheet || activeSheetName;
                        const shouldSoftRefresh = (
                            ['insert_rows', 'delete_rows', 'insert_cols', 'delete_cols'].includes(action) &&
                            nextActiveSheet === activeSheetName
                        );

                        if (shouldSoftRefresh) {
                            if (data.sheet) {
                                renderSheetChunk(data.sheet);
                                setStatus(data.message || '表结构已更新。', 'success');
                                return;
                            }

                            if (data.sheet_meta) {
                                const patchResult = applyOptimisticStructureSheetUpdate(
                                    nextActiveSheet,
                                    action,
                                    extraPayload,
                                    data.sheet_meta
                                );
                                if (patchResult.applied) {
                                    setStatus(
                                        data.message || (patchResult.needsReload ? '表结构已更新，正在后台校准当前区域...' : '表结构已更新。'),
                                        'success'
                                    );
                                    if (patchResult.needsReload) {
                                        window.setTimeout(() => {
                                            loadSheetRange(nextActiveSheet, currentStartRow, currentStartCol, { quiet: true });
                                        }, 0);
                                    } else {
                                        window.setTimeout(() => {
                                            syncCollaborationState({ quiet: true });
                                        }, 0);
                                    }
                                    return;
                                }
                            }

                            setStatus(data.message || '表结构已更新，正在刷新当前区域...', 'success');
                            await loadSheetRange(nextActiveSheet, currentStartRow, currentStartCol);
                            return;
                        }

                        setStatus(data.message || '表结构已更新，正在刷新...', 'success');
                        window.location.href = buildExcelEditorUrl(nextActiveSheet);
                    } catch (error) {
                        setStatus('表结构修改失败：网络异常或服务器不可用。', 'error');
                    }
                }

                function renderSearchResults(data) {
                    const results = data.results || [];
                    const title = results.length
                        ? `搜索 "${escapeHtml(data.keyword)}"：找到 ${results.length} 个结果${data.truncated ? '（仅显示前几项）' : ''}`
                        : `搜索 "${escapeHtml(data.keyword)}"：当前工作表没有匹配结果`;

                    const items = results.length
                        ? `<div class="search-result-list">${results.map(item => `
                            <button class="search-result-item" type="button" data-search-sheet="${escapeHtml(item.sheet)}" data-search-row="${item.row}" data-search-col="${item.col}">
                                <span class="search-result-cell">${escapeHtml(item.cell)}</span>
                                <span>${escapeHtml(item.value || '(空白)')}</span>
                            </button>
                        `).join('')}</div>`
                        : '<div class="search-empty">你可以继续换一个关键词，或者先定位到某个单元格再编辑。</div>';

                    searchResultsEl.innerHTML = `<div class="search-results-title">${title}</div>${items}`;
                    searchResultsEl.hidden = false;
                }

                async function searchCurrentSheet() {
                    const keyword = searchInput.value.trim();
                    if (!keyword) {
                        searchResultsEl.hidden = true;
                        setStatus('请输入要搜索的内容。', 'error');
                        return;
                    }

                    const sheetName = getActiveSheetName();
                    setStatus(`正在搜索 ${sheetName} ...`, 'info');

                    try {
                        const response = await fetch('/api/excel_sheet_search', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                sheet_name: sheetName,
                                keyword,
                                max_results: 30
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            setStatus(data.message || '搜索失败，请稍后重试。', 'error');
                            return;
                        }

                        if (data.mtime_ns) {
                            state.mtimeNs = data.mtime_ns;
                        }
                        renderSearchResults(data);
                        setStatus(`搜索完成：${sheetName} 找到 ${data.results.length} 个结果。`, 'info');
                    } catch (error) {
                        setStatus('搜索失败：网络异常或服务器不可用。', 'error');
                    }
                }

                function parseClipboardMatrix(clipboardText) {
                    const normalized = String(clipboardText || '')
                        .split(carriageReturnChar + newlineChar).join(newlineChar)
                        .split(carriageReturnChar).join(newlineChar);
                    const rows = normalized.split(newlineChar);
                    if (rows.length > 1 && rows[rows.length - 1] === '') {
                        rows.pop();
                    }
                    return rows.map(row => row.split(tabChar));
                }

                function applyClipboardMatrix(targetInput, clipboardText) {
                    const matrix = parseClipboardMatrix(clipboardText);
                    if (!matrix.length || (matrix.length === 1 && matrix[0].length === 1)) {
                        return false;
                    }

                    const sheetName = targetInput.dataset.sheet;
                    const panel = getSheetPanel(sheetName);
                    if (!panel) {
                        return false;
                    }

                    const startRow = Number(targetInput.dataset.row);
                    const startCol = Number(targetInput.dataset.col);
                    const rowCount = Number(panel.dataset.rowCount || startRow);
                    const colCount = Number(panel.dataset.colCount || startCol);
                    let appliedCount = 0;
                    let skippedLocked = 0;
                    let clipped = false;

                    state.isBulkEditing = true;
                    try {
                        matrix.forEach((rowValues, rowOffset) => {
                            rowValues.forEach((cellValue, colOffset) => {
                                const row = startRow + rowOffset;
                                const col = startCol + colOffset;
                                if (row > rowCount || col > colCount) {
                                    clipped = true;
                                    return;
                                }
                                if (applyValueToCell(sheetName, row, col, cellValue, { deferAutoSave: true })) {
                                    appliedCount += 1;
                                } else {
                                    skippedLocked += 1;
                                }
                            });
                        });
                    } finally {
                        state.isBulkEditing = false;
                    }

                    scheduleAutoSave();
                    setActiveCell(sheetName, startRow, startCol);
                    const statusParts = [];
                    if (appliedCount > 0) {
                        statusParts.push(`已批量粘贴 ${appliedCount} 个单元格`);
                    }
                    if (skippedLocked > 0) {
                        statusParts.push(`${skippedLocked} 个单元格正在被其他人编辑，已自动跳过`);
                    }
                    if (clipped) {
                        statusParts.push('超出当前工作表范围的内容已忽略');
                    }
                    if (statusParts.length > 0) {
                        setStatus(statusParts.join('；') + '。', appliedCount > 0 ? 'info' : 'error');
                    }
                    return true;
                }

                function handleCellNavigation(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement) || !target.classList.contains('cell-input')) {
                        return;
                    }

                    const input = target;
                    const sheetName = input.dataset.sheet;
                    const row = Number(input.dataset.row);
                    const col = Number(input.dataset.col);
                    const valueLength = (input.value || '').length;
                    let nextRow = row;
                    let nextCol = col;
                    let shouldMove = false;

                    if (event.key === 'Enter') {
                        event.preventDefault();
                        nextRow += event.shiftKey ? -1 : 1;
                        shouldMove = true;
                    } else if (event.key === 'Tab') {
                        event.preventDefault();
                        nextCol += event.shiftKey ? -1 : 1;
                        shouldMove = true;
                    } else if (event.key === 'ArrowUp') {
                        event.preventDefault();
                        nextRow -= 1;
                        shouldMove = true;
                    } else if (event.key === 'ArrowDown') {
                        event.preventDefault();
                        nextRow += 1;
                        shouldMove = true;
                    } else if (event.key === 'ArrowLeft' && input.selectionStart === 0 && input.selectionEnd === 0) {
                        event.preventDefault();
                        nextCol -= 1;
                        shouldMove = true;
                    } else if (event.key === 'ArrowRight' && input.selectionStart === valueLength && input.selectionEnd === valueLength) {
                        event.preventDefault();
                        nextCol += 1;
                        shouldMove = true;
                    }

                    if (shouldMove) {
                        jumpToCell(sheetName, nextRow, nextCol);
                    }
                }

                async function checkRemoteUpdates() {
                    if (isSaving || !state.filePath) {
                        return;
                    }

                    try {
                        const response = await fetch('/api/excel_file_status', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ file_path: state.filePath })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success || !data.mtime_ns || data.mtime_ns === state.mtimeNs) {
                            return;
                        }

                        if (pendingChanges.size > 0) {
                            setStatus('检测到共享文件已有新版本，你当前还有未保存修改，请先保存或刷新页面。', 'error');
                            return;
                        }

                        state.mtimeNs = data.mtime_ns;
                        const activeSheet = getActiveSheetName();
                        const panel = getSheetPanel(activeSheet);
                        if (panel) {
                            await loadSheetRange(activeSheet, Number(panel.dataset.startRow || 1), Number(panel.dataset.startCol || 1), { quiet: true });
                        }
                        setStatus('检测到共享文件更新，已自动刷新当前区域。', 'info');
                    } catch (error) {
                    }
                }

                async function saveExcelFile(options = {}) {
                    const { auto = false } = options;
                    if (isSaving) {
                        if (auto) {
                            autoSaveRetryPending = true;
                        }
                        return;
                    }
                    if (!state.canEdit) {
                        setStatus('只有管理员才能保存修改。', 'error');
                        return;
                    }

                    const sheets = collectChanges();
                    if (sheets.length === 0) {
                        if (!auto) {
                            setStatus('没有检测到修改内容。', 'info');
                        }
                        return;
                    }

                    cancelAutoSave();
                    isSaving = true;
                    saveBtn.disabled = true;
                    setStatus(auto ? '正在自动保存到共享目录...' : '正在保存 Excel 到共享目录...', 'info');

                    try {
                        const response = await fetch('/api/save_excel_file', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                mtime_ns: state.mtimeNs,
                                sheets
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            setStatus(data.message || '保存失败，请稍后重试。', 'error');
                            if (data.mtime_ns) {
                                state.mtimeNs = data.mtime_ns;
                            }
                            return;
                        }

                        state.mtimeNs = data.mtime_ns;
                        pendingChanges.clear();
                        document.querySelectorAll('.cell-input').forEach(input => {
                            input.dataset.original = input.value ?? '';
                            input.classList.remove('changed');
                        });
                        const saveMessage = auto
                            ? `${data.merged ? '已自动合并并保存' : '已自动保存到共享目录'}（修改 ${data.changed_cells || 0} 个单元格）`
                            : `${data.merged ? '已合并并保存到共享目录' : '已保存到共享目录'}（修改 ${data.changed_cells || 0} 个单元格）`;
                        setStatus(saveMessage, 'success');
                    } catch (error) {
                        setStatus('保存失败：网络异常或服务器不可用。', 'error');
                    } finally {
                        isSaving = false;
                        saveBtn.disabled = !state.editable || !state.canEdit;
                        if (pendingChanges.size > 0 && state.autoSaveEnabled) {
                            const shouldForceRetry = autoSaveRetryPending;
                            autoSaveRetryPending = false;
                            scheduleAutoSave(shouldForceRetry);
                        } else {
                            autoSaveRetryPending = false;
                        }
                    }
                }

                tabButtons.forEach(btn => {
                    btn.addEventListener('click', () => activateSheet(btn.dataset.sheetTarget));
                });

                document.addEventListener('input', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement)) {
                        return;
                    }

                    if (target.classList.contains('cell-input')) {
                        if (!state.editable || !state.canEdit) {
                            return;
                        }
                        if (isCellLockedByOther(target.dataset.sheet, Number(target.dataset.row), Number(target.dataset.col))) {
                            setStatus('这个单元格正在被其他人编辑，当前不能修改。', 'error');
                            return;
                        }
                        updatePendingChangeFromInput(target);
                        setStatus(
                            state.autoSaveEnabled
                                ? '表格已修改，系统会自动保存；也可以按 Ctrl+S 立即保存。'
                                : '表格已修改，按 Ctrl+S 可直接保存到共享目录。',
                            'info'
                        );
                    } else if (target === formulaBar) {
                        if (isSyncingFormulaBar || !state.activeCell || !state.editable || !state.canEdit) {
                            return;
                        }
                        if (!applyValueToCell(state.activeCell.sheet, state.activeCell.row, state.activeCell.col, formulaBar.value)) {
                            setStatus('这个单元格正在被其他人编辑，当前不能修改。', 'error');
                            return;
                        }
                        setStatus(
                            state.autoSaveEnabled
                                ? '当前单元格已修改，系统会自动保存。'
                                : '当前单元格已修改，按 Ctrl+S 可保存。',
                            'info'
                        );
                    }
                });

                document.addEventListener('click', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement)) {
                        return;
                    }
                    const searchResultButton = target.closest('.search-result-item');

                    if (target.classList.contains('load-range-btn')) {
                        loadSheetRange(target.dataset.sheet);
                    } else if (target.classList.contains('nav-rows-prev')) {
                        shiftSheetRange(target.dataset.sheet, -state.rowLimit, 0);
                    } else if (target.classList.contains('nav-rows-next')) {
                        shiftSheetRange(target.dataset.sheet, state.rowLimit, 0);
                    } else if (target.classList.contains('nav-cols-prev')) {
                        shiftSheetRange(target.dataset.sheet, 0, -state.colLimit);
                    } else if (target.classList.contains('nav-cols-next')) {
                        shiftSheetRange(target.dataset.sheet, 0, state.colLimit);
                    } else if (searchResultButton) {
                        jumpToCell(searchResultButton.dataset.searchSheet, Number(searchResultButton.dataset.searchRow), Number(searchResultButton.dataset.searchCol));
                    }
                });

                document.addEventListener('focusin', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement)) {
                        return;
                    }

                    if (target.classList.contains('cell-input')) {
                        setActiveCell(target.dataset.sheet, Number(target.dataset.row), Number(target.dataset.col));
                        if (state.editable && state.canEdit) {
                            updateCollabLockTarget({
                                sheet: target.dataset.sheet,
                                row: Number(target.dataset.row),
                                col: Number(target.dataset.col)
                            });
                        }
                        return;
                    }

                    if (target === formulaBar && state.activeCell && state.editable && state.canEdit) {
                        updateCollabLockTarget({
                            sheet: state.activeCell.sheet,
                            row: state.activeCell.row,
                            col: state.activeCell.col
                        });
                    }
                });

                document.addEventListener('focusout', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement)) {
                        return;
                    }
                    const leavingEditable = target.classList.contains('cell-input') || target === formulaBar;
                    if (!leavingEditable) {
                        return;
                    }
                    window.setTimeout(function() {
                        if (!state.autoSaveEnabled || pendingChanges.size === 0) {
                            return;
                        }
                        if (!isEditingCellOrFormula()) {
                            scheduleAutoSave(true);
                        }
                    }, 0);
                });

                document.addEventListener('paste', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLElement) || !target.classList.contains('cell-input') || !state.editable || !state.canEdit) {
                        return;
                    }

                    const clipboardText = event.clipboardData ? event.clipboardData.getData('text/plain') : '';
                    if (!clipboardText || (
                        clipboardText.indexOf(tabChar) === -1 &&
                        clipboardText.indexOf(newlineChar) === -1 &&
                        clipboardText.indexOf(carriageReturnChar) === -1
                    )) {
                        return;
                    }

                    if (applyClipboardMatrix(target, clipboardText)) {
                        event.preventDefault();
                    }
                });

                saveBtn.addEventListener('click', saveExcelFile);
                if (jumpCellBtn) {
                    jumpCellBtn.addEventListener('click', async function() {
                        const ref = parseCellReference(cellRefInput.value);
                        if (!ref) {
                            setStatus('请输入有效单元格，例如 A1 或 12,5。', 'error');
                            return;
                        }
                        await jumpToCell(getActiveSheetName(), ref.row, ref.col);
                    });
                }

                if (searchBtn) {
                    searchBtn.addEventListener('click', searchCurrentSheet);
                }

                if (searchInput) {
                    searchInput.addEventListener('keydown', function(event) {
                        if (event.key === 'Enter') {
                            event.preventDefault();
                            searchCurrentSheet();
                        }
                    });
                }

                if (cellRefInput) {
                    cellRefInput.addEventListener('keydown', function(event) {
                        if (event.key === 'Enter') {
                            event.preventDefault();
                            jumpCellBtn.click();
                        }
                    });
                }

                if (formulaBar) {
                    formulaBar.addEventListener('keydown', function(event) {
                        if (event.key === 'Enter' && state.activeCell) {
                            event.preventDefault();
                            jumpToCell(state.activeCell.sheet, state.activeCell.row + 1, state.activeCell.col);
                        }
                    });
                }

                if (autoSaveBtn) {
                    autoSaveBtn.addEventListener('click', function() {
                        if (!state.canEdit || !state.editable) {
                            return;
                        }
                        state.autoSaveEnabled = !state.autoSaveEnabled;
                        updateAutoSaveButton();
                        if (state.autoSaveEnabled) {
                            scheduleAutoSave();
                            setStatus('已开启自动保存。', 'info');
                        } else {
                            cancelAutoSave();
                            setStatus('已关闭自动保存。', 'info');
                        }
                    });
                }

                if (insertRowsBtn) {
                    insertRowsBtn.addEventListener('click', function() {
                        runStructureOperation('insert_rows', {
                            row: parsePositiveInputValue(rowOpIndexInput, 1),
                            amount: parsePositiveInputValue(rowOpAmountInput, 1)
                        });
                    });
                }

                if (deleteRowsBtn) {
                    deleteRowsBtn.addEventListener('click', function() {
                        const row = parsePositiveInputValue(rowOpIndexInput, 1);
                        const amount = parsePositiveInputValue(rowOpAmountInput, 1);
                        runStructureOperation(
                            'delete_rows',
                            { row, amount },
                            { confirmMessage: `确认删除当前工作表从第 ${row} 行开始的 ${amount} 行吗？` }
                        );
                    });
                }

                if (insertColsBtn) {
                    insertColsBtn.addEventListener('click', function() {
                        runStructureOperation('insert_cols', {
                            col: parseColumnInputValue(colOpIndexInput?.value),
                            amount: parsePositiveInputValue(colOpAmountInput, 1)
                        });
                    });
                }

                if (deleteColsBtn) {
                    deleteColsBtn.addEventListener('click', function() {
                        const col = parseColumnInputValue(colOpIndexInput?.value);
                        const amount = parsePositiveInputValue(colOpAmountInput, 1);
                        runStructureOperation(
                            'delete_cols',
                            { col, amount },
                            { confirmMessage: `确认删除当前工作表从第 ${columnNumberToName(col)} 列开始的 ${amount} 列吗？` }
                        );
                    });
                }

                if (addSheetBtn) {
                    addSheetBtn.addEventListener('click', function() {
                        runStructureOperation('add_sheet', {
                            new_name: (sheetNameInput?.value || '').trim() || 'Sheet'
                        });
                    });
                }

                if (renameSheetBtn) {
                    renameSheetBtn.addEventListener('click', function() {
                        const newName = (sheetNameInput?.value || '').trim();
                        if (!newName) {
                            setStatus('请输入新的工作表名称。', 'error');
                            return;
                        }
                        runStructureOperation('rename_sheet', { new_name: newName });
                    });
                }

                if (deleteSheetBtn) {
                    deleteSheetBtn.addEventListener('click', function() {
                        const sheetName = getActiveSheetName();
                        runStructureOperation(
                            'delete_sheet',
                            {},
                            { confirmMessage: `确认删除工作表“${sheetName}”吗？此操作不可撤销。` }
                        );
                    });
                }

                document.addEventListener('keydown', function(event) {
                    if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 's') {
                        event.preventDefault();
                        saveExcelFile();
                        return;
                    }

                    if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 'f') {
                        event.preventDefault();
                        searchInput.focus();
                        searchInput.select();
                        return;
                    }

                    if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 'g') {
                        event.preventDefault();
                        cellRefInput.focus();
                        cellRefInput.select();
                        return;
                    }

                    handleCellNavigation(event);
                });

                window.addEventListener('beforeunload', function(event) {
                    if (!suppressBeforeUnload && pendingChanges.size > 0) {
                        event.preventDefault();
                        event.returnValue = '';
                    }
                });
                window.addEventListener('pagehide', function() {
                    releaseCollaborationState(true);
                });
                if (backToListBtn) {
                    backToListBtn.addEventListener('click', returnToList);
                }

                updateAutoSaveButton();
                activateSheet(state.activeSheet || state.sheetNames[0] || '');
                if (state.sheetNames.length > 0) {
                    const firstInput = getFirstVisibleInput(getActiveSheetName());
                    if (firstInput) {
                        setActiveCell(getActiveSheetName(), Number(firstInput.dataset.row), Number(firstInput.dataset.col));
                    }
                }
                if (state.filePath) {
                    syncCollaborationState({ quiet: true });
                    remoteCheckTimer = window.setInterval(function() {
                        syncCollaborationState({ quiet: true });
                    }, state.collabPollInterval);
                }
            </script>
        </body>
        </html>
        ''',
        filename=os.path.basename(filename),
        file_path=filename,
        back_url=back_url,
        editable=preview['editable'],
        can_edit=can_edit,
        mtime_ns=preview['mtime_ns'],
        file_size=get_file_size(preview['file_size']),
        sheet_count=preview['sheet_count'],
        sheets=preview['sheets'],
        sheet_names=[sheet['name'] for sheet in preview['sheets']],
        active_sheet=active_sheet_name,
        backend=preview['backend'],
        excel_large_file=preview.get('large_file', False),
        excel_row_limit=preview['row_limit'],
        excel_col_limit=preview['col_limit'],
        collab_client_id=get_excel_editor_client_id(),
        collab_username=get_current_username(),
        collab_poll_interval_ms=EXCEL_COLLAB_POLL_INTERVAL_MS,
        warning_message=warning_message,
        initial_status=initial_status
        )
    except Exception as e:
        print(f"打开 Excel 编辑器失败: {str(e)}")
        flash(f'打开 Excel 编辑器失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/edit_docx/<path:filename>')
def edit_docx_file(filename):
    """在线编辑 DOCX 文件。"""
    try:
        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], filename)
        if not success:
            flash(f'路径错误: {error}', 'danger')
            return redirect(url_for('index'))

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            flash('文件不存在', 'danger')
            return redirect(url_for('index'))

        if not is_word_editable_file(filename):
            return redirect(url_for('preview_file', filename=filename))

        preview = load_docx_file_preview(filepath)
        can_edit = user_can_edit_files()
        parent_path = os.path.dirname(filename).replace('\\', '/')
        back_url = url_for('index', subpath=parent_path) if parent_path else url_for('index')

        warning_message = '当前在线编辑器支持修改正文、页眉页脚、批注、文本框、图片说明以及其中的表格单元格，并写回原 DOCX。'
        if not preview['editable']:
            warning_message += (
                f' 当前文件大小为 {get_file_size(preview["file_size"])}，超过在线编辑上限 '
                f'{get_file_size(MAX_EDITABLE_DOCX_FILE_SIZE)}，当前仅支持只读预览。'
            )
        elif not can_edit:
            warning_message += ' 当前昵称可以预览此文件，但只有管理员才能保存修改。'

        initial_status = '可直接修改正文、页眉页脚、批注、文本框和图片说明，快捷键：Ctrl+S。'
        if not can_edit:
            initial_status = '当前为只读模式，只有管理员才能保存修改。'
        if not preview['editable']:
            initial_status = '文件过大，当前仅支持只读预览。'

        return render_template_string('''
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>预览/编辑: {{ filename }}</title>
            <style>
                :root {
                    color-scheme: dark;
                }
                body {
                    margin: 0;
                    min-height: 100vh;
                    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
                    background:
                        radial-gradient(circle at top left, rgba(16, 185, 129, 0.16), transparent 28%),
                        radial-gradient(circle at top right, rgba(59, 130, 246, 0.16), transparent 34%),
                        #0f172a;
                    color: #e2e8f0;
                }
                .page {
                    max-width: 1480px;
                    margin: 0 auto;
                    padding: 24px;
                }
                .header, .card {
                    background: rgba(15, 23, 42, 0.84);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    backdrop-filter: blur(14px);
                    border-radius: 20px;
                    box-shadow: 0 20px 45px rgba(2, 6, 23, 0.28);
                }
                .header {
                    padding: 22px 24px;
                    margin-bottom: 18px;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    gap: 16px;
                    flex-wrap: wrap;
                }
                .title h1 {
                    margin: 0 0 8px;
                    font-size: 28px;
                    color: #f8fafc;
                }
                .title p {
                    margin: 0;
                    color: #94a3b8;
                    word-break: break-all;
                }
                .actions {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                }
                .btn {
                    border: none;
                    border-radius: 12px;
                    padding: 12px 18px;
                    font-size: 14px;
                    font-weight: 600;
                    cursor: pointer;
                    text-decoration: none;
                    transition: transform 0.2s ease, opacity 0.2s ease, background 0.2s ease;
                }
                .btn:hover {
                    transform: translateY(-1px);
                }
                .btn:disabled {
                    cursor: not-allowed;
                    opacity: 0.55;
                    transform: none;
                }
                .btn-primary {
                    background: linear-gradient(135deg, #10b981, #2563eb);
                    color: white;
                }
                .btn-secondary {
                    background: rgba(30, 41, 59, 0.95);
                    color: #e2e8f0;
                    border: 1px solid rgba(148, 163, 184, 0.18);
                }
                .card {
                    padding: 22px;
                }
                .meta {
                    display: flex;
                    gap: 16px;
                    flex-wrap: wrap;
                    margin-bottom: 14px;
                    font-size: 13px;
                    color: #94a3b8;
                }
                .notice {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(245, 158, 11, 0.16);
                    border: 1px solid rgba(245, 158, 11, 0.28);
                    color: #fde68a;
                    line-height: 1.7;
                }
                .guide {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(16, 185, 129, 0.1);
                    border: 1px solid rgba(16, 185, 129, 0.2);
                    color: #d1fae5;
                    line-height: 1.7;
                }
                .collab-panel {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    justify-content: space-between;
                    margin-bottom: 16px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(59, 130, 246, 0.1);
                    border: 1px solid rgba(59, 130, 246, 0.22);
                    color: #dbeafe;
                }
                .collab-users {
                    display: flex;
                    gap: 8px;
                    flex-wrap: wrap;
                    align-items: center;
                }
                .collab-user {
                    display: inline-flex;
                    align-items: center;
                    gap: 6px;
                    padding: 6px 10px;
                    border-radius: 999px;
                    background: rgba(15, 23, 42, 0.72);
                    border: 1px solid rgba(59, 130, 246, 0.22);
                    font-size: 12px;
                    color: #eff6ff;
                }
                .collab-user.self {
                    border-color: rgba(16, 185, 129, 0.44);
                    background: rgba(15, 23, 42, 0.95);
                }
                .collab-hint {
                    font-size: 12px;
                    color: #bfdbfe;
                }
                .docx-stage {
                    padding-top: 8px;
                }
                .document-paper {
                    max-width: 980px;
                    margin: 0 auto;
                    border-radius: 26px;
                    padding: 26px 34px 30px;
                    background:
                        linear-gradient(180deg, rgba(15, 23, 42, 0.94), rgba(12, 18, 34, 0.9)),
                        radial-gradient(circle at top right, rgba(45, 212, 191, 0.12), transparent 30%);
                    color: #e5eefc;
                    border: 1px solid rgba(96, 165, 250, 0.16);
                    box-shadow:
                        0 28px 55px rgba(2, 6, 23, 0.34),
                        inset 0 1px 0 rgba(255, 255, 255, 0.04),
                        inset 0 0 0 1px rgba(15, 23, 42, 0.55);
                }
                .document-topline {
                    display: flex;
                    justify-content: space-between;
                    gap: 12px;
                    flex-wrap: wrap;
                    margin-bottom: 14px;
                    padding-bottom: 10px;
                    border-bottom: 1px solid rgba(148, 163, 184, 0.16);
                    font-size: 12px;
                    color: #8ba2c9;
                    letter-spacing: 0.04em;
                    text-transform: uppercase;
                }
                .editor-surface {
                    display: flex;
                    flex-direction: column;
                    gap: 4px;
                }
                .docx-block {
                    position: relative;
                }
                .docx-block.scoped-flow,
                .docx-block[data-block-type="image_meta"] {
                    margin-top: 12px;
                }
                .docx-block.compact-placeholder {
                    margin-top: 8px;
                }
                .block-header {
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    gap: 10px;
                    flex-wrap: wrap;
                    margin: 10px 0 6px;
                    padding: 0;
                    background: transparent;
                    border-bottom: none;
                    color: #8ba2c9;
                    font-size: 11px;
                }
                .block-title {
                    display: inline-flex;
                    align-items: center;
                    padding: 3px 10px;
                    border-radius: 999px;
                    background: rgba(59, 130, 246, 0.12);
                    color: #dbeafe;
                    font-weight: 700;
                    border: 1px solid rgba(96, 165, 250, 0.2);
                }
                .block-meta {
                    color: #7f93b7;
                }
                .paragraph-wrap {
                    padding: 0;
                }
                .docx-input {
                    width: 100%;
                    box-sizing: border-box;
                    outline: none;
                    resize: none;
                    overflow: hidden;
                    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
                    transition: background 0.18s ease, border-color 0.18s ease, box-shadow 0.18s ease;
                }
                .docx-input:focus {
                    box-shadow: none;
                }
                .docx-input[readonly] {
                    cursor: default;
                }
                .docx-paragraph-input {
                    min-height: 34px;
                    border: none;
                    border-radius: 14px;
                    background: transparent;
                    color: #edf4ff;
                    padding: 6px 8px;
                    line-height: 1.85;
                    font-size: 16px;
                }
                .docx-paragraph-input:focus {
                    background: rgba(59, 130, 246, 0.08);
                    box-shadow: inset 0 0 0 1px rgba(96, 165, 250, 0.22);
                }
                .table-wrap {
                    overflow: auto;
                    margin: 10px 0 16px;
                    border-radius: 18px;
                    border: 1px solid rgba(96, 165, 250, 0.18);
                    background: rgba(8, 15, 30, 0.72);
                    box-shadow: inset 0 1px 0 rgba(255, 255, 255, 0.03);
                }
                table {
                    width: 100%;
                    border-collapse: separate;
                    border-spacing: 0;
                }
                td {
                    min-width: 150px;
                    padding: 0;
                    vertical-align: top;
                    border-right: 1px solid rgba(148, 163, 184, 0.12);
                    border-bottom: 1px solid rgba(148, 163, 184, 0.12);
                    background: rgba(15, 23, 42, 0.56);
                }
                tr:last-child td {
                    border-bottom: none;
                }
                td:last-child {
                    border-right: none;
                }
                .cell-label {
                    padding: 8px 10px 0;
                    font-size: 11px;
                    color: #8ba2c9;
                    font-weight: 700;
                    letter-spacing: 0.04em;
                }
                .table-cell-input {
                    min-height: 52px;
                    border: none;
                    background: transparent;
                    color: #edf4ff;
                    padding: 8px 10px 12px;
                    line-height: 1.7;
                    font-size: 14px;
                    border-radius: 0;
                }
                .table-cell-input:focus {
                    background: rgba(59, 130, 246, 0.08);
                    box-shadow: inset 0 0 0 2px rgba(96, 165, 250, 0.16);
                }
                .docx-input.locked {
                    background: rgba(248, 113, 113, 0.08);
                    box-shadow: inset 0 0 0 2px rgba(248, 113, 113, 0.2);
                }
                .docx-input.remote-updated {
                    background: rgba(45, 212, 191, 0.08);
                }
                .empty-state {
                    padding: 28px 18px;
                    text-align: center;
                    color: #8ba2c9;
                    border: 1px dashed rgba(148, 163, 184, 0.22);
                    border-radius: 16px;
                    background: rgba(15, 23, 42, 0.45);
                }
                .docx-block.body-flow[data-block-type="paragraph"] .block-header {
                    display: none;
                }
                .docx-block.body-flow[data-block-type="paragraph"] + .docx-block.body-flow[data-block-type="paragraph"] {
                    margin-top: 2px;
                }
                .docx-block.body-flow[data-block-type="paragraph"] + .docx-block.body-flow[data-block-type="table"] {
                    margin-top: 12px;
                }
                .docx-block.body-flow[data-block-type="table"] + .docx-block.body-flow[data-block-type="paragraph"] {
                    margin-top: 12px;
                }
                .docx-block.compact-placeholder .paragraph-wrap {
                    max-width: 340px;
                }
                .docx-block.compact-placeholder .docx-paragraph-input {
                    min-height: 22px;
                    padding: 8px 10px;
                    line-height: 1.55;
                    font-size: 14px;
                    color: #dbeafe;
                    border-radius: 12px;
                    background: rgba(15, 23, 42, 0.52);
                    box-shadow: inset 0 0 0 1px rgba(96, 165, 250, 0.12);
                }
                .docx-block.compact-placeholder .docx-paragraph-input::placeholder {
                    color: #7f93b7;
                }
                .docx-block.compact-placeholder .block-meta {
                    color: #6f86ac;
                }
                .image-meta-card {
                    display: grid;
                    gap: 12px;
                    padding: 14px;
                    border-radius: 18px;
                    background: rgba(8, 15, 30, 0.52);
                    border: 1px solid rgba(96, 165, 250, 0.14);
                }
                .image-meta-field {
                    display: flex;
                    flex-direction: column;
                    gap: 6px;
                }
                .image-meta-label {
                    font-size: 12px;
                    color: #93c5fd;
                    letter-spacing: 0.04em;
                }
                .image-meta-title-input,
                .image-meta-description-input {
                    border: none;
                    border-radius: 14px;
                    background: rgba(15, 23, 42, 0.66);
                    color: #edf4ff;
                    padding: 10px 12px;
                    line-height: 1.7;
                }
                .image-meta-title-input:focus,
                .image-meta-description-input:focus {
                    background: rgba(59, 130, 246, 0.08);
                    box-shadow: inset 0 0 0 2px rgba(96, 165, 250, 0.16);
                }
                .status {
                    min-height: 22px;
                    margin-top: 14px;
                    color: #cbd5e1;
                    font-size: 14px;
                }
                .status.success { color: #86efac; }
                .status.error { color: #fca5a5; }
                .status.info { color: #cbd5e1; }
                @media (max-width: 768px) {
                    .page { padding: 14px; }
                    .header, .card { padding: 16px; border-radius: 16px; }
                    .title h1 { font-size: 22px; }
                    td { min-width: 120px; }
                    .document-paper { padding: 20px 16px 24px; }
                    .docx-paragraph-input { font-size: 15px; line-height: 1.8; padding-inline: 4px; }
                }
                @media (max-width: 960px) {
                    .document-paper {
                        border-radius: 18px;
                    }
                }
            </style>
        </head>
        <body>
            <div class="page">
                <div class="header">
                    <div class="title">
                        <h1>{{ filename }}</h1>
                        <p>{{ file_path }}</p>
                    </div>
                    <div class="actions">
                        <a href="{{ back_url }}" id="backToListBtn" class="btn btn-secondary">返回列表</a>
                        <a href="{{ url_for('download_file', filename=file_path) }}" class="btn btn-secondary">下载原文件</a>
                        <button id="saveBtn" class="btn btn-primary" {% if not editable or not can_edit %}disabled{% endif %}>保存修改</button>
                    </div>
                </div>
                <div class="card">
                    <div class="meta">
                        <span>大小: {{ file_size }}</span>
                        <span>可编辑块: {{ block_count }}</span>
                        <span>段落: {{ paragraph_count }}</span>
                        <span>表格: {{ table_count }}</span>
                        <span>页眉/页脚: {{ header_count }}/{{ footer_count }}</span>
                        <span>批注: {{ comment_count }}</span>
                        <span>文本框: {{ textbox_count }}</span>
                        <span>图片说明: {{ image_count }}</span>
                        <span>快捷键: Ctrl+S</span>
                    </div>
                    <div class="notice">{{ warning_message }}</div>
                    <div class="guide">正文仍然按连续文档的方式铺开显示；页眉页脚、批注、文本框和图片说明会带着上下文标签显示，方便多人协作时知道自己正在修改哪一块。</div>
                    <div class="collab-panel">
                        <div class="collab-users" id="collabUsers">协同连接中...</div>
                        <div class="collab-hint" id="collabHint">正在同步在线成员、段落锁和远端修改。</div>
                    </div>
                    <div class="docx-stage">
                        <div class="document-paper">
                            <div class="document-topline">
                                <span>连续文档视图</span>
                                <span>多人协作已开启</span>
                            </div>
                            <div class="editor-surface" id="docxEditor">
                        {% if blocks %}
                            {% for block in blocks %}
                                {% if block.type == 'paragraph' %}
                                <section class="docx-block {% if block.story_kind == 'body' %}body-flow{% else %}scoped-flow{% endif %}{% if block.is_placeholder %} compact-placeholder{% endif %}" data-block-id="{{ block.id }}" data-block-type="paragraph" data-block-label="{{ block.target_label }}" data-story-kind="{{ block.story_kind }}">
                                    <div class="block-header">
                                        <span class="block-title">{{ block.scope_label }}</span>
                                        <span class="block-meta">{% if block.is_placeholder %}当前为空{% else %}{{ block.label }}{% if block.style %} · {{ block.style }}{% endif %}{% endif %}{% if block.scope_meta %} · {{ block.scope_meta }}{% endif %}</span>
                                    </div>
                                    <div class="paragraph-wrap">
                                        <textarea class="docx-input docx-paragraph-input" spellcheck="false" placeholder="{{ block.placeholder_text }}" data-block-id="{{ block.id }}" data-target-key="paragraph::{{ block.id }}" data-target-kind="docx_paragraph" data-target-label="{{ block.target_label }}" data-original-text="{{ block.text|e }}" {% if not editable or not can_edit %}readonly{% endif %}>{{ block.text }}</textarea>
                                    </div>
                                </section>
                                {% elif block.type == 'table' %}
                                <section class="docx-block {% if block.story_kind == 'body' %}body-flow{% else %}scoped-flow{% endif %}" data-block-id="{{ block.id }}" data-block-type="table" data-block-label="{{ block.target_label }}" data-story-kind="{{ block.story_kind }}">
                                    <div class="block-header">
                                        <span class="block-title">{{ block.scope_label }}</span>
                                        <span class="block-meta">{{ block.label }} · {{ block.row_count }} 行 / {{ block.col_count }} 列{% if block.style %} · {{ block.style }}{% endif %}{% if block.scope_meta %} · {{ block.scope_meta }}{% endif %}</span>
                                    </div>
                                    <div class="table-wrap">
                                        <table>
                                            <tbody>
                                                {% for row in block.rows %}
                                                <tr class="docx-table-row">
                                                    {% for cell in row %}
                                                    <td>
                                                        <div class="cell-label">{{ cell.label }}</div>
                                                        <textarea class="docx-input table-cell-input docx-table-cell-input" spellcheck="false" data-block-id="{{ block.id }}" data-row="{{ cell.row }}" data-col="{{ cell.col }}" data-cell-label="{{ cell.label }}" data-target-key="table_cell::{{ block.id }}::{{ cell.row }}::{{ cell.col }}" data-target-kind="docx_table_cell" data-target-label="{{ block.target_label }} / {{ cell.label }}" data-original-text="{{ cell.text|e }}" {% if not editable or not can_edit %}readonly{% endif %}>{{ cell.text }}</textarea>
                                                    </td>
                                                    {% endfor %}
                                                </tr>
                                                {% endfor %}
                                            </tbody>
                                        </table>
                                    </div>
                                </section>
                                {% elif block.type == 'image_meta' %}
                                <section class="docx-block scoped-flow" data-block-id="{{ block.id }}" data-block-type="image_meta" data-block-label="{{ block.target_label }}" data-story-kind="{{ block.story_kind }}">
                                    <div class="block-header">
                                        <span class="block-title">{{ block.scope_label }}</span>
                                        <span class="block-meta">{{ block.label }}{% if block.name %} · {{ block.name }}{% endif %}{% if block.scope_meta %} · {{ block.scope_meta }}{% endif %}</span>
                                    </div>
                                    <div class="image-meta-card">
                                        <div class="image-meta-field">
                                            <div class="image-meta-label">图片标题</div>
                                            <textarea class="docx-input image-meta-title-input" spellcheck="false" data-block-id="{{ block.id }}" data-target-key="image_meta::{{ block.id }}::title" data-target-kind="docx_image_meta" data-target-field="title" data-target-label="{{ block.target_label }} / 标题" data-original-text="{{ block.title|e }}" {% if not editable or not can_edit %}readonly{% endif %}>{{ block.title }}</textarea>
                                        </div>
                                        <div class="image-meta-field">
                                            <div class="image-meta-label">图片说明</div>
                                            <textarea class="docx-input image-meta-description-input" spellcheck="false" data-block-id="{{ block.id }}" data-target-key="image_meta::{{ block.id }}::description" data-target-kind="docx_image_meta" data-target-field="description" data-target-label="{{ block.target_label }} / 说明" data-original-text="{{ block.description|e }}" {% if not editable or not can_edit %}readonly{% endif %}>{{ block.description }}</textarea>
                                        </div>
                                    </div>
                                </section>
                                {% endif %}
                            {% endfor %}
                        {% else %}
                            <div class="empty-state">这个 DOCX 正文里没有可直接编辑的段落或表格内容。</div>
                        {% endif %}
                            </div>
                        </div>
                    </div>
                    <div class="status info" id="status">{{ initial_status }}</div>
                </div>
            </div>
            <script>
                const state = {
                    filePath: {{ file_path|tojson }},
                    mtimeNs: {{ mtime_ns|tojson }},
                    editable: {{ 'true' if editable else 'false' }},
                    canEdit: {{ 'true' if can_edit else 'false' }},
                    collabClientId: {{ collab_client_id|tojson }},
                    collabUsername: {{ collab_username|tojson }},
                    collabPollInterval: {{ collab_poll_interval_ms }},
                    collabEditors: [],
                    collabLockMap: new Map(),
                    collabOwnLockKey: '',
                    collabLockTarget: null,
                    activeTarget: null,
                    collabSyncInFlight: false,
                    snapshotInFlight: false
                };

                const saveBtn = document.getElementById('saveBtn');
                const backToListBtn = document.getElementById('backToListBtn');
                const statusEl = document.getElementById('status');
                const collabUsersEl = document.getElementById('collabUsers');
                const collabHintEl = document.getElementById('collabHint');
                const dirtyTargets = new Set();
                let isSaving = false;
                let hasPendingChanges = false;
                let remoteCheckTimer = null;
                let focusReleaseTimer = null;
                let suppressBeforeUnload = false;

                function setStatus(message, type = 'info') {
                    statusEl.textContent = message;
                    statusEl.className = `status ${type}`;
                }

                function escapeHtml(value) {
                    return String(value ?? '')
                        .replace(/&/g, '&amp;')
                        .replace(/</g, '&lt;')
                        .replace(/>/g, '&gt;')
                        .replace(/"/g, '&quot;')
                        .replace(/'/g, '&#39;');
                }

                function setCollabHint(message) {
                    if (collabHintEl) {
                        collabHintEl.textContent = message;
                    }
                }

                function autosizeTextarea(textarea) {
                    if (!(textarea instanceof HTMLTextAreaElement)) {
                        return;
                    }
                    textarea.style.height = 'auto';
                    let minHeight = 34;
                    if (textarea.closest('.compact-placeholder')) {
                        minHeight = 22;
                    } else if (textarea.classList.contains('table-cell-input')) {
                        minHeight = 52;
                    } else if (textarea.classList.contains('image-meta-description-input')) {
                        minHeight = 64;
                    }
                    textarea.style.height = `${Math.max(textarea.scrollHeight, minHeight)}px`;
                }

                function autosizeAllTextareas() {
                    document.querySelectorAll('.docx-input').forEach(autosizeTextarea);
                }

                function getTargetKey(input) {
                    return input instanceof HTMLElement ? String(input.dataset.targetKey || '') : '';
                }

                function buildTargetFromInput(input) {
                    if (!(input instanceof HTMLElement)) {
                        return null;
                    }
                    const kind = String(input.dataset.targetKind || '').trim();
                    const blockId = String(input.dataset.blockId || '').trim();
                    const label = String(input.dataset.targetLabel || '').trim();
                    if (!kind || !blockId) {
                        return null;
                    }
                    if (kind === 'docx_paragraph') {
                        return { kind, block_id: blockId, label };
                    }
                    if (kind === 'docx_image_meta') {
                        return {
                            kind,
                            block_id: blockId,
                            field: String(input.dataset.targetField || '').trim(),
                            label
                        };
                    }
                    if (kind !== 'docx_table_cell') {
                        return null;
                    }
                    return {
                        kind,
                        block_id: blockId,
                        row: Number(input.dataset.row),
                        col: Number(input.dataset.col),
                        label
                    };
                }

                function getLockInfo(input) {
                    return state.collabLockMap.get(getTargetKey(input)) || null;
                }

                function applyCollaborationLocks() {
                    const forceReadonly = !state.editable || !state.canEdit;
                    document.querySelectorAll('.docx-input').forEach(input => {
                        const lockInfo = getLockInfo(input);
                        const lockedByOther = Boolean(lockInfo && lockInfo.client_id !== state.collabClientId);
                        input.readOnly = forceReadonly || lockedByOther;
                        input.classList.toggle('locked', lockedByOther);
                        if (lockedByOther) {
                            input.title = `${lockInfo.username} 正在编辑 ${lockInfo.label || '当前内容'}`;
                            if (document.activeElement === input) {
                                input.blur();
                            }
                        } else {
                            input.removeAttribute('title');
                        }
                    });
                    if (saveBtn) {
                        saveBtn.disabled = forceReadonly || isSaving;
                    }
                }

                function renderCollaborationUsers() {
                    if (!collabUsersEl) {
                        return;
                    }

                    const editors = Array.isArray(state.collabEditors) ? state.collabEditors : [];
                    if (editors.length === 0) {
                        collabUsersEl.innerHTML = '<span class="collab-user self">只有你</span>';
                        setCollabHint('当前没有其他在线协作者。');
                        return;
                    }

                    collabUsersEl.innerHTML = editors.map(editor => `
                        <span class="collab-user${editor.is_self ? ' self' : ''}">
                            <strong>${escapeHtml(editor.username || '匿名用户')}</strong>
                            <span>${escapeHtml(editor.active_target_label || '正在查看')}</span>
                        </span>
                    `).join('');

                    const ownLock = state.collabOwnLockKey ? state.collabLockMap.get(state.collabOwnLockKey) : null;
                    if (ownLock) {
                        setCollabHint(`你当前锁定了 ${ownLock.label || '当前内容'}，其他人会看到这一块暂时只读。`);
                    } else if (editors.length > 1) {
                        setCollabHint('支持多人同时修改不同段落、表格单元格或图片说明，其他人保存后会自动同步到这一页。');
                    } else {
                        setCollabHint('当前没有其他在线协作者。');
                    }
                }

                function updateDirtyState(input) {
                    if (!(input instanceof HTMLTextAreaElement) || !state.editable || !state.canEdit) {
                        return;
                    }
                    const targetKey = getTargetKey(input);
                    if (!targetKey) {
                        return;
                    }
                    const originalValue = String(input.dataset.originalText || '');
                    if ((input.value || '') === originalValue) {
                        dirtyTargets.delete(targetKey);
                    } else {
                        dirtyTargets.add(targetKey);
                    }
                    hasPendingChanges = dirtyTargets.size > 0;
                    if (hasPendingChanges) {
                        setStatus('文档已修改，按 Ctrl+S 可直接保存到共享目录。', 'info');
                    }
                }

                function collectDocBlocks() {
                    return Array.from(document.querySelectorAll('.docx-block')).map(block => {
                        const blockId = String(block.dataset.blockId || '');
                        const blockType = String(block.dataset.blockType || '');
                        const blockLabel = String(block.dataset.blockLabel || blockId);
                        if (blockType === 'paragraph') {
                            const input = block.querySelector('.docx-paragraph-input');
                            return {
                                id: blockId,
                                type: blockType,
                                label: blockLabel,
                                text: input?.value || '',
                                original_text: input?.dataset.originalText || ''
                            };
                        }

                        if (blockType === 'image_meta') {
                            const titleInput = block.querySelector('.image-meta-title-input');
                            const descriptionInput = block.querySelector('.image-meta-description-input');
                            return {
                                id: blockId,
                                type: blockType,
                                label: blockLabel,
                                title: titleInput?.value || '',
                                original_title: titleInput?.dataset.originalText || '',
                                description: descriptionInput?.value || '',
                                original_description: descriptionInput?.dataset.originalText || ''
                            };
                        }

                        const rows = Array.from(block.querySelectorAll('.docx-table-row')).map(row => (
                            Array.from(row.querySelectorAll('.docx-table-cell-input')).map(input => ({
                                row: Number(input.dataset.row),
                                col: Number(input.dataset.col),
                                label: input.dataset.cellLabel || '',
                                text: input.value || '',
                                original_text: input.dataset.originalText || ''
                            }))
                        ));

                        return {
                            id: blockId,
                            type: blockType,
                            label: blockLabel,
                            rows
                        };
                    });
                }

                function updateOriginalValuesFromCurrent() {
                    document.querySelectorAll('.docx-input').forEach(input => {
                        input.dataset.originalText = input.value || '';
                        input.classList.remove('remote-updated');
                    });
                    dirtyTargets.clear();
                    hasPendingChanges = false;
                }

                async function refreshRemoteDocx() {
                    if (!state.filePath || state.snapshotInFlight) {
                        return;
                    }

                    state.snapshotInFlight = true;
                    try {
                        const response = await fetch('/api/docx_file_snapshot', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ file_path: state.filePath })
                        });
                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            return;
                        }

                        const blockMap = new Map((data.blocks || []).map(block => [block.id, block]));
                        let updatedCount = 0;
                        let skippedCount = 0;

                        document.querySelectorAll('.docx-block').forEach(block => {
                            const blockId = String(block.dataset.blockId || '');
                            const blockType = String(block.dataset.blockType || '');
                            const remoteBlock = blockMap.get(blockId);
                            if (!remoteBlock) {
                                return;
                            }

                            if (blockType === 'paragraph') {
                                const input = block.querySelector('.docx-paragraph-input');
                                if (!(input instanceof HTMLTextAreaElement)) {
                                    return;
                                }
                                const targetKey = getTargetKey(input);
                                if (dirtyTargets.has(targetKey)) {
                                    skippedCount += 1;
                                    return;
                                }
                                const nextText = String(remoteBlock.text || '');
                                if (input.value !== nextText) {
                                    input.value = nextText;
                                    input.classList.add('remote-updated');
                                    autosizeTextarea(input);
                                    updatedCount += 1;
                                }
                                input.dataset.originalText = nextText;
                                return;
                            }

                            if (blockType === 'image_meta') {
                                const imageInputs = Array.from(block.querySelectorAll('.image-meta-title-input, .image-meta-description-input'));
                                imageInputs.forEach(input => {
                                    if (!(input instanceof HTMLTextAreaElement)) {
                                        return;
                                    }
                                    const targetKey = getTargetKey(input);
                                    if (dirtyTargets.has(targetKey)) {
                                        skippedCount += 1;
                                        return;
                                    }

                                    const fieldName = String(input.dataset.targetField || '').trim();
                                    const nextText = fieldName === 'title'
                                        ? String(remoteBlock.title || '')
                                        : String(remoteBlock.description || '');
                                    if (input.value !== nextText) {
                                        input.value = nextText;
                                        input.classList.add('remote-updated');
                                        autosizeTextarea(input);
                                        updatedCount += 1;
                                    }
                                    input.dataset.originalText = nextText;
                                });
                                return;
                            }

                            block.querySelectorAll('.docx-table-cell-input').forEach(input => {
                                const rowIndex = Number(input.dataset.row);
                                const colIndex = Number(input.dataset.col);
                                const remoteCell = (((remoteBlock.rows || [])[rowIndex] || [])[colIndex]) || null;
                                if (!remoteCell) {
                                    return;
                                }
                                const targetKey = getTargetKey(input);
                                if (dirtyTargets.has(targetKey)) {
                                    skippedCount += 1;
                                    return;
                                }
                                const nextText = String(remoteCell.text || '');
                                if (input.value !== nextText) {
                                    input.value = nextText;
                                    input.classList.add('remote-updated');
                                    autosizeTextarea(input);
                                    updatedCount += 1;
                                }
                                input.dataset.originalText = nextText;
                            });
                        });

                        state.mtimeNs = data.mtime_ns || state.mtimeNs;
                        if (updatedCount > 0 && skippedCount > 0) {
                            setStatus(`已同步其他人保存的最新内容，刷新了 ${updatedCount} 处；你本地还有 ${skippedCount} 处未保存修改，已保留本地改动。`, 'info');
                        } else if (updatedCount > 0) {
                            setStatus(`已同步其他人保存的最新内容，共刷新 ${updatedCount} 处。`, 'info');
                        } else if (skippedCount > 0) {
                            setStatus('检测到其他人保存了文档，但你当前也有未保存修改，已保留本地改动。', 'error');
                        }
                    } catch (error) {
                    } finally {
                        state.snapshotInFlight = false;
                    }
                }

                async function syncCollaborationState(options = {}) {
                    const { quiet = true, releaseLock = false } = options;
                    if (!state.filePath || state.collabSyncInFlight) {
                        return;
                    }

                    state.collabSyncInFlight = true;
                    try {
                        const response = await fetch('/api/document_collaboration_sync', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                editor_type: 'docx',
                                active_target: state.activeTarget,
                                lock_target: releaseLock ? null : state.collabLockTarget,
                                release_lock: releaseLock || !state.collabLockTarget
                            })
                        });
                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (!quiet) {
                                setStatus(data.message || '协同状态同步失败。', 'error');
                            }
                            return;
                        }

                        if (data.client_id) {
                            state.collabClientId = data.client_id;
                        }
                        state.collabEditors = Array.isArray(data.editors) ? data.editors : [];
                        state.collabLockMap = new Map();
                        (data.locks || []).forEach(lock => {
                            state.collabLockMap.set(lock.key, lock);
                        });
                        state.collabOwnLockKey = data.own_lock ? data.own_lock.key : '';

                        renderCollaborationUsers();
                        applyCollaborationLocks();

                        if (data.lock_denied) {
                            state.collabLockTarget = null;
                            applyCollaborationLocks();
                            setStatus(`${data.lock_denied.username} 正在编辑 ${data.lock_denied.label || '当前内容'}，请先换一个位置再试。`, 'error');
                        }

                        if (data.mtime_ns && data.mtime_ns !== state.mtimeNs) {
                            refreshRemoteDocx();
                        }
                    } catch (error) {
                        if (!quiet) {
                            setStatus('协同状态同步失败：网络异常或服务器不可用。', 'error');
                        }
                    } finally {
                        state.collabSyncInFlight = false;
                    }
                }

                function releaseCollaborationState(removeSession = false) {
                    if (!state.filePath) {
                        return;
                    }

                    const payload = JSON.stringify({
                        file_path: state.filePath,
                        remove_session: removeSession
                    });
                    if (navigator.sendBeacon) {
                        const blob = new Blob([payload], { type: 'application/json' });
                        navigator.sendBeacon('/api/document_collaboration_release', blob);
                        return;
                    }

                    fetch('/api/document_collaboration_release', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: payload,
                        keepalive: true
                    }).catch(() => {});
                }

                async function saveDocxFile() {
                    if (isSaving) {
                        return;
                    }
                    if (!state.editable) {
                        setStatus('当前文件过大，当前仅支持只读预览。', 'error');
                        return;
                    }
                    if (!state.canEdit) {
                        setStatus('只有管理员才能保存文件修改。', 'error');
                        return;
                    }

                    isSaving = true;
                    if (saveBtn) {
                        saveBtn.disabled = true;
                    }
                    setStatus('正在保存 Word 文档到共享目录...', 'info');

                    try {
                        const response = await fetch('/api/save_docx_file', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                mtime_ns: state.mtimeNs,
                                blocks: collectDocBlocks()
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (data.mtime_ns) {
                                state.mtimeNs = data.mtime_ns;
                            }
                            if (response.status === 409) {
                                refreshRemoteDocx();
                            }
                            setStatus(data.message || '保存失败，请稍后重试。', 'error');
                            return;
                        }

                        state.mtimeNs = data.mtime_ns || state.mtimeNs;
                        updateOriginalValuesFromCurrent();
                        setStatus(`已保存到共享目录（段落 ${data.changed_paragraphs || 0} 处，表格单元格 ${data.changed_cells || 0} 处，图片说明 ${data.changed_image_notes || 0} 处）。`, 'success');
                        syncCollaborationState({ quiet: true });
                    } catch (error) {
                        setStatus('保存失败：网络异常或服务器不可用。', 'error');
                    } finally {
                        isSaving = false;
                        applyCollaborationLocks();
                    }
                }

                function returnToList(event) {
                    if (!backToListBtn) {
                        return;
                    }

                    const backUrl = backToListBtn.href;
                    if (!window.opener || window.opener.closed) {
                        return;
                    }

                    event.preventDefault();
                    suppressBeforeUnload = true;
                    try {
                        window.opener.location.href = backUrl;
                        if (typeof window.opener.focus === 'function') {
                            window.opener.focus();
                        }
                    } catch (error) {
                    }

                    releaseCollaborationState(true);
                    window.close();
                    window.setTimeout(function() {
                        window.location.href = backUrl;
                    }, 120);
                }

                document.addEventListener('input', function(event) {
                    const target = event.target;
                    if (target instanceof HTMLTextAreaElement && target.classList.contains('docx-input')) {
                        autosizeTextarea(target);
                        updateDirtyState(target);
                    }
                });

                document.addEventListener('focusin', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLTextAreaElement) || !target.classList.contains('docx-input')) {
                        return;
                    }
                    if (focusReleaseTimer) {
                        window.clearTimeout(focusReleaseTimer);
                        focusReleaseTimer = null;
                    }
                    state.activeTarget = buildTargetFromInput(target);
                    if (state.editable && state.canEdit) {
                        state.collabLockTarget = state.activeTarget;
                    }
                    syncCollaborationState({ quiet: true });
                });

                document.addEventListener('focusout', function(event) {
                    const target = event.target;
                    if (!(target instanceof HTMLTextAreaElement) || !target.classList.contains('docx-input')) {
                        return;
                    }
                    focusReleaseTimer = window.setTimeout(function() {
                        const activeElement = document.activeElement;
                        if (!(activeElement instanceof HTMLTextAreaElement) || !activeElement.classList.contains('docx-input')) {
                            state.activeTarget = null;
                            state.collabLockTarget = null;
                            syncCollaborationState({ quiet: true, releaseLock: true });
                        }
                    }, 120);
                });

                if (saveBtn) {
                    saveBtn.addEventListener('click', saveDocxFile);
                }
                if (backToListBtn) {
                    backToListBtn.addEventListener('click', returnToList);
                }

                document.addEventListener('keydown', function(event) {
                    if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 's') {
                        event.preventDefault();
                        saveDocxFile();
                    }
                });

                window.addEventListener('beforeunload', function(event) {
                    if (!suppressBeforeUnload && hasPendingChanges) {
                        event.preventDefault();
                        event.returnValue = '';
                    }
                });

                window.addEventListener('pagehide', function() {
                    releaseCollaborationState(true);
                });

                document.addEventListener('visibilitychange', function() {
                    if (!document.hidden) {
                        syncCollaborationState({ quiet: true });
                    }
                });

                autosizeAllTextareas();
                applyCollaborationLocks();
                if (state.filePath) {
                    syncCollaborationState({ quiet: true });
                    remoteCheckTimer = window.setInterval(function() {
                        syncCollaborationState({ quiet: true });
                    }, state.collabPollInterval);
                }
            </script>
        </body>
        </html>
        ''',
        filename=os.path.basename(filename),
        file_path=filename,
        back_url=back_url,
        editable=preview['editable'],
        can_edit=can_edit,
        mtime_ns=str(preview['mtime_ns']),
        file_size=get_file_size(preview['file_size']),
        paragraph_count=preview['paragraph_count'],
        table_count=preview['table_count'],
        header_count=preview['header_count'],
        footer_count=preview['footer_count'],
        comment_count=preview['comment_count'],
        textbox_count=preview['textbox_count'],
        image_count=preview['image_count'],
        block_count=preview['block_count'],
        blocks=preview['blocks'],
        collab_client_id=get_document_editor_client_id(),
        collab_username=get_current_username(),
        collab_poll_interval_ms=DOCUMENT_COLLAB_POLL_INTERVAL_MS,
        warning_message=warning_message,
        initial_status=initial_status
        )
    except Exception as e:
        print(f"打开 DOCX 编辑器失败: {str(e)}")
        flash(f'打开 DOCX 编辑器失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/edit/<path:filename>')
def edit_text_file(filename):
    """在线编辑文本文件。"""
    try:
        success, filepath, error = safe_join_path(app.config['UPLOAD_FOLDER'], filename)
        if not success:
            flash(f'路径错误: {error}', 'danger')
            return redirect(url_for('index'))

        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            flash('文件不存在', 'danger')
            return redirect(url_for('index'))

        if not is_text_previewable_file(filename):
            return redirect(url_for('preview_file', filename=filename))

        preview = load_text_file_preview(filepath)
        can_edit = user_can_edit_files()
        parent_path = os.path.dirname(filename).replace('\\', '/')
        back_url = url_for('index', subpath=parent_path) if parent_path else url_for('index')

        warning_message = None
        if not preview['editable']:
            warning_message = (
                f'当前文件大小为 {get_file_size(preview["file_size"])}，'
                f'仅显示前 {get_file_size(TEXT_PREVIEW_FALLBACK_BYTES)}，暂不支持在线编辑。'
            )
        elif not can_edit:
            warning_message = '当前昵称可以预览此文件，但只有管理员才能保存修改。'

        initial_status = '可直接编辑并保存到共享目录，快捷键：Ctrl+S。'
        if not can_edit:
            initial_status = '当前为只读模式，只有管理员才能保存修改。'
        if not preview['editable']:
            initial_status = '文件过大，当前仅支持只读预览。'

        return render_template_string('''
        <!DOCTYPE html>
        <html lang="zh-CN">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>预览/编辑: {{ filename }}</title>
            <style>
                :root {
                    color-scheme: dark;
                }
                body {
                    margin: 0;
                    min-height: 100vh;
                    font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
                    background:
                        radial-gradient(circle at top left, rgba(14, 165, 233, 0.18), transparent 30%),
                        radial-gradient(circle at top right, rgba(59, 130, 246, 0.16), transparent 35%),
                        #0f172a;
                    color: #e2e8f0;
                }
                .page {
                    max-width: 1400px;
                    margin: 0 auto;
                    padding: 24px;
                }
                .header {
                    background: rgba(15, 23, 42, 0.82);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    backdrop-filter: blur(14px);
                    padding: 22px 24px;
                    border-radius: 20px;
                    margin-bottom: 18px;
                    display: flex;
                    justify-content: space-between;
                    align-items: center;
                    gap: 16px;
                    flex-wrap: wrap;
                }
                .title h1 {
                    margin: 0 0 8px;
                    font-size: 28px;
                    color: #f8fafc;
                }
                .title p {
                    margin: 0;
                    color: #94a3b8;
                    word-break: break-all;
                }
                .actions {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                }
                .btn {
                    border: none;
                    border-radius: 12px;
                    padding: 12px 18px;
                    font-size: 14px;
                    font-weight: 600;
                    cursor: pointer;
                    text-decoration: none;
                    transition: transform 0.2s ease, opacity 0.2s ease, background 0.2s ease;
                }
                .btn:hover {
                    transform: translateY(-1px);
                }
                .btn:disabled {
                    cursor: not-allowed;
                    opacity: 0.55;
                    transform: none;
                }
                .btn-primary {
                    background: linear-gradient(135deg, #0ea5e9, #2563eb);
                    color: white;
                }
                .btn-secondary {
                    background: rgba(30, 41, 59, 0.95);
                    color: #e2e8f0;
                    border: 1px solid rgba(148, 163, 184, 0.18);
                }
                .card {
                    background: rgba(15, 23, 42, 0.82);
                    border: 1px solid rgba(148, 163, 184, 0.18);
                    backdrop-filter: blur(14px);
                    border-radius: 20px;
                    padding: 22px;
                    box-shadow: 0 20px 45px rgba(2, 6, 23, 0.28);
                }
                .meta {
                    display: flex;
                    gap: 16px;
                    flex-wrap: wrap;
                    margin-bottom: 14px;
                    font-size: 13px;
                    color: #94a3b8;
                }
                .notice {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(245, 158, 11, 0.16);
                    border: 1px solid rgba(245, 158, 11, 0.28);
                    color: #fde68a;
                    line-height: 1.6;
                }
                .guide {
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(14, 165, 233, 0.12);
                    border: 1px solid rgba(14, 165, 233, 0.24);
                    color: #bae6fd;
                    line-height: 1.6;
                }
                .collab-panel {
                    display: flex;
                    gap: 12px;
                    flex-wrap: wrap;
                    align-items: center;
                    justify-content: space-between;
                    margin-bottom: 14px;
                    padding: 12px 14px;
                    border-radius: 12px;
                    background: rgba(34, 197, 94, 0.1);
                    border: 1px solid rgba(34, 197, 94, 0.22);
                    color: #dcfce7;
                }
                .collab-users {
                    display: flex;
                    gap: 8px;
                    flex-wrap: wrap;
                    align-items: center;
                }
                .collab-user {
                    display: inline-flex;
                    align-items: center;
                    gap: 6px;
                    padding: 6px 10px;
                    border-radius: 999px;
                    background: rgba(15, 23, 42, 0.72);
                    border: 1px solid rgba(34, 197, 94, 0.18);
                    font-size: 12px;
                    color: #f0fdf4;
                }
                .collab-user.self {
                    border-color: rgba(59, 130, 246, 0.4);
                    background: rgba(30, 41, 59, 0.95);
                }
                .collab-hint {
                    font-size: 12px;
                    color: #bbf7d0;
                }
                textarea {
                    width: 100%;
                    min-height: calc(100vh - 280px);
                    border-radius: 16px;
                    border: 1px solid rgba(51, 65, 85, 0.95);
                    background: #020617;
                    color: #e2e8f0;
                    padding: 18px;
                    box-sizing: border-box;
                    font-family: 'Cascadia Code', 'Consolas', 'Monaco', monospace;
                    font-size: 14px;
                    line-height: 1.65;
                    resize: vertical;
                    outline: none;
                }
                textarea[readonly] {
                    opacity: 0.9;
                }
                textarea.locked {
                    border-color: rgba(248, 113, 113, 0.42);
                    box-shadow: 0 0 0 2px rgba(248, 113, 113, 0.12);
                }
                .status {
                    min-height: 22px;
                    margin-top: 12px;
                    color: #cbd5e1;
                    font-size: 14px;
                }
                .status.success {
                    color: #86efac;
                }
                .status.error {
                    color: #fca5a5;
                }
                .status.info {
                    color: #cbd5e1;
                }
                @media (max-width: 768px) {
                    .page {
                        padding: 14px;
                    }
                    .header, .card {
                        padding: 16px;
                        border-radius: 16px;
                    }
                    .title h1 {
                        font-size: 22px;
                    }
                    textarea {
                        min-height: calc(100vh - 240px);
                        padding: 14px;
                    }
                }
            </style>
        </head>
        <body>
            <div class="page">
                <div class="header">
                    <div class="title">
                        <h1>{{ filename }}</h1>
                        <p>{{ file_path }}</p>
                    </div>
                    <div class="actions">
                        <a href="{{ back_url }}" id="backToListBtn" class="btn btn-secondary">返回列表</a>
                        <a href="{{ url_for('download_file', filename=file_path) }}" class="btn btn-secondary">下载原文件</a>
                        <button id="saveBtn" class="btn btn-primary" {% if not editable or not can_edit %}disabled{% endif %}>保存修改</button>
                    </div>
                </div>
                <div class="card">
                    <div class="meta">
                        <span>大小: {{ file_size }}</span>
                        <span>编码: <span id="encodingText">{{ encoding }}</span></span>
                        <span>快捷键: Ctrl+S</span>
                    </div>
                    {% if warning_message %}
                    <div class="notice">{{ warning_message }}</div>
                    {% endif %}
                    <div class="guide">多人可以同时输入，系统会自动推送最新内容，并尽量合并不同位置的修改。</div>
                    <div class="collab-panel">
                        <div class="collab-users" id="collabUsers">协同连接中...</div>
                        <div class="collab-hint" id="collabHint">正在连接实时协作通道并同步在线成员。</div>
                    </div>
                    <textarea id="editor" spellcheck="false" {% if not editable or not can_edit %}readonly{% endif %}>{{ content }}</textarea>
                    <div class="status info" id="status">{{ initial_status }}</div>
                </div>
            </div>
            <script>
                const state = {
                    filePath: {{ file_path|tojson }},
                    encoding: {{ encoding|tojson }},
                    newline: {{ newline|tojson }},
                    mtimeNs: {{ mtime_ns|tojson }},
                    editable: {{ 'true' if editable else 'false' }},
                    canEdit: {{ 'true' if can_edit else 'false' }},
                    collabClientId: {{ collab_client_id|tojson }},
                    collabUsername: {{ collab_username|tojson }},
                    collabPollInterval: {{ collab_poll_interval_ms }},
                    collabEditors: [],
                    collabSyncInFlight: false,
                    snapshotInFlight: false,
                    realtimeRevision: 0,
                    syncedContent: '',
                    realtimeEventSource: null,
                    realtimeSendTimer: null,
                    realtimeSendInFlight: false,
                    realtimeSendPromise: null,
                    realtimeResendAfterFlight: false,
                    realtimeConnected: false,
                    realtimeDebounce: {{ text_realtime_debounce_ms }},
                    collabSyncTimer: null
                };

                const editor = document.getElementById('editor');
                const saveBtn = document.getElementById('saveBtn');
                const backToListBtn = document.getElementById('backToListBtn');
                const statusEl = document.getElementById('status');
                const encodingText = document.getElementById('encodingText');
                const collabUsersEl = document.getElementById('collabUsers');
                const collabHintEl = document.getElementById('collabHint');
                let isSaving = false;
                let hasPendingChanges = false;
                let remoteCheckTimer = null;
                let suppressBeforeUnload = false;
                state.syncedContent = editor.value;

                function setStatus(message, type = 'info') {
                    statusEl.textContent = message;
                    statusEl.className = `status ${type}`;
                }

                function escapeHtml(value) {
                    return String(value ?? '')
                        .replace(/&/g, '&amp;')
                        .replace(/</g, '&lt;')
                        .replace(/>/g, '&gt;')
                        .replace(/"/g, '&quot;')
                        .replace(/'/g, '&#39;');
                }

                function setCollabHint(message) {
                    if (collabHintEl) {
                        collabHintEl.textContent = message;
                    }
                }

                function commonPrefixLength(left, right) {
                    const maxLength = Math.min(left.length, right.length);
                    let index = 0;
                    while (index < maxLength && left[index] === right[index]) {
                        index += 1;
                    }
                    return index;
                }

                function replaceEditorValue(nextValue, preserveSelection = true) {
                    const previousValue = editor.value;
                    const previousStart = Number(editor.selectionStart || 0);
                    const previousEnd = Number(editor.selectionEnd || 0);
                    editor.value = nextValue;

                    if (!preserveSelection) {
                        return;
                    }

                    const sharedPrefix = commonPrefixLength(previousValue, nextValue);
                    const lengthDelta = nextValue.length - previousValue.length;
                    const nextStart = previousStart <= sharedPrefix
                        ? previousStart
                        : Math.max(sharedPrefix, Math.min(nextValue.length, previousStart + lengthDelta));
                    const nextEnd = previousEnd <= sharedPrefix
                        ? previousEnd
                        : Math.max(sharedPrefix, Math.min(nextValue.length, previousEnd + lengthDelta));

                    try {
                        editor.setSelectionRange(nextStart, nextEnd);
                    } catch (error) {
                    }
                }

                function buildActiveTarget() {
                    const cursorIndex = Number(editor.selectionStart || 0);
                    const textBeforeCursor = editor.value.slice(0, cursorIndex);
                    const lines = textBeforeCursor.split('\\n');
                    const line = lines.length;
                    const column = (lines[lines.length - 1] || '').length + 1;
                    return {
                        kind: 'text_document',
                        label: `全文 / 第 ${line} 行 第 ${column} 列`
                    };
                }

                function renderCollaborationUsers() {
                    if (!collabUsersEl) {
                        return;
                    }

                    const editors = Array.isArray(state.collabEditors) ? state.collabEditors : [];
                    if (editors.length === 0) {
                        collabUsersEl.innerHTML = '<span class="collab-user self">只有你</span>';
                        setCollabHint('当前没有其他在线协作者。');
                        return;
                    }

                    collabUsersEl.innerHTML = editors.map(editorItem => `
                        <span class="collab-user${editorItem.is_self ? ' self' : ''}">
                            <strong>${escapeHtml(editorItem.username || '匿名用户')}</strong>
                            <span>${escapeHtml(editorItem.active_target_label || '正在查看')}</span>
                        </span>
                    `).join('');

                    if (state.realtimeConnected && editors.length > 1) {
                        setCollabHint('实时协作已连接，大家的改动会自动推送并尽量合并。');
                    } else if (state.realtimeConnected) {
                        setCollabHint('实时协作已连接，当前只有你在编辑。');
                    } else if (editors.length > 1) {
                        setCollabHint('实时通道正在重连，在线成员仍会通过轮询同步。');
                    } else {
                        setCollabHint('当前没有其他在线协作者。');
                    }
                }

                function applyCollaborationState() {
                    const forceReadonly = !state.editable || !state.canEdit;
                    editor.readOnly = forceReadonly;
                    editor.classList.remove('locked');
                    editor.removeAttribute('title');
                    if (saveBtn) {
                        saveBtn.disabled = forceReadonly || isSaving;
                    }
                }

                function buildTextChangeSpans(baseText, modifiedText) {
                    if (baseText === modifiedText) {
                        return [];
                    }

                    let prefix = 0;
                    while (prefix < baseText.length && prefix < modifiedText.length && baseText[prefix] === modifiedText[prefix]) {
                        prefix += 1;
                    }

                    let baseSuffix = baseText.length;
                    let modifiedSuffix = modifiedText.length;
                    while (
                        baseSuffix > prefix &&
                        modifiedSuffix > prefix &&
                        baseText[baseSuffix - 1] === modifiedText[modifiedSuffix - 1]
                    ) {
                        baseSuffix -= 1;
                        modifiedSuffix -= 1;
                    }

                    return [{
                        start: prefix,
                        end: baseSuffix,
                        replacement: modifiedText.slice(prefix, modifiedSuffix)
                    }];
                }

                function applyTextChangeSpans(baseText, changes) {
                    let result = baseText;
                    let offset = 0;
                    (changes || []).forEach(change => {
                        const start = Number(change.start || 0);
                        const end = Number(change.end ?? start);
                        const replacement = String(change.replacement || '');
                        const actualStart = Math.max(0, start + offset);
                        const actualEnd = Math.max(actualStart, end + offset);
                        result = `${result.slice(0, actualStart)}${replacement}${result.slice(actualEnd)}`;
                        offset += replacement.length - (end - start);
                    });
                    return result;
                }

                function isInsertChange(change) {
                    return Number(change.start || 0) === Number(change.end || 0);
                }

                function changeTouchesCluster(change, clusterStart, clusterEnd) {
                    const start = Number(change.start || 0);
                    const end = Number(change.end ?? start);
                    if (clusterStart === clusterEnd && start === end && start === clusterStart) {
                        return true;
                    }
                    return start < clusterEnd && end > clusterStart;
                }

                function mergeTextVersions(baseText, oursText, theirsText) {
                    const oursChanges = buildTextChangeSpans(baseText, oursText);
                    const theirsChanges = buildTextChangeSpans(baseText, theirsText);

                    if (oursChanges.length === 0) {
                        return { content: theirsText, merged: false };
                    }
                    if (theirsChanges.length === 0) {
                        return { content: oursText, merged: false };
                    }

                    const mergedChanges = [];
                    let oursIndex = 0;
                    let theirsIndex = 0;
                    let merged = false;

                    while (oursIndex < oursChanges.length || theirsIndex < theirsChanges.length) {
                        const oursChange = oursChanges[oursIndex] || null;
                        const theirsChange = theirsChanges[theirsIndex] || null;

                        if (!oursChange) {
                            mergedChanges.push(theirsChange);
                            theirsIndex += 1;
                            continue;
                        }
                        if (!theirsChange) {
                            mergedChanges.push(oursChange);
                            oursIndex += 1;
                            continue;
                        }

                        const oursStart = Number(oursChange.start || 0);
                        const oursEnd = Number(oursChange.end ?? oursStart);
                        const theirsStart = Number(theirsChange.start || 0);
                        const theirsEnd = Number(theirsChange.end ?? theirsStart);
                        const samePointInsert = isInsertChange(oursChange) && isInsertChange(theirsChange) && oursStart === theirsStart;

                        if (oursEnd < theirsStart || (oursEnd === theirsStart && !samePointInsert)) {
                            mergedChanges.push(oursChange);
                            oursIndex += 1;
                            continue;
                        }

                        if (theirsEnd < oursStart || (theirsEnd === oursStart && !samePointInsert)) {
                            mergedChanges.push(theirsChange);
                            theirsIndex += 1;
                            continue;
                        }

                        let clusterStart = Math.min(oursStart, theirsStart);
                        let clusterEnd = Math.max(oursEnd, theirsEnd);
                        const clusterOurs = [oursChange];
                        const clusterTheirs = [theirsChange];
                        oursIndex += 1;
                        theirsIndex += 1;

                        while (oursIndex < oursChanges.length) {
                            const nextChange = oursChanges[oursIndex];
                            if (!changeTouchesCluster(nextChange, clusterStart, clusterEnd)) {
                                break;
                            }
                            clusterOurs.push(nextChange);
                            clusterStart = Math.min(clusterStart, Number(nextChange.start || 0));
                            clusterEnd = Math.max(clusterEnd, Number(nextChange.end ?? nextChange.start ?? 0));
                            oursIndex += 1;
                        }

                        while (theirsIndex < theirsChanges.length) {
                            const nextChange = theirsChanges[theirsIndex];
                            if (!changeTouchesCluster(nextChange, clusterStart, clusterEnd)) {
                                break;
                            }
                            clusterTheirs.push(nextChange);
                            clusterStart = Math.min(clusterStart, Number(nextChange.start || 0));
                            clusterEnd = Math.max(clusterEnd, Number(nextChange.end ?? nextChange.start ?? 0));
                            theirsIndex += 1;
                        }

                        const identicalCluster =
                            clusterOurs.length === clusterTheirs.length &&
                            clusterOurs.every((change, index) => {
                                const otherChange = clusterTheirs[index];
                                return (
                                    Number(change.start || 0) === Number(otherChange.start || 0) &&
                                    Number(change.end ?? change.start ?? 0) === Number(otherChange.end ?? otherChange.start ?? 0) &&
                                    String(change.replacement || '') === String(otherChange.replacement || '')
                                );
                            });

                        if (identicalCluster) {
                            clusterOurs.forEach(change => {
                                mergedChanges.push(change);
                            });
                            continue;
                        }

                        const samePositionInserts =
                            clusterOurs.every(isInsertChange) &&
                            clusterTheirs.every(isInsertChange) &&
                            new Set(clusterOurs.concat(clusterTheirs).map(change => Number(change.start || 0))).size === 1;

                        if (samePositionInserts) {
                            let replacement = '';
                            clusterTheirs.forEach(change => { replacement += String(change.replacement || ''); });
                            clusterOurs.forEach(change => { replacement += String(change.replacement || ''); });
                            mergedChanges.push({
                                start: clusterStart,
                                end: clusterEnd,
                                replacement
                            });
                            merged = true;
                            continue;
                        }

                        if (
                            clusterOurs.length === 1 &&
                            clusterTheirs.length === 1 &&
                            clusterOurs[0].start === clusterTheirs[0].start &&
                            clusterOurs[0].end === clusterTheirs[0].end &&
                            clusterOurs[0].replacement === clusterTheirs[0].replacement
                        ) {
                            mergedChanges.push(clusterOurs[0]);
                            continue;
                        }

                        const clusterBase = baseText.slice(clusterStart, clusterEnd);
                        const normalizedOurs = clusterOurs.map(change => ({
                            start: Number(change.start || 0) - clusterStart,
                            end: Number(change.end ?? change.start ?? 0) - clusterStart,
                            replacement: String(change.replacement || '')
                        }));
                        mergedChanges.push({
                            start: clusterStart,
                            end: clusterEnd,
                            replacement: applyTextChangeSpans(clusterBase, normalizedOurs)
                        });
                        merged = true;
                    }

                    return {
                        content: applyTextChangeSpans(baseText, mergedChanges),
                        merged
                    };
                }

                function applyRealtimePayload(payload, source = 'remote') {
                    const nextSyncedContent = String(payload.content || '');
                    const payloadRevision = Number(payload.revision ?? state.realtimeRevision ?? 0);
                    const currentRevision = Number(state.realtimeRevision || 0);
                    if (source !== 'local') {
                        if (payloadRevision < currentRevision) {
                            return;
                        }
                        if (payloadRevision === currentRevision && nextSyncedContent === state.syncedContent) {
                            state.mtimeNs = payload.mtime_ns || state.mtimeNs;
                            state.encoding = payload.encoding || state.encoding;
                            state.newline = payload.newline || state.newline;
                            encodingText.textContent = state.encoding;
                            return;
                        }
                    }
                    const previousSyncedContent = state.syncedContent;
                    const currentValue = editor.value;
                    let nextEditorValue = nextSyncedContent;
                    let merged = false;

                    if (source === 'local') {
                        state.syncedContent = nextSyncedContent;
                        state.realtimeRevision = payloadRevision;
                        state.mtimeNs = payload.mtime_ns || state.mtimeNs;
                        state.encoding = payload.encoding || state.encoding;
                        state.newline = payload.newline || state.newline;
                        encodingText.textContent = state.encoding;

                        if (currentValue === previousSyncedContent && currentValue !== nextSyncedContent) {
                            replaceEditorValue(nextSyncedContent, true);
                        } else if (
                            payload.merged &&
                            currentValue !== nextSyncedContent &&
                            previousSyncedContent !== nextSyncedContent
                        ) {
                            const mergeResult = mergeTextVersions(previousSyncedContent, currentValue, nextSyncedContent);
                            nextEditorValue = mergeResult.content;
                            merged = mergeResult.merged || nextEditorValue !== nextSyncedContent;
                            if (nextEditorValue !== currentValue) {
                                replaceEditorValue(nextEditorValue, true);
                            }
                        }

                        hasPendingChanges = editor.value !== state.syncedContent;
                        applyCollaborationState();
                        if (payload.merged && merged && hasPendingChanges) {
                            setStatus('已将你的最新修改与其他人的实时更新自动合并。', 'success');
                        }
                        if (hasPendingChanges) {
                            scheduleRealtimeUpdate(80);
                        }
                        return;
                    }

                    if (currentValue !== previousSyncedContent) {
                        const mergeResult = mergeTextVersions(previousSyncedContent, currentValue, nextSyncedContent);
                        nextEditorValue = mergeResult.content;
                        merged = mergeResult.merged || nextEditorValue !== nextSyncedContent;
                    }

                    replaceEditorValue(nextEditorValue, true);
                    state.syncedContent = nextSyncedContent;
                    state.realtimeRevision = payloadRevision;
                    state.mtimeNs = payload.mtime_ns || state.mtimeNs;
                    state.encoding = payload.encoding || state.encoding;
                    state.newline = payload.newline || state.newline;
                    encodingText.textContent = state.encoding;
                    hasPendingChanges = editor.value !== state.syncedContent;
                    applyCollaborationState();

                    if (source === 'remote') {
                        if (merged && hasPendingChanges) {
                            setStatus(`已收到 ${payload.author || '其他协作者'} 的实时更新，并和你当前输入自动合并。`, 'info');
                        } else {
                            setStatus(`已收到 ${payload.author || '其他协作者'} 的实时更新。`, 'info');
                        }
                    } else if (source === 'snapshot' && merged) {
                        setStatus('已将服务器最新内容合并到当前文本。', 'info');
                    }

                    if (hasPendingChanges) {
                        scheduleRealtimeUpdate(80);
                    }
                }

                async function refreshTextSnapshot() {
                    if (!state.filePath || state.snapshotInFlight) {
                        return;
                    }

                    state.snapshotInFlight = true;
                    try {
                        const response = await fetch('/api/text_file_snapshot', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ file_path: state.filePath })
                        });
                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            return;
                        }
                        applyRealtimePayload(data, 'snapshot');
                    } catch (error) {
                    } finally {
                        state.snapshotInFlight = false;
                    }
                }

                function scheduleRealtimeUpdate(delay = state.realtimeDebounce) {
                    if (!state.editable || !state.canEdit) {
                        return;
                    }
                    if (state.realtimeSendTimer) {
                        window.clearTimeout(state.realtimeSendTimer);
                    }
                    state.realtimeSendTimer = window.setTimeout(function() {
                        flushRealtimeUpdate();
                    }, delay);
                }

                function scheduleCollaborationSync(delay = 160) {
                    if (!state.filePath) {
                        return;
                    }
                    if (state.collabSyncTimer) {
                        window.clearTimeout(state.collabSyncTimer);
                    }
                    state.collabSyncTimer = window.setTimeout(function() {
                        state.collabSyncTimer = null;
                        syncCollaborationState({ quiet: true });
                    }, delay);
                }

                async function flushRealtimeUpdate(force = false) {
                    if (!state.editable || !state.canEdit) {
                        return true;
                    }

                    if (state.realtimeSendTimer) {
                        window.clearTimeout(state.realtimeSendTimer);
                        state.realtimeSendTimer = null;
                    }

                    if (editor.value === state.syncedContent && !force) {
                        return true;
                    }

                    if (state.realtimeSendInFlight) {
                        state.realtimeResendAfterFlight = true;
                        return state.realtimeSendPromise || true;
                    }

                    const contentToSend = editor.value;
                    const baseRevision = Number(state.realtimeRevision || 0);
                    state.realtimeSendInFlight = true;
                    state.realtimeSendPromise = (async function() {
                        let shouldRetry = false;
                        try {
                            const response = await fetch('/api/text_realtime_update', {
                                method: 'POST',
                                headers: { 'Content-Type': 'application/json' },
                                body: JSON.stringify({
                                    file_path: state.filePath,
                                    content: contentToSend,
                                    base_revision: baseRevision
                                })
                            });
                            const data = await response.json();
                            if (!response.ok || !data.success) {
                                if (data.content !== undefined) {
                                    applyRealtimePayload(data, 'snapshot');
                                    shouldRetry = editor.value !== state.syncedContent;
                                }
                                setStatus(data.message || '实时同步失败，请稍后重试。', 'error');
                                return false;
                            }

                            applyRealtimePayload(data, 'local');
                            shouldRetry = editor.value !== state.syncedContent;
                            if (data.merged) {
                                setStatus('已将你的最新修改与其他人的实时更新自动合并。', 'success');
                            }
                            return true;
                        } catch (error) {
                            setStatus('实时同步失败：网络异常或服务器不可用。', 'error');
                            return false;
                        } finally {
                            state.realtimeSendInFlight = false;
                            state.realtimeSendPromise = null;
                            shouldRetry = shouldRetry || state.realtimeResendAfterFlight;
                            state.realtimeResendAfterFlight = false;
                            if (shouldRetry) {
                                scheduleRealtimeUpdate(80);
                            }
                        }
                    })();

                    return state.realtimeSendPromise;
                }

                function openRealtimeStream() {
                    if (document.hidden || !state.filePath || state.realtimeEventSource || !window.EventSource) {
                        return;
                    }

                    const streamUrl = `/api/document_realtime_stream?file_path=${encodeURIComponent(state.filePath)}&editor_type=text`;
                    const eventSource = new EventSource(streamUrl);
                    state.realtimeEventSource = eventSource;

                    eventSource.onopen = function() {
                        state.realtimeConnected = true;
                        renderCollaborationUsers();
                    };

                    eventSource.addEventListener('text_init', function(event) {
                        try {
                            const payload = JSON.parse(event.data);
                            applyRealtimePayload(payload, 'snapshot');
                        } catch (error) {
                        }
                    });

                    eventSource.addEventListener('text_update', function(event) {
                        try {
                            const payload = JSON.parse(event.data);
                            applyRealtimePayload(payload, 'remote');
                        } catch (error) {
                        }
                    });

                    eventSource.onerror = function() {
                        state.realtimeConnected = false;
                        renderCollaborationUsers();
                    };
                }

                function closeRealtimeStream() {
                    if (state.realtimeEventSource) {
                        state.realtimeEventSource.close();
                        state.realtimeEventSource = null;
                    }
                    state.realtimeConnected = false;
                    renderCollaborationUsers();
                }

                async function syncCollaborationState(options = {}) {
                    const { quiet = true } = options;
                    if (!state.filePath || state.collabSyncInFlight) {
                        return;
                    }

                    state.collabSyncInFlight = true;
                    try {
                        const response = await fetch('/api/document_collaboration_sync', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                editor_type: 'text',
                                active_target: buildActiveTarget(),
                                lock_target: null,
                                release_lock: true
                            })
                        });
                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (!quiet) {
                                setStatus(data.message || '协同状态同步失败。', 'error');
                            }
                            return;
                        }

                        if (data.client_id) {
                            state.collabClientId = data.client_id;
                        }
                        state.collabEditors = Array.isArray(data.editors) ? data.editors : [];
                        renderCollaborationUsers();
                        applyCollaborationState();

                        if (data.mtime_ns && data.mtime_ns !== state.mtimeNs && !state.realtimeConnected) {
                            refreshTextSnapshot();
                        }
                    } catch (error) {
                        if (!quiet) {
                            setStatus('协同状态同步失败：网络异常或服务器不可用。', 'error');
                        }
                    } finally {
                        state.collabSyncInFlight = false;
                    }
                }

                function releaseCollaborationState(removeSession = false) {
                    if (!state.filePath) {
                        return;
                    }

                    if (state.collabSyncTimer) {
                        window.clearTimeout(state.collabSyncTimer);
                        state.collabSyncTimer = null;
                    }
                    closeRealtimeStream();

                    const payload = JSON.stringify({
                        file_path: state.filePath,
                        remove_session: removeSession
                    });
                    if (navigator.sendBeacon) {
                        const blob = new Blob([payload], { type: 'application/json' });
                        navigator.sendBeacon('/api/document_collaboration_release', blob);
                        return;
                    }

                    fetch('/api/document_collaboration_release', {
                        method: 'POST',
                        headers: { 'Content-Type': 'application/json' },
                        body: payload,
                        keepalive: true
                    }).catch(() => {});
                }

                function returnToList(event) {
                    if (!backToListBtn) {
                        return;
                    }

                    const backUrl = backToListBtn.href;
                    if (!window.opener || window.opener.closed) {
                        return;
                    }

                    event.preventDefault();
                    suppressBeforeUnload = true;
                    try {
                        window.opener.location.href = backUrl;
                        if (typeof window.opener.focus === 'function') {
                            window.opener.focus();
                        }
                    } catch (error) {
                    }

                    releaseCollaborationState(true);
                    window.close();
                    window.setTimeout(function() {
                        window.location.href = backUrl;
                    }, 120);
                }

                async function saveFile() {
                    if (isSaving) {
                        return;
                    }
                    if (!state.editable) {
                        setStatus('当前文件过大，只支持只读预览。', 'error');
                        return;
                    }
                    if (!state.canEdit) {
                        setStatus('只有管理员才能保存修改。', 'error');
                        return;
                    }

                    if (state.realtimeSendTimer) {
                        window.clearTimeout(state.realtimeSendTimer);
                        state.realtimeSendTimer = null;
                    }
                    if (state.realtimeSendInFlight && state.realtimeSendPromise) {
                        await state.realtimeSendPromise;
                    }
                    isSaving = true;
                    if (saveBtn) {
                        saveBtn.disabled = true;
                    }
                    setStatus('正在保存到共享目录...', 'info');

                    try {
                        const response = await fetch('/api/save_text_file', {
                            method: 'POST',
                            headers: {
                                'Content-Type': 'application/json'
                            },
                            body: JSON.stringify({
                                file_path: state.filePath,
                                content: editor.value,
                                encoding: state.encoding,
                                newline: state.newline,
                                mtime_ns: state.mtimeNs,
                                realtime_revision: state.realtimeRevision
                            })
                        });

                        const data = await response.json();
                        if (!response.ok || !data.success) {
                            if (data.content !== undefined) {
                                applyRealtimePayload(data, 'snapshot');
                            }
                            setStatus(data.message || '保存失败，请稍后重试。', 'error');
                            return;
                        }

                        state.mtimeNs = data.mtime_ns;
                        state.encoding = data.encoding || state.encoding;
                        encodingText.textContent = state.encoding;
                        if (data.realtime_revision !== undefined) {
                            state.realtimeRevision = Number(data.realtime_revision || state.realtimeRevision);
                        }
                        if (data.content !== undefined) {
                            applyRealtimePayload({
                                content: String(data.content),
                                revision: state.realtimeRevision,
                                encoding: state.encoding,
                                newline: state.newline,
                                mtime_ns: state.mtimeNs
                            }, 'local');
                        } else {
                            state.syncedContent = editor.value;
                        }
                        state.syncedContent = editor.value;
                        hasPendingChanges = false;
                        const saveMessage = data.merged
                            ? `已保存到共享目录 (${data.size || '未知大小'})，并合并了其他人的最新修改。`
                            : `已保存到共享目录 (${data.size || '未知大小'})`;
                        setStatus(saveMessage, 'success');
                    } catch (error) {
                        setStatus('保存失败：网络异常或服务器不可用。', 'error');
                    } finally {
                        isSaving = false;
                        applyCollaborationState();
                    }
                }

                if (state.editable && state.canEdit) {
                    editor.addEventListener('input', function() {
                        hasPendingChanges = editor.value !== state.syncedContent;
                        if (hasPendingChanges) {
                            setStatus('内容已修改，只会实时同步给在线协作者；按 Ctrl+S 才会保存到文件。', 'info');
                        }
                        scheduleRealtimeUpdate();
                        scheduleCollaborationSync(220);
                    });
                }

                editor.addEventListener('focus', function() {
                    if (document.activeElement === editor) {
                        syncCollaborationState({ quiet: true });
                    }
                });

                ['keyup', 'mouseup'].forEach(eventName => {
                    editor.addEventListener(eventName, function() {
                        if (document.activeElement === editor) {
                            scheduleCollaborationSync(140);
                        }
                    });
                });

                if (saveBtn) {
                    saveBtn.addEventListener('click', saveFile);
                }
                if (backToListBtn) {
                    backToListBtn.addEventListener('click', returnToList);
                }
                document.addEventListener('keydown', function(event) {
                    if ((event.ctrlKey || event.metaKey) && event.key.toLowerCase() === 's') {
                        event.preventDefault();
                        saveFile();
                    }
                });

                window.addEventListener('beforeunload', function(event) {
                    if (!suppressBeforeUnload && hasPendingChanges) {
                        event.preventDefault();
                        event.returnValue = '';
                    }
                });

                window.addEventListener('pagehide', function() {
                    releaseCollaborationState(true);
                });

                document.addEventListener('visibilitychange', function() {
                    if (document.hidden) {
                        closeRealtimeStream();
                    } else {
                        openRealtimeStream();
                        syncCollaborationState({ quiet: true });
                    }
                });

                applyCollaborationState();
                openRealtimeStream();
                if (state.filePath) {
                    syncCollaborationState({ quiet: true });
                    remoteCheckTimer = window.setInterval(function() {
                        syncCollaborationState({ quiet: true });
                    }, state.collabPollInterval);
                }
            </script>
        </body>
        </html>
        ''',
        filename=os.path.basename(filename),
        file_path=filename,
        back_url=back_url,
        content=preview['content'],
        encoding=preview['encoding'],
        newline=preview['newline'],
        editable=preview['editable'],
        can_edit=can_edit,
        mtime_ns=get_file_mtime_token(filepath),
        file_size=get_file_size(preview['file_size']),
        collab_client_id=get_document_editor_client_id(),
        collab_username=get_current_username(),
        collab_poll_interval_ms=DOCUMENT_COLLAB_POLL_INTERVAL_MS,
        text_realtime_debounce_ms=TEXT_REALTIME_EDIT_DEBOUNCE_MS,
        warning_message=warning_message,
        initial_status=initial_status
        )
    except Exception as e:
        print(f"打开在线编辑器失败: {str(e)}")
        flash(f'打开编辑器失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

@app.route('/preview/<path:filename>')
def preview_file(filename):
    """文件预览"""
    try:
        actual_target = resolve_macos_metadata_target(filename)
        if actual_target:
            return redirect(url_for('preview_file', filename=actual_target))
        if is_macos_metadata_file(filename):
            flash('这个以 ._ 开头的文件是 macOS 生成的元数据文件，不是真正的可预览文件。请打开同目录下不带 ._ 前缀的原文件。', 'warning')
            parent_path = os.path.dirname(filename).replace('\\', '/')
            return redirect(url_for('index', subpath=parent_path) if parent_path else url_for('index'))

        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        if not os.path.exists(filepath) or not os.path.isfile(filepath):
            flash('文件不存在', 'danger')
            return redirect(url_for('index'))
        
        parent_path = os.path.dirname(filename).replace('\\', '/')
        back_url = url_for('index', subpath=parent_path) if parent_path else url_for('index')

        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''
        
        # 图片预览
        if ext in ['jpg', 'jpeg', 'png', 'gif', 'bmp', 'webp', 'svg', 'ico']:
            return send_file(filepath, mimetype=f'image/{ext}')
        
        # 文本文件预览
        elif ext in ['txt', 'py', 'js', 'html', 'css', 'json', 'xml', 'md', 'csv', 'log', 'sql', 'sh', 'bat', 'java', 'cpp', 'c', 'h', 'ts', 'yaml', 'yml']:
            try:
                with open(filepath, 'r', encoding='utf-8') as f:
                    content = f.read(1024 * 100)  # 最多读取 100KB
                
                return render_template_string('''
                <!DOCTYPE html>
                <html>
                <head>
                    <title>预览: {{ filename }}</title>
                    <style>
                        body {
                            margin: 0;
                            padding: 20px;
                            font-family: 'Consolas', 'Monaco', monospace;
                            background: #1e1e1e;
                            color: #d4d4d4;
                        }
                        .header {
                            background: #2d2d30;
                            padding: 15px;
                            border-radius: 8px;
                            margin-bottom: 20px;
                            display: flex;
                            justify-content: space-between;
                            align-items: center;
                        }
                        h2 { margin: 0; color: #fff; }
                        .btn {
                            padding: 8px 16px;
                            background: #0e639c;
                            color: white;
                            text-decoration: none;
                            border-radius: 4px;
                            font-size: 14px;
                        }
                        .btn:hover { background: #1177bb; }
                        pre {
                            background: #2d2d30;
                            padding: 20px;
                            border-radius: 8px;
                            overflow-x: auto;
                            white-space: pre-wrap;
                            word-wrap: break-word;
                            line-height: 1.5;
                        }
                    </style>
                </head>
                <body>
                    <div class="header">
                        <h2>📄 {{ filename }}</h2>
                        <a href="{{ back_url }}" id="backToListBtn" class="btn">返回</a>
                    </div>
                    <pre>{{ content }}</pre>
                    <script>
                        const backToListBtn = document.getElementById('backToListBtn');
                        function returnToList(event) {
                            if (!backToListBtn) {
                                return;
                            }

                            const backUrl = backToListBtn.href;
                            if (!window.opener || window.opener.closed) {
                                return;
                            }

                            event.preventDefault();
                            try {
                                window.opener.location.href = backUrl;
                                if (typeof window.opener.focus === 'function') {
                                    window.opener.focus();
                                }
                            } catch (error) {
                            }

                            window.close();
                            window.setTimeout(function() {
                                window.location.href = backUrl;
                            }, 120);
                        }

                        if (backToListBtn) {
                            backToListBtn.addEventListener('click', returnToList);
                        }
                    </script>
                </body>
                </html>
                ''', filename=os.path.basename(filename), content=content, back_url=back_url)
            except:
                flash('无法预览此文件', 'danger')
                return redirect(url_for('index'))

        # DOCX 在线编辑/预览
        elif is_word_editable_file(filename):
            return redirect(url_for('edit_docx_file', filename=filename))
        
        # PDF预览
        elif ext == 'pdf':
            return send_file(filepath, mimetype='application/pdf')
        
        # 视频预览
        elif ext in ['mp4', 'webm', 'ogg']:
            return render_template_string('''
            <!DOCTYPE html>
            <html>
            <head>
                <title>预览: {{ filename }}</title>
                <style>
                    body {
                        margin: 0;
                        padding: 20px;
                        background: #000;
                        display: flex;
                        flex-direction: column;
                        align-items: center;
                        min-height: 100vh;
                    }
                    .header {
                        width: 100%;
                        max-width: 1200px;
                        background: #1e1e1e;
                        padding: 15px;
                        border-radius: 8px;
                        margin-bottom: 20px;
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                    }
                    h2 { margin: 0; color: #fff; }
                    .btn {
                        padding: 8px 16px;
                        background: #0e639c;
                        color: white;
                        text-decoration: none;
                        border-radius: 4px;
                    }
                    video {
                        max-width: 100%;
                        max-height: 80vh;
                        border-radius: 8px;
                    }
                </style>
            </head>
            <body>
                <div class="header">
                    <h2>🎬 {{ filename }}</h2>
                    <a href="{{ back_url }}" id="backToListBtn" class="btn">返回</a>
                </div>
                <video controls autoplay>
                    <source src="{{ url_for('stream_file', filename=filepath) }}" type="{{ media_mimetype }}">
                    您的浏览器不支持视频播放
                </video>
                <script>
                    const backToListBtn = document.getElementById('backToListBtn');
                    function returnToList(event) {
                        if (!backToListBtn) {
                            return;
                        }

                        const backUrl = backToListBtn.href;
                        if (!window.opener || window.opener.closed) {
                            return;
                        }

                        event.preventDefault();
                        try {
                            window.opener.location.href = backUrl;
                            if (typeof window.opener.focus === 'function') {
                                window.opener.focus();
                            }
                        } catch (error) {
                        }

                        window.close();
                        window.setTimeout(function() {
                            window.location.href = backUrl;
                        }, 120);
                    }

                    if (backToListBtn) {
                        backToListBtn.addEventListener('click', returnToList);
                    }
                </script>
            </body>
            </html>
            ''', filename=os.path.basename(filename), filepath=filename, ext=ext, back_url=back_url, media_mimetype=guess_inline_mimetype(filename))
        
        # 音频预览
        elif ext in ['mp3', 'wav', 'ogg', 'm4a', 'flac', 'aac']:
            return render_template_string('''
            <!DOCTYPE html>
            <html>
            <head>
                <title>预览: {{ filename }}</title>
                <style>
                    body {
                        margin: 0;
                        padding: 20px;
                        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                        display: flex;
                        justify-content: center;
                        align-items: center;
                        min-height: 100vh;
                        font-family: Arial, sans-serif;
                    }
                    .player {
                        background: white;
                        padding: 40px;
                        border-radius: 15px;
                        box-shadow: 0 10px 40px rgba(0,0,0,0.3);
                        text-align: center;
                        max-width: 500px;
                    }
                    h2 { color: #333; margin-bottom: 30px; }
                    audio {
                        width: 100%;
                        margin-bottom: 20px;
                    }
                    .btn {
                        padding: 10px 20px;
                        background: #667eea;
                        color: white;
                        text-decoration: none;
                        border-radius: 8px;
                        display: inline-block;
                    }
                </style>
            </head>
            <body>
                <div class="player">
                    <h2>🎵 {{ filename }}</h2>
                    <audio controls autoplay>
                        <source src="{{ url_for('stream_file', filename=filepath) }}" type="{{ media_mimetype }}">
                        您的浏览器不支持音频播放
                    </audio>
                    <a href="{{ back_url }}" id="backToListBtn" class="btn">返回</a>
                </div>
                <script>
                    const backToListBtn = document.getElementById('backToListBtn');
                    function returnToList(event) {
                        if (!backToListBtn) {
                            return;
                        }

                        const backUrl = backToListBtn.href;
                        if (!window.opener || window.opener.closed) {
                            return;
                        }

                        event.preventDefault();
                        try {
                            window.opener.location.href = backUrl;
                            if (typeof window.opener.focus === 'function') {
                                window.opener.focus();
                            }
                        } catch (error) {
                        }

                        window.close();
                        window.setTimeout(function() {
                            window.location.href = backUrl;
                        }, 120);
                    }

                    if (backToListBtn) {
                        backToListBtn.addEventListener('click', returnToList);
                    }
                </script>
            </body>
            </html>
            ''', filename=os.path.basename(filename), filepath=filename, ext=ext, back_url=back_url, media_mimetype=guess_inline_mimetype(filename))
        
        else:
            flash('不支持预览此文件类型', 'warning')
            return redirect(url_for('index'))
    
    except Exception as e:
        print(f"预览文件失败: {str(e)}")
        flash(f'预览失败: {str(e)}', 'danger')
        return redirect(url_for('index'))

def main():
    """Docstring."""
    local_ip = get_local_ip()
    port = SERVER_PORT  # 使用配置文件中的端口
    
    print("=" * 60)
    print("     局域网文件共享服务器已启动")
    print("=" * 60)
    print(f"\n本机访问地址: http://127.0.0.1:{port}")
    if ALLOW_LAN:
        print(f"局域网访问地址: http://{local_ip}:{port}")
    else:
        print("仅允许本机访问（局域网访问已禁用）")
    print(f"\n共享文件夹: {os.path.abspath(UPLOAD_FOLDER)}")
    print("最大文件大小: 无限制（已绕过 Werkzeug 默认限制）")
    
    try:
        import werkzeug
        # 尝试获取版本号，不同版本的属性名可能不同
        werkzeug_version = getattr(werkzeug, '__version__', 
                                   getattr(werkzeug, 'version', 
                                          getattr(werkzeug, '__version_info__', '未知')))
        print(f"Werkzeug 版本: {werkzeug_version}")
    except Exception as e:
        print("Werkzeug 版本: 无法获取")
    
    print("\n使用说明:")
    print("   1. 请确保本机和其他设备在同一局域网内")
    print("   2. 其他设备可通过局域网地址访问此页面")
    print("   3. 支持上传、下载和删除文件")
    print("   4. 支持大文件上传")
    print("   5. 按 Ctrl+C 停止服务器")
    print("\n提示：大文件上传可能需要较长时间，请耐心等待")
    print("\n" + "=" * 60)
    
    try:
        from waitress import serve
        print("\n" + "=" * 60)
        print("使用 Waitress 服务器（支持大文件上传）")
        print("=" * 60)
        print("\n服务器配置详情：")
        print("   - 最大请求体大小: 20GB")
        print("   - 最大请求头大小: 5MB")
        print("   - 输入缓冲区上限: 2MB")
        print("   - 工作线程数: 24")
        print("   - 通道超时: 1800 秒（30 分钟）")
        print("   - 接收缓冲区: 256KB")
        print("   - 发送缓冲区: 1MB")
        print("   - 输出缓冲区上限: 8MB")
        print("\n如需停止服务器，请按 Ctrl+C")
        print("=" * 60 + "\n")

        try:
            import pkg_resources
            waitress_version = pkg_resources.get_distribution('waitress').version
            print(f"Waitress 版本: {waitress_version}\n")
        except Exception:
            print("Waitress 版本: 已安装（版本未知）\n")

        # 协同编辑页会产生长连接与轮询，这类排队提醒很常见，但并不代表请求失败。
        # 将其降级，避免控制台被同类提示刷屏。
        logging.getLogger('waitress.queue').setLevel(logging.ERROR)

        host = '0.0.0.0' if ALLOW_LAN else '127.0.0.1'
        serve(
            app,
            host=host,
            port=port,
            threads=24,
            channel_timeout=1800,
            max_request_body_size=21474836480,
            max_request_header_size=5242880,
            recv_bytes=262144,
            send_bytes=1048576,
            inbuf_overflow=2097152,
            outbuf_overflow=8388608
        )
    except ImportError:
        print("\n未安装 Waitress，改用 Flask 开发服务器")
        print("建议安装 Waitress 以支持更稳定的大文件传输：pip install waitress")
        print("如需停止服务器，请按 Ctrl+C\n")
        host = '0.0.0.0' if ALLOW_LAN else '127.0.0.1'
        app.run(host=host, port=port, debug=False, threaded=True)

if __name__ == '__main__':
    main()
